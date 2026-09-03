"""The marks a session has made, written to a file and put back.

The workbook is the book of record and stays that way (§2): everything the
marking and market-maker screens do -- a re-marked backbone, a correlation
term structure, an event schedule, a tenor overwrite, a smile parameter
overwrite, a wing shift out of a broker run, a band treatment -- lives on the
loaded book in memory, and a reload discards it.  That is the right default
for a tool whose primary file is somebody else's spreadsheet, and it is the
wrong thing to do to a morning's work at 5pm.

So a session can be *saved beside* the workbook rather than into it.  The file
is the tool's own, like the knowledge bank: JSON, atomically written, and
readable by the person whose marks are in it.

Three things are decided here once.

**The file holds what the screen shows.**  Volatility numbers are in
volatility points, the way they are typed and read on the marking screen; the
shape parameters, decays and correlations are the raw numbers those fields
carry.  ``curve_params`` and ``set_curve_params`` are the one conversion, and
the marking screen uses the same two functions, so the file cannot drift out
of step with the panel that wrote it (§4, "volatility points at the edges").

**Loading replaces, and says what it replaced.**  Overwrites and events are
cleared before the saved ones go on, because a merge would leave a tenor
overwritten by a session nobody remembers.  A pair in the file that this
workbook does not build is reported, not skipped quietly; a pair in the book
that the file does not mention is left exactly as it is, and that is reported
too.

**A pair that will not take its marks does not take the rest down with it.**
Each pair is applied inside its own guard and each failure is collected with
the pair's name on it, in the same spirit as every other table in this
project: the row keeps its place and carries its reason.

Pairs are applied in the book's own build order, so a cross gets its legs'
marks before it recalibrates against them.
"""

from __future__ import annotations

import json
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

from .banded import BandTreatment
from .cross import is_cross
from .cross import CrossAtmCurve
from .paths import WRITE_ENCODING, app_dir, read_text
from .surface import (PARAM_NAMES, QUOTE_FIELDS, QUOTE_LABELS, RATIO_WINGS,
                      TERM_COEFFS, WING_RATIOS_SHEET, check_ratio,
                      load_wing_ratios)
from .events import event_entries
from .timeutil import UTC, as_utc, parse_datetime

#: Bumped when the shape changes in a way an older build cannot read.
SESSION_VERSION = 1

#: The default name, beside the workbook -- a session is the user's data.
SESSION_FILENAME = "vol_session.json"

#: What every block of a pair may carry.  Listed once so the reader, the
#: writer and the tests all agree on the vocabulary.
PAIR_BLOCKS = ("curve", "events", "atm_overwrites", "smile_overwrites",
               "smile_term", "param_shifts", "anchor_tenors", "band",
               "quote_overwrites", "wing_ratios")

#: The suffix a workbook copy gets when a session is exported without a
#: named destination: ``vol_marks.xlsx`` -> ``vol_marks_marked.xlsx``.
EXPORT_SUFFIX = "_marked"

#: What an in-place export leaves behind before it writes over the book of
#: record: ``vol_marks.xlsx`` -> ``vol_marks.bak-20260901-142530.xlsx``, the
#: bytes exactly as they were read.  openpyxl does not carry images, charts
#: or pivots through a round trip, and an export cannot be undone from the
#: session file alone, so the file it replaced is kept beside it.
BACKUP_INFIX = ".bak-"

#: How many of those backups to keep.  It used to be all of them, which was
#: right while an export was a thing a desk did occasionally and the workbook
#: was the book of record.  As the *store*, a save is a save: a full copy of
#: the workbook per write, kept forever, fills a folder with files nobody
#: reads and buries the one somebody wants.  The most recent are the ones an
#: undo reaches for, so those are what is kept.
BACKUP_KEEP = 20


#: What a pair's PARAMS column starts at when the screen creates it: the
#: shape of a curve, not a mark.  A new pair has to have *something* in these
#: cells -- the reader refuses a column that is missing them -- and inventing
#: a level would be a mark nobody made, so it starts at the one number the
#: person adding the pair is asked for and the rest is the model's own
#: neutral shape: flat, no short-dated add-on, no rate effects.
NEW_PAIR_CURVE = {"initial": None, "long term": None, "MR": 5.0, "addon": 0.0,
                  "short decay": 50.0, "ratevol": 0.0, "rate corr": 0.0}


def add_pair(workbook, pair: str, *, atm: float, quotes: dict | None = None,
             expect: str = "", force: bool = False) -> dict:
    """Put a new pair into the workbook: CONFIG, PARAMS and its quote sheet.

    The three places a pair exists, written together, because a pair that is
    in one of them and not the others is exactly the half-built state the
    reader spends its error messages on -- ``CONFIG lists 'X' but the workbook
    has no 'X' sheet``.

    ``atm`` is the level its curve starts flat at -- volatility points for a
    dollar pair, a **correlation** in [-1, 1] for a cross, which is what those
    same PARAMS cells mean there.  It is the one number that cannot be
    defaulted, because a volatility this tool made up is a mark nobody made.  ``quotes`` is ``{TENOR: {rr_25, st_25, ...}}`` in
    points for the tenors being quoted; a tenor with no entry gets no row, and
    a pair with no rows at all is created and reported as not yet quoted --
    which is a real state now that pairs are made here rather than in Excel.

    A **cross** is not marked on a backbone: its PARAMS column is the
    correlation between its dollar legs, and the same three cells mean
    initial / final / decay there.  It is written the same way and the reader
    tells them apart by the name, as it always has.
    """
    import io
    import openpyxl
    from .cross import dollar_legs, is_cross
    from .marketdata import CONFIG_PAIR_COLUMNS, PARAM_ROWS, SMILE_COLUMNS, _norm

    name = str(pair).strip().upper()
    if len(name) != 6 or not name.isalpha():
        raise SessionError(f"{pair!r} is not a six-letter currency pair")
    level = float(atm)
    cross = is_cross(name)
    # A cross's first three PARAMS cells are its correlation term structure,
    # not a volatility backbone, and the reader does not rescale them.  The
    # two are different numbers in different units and are checked as such.
    if cross and not -1.0 < level < 1.0:
        raise SessionError(f"{name} is a cross, so its curve is the correlation between "
                           f"{' and '.join(dollar_legs(name))}; {atm!r} is not a correlation")
    if not cross and not 0 < level < 200:
        raise SessionError(f"{name}: a starting at-the-money of {atm!r} points is not a "
                           f"volatility; it is the level the curve starts flat at")
    src = Path(workbook)
    if not src.exists():
        raise SessionError(f"workbook not found: {src}")
    now = workbook_stamp(src)
    if expect and now and expect != now and not force:
        raise SessionError(f"{src.name} has changed since it was read; reload the workbook "
                           f"and add the pair to what it says now")
    with src.open("rb") as fh:
        blob = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    cached = _formula_cache(wb, openpyxl.load_workbook(io.BytesIO(blob), data_only=True))
    notes: list[str] = []

    # -- CONFIG ------------------------------------------------------------
    if "CONFIG" not in wb.sheetnames:
        raise SessionError(f"{src.name} has no CONFIG sheet")
    cfg = wb["CONFIG"]
    cols = {}
    for c in range(1, cfg.max_column + 1):
        v = cfg.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            cols[_norm(v)] = c
    pair_col = next((cols[k] for k in CONFIG_PAIR_COLUMNS if k in cols), None)
    if pair_col is None:
        raise SessionError("CONFIG has no PAIRS column listing the pairs to build")
    listed = []
    last = 1
    for r in range(2, cfg.max_row + 1):
        v = cfg.cell(row=r, column=pair_col).value
        text = "" if v is None else str(v).strip().upper()
        if text:
            listed.append(text)
            last = r
    if name in listed:
        raise SessionError(f"CONFIG already lists {name}")
    cfg.cell(row=last + 1, column=pair_col, value=name)
    notes.append(f"CONFIG: {name} added ({'a cross' if cross else 'a dollar pair'})")

    # -- PARAMS ------------------------------------------------------------
    ws = wb["PARAMS"]
    header = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            header[str(v).strip().upper()] = c
    # A pair taken out of CONFIG keeps its column and its sheet, so adding one
    # back finds what it had.  Reused rather than refused, and reused whole:
    # what is there is somebody's curve, not a leftover to be reset.
    reused = name in header
    col = header[name] if reused else ws.max_column + 1
    if not reused:
        ws.cell(row=1, column=col, value=name)
    label_row = {}
    nrow = 1
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label is None or not str(label).strip():
            continue
        nrow = r
        key = PARAM_ROWS.get(_norm(label))
        if key:
            label_row.setdefault(key, r)
    if reused:
        notes.append(f"PARAMS: {name} already had a column and it was left as it is")
    else:
        for label, value in NEW_PAIR_CURVE.items():
            key = PARAM_ROWS[_norm(label)]
            r = label_row.get(key)
            if r is None:
                nrow += 1
                ws.cell(row=nrow, column=1, value=label)
                label_row[key] = r = nrow
            ws.cell(row=r, column=col, value=(level if value is None else value))
        notes.append(f"PARAMS: a column for {name}, flat at {level:g}"
                     + (" correlation" if cross else " volatility point(s)"))

    # -- the quote sheet ---------------------------------------------------
    existing = next((sn for sn in wb.sheetnames if sn.strip().upper() == name), None)
    order = [c for c in SMILE_COLUMNS if c != "expiry"]
    if existing is not None:
        if quotes:
            raise SessionError(
                f"the workbook already has a {existing!r} sheet; the pair was added back "
                f"onto it, so mark its quotes on the screen rather than here")
        sheet = wb[existing]
        r = sheet.max_row + 1
        notes.append(f"{name}: its {existing!r} sheet was already there and was kept")
    else:
        sheet = wb.create_sheet(name)
        sheet.cell(row=1, column=1, value="expiry")
        for i, label in enumerate(order, start=2):
            sheet.cell(row=1, column=i, value=label.upper())
        r = 2
    for tenor, vals in (quotes or {}).items():
        row = {f: vals.get(f) for f in QUOTE_FIELDS}
        if any(v is None for v in row.values()):
            raise SessionError(
                f"{name} {tenor}: a quoted tenor is all four of its numbers or none; "
                f"{', '.join(QUOTE_LABELS[f] for f in QUOTE_FIELDS if row[f] is None)} "
                f"{'is' if sum(v is None for v in row.values()) == 1 else 'are'} blank")
        sheet.cell(row=r, column=1, value=str(tenor).upper())
        for i, label in enumerate(order, start=2):
            sheet.cell(row=r, column=i, value=float(row[SMILE_COLUMNS[label]]))
        r += 1
    quoted = len(quotes or {})
    if existing is None:
        notes.append(f"{name}: a quote sheet with {quoted} tenor(s)" if quoted else
                     f"{name}: a quote sheet with no tenors yet, so it has no smile until "
                     f"one is quoted on the marking screen")

    bak = _backup_path(src)
    with open(bak, "wb") as fh:
        fh.write(blob)
    notes.append(f"the workbook as it was is kept at {bak.name}")
    pruned = prune_backups(src)
    if pruned:
        notes.append(f"{len(pruned)} older backup(s) removed, {BACKUP_KEEP} kept")
    _save_workbook(wb, src, cached)
    return {"written": str(src), "backup": str(bak), "pair": name, "quoted": quoted,
            "notes": notes, "stamp": workbook_stamp(src)}


def remove_pair(workbook, pair: str, *, expect: str = "", force: bool = False) -> dict:
    """Take a pair out of CONFIG, and say what is left behind.

    Only the CONFIG row: the PARAMS column and the quote sheet stay where they
    are.  A pair is dropped from a run for a morning as often as forever, and
    deleting a sheet of somebody's quotes to answer "stop building this" is a
    great deal more than was asked for.  The workbook then holds a column and a
    sheet nothing reads, which is said here so it is a choice and not a
    surprise; both are still there to be deleted in Excel, or the pair added
    back with what it had.
    """
    import io
    import openpyxl
    from .marketdata import CONFIG_PAIR_COLUMNS, _norm

    name = str(pair).strip().upper()
    src = Path(workbook)
    if not src.exists():
        raise SessionError(f"workbook not found: {src}")
    now = workbook_stamp(src)
    if expect and now and expect != now and not force:
        raise SessionError(f"{src.name} has changed since it was read; reload the workbook "
                           f"and make the change on what it says now")
    with src.open("rb") as fh:
        blob = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    cached = _formula_cache(wb, openpyxl.load_workbook(io.BytesIO(blob), data_only=True))
    if "CONFIG" not in wb.sheetnames:
        raise SessionError(f"{src.name} has no CONFIG sheet")
    cfg = wb["CONFIG"]
    cols = {}
    for c in range(1, cfg.max_column + 1):
        v = cfg.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            cols[_norm(v)] = c
    targets = [cols[k] for k in list(CONFIG_PAIR_COLUMNS) + ["cor"] if k in cols]
    hit = 0
    for col in targets:
        for r in range(2, cfg.max_row + 1):
            v = cfg.cell(row=r, column=col).value
            if v is not None and str(v).strip().upper() == name:
                cfg.cell(row=r, column=col).value = None
                hit += 1
    if not hit:
        raise SessionError(f"CONFIG does not list {name}")
    notes = [f"CONFIG: {name} removed"]
    kept = [w for w in ("its PARAMS column",
                        f"its {name} sheet" if any(sn.strip().upper() == name
                                                   for sn in wb.sheetnames) else "")
            if w]
    if kept:
        notes.append(f"{name}: {' and '.join(kept)} left in place, so nothing reads "
                     f"{'them' if len(kept) > 1 else 'it'} now; adding the pair back finds "
                     f"what it had")
    bak = _backup_path(src)
    with open(bak, "wb") as fh:
        fh.write(blob)
    notes.append(f"the workbook as it was is kept at {bak.name}")
    pruned = prune_backups(src)
    if pruned:
        notes.append(f"{len(pruned)} older backup(s) removed, {BACKUP_KEEP} kept")
    _save_workbook(wb, src, cached)
    return {"written": str(src), "backup": str(bak), "pair": name, "notes": notes,
            "stamp": workbook_stamp(src)}


def write_config_tabs(workbook, tabs: dict, *, expect: str = "",
                      force: bool = False) -> dict:
    """Write configuration tabs into the workbook, whole.

    The second thing in this package that writes a workbook, and deliberately
    not part of the session export: a peg band, a kACE pillar, a holiday and a
    wing ratio are not *marks*.  They are what the tool is configured with, a
    session does not carry them, and a desk that changes one means it to apply
    to every session from now on -- so they go straight in, with the same
    backup, the same staleness check and the same pruning the export has.

    ``tabs`` is ``{SHEET: [row dicts]}``, each sheet one of
    ``configsheets.EDITABLE``.  A tab is replaced by what is given for it and
    the ones not named are untouched.
    """
    import io
    import openpyxl
    from . import configsheets

    src = Path(workbook)
    if not src.exists():
        raise SessionError(f"workbook not found: {src}")
    unknown = [k for k in tabs if k not in configsheets.EDITABLE]
    if unknown:
        raise SessionError(
            f"{', '.join(unknown)} is not a configuration tab this writes "
            f"({', '.join(configsheets.EDITABLE)}); the rest of the workbook is edited "
            f"through the marking screens")
    now = workbook_stamp(src)
    if expect and now and expect != now and not force:
        raise SessionError(
            f"{src.name} has changed since it was read -- another volkit, or somebody "
            f"saving it from Excel. Reload the workbook and make the change on what it "
            f"says now")
    with src.open("rb") as fh:
        blob = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    cached = _formula_cache(wb, openpyxl.load_workbook(io.BytesIO(blob), data_only=True))
    # ``columns_for`` rather than the fixed list: ``Vega Weights`` carries a
    # column per pair with its own curve shape, and a writer that knew only
    # the fixed columns would drop every one of them on the first save.
    written = {sheet: configsheets.columns_for(sheet, rows)
               for sheet, rows in sorted(tabs.items())}
    for sheet, columns in written.items():
        configsheets.check_open_columns(sheet, columns)
    notes = [configsheets.write_rows(wb, sheet, written[sheet], rows,
                                     header=configsheets.EDITABLE[sheet][1])
             for sheet, rows in sorted(tabs.items())]
    bak = _backup_path(src)
    with open(bak, "wb") as fh:
        fh.write(blob)
    notes.append(f"the workbook as it was is kept at {bak.name}")
    pruned = prune_backups(src)
    if pruned:
        notes.append(f"{len(pruned)} older backup(s) removed, {BACKUP_KEEP} kept")
    _save_workbook(wb, src, cached)
    locked = excel_lock(src)
    if locked is not None:
        notes.append(f"{locked.name} is beside the workbook, so Excel probably has it open; "
                     f"anything it saves afterwards is written over this")
    return {"written": str(src), "backup": str(bak), "tabs": sorted(tabs),
            "notes": notes, "pruned": pruned, "stamp": workbook_stamp(src)}


def round_trip_losses(path) -> list[str]:
    """What a write would drop from this workbook, said before the first one.

    openpyxl carries cell values, formulas and their cached results; it does
    not carry images, charts, pivot tables or conditional-formatting rules
    that came from parts it does not model.  While the workbook was a book of
    record nobody wrote to, that cost nothing.  As the store the screens write
    on every save, a chart somebody built is gone at the first save and the
    ``.bak`` is the only place it still exists -- which is a thing to be told
    once, not to discover.
    """
    import zipfile
    out: list[str] = []
    try:
        with zipfile.ZipFile(Path(path)) as z:
            names = z.namelist()
    except (OSError, zipfile.BadZipFile) as exc:
        return [f"cannot look inside {Path(path).name}: {exc}"]
    for what, prefix in (("image", "xl/media/"), ("chart", "xl/charts/chart"),
                         ("pivot table", "xl/pivotTables/pivotTable"),
                         ("drawing", "xl/drawings/drawing")):
        n = sum(1 for name in names if name.startswith(prefix))
        if n:
            out.append(f"{n} {what}{'' if n == 1 else 's'}")
    return out


def workbook_stamp(path) -> str:
    """What identifies a workbook file, cheaply: size and modification time.

    Recorded when a book is loaded and checked before it is written over.  Not
    a hash -- the point is to notice that the file is not the one this session
    read, and a rewritten workbook always changes both of these; hashing tens
    of megabytes on every save to catch a rewrite that happened to preserve
    them is paying a great deal for very little.
    """
    p = Path(path)
    try:
        st = p.stat()
    except OSError:
        return ""
    return f"{st.st_size}:{st.st_mtime_ns}"


def excel_lock(path) -> Path | None:
    """Excel's own lock file beside a workbook it has open, if there is one.

    Excel writes ``~$vol_marks.xlsx`` next to the file while it is open.  It is
    a hint and not a guarantee -- a crash leaves one behind, and a file opened
    read-only may not make one -- so what is done with it is to *say so*, not
    to refuse.  On Windows an open workbook is one the replace below cannot
    make, and "permission denied" is a much worse way to find that out.
    """
    p = Path(path)
    lock = p.with_name("~$" + p.name)
    return lock if lock.exists() else None


def prune_backups(src, keep: int = BACKUP_KEEP) -> list[str]:
    """Delete all but the newest ``keep`` backups of one workbook.

    Returns what it removed.  Newest by name, which is newest by time: the
    stamp is written so that it sorts.
    """
    p = Path(src)
    found = sorted(p.parent.glob(f"{p.stem}{BACKUP_INFIX}*{p.suffix}"))
    gone: list[str] = []
    for old in found[:max(0, len(found) - max(0, int(keep)))]:
        try:
            old.unlink()
            gone.append(old.name)
        except OSError:
            continue
    return gone

#: How the workbook spells each curve parameter's PARAMS row, for a row the
#: sheet does not have yet.  The reader accepts these (``marketdata.PARAM_ROWS``).
_CURVE_ROW_LABEL = {
    "initial_vol": "initial", "long_term_vol": "long term", "mean_reversion": "MR",
    "short_addon": "addon", "short_decay": "short decay", "rate_vol": "ratevol",
    "rate_corr": "rate corr",
    # a cross: the same three cells mean the correlation term structure
    "corr_initial": "initial", "corr_final": "long term", "corr_decay": "MR",
}

UNITS_NOTE = ("volatility numbers are in volatility points, as they are typed on the "
              "marking screen; shape parameters, decays, correlations and smile "
              "parameters are the raw numbers those fields carry")


class SessionError(ValueError):
    """A session file that cannot be read, or one that is not a session file."""


def default_path() -> Path:
    """Beside the user's workbook, like the knowledge bank."""
    return app_dir() / SESSION_FILENAME


# --------------------------------------------------------------------------
# the ATM curve, in the units the screen uses
# --------------------------------------------------------------------------
def curve_params(atm) -> dict:
    """The backbone (or a cross's correlation) as the marking screen shows it.

    Volatility fields in points, everything else raw.  The screen and the
    session file read the same function so the two cannot disagree about what
    a number means.
    """
    if isinstance(atm, CrossAtmCurve):
        return {
            "corr_initial": atm.correlation.initial,
            "corr_final": atm.correlation.final,
            "corr_decay": atm.correlation.decay,
            "short_addon": atm.params.short_addon * 100.0,
            "short_decay": atm.params.short_decay,
        }
    p = atm.params
    return {
        "initial_vol": p.initial_vol * 100.0,
        "long_term_vol": p.long_term_vol * 100.0,
        "mean_reversion": p.mean_reversion,
        "short_addon": p.short_addon * 100.0,
        "short_decay": p.short_decay,
        "rate_vol": p.rate_vol * 100.0,
        "rate_corr": p.rate_corr,
    }


def set_curve_params(atm, vals: dict) -> list[str]:
    """Put the screen's own numbers back on a curve.  Returns any problems.

    A field left out keeps its current value rather than becoming zero, which
    is what lets a saved file written by an older build -- or a partial edit
    from the panel -- move only what it actually names.
    """
    if isinstance(atm, CrossAtmCurve):
        problems = atm.set_correlation(
            vals.get("corr_initial", atm.correlation.initial),
            vals.get("corr_final", atm.correlation.final),
            vals.get("corr_decay", atm.correlation.decay))
        if problems:
            return problems
        return atm.set_params(
            short_addon=vals.get("short_addon", atm.params.short_addon * 100.0) / 100.0,
            short_decay=vals.get("short_decay", atm.params.short_decay))
    changes: dict[str, float] = {}
    for key in ("initial_vol", "long_term_vol", "short_addon", "rate_vol"):
        if key in vals:
            changes[key] = vals[key] / 100.0
    for key in ("mean_reversion", "short_decay", "rate_corr"):
        if key in vals:
            changes[key] = vals[key]
    return atm.set_params(**changes)


# --------------------------------------------------------------------------
# capture
# --------------------------------------------------------------------------
def capture_pair(book, pair: str) -> dict:
    """Everything a session has marked on one pair."""
    surface = book[pair]
    atm = surface.atm
    block: dict = {
        "curve": curve_params(atm),
        # The parts and the total, in points: the total is what an old file
        # holds and what a reader wants, the parts are what was marked.
        "events": [{"when": e.when.strftime("%Y-%m-%dT%H:%M"),
                    "bump": e.bump * 100.0,
                    "weights": {c: v * 100.0 for c, v in e.weights.items()},
                    "adjust": (e.adjust or 0.0) * 100.0,
                    "label": e.label} for e in atm.events.events],
        "atm_overwrites": {k: v * 100.0 for k, v in sorted(atm.tenor_overwrites.items())},
        "smile_overwrites": {name: dict(sorted(ow.items()))
                             for name, ow in sorted(surface.param_overwrites.items()) if ow},
        # The quotes typed over the sheet's, in points like every other
        # volatility here.  Only the boxes somebody filled: what the sheet
        # already holds is the sheet's, and a session that copied it out would
        # freeze one morning's quotes into every workbook it was put back onto.
        "quote_overwrites": {tenor: {f: v * 100.0 for f, v in sorted(edits.items())}
                             for tenor, edits in sorted(surface.quote_overwrites.items())
                             if edits},
        # The wing multipliers the screen changed.  Raw numbers, not points:
        # a ratio is a multiple, not a volatility.  ``None`` is kept and means
        # a wing taken off its ratio and quoted in its own right -- the
        # difference between that and "no opinion" is the whole point of the
        # tab, so it survives the file.
        "wing_ratios": {tenor: {w: (None if v is None else float(v))
                                for w, v in sorted(edits.items())}
                        for tenor, edits in sorted(surface.ratio_overwrites.items())
                        if edits},
        # A marked term structure replaces the fitted one for that parameter,
        # so what is written is the three coefficients themselves, raw.  Only
        # the parameters somebody marked appear; the rest are the fit's and
        # are not a session's to carry.
        "smile_term": {name: {c: getattr(curve, c) for c in TERM_COEFFS}
                       for name, curve in sorted(surface.term_marks.items())},
        "param_shifts": {k: float(v) for k, v in sorted(surface.param_shifts.items()) if v},
        "anchor_tenors": bool(surface.anchor_tenors),
    }
    # The treatment is only written for a pair that has a band or has been
    # marked away from the default.  Writing it for every free floater would
    # put a hazard rate in the file for pairs that have no peg to break.
    treatment = surface.band_treatment
    if surface.band is not None or treatment != BandTreatment():
        block["band"] = treatment.to_request()
    return block


def capture(book, pairs=None, *, note: str = "") -> dict:
    """The whole session as a document, ready to be written.

    ``pairs`` narrows it; the default is every pair the book built.  The
    header records the workbook and the valuation clock the marks were made
    against -- neither is enforced on the way back in, because a marker
    re-opening yesterday's marks against today's clock is the normal case, but
    a file that came from a different workbook is worth being told about.
    """
    if book is None:
        raise SessionError("there is no loaded book to save marks from")
    wanted = [p for p in book.pairs if pairs is None or p.upper() in
              {str(x).upper() for x in pairs}]
    unknown = ([] if pairs is None else
               [str(x).upper() for x in pairs if str(x).upper() not in
                {p.upper() for p in book.pairs}])
    if unknown:
        raise SessionError(f"{', '.join(unknown)} is not built in this book; it holds "
                           f"{', '.join(book.pairs)}")
    return {
        "kind": "volkit session",
        "version": SESSION_VERSION,
        "saved": datetime.now().astimezone().replace(microsecond=0).isoformat(),
        "valuation": book.clock.now.isoformat(),
        "workbook": str(getattr(book.data, "source", "") or ""),
        # What that workbook was when these marks were made.  Checked before
        # a write over it, so a file that somebody else has saved since is
        # refused rather than quietly overwritten.  Advisory on the way back
        # *in* -- a marker re-opening yesterday's marks is the normal case.
        "workbook_stamp": workbook_stamp(getattr(book.data, "source", "") or ""),
        "note": str(note or ""),
        "units": UNITS_NOTE,
        "pairs": {p: capture_pair(book, p) for p in wanted},
        # The event table is the book's, not a pair's: one row per release,
        # a weight per currency and an adjustment per pair, in points.  This
        # is the authority on the way back in; a pair block's ``events`` is
        # that pair's resolved schedule, kept because a re-mark of an event
        # is a per-pair thing to record (``remarks.SECTIONS``).
        "event_table": capture_events(book),
    }


def capture_events(book) -> list[dict]:
    """The event table as the file holds it: points, and times in UTC."""
    return [{"when": r.when.strftime("%Y-%m-%dT%H:%M"),
             "label": r.label,
             "weights": {c: v * 100.0 for c, v in sorted(r.weights.items())},
             "adjust": {p: v * 100.0 for p, v in sorted(r.adjust.items())}}
            for r in sorted(book.events.rows, key=lambda r: r.when)]


def event_table_from_doc(doc: dict) -> tuple[list, list[str], list[str]]:
    """The document's event table as ``EventRow``s, in decimals.

    A file written by this version carries ``event_table``.  An older one has
    the same events spread across its pairs, so they are unioned back into a
    table -- a currency weight was shared even then, and two pairs that
    disagree about one are reported rather than averaged.
    """
    from .events import EventRow

    problems: list[str] = []
    notes: list[str] = []
    rows: dict[datetime, EventRow] = {}

    def at(when: datetime, label: str) -> "EventRow":
        key = when.replace(second=0, microsecond=0)
        row = rows.get(key)
        if row is None:
            row = rows[key] = EventRow(key, label)
        elif label and not row.label:
            row.label = label
        return row

    table = doc.get("event_table")
    if isinstance(table, list):
        for i, item in enumerate(table, start=1):
            if not isinstance(item, dict):
                problems.append(f"event table row {i} is not an object")
                continue
            try:
                when = parse_datetime(str(item.get("when")))
            except ValueError as exc:
                problems.append(f"event table row {i}: {exc}")
                continue
            row = at(when, str(item.get("label") or ""))
            try:
                for c, v in (item.get("weights") or {}).items():
                    if float(v or 0.0):
                        row.weights[str(c).upper()] = float(v) / 100.0
                for pr, v in (item.get("adjust") or {}).items():
                    if float(v or 0.0):
                        row.adjust[str(pr).upper()] = float(v) / 100.0
            except (TypeError, ValueError):
                problems.append(f"event table row {i}: a weight must be a number")
        return list(rows.values()), problems, notes

    # An older file: the table is the union of what its pairs were given.
    seen_pairs = False
    for name, block in (doc.get("pairs") or {}).items():
        if not isinstance(block, dict) or "events" not in block:
            continue
        pair = str(name).upper()
        entries, bad = event_entries([r if isinstance(r, dict) else {}
                                      for r in (block.get("events") or [])])
        problems.extend(f"{pair}: {b}" for b in bad)
        for e in entries:
            # Resolved for the pair, so a row typed as one number lands as
            # that pair's adjustment -- which is where a total with no parts
            # has always gone (``EventEntry.resolve``).
            e = e.resolve(pair)
            seen_pairs = True
            row = at(as_utc(e.when), e.label)
            for ccy, w in (e.weights or {}).items():
                if not w:
                    continue
                prior = row.weights.get(ccy)
                if prior is not None and abs(prior - w) > 1e-12:
                    problems.append(
                        f"{pair}: event {e.when:%Y-%m-%d %H:%M}Z weight {ccy} "
                        f"{w * 100:.4g} disagrees with {prior * 100:.4g} from another "
                        "pair; a currency weight is shared")
                row.weights[ccy] = w
            if e.adjust:
                row.adjust[pair] = float(e.adjust)
    if seen_pairs:
        notes.append("the file has no event table, so one was rebuilt from its pairs' "
                     "schedules; a currency weight is shared by every pair with that "
                     "currency")
    return list(rows.values()), problems, notes


# --------------------------------------------------------------------------
# apply
# --------------------------------------------------------------------------
def apply_block(surface, block: dict) -> list[str]:
    """One pair's marks, in the order the screen would make them.

    Also what the book calls for the marks a workbook holds in its own
    session rows (``marketdata.overlay_label``), which are read into exactly
    this shape: one reader of the vocabulary, wherever the block came from.

    Curve first, because the events calibrate against it; then the overwrites,
    which sit on top of whatever the curve produced; then the wing shifts and
    the band treatment.  Every step that can refuse a value returns its
    reasons and they are collected rather than raised, so a file with one bad
    number still restores everything else and says what it could not.
    """
    problems: list[str] = []
    if isinstance(block.get("curve"), dict):
        vals = {}
        for k, v in block["curve"].items():
            if v in (None, ""):
                continue
            try:
                vals[k] = float(v)
            except (TypeError, ValueError):
                problems.append(f"curve parameter {k}: {v!r} is not a number")
        problems.extend(set_curve_params(surface.atm, vals))

    # ``events`` is deliberately not applied here.  An event is not a pair's
    # to hold any more: a currency weight is shared, so the schedule comes
    # from the document's one event table (``load``), and a pair block's
    # ``events`` is the record of what that pair's schedule *was* -- which is
    # what a re-mark of an event is diffed against (``remarks.SECTIONS``).

    if "atm_overwrites" in block:
        surface.atm.clear_overwrite()
        for tenor, value in (block.get("atm_overwrites") or {}).items():
            try:
                surface.atm.overwrite_tenor(str(tenor), float(value) / 100.0)
            except (TypeError, ValueError):
                problems.append(f"ATM overwrite {tenor}: {value!r} is not a number")

    if "smile_overwrites" in block:
        surface.clear_param_overwrites()
        for name, rows in (block.get("smile_overwrites") or {}).items():
            if name not in PARAM_NAMES:
                problems.append(f"smile overwrite: {name!r} is not a smile parameter "
                                f"({', '.join(PARAM_NAMES)})")
                continue
            for tenor, value in (rows or {}).items():
                try:
                    surface.overwrite_param(name, str(tenor), float(value))
                except (TypeError, ValueError):
                    problems.append(f"smile overwrite {name} {tenor}: {value!r} is not a number")

    requote = False
    if "wing_ratios" in block:
        requote = requote or bool(surface.ratio_overwrites)
        surface.clear_ratio_overwrite()
        for tenor, edits in (block.get("wing_ratios") or {}).items():
            if not isinstance(edits, dict):
                problems.append(f"wing ratio {tenor}: expected the wings as an object, "
                                f"got {edits!r}")
                continue
            for wing, value in edits.items():
                if wing not in RATIO_WINGS:
                    problems.append(f"wing ratio {tenor}: {wing!r} is not a wing "
                                    f"({', '.join(RATIO_WINGS)})")
                    continue
                try:
                    surface.overwrite_ratio(str(tenor), wing,
                                            None if value is None else float(value))
                except (TypeError, ValueError) as exc:
                    problems.append(f"wing ratio {wing} {tenor}: {exc}")
        requote = requote or bool(surface.ratio_overwrites)

    if "quote_overwrites" in block:
        # Whether the quotes moved at all, which decides whether the smile has
        # to be refitted below.  A block that carries no quote edits and lands
        # on a surface that had none is not a re-quote, and refitting it would
        # rebuild every smile against whatever the curve has just been marked
        # to -- a different surface from the one that was saved, arriving
        # through a code path nobody asked to run.
        requote = bool(surface.quote_overwrites)
        surface.clear_quote_overwrite()
        for tenor, edits in (block.get("quote_overwrites") or {}).items():
            if not isinstance(edits, dict):
                problems.append(f"quote {tenor}: expected the quotes as an object, "
                                f"got {edits!r}")
                continue
            for name, value in edits.items():
                if name not in QUOTE_FIELDS:
                    problems.append(f"quote {tenor}: {name!r} is not one of the four quotes "
                                    f"({', '.join(QUOTE_FIELDS)})")
                    continue
                try:
                    surface.overwrite_quote(str(tenor), name, float(value) / 100.0)
                except (TypeError, ValueError) as exc:
                    problems.append(f"quote {QUOTE_LABELS[name]} {tenor}: {exc}")
        requote = requote or bool(surface.quote_overwrites)

    if "smile_term" in block:
        surface.clear_param_terms()
        for name, coeffs in (block.get("smile_term") or {}).items():
            if name not in PARAM_NAMES:
                problems.append(f"smile term structure: {name!r} is not a smile parameter "
                                f"({', '.join(PARAM_NAMES)})")
                continue
            if not isinstance(coeffs, dict):
                problems.append(f"smile term structure {name}: expected the three "
                                f"coefficients, got {coeffs!r}")
                continue
            vals: dict[str, float] = {}
            for coeff in TERM_COEFFS:
                try:
                    vals[coeff] = float(coeffs[coeff])
                except (KeyError, TypeError, ValueError):
                    problems.append(f"smile term structure {name}: {coeff} is missing or "
                                    f"is not a number")
                    break
            else:
                problems.extend(f"smile term structure: {m}"
                                for m in surface.set_param_term(name, **vals))

    if "param_shifts" in block:
        shifts = {}
        for name, value in (block.get("param_shifts") or {}).items():
            if name not in PARAM_NAMES:
                problems.append(f"wing shift: {name!r} is not a smile parameter "
                                f"({', '.join(PARAM_NAMES)})")
                continue
            try:
                shifts[name] = float(value)
            except (TypeError, ValueError):
                problems.append(f"wing shift {name}: {value!r} is not a number")
        problems.extend(surface.set_param_shifts(shifts))

    if "anchor_tenors" in block:
        surface.anchor_tenors = bool(block["anchor_tenors"])

    if isinstance(block.get("band"), dict):
        try:
            problems.extend(surface.set_band_treatment(
                BandTreatment.from_request(block["band"])))
        except (TypeError, ValueError) as exc:
            problems.append(f"band treatment: {exc}")

    # A typed quote is an input to the fit and not a mark applied on top of
    # one, so a block that moved one has to refit the smile before anything
    # reads it.  Only that case: on the way up from the workbook this runs
    # before the book calibrates anything, and fitting an empty surface here
    # would leave it with no parameter term structure at all -- and a refit
    # on a block that touched no quote would rebuild the smiles against a
    # curve the saved surface never fitted against, which is a different
    # surface from the one somebody marked.
    if requote and surface.fits:
        surface.calibrate()
        problems.extend(surface.warnings)
        surface.warnings.clear()
    surface.invalidate()
    return problems


def apply_document(book, doc: dict, pairs=None) -> dict:
    """Put a saved session back onto a loaded book.

    Returns what happened rather than raising: which pairs took their marks,
    which the workbook does not build, which it built and the file never
    mentioned, and every problem with the pair's name on it.
    """
    if book is None:
        raise SessionError("there is no loaded book to put marks onto")
    if not isinstance(doc, dict) or not isinstance(doc.get("pairs"), dict):
        raise SessionError("that is not a volkit session file: it has no 'pairs' object")

    version = int(doc.get("version") or 1)
    notes: list[str] = []
    if version > SESSION_VERSION:
        notes.append(
            f"this file was written by a newer volkit (format {version}, this build reads "
            f"{SESSION_VERSION}); anything it does not understand is left alone")
    saved_book = str(doc.get("workbook") or "")
    here = str(getattr(book.data, "source", "") or "")
    if saved_book and here and Path(saved_book).name != Path(here).name:
        notes.append(f"these marks were saved against {Path(saved_book).name}, and the loaded "
                     f"workbook is {Path(here).name}; the tenors and pairs may not line up")

    wanted = {k.upper() for k in doc["pairs"]} if pairs is None else \
        {str(x).upper() for x in pairs}
    have = {p.upper(): p for p in book.pairs}
    problems: list[str] = []
    applied: list[str] = []

    missing = sorted(k for k in doc["pairs"] if k.upper() in wanted and k.upper() not in have)
    if missing:
        problems.append(f"the file has marks for {', '.join(missing)}, which this workbook "
                        f"does not build; they were not applied")

    # Build order, so a cross recalibrates against legs that already have
    # their saved marks rather than against the workbook's.
    order = [p for p in book.build_order() if p in have.values()]
    order += [p for p in book.pairs if p not in order]
    by_upper = {k.upper(): v for k, v in doc["pairs"].items()}
    for pair in order:
        key = pair.upper()
        if key not in wanted or key not in by_upper:
            continue
        block = by_upper[key]
        if not isinstance(block, dict):
            problems.append(f"{pair}: its entry in the file is not an object")
            continue
        try:
            problems.extend(f"{pair}: {x}" for x in apply_block(book[pair], block))
        except Exception as exc:  # noqa: BLE001 - one pair, not the whole file
            problems.append(f"{pair}: {type(exc).__name__}: {exc}")
            continue
        applied.append(pair)

    # The event table, once, for the whole book: a row is one release and its
    # currency weights belong to every pair with that currency.  Applied
    # after the pairs so it is the last word on every curve's schedule.
    rows, bad, table_notes = event_table_from_doc(doc)
    problems.extend(bad)
    notes.extend(table_notes)
    if rows or "event_table" in doc:
        book.events.rows = rows
        book.events.sort()
        book.events.source = str(doc.get("saved") or "session file")
        problems.extend(book.apply_events())

    untouched = [p for p in book.pairs if p.upper() not in by_upper]
    if untouched:
        notes.append(f"the file says nothing about {', '.join(untouched)}; "
                     f"{'they were' if len(untouched) > 1 else 'it was'} left as the workbook "
                     f"has {'them' if len(untouched) > 1 else 'it'}")
    return {"applied": applied, "problems": problems, "notes": notes,
            "saved": str(doc.get("saved") or ""), "note": str(doc.get("note") or ""),
            "workbook": saved_book, "valuation": str(doc.get("valuation") or "")}


# --------------------------------------------------------------------------
# io
# --------------------------------------------------------------------------
def load(path: str | Path) -> dict:
    """Read a session file.  A missing or unreadable one is an error, not an
    empty session: somebody asked for these marks by name."""
    p = Path(path)
    if not p.exists():
        raise SessionError(f"no session file at {p}")
    try:
        raw = json.loads(read_text(p))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise SessionError(f"cannot read the session file at {p}: {exc}") from None
    if not isinstance(raw, dict) or "pairs" not in raw:
        raise SessionError(f"{p} is not a volkit session file; expected an object with a "
                           f"'pairs' key, got {type(raw).__name__}")
    return raw


def write(doc: dict, path: str | Path | None = None) -> str:
    """Write a document atomically, so an interrupted save cannot lose the file."""
    p = Path(path) if path else default_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=str(p.parent), prefix=".vol_session", suffix=".json")
    try:
        with os.fdopen(fd, "w", encoding=WRITE_ENCODING) as fh:
            json.dump(doc, fh, indent=2)
            fh.write("\n")
        os.replace(tmp, p)
    except BaseException:
        Path(tmp).unlink(missing_ok=True)
        raise
    return str(p)


def save(book, path: str | Path | None = None, pairs=None, *, note: str = "") -> str:
    """Capture the session and write it.  Returns the file written."""
    return write(capture(book, pairs, note=note), path)


def restore(book, path: str | Path, pairs=None) -> dict:
    """Read a session file and put it onto the book, in one step."""
    out = apply_document(book, load(path), pairs)
    out["path"] = str(path)
    return out


# --------------------------------------------------------------------------
# export: a session file into a workbook
# --------------------------------------------------------------------------
def export_workbook(doc: dict, workbook: str | Path, out: str | Path | None = None,
                    pairs=None, *, in_place: bool = False, expect: str = "",
                    force: bool = False) -> dict:
    """Write a session's marks into the cells a workbook keeps them in.

    The workbook is the book of record and nothing else in this package
    writes to it (§2).  This is the one deliberate exception, and it is an
    *export*: it is never run off the live book, only off a session file
    somebody saved first, and it says what it wrote and what it could not.
    ``in_place`` (the marking screen's button, and ``--in-place``) writes
    over the workbook itself; without it a **copy** goes beside the original
    (``<name>_marked.xlsx``).  Writing over the original keeps the bytes it
    replaced as ``<name>.bak-<stamp>.xlsx`` beside it, because openpyxl does
    not carry images or charts through a round trip and an export is
    otherwise not undoable.

    What goes where.  The curve parameters go into the PARAMS rows the
    reader already reads (a cross's ``corr_*`` into the same three cells the
    workbook has always used for a correlation).  Events go onto the
    **EVENTS** sheet, one row per release in the workbook's own clock (Hong
    Kong, ``event_tz_offset`` hours ahead of the file's UTC): weights into
    the currency columns and each pair's adjustment into its own column.  A
    row, a column or the sheet itself is added if the workbook lacks it, and
    the sheet is written whole -- the table is the book's, not any one
    pair's.  The marks the workbook had no cell for -- ATM
    and smile overwrites, wing shifts, the anchor switch -- go into rows
    ``marketdata.overlay_label`` reads (``atm 1m``, ``slog25 3m``,
    ``term rho10 decay``, ``shift rho25``, ``anchor``), and the band
    treatment into a ``BANDS``
    sheet.  Every one of them is read back by ``ExcelSource``, so a workbook
    written here loads as the session it came from; writing a cell the tool
    would not read is the silent zero this project exists to remove.

    A pair is replaced, never merged: its overwrites and shifts in the file
    are the whole of what the workbook then holds for it.  Events are not
    per pair at all: the file's event table replaces the EVENTS sheet
    outright, because a currency weight belongs to every pair with that
    currency and a half-written table is a bump somebody else silently
    inherits.  ``pairs`` narrows which pairs' *columns* are written, not
    which events exist.

    Formulas and the other sheets are kept as they are; images and charts,
    if any, are not (openpyxl does not carry them), which is why an in-place
    write keeps a backup and why the CLI's default is still a copy.

    ``expect`` is the workbook's :func:`workbook_stamp` as it was when the
    marks being written were made.  An in-place write whose file no longer
    matches it is **refused**: something else has written the workbook since,
    and a store that lets the second writer win silently loses the first one's
    morning.  ``force`` writes anyway, for a person who has looked and decided.
    Returns ``{"written", "backup", "pairs", "problems", "notes", "in_place",
    "stamp", "stale", "pruned", "locked"}``.
    """
    import openpyxl
    from datetime import timedelta
    from .marketdata import BANDS_SHEET, PARAM_ROWS, _norm, overlay_label
    from .surface import PARAM_NAMES as _SMILE, TERM_COEFFS as _TERM

    if not isinstance(doc, dict) or not isinstance(doc.get("pairs"), dict):
        raise SessionError("that is not a volkit session file: it has no 'pairs' object")
    src = Path(workbook)
    if not src.exists():
        raise SessionError(f"workbook not found: {src}")
    # ``in_place`` is the destination, not just permission to be it: naming
    # no output and asking for in place used to write ``_marked`` and report
    # it as having gone into the workbook.
    dst = Path(out) if out else (
        src if in_place else src.with_name(src.stem + EXPORT_SUFFIX + src.suffix))
    if dst.resolve() == src.resolve() and not in_place:
        raise SessionError(
            f"{src.name} is the workbook itself; writing into it needs in_place "
            "(--in-place), or name another file")
    if dst.suffix.lower() != ".xlsx":
        raise SessionError(f"{dst.name}: a workbook is written as .xlsx")

    # -- is this still the file the marks were made against? ---------------
    # Only for a write over the original: a copy takes nothing away from
    # anybody, and refusing one would be a check with no failure to prevent.
    writing_over = dst.exists() and dst.resolve() == src.resolve()
    now = workbook_stamp(src)
    stale = bool(writing_over and expect and now and expect != now)
    if stale and not force:
        raise SessionError(
            f"{src.name} has changed since these marks were made -- another volkit, or "
            f"somebody saving it from Excel. Writing now would drop whatever they did. "
            f"Reload the workbook and put the marks back on it, or write a copy instead")
    locked = str(excel_lock(dst) or "") if writing_over else ""

    # Read whole and closed before parsing, like every workbook here
    # (``marketdata.open_workbook``): an .xlsx held open is one Excel cannot
    # save.
    import io
    with src.open("rb") as fh:
        blob = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    # openpyxl keeps a formula and drops the value Excel last computed for
    # it, and pandas reads a formula with no cached value as blank -- so a
    # copy saved by openpyxl alone has smile sheets full of ``=C2*3`` that
    # read as no quote at all until Excel has opened and saved it.  The
    # cached values are read here and put back into the saved file.
    values = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    cached = _formula_cache(wb, values)
    if "PARAMS" not in wb.sheetnames:
        raise SessionError(f"{src.name} has no PARAMS sheet")
    ws = wb["PARAMS"]
    tz_shift = timedelta(hours=8.0)  # ExcelSource's default event_tz_offset_hours

    problems: list[str] = []
    notes: list[str] = []
    written: list[str] = []
    #: The multipliers as they end up, by pair, for the tab written at the end.
    written_ratios: dict[str, dict[str, dict[str, float | None]]] = {}
    # The wing multipliers as the workbook holds them, so a quote written into
    # a sheet can be written with the wing it derives beside it.  Read from the
    # source and not from the file being built: the tab is rewritten below.
    try:
        sheet_ratios = load_wing_ratios(src)
    except (OSError, ValueError) as exc:
        sheet_ratios = {}
        problems.append(f"{WING_RATIOS_SHEET}: {exc}; wings were written as they stand")

    # -- the sheet as it stands -------------------------------------------
    header = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            header[str(v).strip().upper()] = c
    ncol = [ws.max_column]

    def column_for(name: str, *, create: bool) -> int | None:
        key = name.upper()
        if key in header:
            return header[key]
        if not create:
            return None
        ncol[0] += 1
        ws.cell(row=1, column=ncol[0], value=key)
        header[key] = ncol[0]
        return ncol[0]

    rows_by_key: dict[str, int] = {}
    overlay_rows: dict[tuple, int] = {}
    # Rows are appended after the last *labelled* one: ``max_row`` counts
    # formatted-but-empty rows, and a gap of blank labels is what the reader
    # then meets as a row that is neither a parameter nor a date.
    nrow = [1]
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label is None or not str(label).strip():
            continue
        nrow[0] = r
        key = PARAM_ROWS.get(_norm(label))
        if key is not None:
            rows_by_key[key] = r
            continue
        ov = overlay_label(label)
        if ov is not None:
            overlay_rows[ov] = r

    def row_for(key: str, label: str) -> int:
        if key in rows_by_key:
            return rows_by_key[key]
        nrow[0] += 1
        ws.cell(row=nrow[0], column=1, value=label)
        rows_by_key[key] = nrow[0]
        return nrow[0]

    def overlay_row(ov: tuple) -> int:
        if ov in overlay_rows:
            return overlay_rows[ov]
        nrow[0] += 1
        # the labels ``marketdata.overlay_label`` reads: a smile overwrite is
        # spelled by its parameter and tenor alone
        ws.cell(row=nrow[0], column=1, value=" ".join(ov[1:] if ov[0] == "smile" else ov))
        overlay_rows[ov] = nrow[0]
        return nrow[0]

    def clear(rows, col: int) -> None:
        # ``ws.cell(row, column, value=None)`` does **not** blank a cell:
        # openpyxl assigns only when ``value is not None``, so the call is a
        # no-op and every clear here was one until 2026-09-02.  An overwrite
        # taken off on the screen therefore stayed in the workbook and came
        # back on the next load -- a mark nobody could see undoing a decision
        # somebody had made.  Assigned through the cell, which does blank it.
        for r in rows:
            ws.cell(row=r, column=col).value = None

    def number(what: str, v):
        try:
            return float(v)
        except (TypeError, ValueError):
            problems.append(f"{what}: {v!r} is not a number")
            return None

    wanted = {k.upper() for k in doc["pairs"]} if pairs is None else \
        {str(x).upper() for x in pairs}
    for raw_name, block in doc["pairs"].items():
        name = str(raw_name).upper()
        if name not in wanted:
            continue
        if not isinstance(block, dict):
            problems.append(f"{name}: its entry in the file is not an object")
            continue
        col = column_for(name, create=False)
        if col is None:
            problems.append(f"{name}: the workbook has no PARAMS column for it (list the pair "
                            "in CONFIG and give it a PARAMS column first); not written")
            continue
        wrote: list[str] = []

        # curve -----------------------------------------------------------
        curve = block.get("curve")
        if isinstance(curve, dict):
            for k, v in curve.items():
                if v in (None, ""):
                    continue
                label = _CURVE_ROW_LABEL.get(k)
                if label is None:
                    problems.append(f"{name}: curve parameter {k!r} is not one the workbook holds")
                    continue
                fv = number(f"{name} curve {k}", v)
                if fv is None:
                    continue
                key = PARAM_ROWS[_norm(label)]
                ws.cell(row=row_for(key, label), column=col, value=fv)
            wrote.append("curve")

        # overwrites, shifts, anchor ---------------------------------------
        if "atm_overwrites" in block:
            clear((r for ov, r in overlay_rows.items() if ov[0] == "atm"), col)
            for tenor, v in (block.get("atm_overwrites") or {}).items():
                fv = number(f"{name} ATM overwrite {tenor}", v)
                if fv is not None:
                    ws.cell(row=overlay_row(("atm", str(tenor).lower())), column=col, value=fv)
            wrote.append(f"{len(block.get('atm_overwrites') or {})} ATM overwrite(s)")
        if "smile_overwrites" in block:
            clear((r for ov, r in overlay_rows.items() if ov[0] == "smile"), col)
            n = 0
            for pname, tenors in (block.get("smile_overwrites") or {}).items():
                if pname not in _SMILE:
                    problems.append(f"{name}: smile overwrite {pname!r} is not a smile parameter")
                    continue
                for tenor, v in (tenors or {}).items():
                    fv = number(f"{name} smile overwrite {pname} {tenor}", v)
                    if fv is not None:
                        ws.cell(row=overlay_row(("smile", pname, str(tenor).lower())),
                                column=col, value=fv)
                        n += 1
            wrote.append(f"{n} smile overwrite(s)")
        if "smile_term" in block:
            clear((r for ov, r in overlay_rows.items() if ov[0] == "term"), col)
            n = 0
            for pname, coeffs in (block.get("smile_term") or {}).items():
                if pname not in _SMILE:
                    problems.append(f"{name}: smile term structure {pname!r} is not a smile "
                                    f"parameter")
                    continue
                if not isinstance(coeffs, dict):
                    problems.append(f"{name}: smile term structure {pname}: expected the "
                                    f"three coefficients, got {coeffs!r}")
                    continue
                # All three or none: two thirds of a curve on the sheet is a
                # set the reader would refuse, and a refusal on the way back
                # in is a worse place to find out than here.
                vals = [number(f"{name} smile term {pname} {c}", coeffs.get(c))
                        for c in _TERM]
                if any(v is None for v in vals):
                    continue
                for coeff, fv in zip(_TERM, vals):
                    ws.cell(row=overlay_row(("term", pname, coeff)), column=col, value=fv)
                n += 1
            wrote.append(f"{n} smile term structure(s)")
        if "param_shifts" in block:
            clear((r for ov, r in overlay_rows.items() if ov[0] == "shift"), col)
            n = 0
            for pname, v in (block.get("param_shifts") or {}).items():
                if pname not in _SMILE:
                    problems.append(f"{name}: wing shift {pname!r} is not a smile parameter")
                    continue
                fv = number(f"{name} wing shift {pname}", v)
                if fv:
                    ws.cell(row=overlay_row(("shift", pname)), column=col, value=fv)
                    n += 1
            wrote.append(f"{n} wing shift(s)")
        if "anchor_tenors" in block:
            # Same trap as ``clear`` above: switching the anchor off has to
            # blank the cell, and passing None to ``cell()`` does not.
            ws.cell(row=overlay_row(("anchor",)), column=col).value = (
                1 if block["anchor_tenors"] else None)
            if block["anchor_tenors"]:
                wrote.append("anchored")

        # the wing multipliers, and the quotes they derive ------------------
        # One step, because they are one statement about the sheet: the tab
        # says which wings are a multiple of the 25-delta, and the sheet is
        # then written as numbers throughout so nothing recalculates.
        live_ratios: dict[str, dict[str, float | None]] = {}
        for tenor, ratio in (sheet_ratios.get(name) or {}).items():
            live_ratios[str(tenor).upper()] = {w: ratio.get(w) for w in RATIO_WINGS}
        if "wing_ratios" in block:
            for tenor, edits in (block.get("wing_ratios") or {}).items():
                if not isinstance(edits, dict):
                    problems.append(f"{name}: wing ratio {tenor}: expected the wings as an "
                                    f"object, got {edits!r}")
                    continue
                row = live_ratios.setdefault(str(tenor).upper(),
                                             {w: None for w in RATIO_WINGS})
                for wing, v in edits.items():
                    if wing not in RATIO_WINGS:
                        problems.append(f"{name}: wing ratio {tenor}: {wing!r} is not a wing "
                                        f"({', '.join(RATIO_WINGS)})")
                        continue
                    if v is None:
                        row[wing] = None
                        continue
                    fv = number(f"{name} wing ratio {wing} {tenor}", v)
                    if fv is None:
                        continue
                    try:
                        row[wing] = check_ratio(wing, f"{name} {tenor}", fv)
                    except ValueError as exc:
                        problems.append(f"{name}: {exc}")

        # quotes, into the pair's own sheet --------------------------------
        if "quote_overwrites" in block:
            clean: dict[str, dict[str, float]] = {}
            for tenor, fields in (block.get("quote_overwrites") or {}).items():
                if not isinstance(fields, dict):
                    problems.append(f"{name}: quote {tenor}: expected the quotes as an "
                                    f"object, got {fields!r}")
                    continue
                vals: dict[str, float] = {}
                for f, v in fields.items():
                    if f not in QUOTE_FIELDS:
                        problems.append(f"{name}: quote {tenor}: {f!r} is not one of the four "
                                        f"quotes ({', '.join(QUOTE_FIELDS)})")
                        continue
                    fv = number(f"{name} quote {QUOTE_LABELS[f]} {tenor}", v)
                    if fv is not None:
                        vals[f] = fv
                if vals:
                    clean[str(tenor)] = vals
            if clean or any(any(v is not None for v in r.values())
                            for r in live_ratios.values()):
                n, probs, ns = _write_quote_cells(wb, values, name, clean, live_ratios)
                problems.extend(probs)
                notes.extend(ns)
                if n:
                    wrote.append(f"{n} quote(s)")
        if live_ratios:
            written_ratios[name] = live_ratios

        # band ------------------------------------------------------------
        if isinstance(block.get("band"), dict):
            try:
                treatment = BandTreatment.from_request(block["band"])
            except (TypeError, ValueError) as exc:
                problems.append(f"{name}: band treatment: {exc}")
            else:
                _write_band_row(wb, BANDS_SHEET, name, treatment.to_request())
                wrote.append(f"band {treatment.mode}")

        written.append(name)
        notes.append(f"{name}: " + ", ".join(wrote))

    # -- EVENTS ------------------------------------------------------------
    # The whole sheet, written once from the file's one table.  Not per pair:
    # a currency weight belongs to every pair with that currency, so writing
    # one pair's view of it would leave the rest of the column reading a
    # number nobody marked.
    rows, bad, table_notes = event_table_from_doc(doc)
    problems.extend(bad)
    notes.extend(table_notes)
    if written and (rows or "event_table" in doc):
        notes.append(_write_events_sheet(wb, rows, sorted(header), tz_shift))

    # -- WING_RATIOS -------------------------------------------------------
    if written and written_ratios:
        notes.append(_write_wing_ratios_sheet(wb, written_ratios))

    backup = ""
    pruned: list[str] = []
    if not written:
        problems.append("nothing was written")
    else:
        # Over the book of record: keep what it replaced.  ``blob`` is the
        # file as it was read, so the backup is the bytes themselves and not
        # a re-read that a second writer could have moved underneath us.
        if dst.exists() and dst.resolve() == src.resolve():
            bak = _backup_path(src)
            with open(bak, "wb") as fh:
                fh.write(blob)
            backup = str(bak)
            notes.append(f"the workbook as it was is kept at {bak.name}")
            pruned = prune_backups(src)
            if pruned:
                notes.append(f"{len(pruned)} older backup(s) removed, {BACKUP_KEEP} kept")
        _save_workbook(wb, dst, cached)
        n = sum(len(v) for v in cached.values())
        if n:
            notes.append(f"{n} formula cell(s) kept, with the values Excel last computed "
                         "for them; a formula reading a cell written here shows its old "
                         "value until Excel recalculates")
    if locked:
        notes.append(f"{Path(locked).name} is beside the workbook, so Excel probably has it "
                     f"open. What is written here is what the tool holds; anything Excel "
                     f"saves afterwards is written over it")
    return {"written": str(dst) if written else "", "backup": backup,
            "pairs": written, "problems": problems, "notes": notes,
            "in_place": bool(in_place), "stamp": workbook_stamp(dst) if written else "",
            "stale": stale, "pruned": pruned, "locked": locked}


def _backup_path(src: Path) -> Path:
    """Where the bytes an in-place export replaces are kept.  Stamped to the
    second in local time, like the session file's own ``saved``; two exports
    inside one second do not overwrite each other's backup."""
    stamp = datetime.now().astimezone().strftime("%Y%m%d-%H%M%S")
    bak = src.with_name(f"{src.stem}{BACKUP_INFIX}{stamp}{src.suffix}")
    n = 2
    while bak.exists():
        bak = src.with_name(f"{src.stem}{BACKUP_INFIX}{stamp}-{n}{src.suffix}")
        n += 1
    return bak


def _write_events_sheet(wb, rows, pairs, tz_shift) -> str:
    """Write the whole EVENTS sheet: a row per release, a column per
    currency and per pair.

    Written whole rather than cell by cell.  A currency weight is shared by
    every pair with that currency, so leaving an old row or an old column
    behind would hand a pair a bump nobody in this session marked -- the
    exact failure the sheet exists to prevent.  Times go in the workbook's
    own clock (Hong Kong), which is what every event in this tool has always
    been typed in and what ``ExcelSource`` reads back.

    ``pairs`` is the PARAMS header, so a pair the workbook knows gets its
    column even when no event adjusts it: an empty column is where the next
    adjustment is typed, and a column that is not there is one nobody can.
    """
    from .marketdata import EVENTS_SHEET

    ccys = sorted({c for r in rows for c in r.weights})
    cols = list(ccys) + [p for p in pairs if len(p) == 6 and p.isalpha()]
    # A named release keeps its name.  The column only appears when something
    # is in it, and it goes second, which is where the reader looks for it.
    labelled = any(r.label for r in rows)
    first = 3 if labelled else 2
    if EVENTS_SHEET in wb.sheetnames:
        # Emptied in place rather than removed and remade: a sheet dropped
        # and re-added moves to the end of the book, and where a trader keeps
        # a tab is theirs.
        ws = wb[EVENTS_SHEET]
        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)
    else:
        ws = wb.create_sheet(EVENTS_SHEET)
    if labelled:
        ws.cell(row=1, column=2, value="label")
    for i, name in enumerate(cols, start=first):
        ws.cell(row=1, column=i, value=name)
    for r, row in enumerate(sorted(rows, key=lambda x: x.when), start=2):
        local = (as_utc(row.when) + tz_shift).replace(tzinfo=None, second=0, microsecond=0)
        cell = ws.cell(row=r, column=1, value=local)
        cell.number_format = "yyyy-mm-dd hh:mm"
        if labelled and row.label:
            ws.cell(row=r, column=2, value=row.label)
        for i, name in enumerate(cols, start=first):
            v = row.weights.get(name) if name in row.weights else row.adjust.get(name)
            if v:
                ws.cell(row=r, column=i, value=float(v) * 100.0)
    return (f"{EVENTS_SHEET}: {len(rows)} event(s) written whole, "
            f"{len(ccys)} currency column(s) ({', '.join(ccys) or 'none'}); "
            "the sheet is the book's table, not one pair's")


def _write_band_row(wb, sheet: str, pair: str, request: dict) -> None:
    """The pair's row on the BANDS sheet, replaced.  The header is the
    request's own keys, so the reader's column matching is against the same
    spelling ``to_request`` uses; a column the sheet already has keeps its
    place and any it lacks is appended."""
    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.create_sheet(sheet)
        ws.cell(row=1, column=1, value="pair")
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            header[str(v).strip().lower().replace(" ", "_")] = c
    if "pair" not in header:
        raise SessionError(f"sheet {sheet!r} has no 'pair' column; not touching it")
    row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=header["pair"]).value
        if v is not None and str(v).strip().upper() == pair:
            row = r
            break
    if row is None:
        row = ws.max_row + 1 if ws.max_row > 1 or ws.cell(row=1, column=1).value else 2
        # a fresh sheet reports max_row 1; an emptied one may report more
        while ws.cell(row=row, column=header["pair"]).value not in (None, ""):
            row += 1
    ws.cell(row=row, column=header["pair"], value=pair)
    for key, value in request.items():
        if key not in header:
            c = ws.max_column + 1
            ws.cell(row=1, column=c, value=key)
            header[key] = c
        if isinstance(value, bool):
            value = 1 if value else 0
        ws.cell(row=row, column=header[key], value=value)


#: ``=C2*3``, ``=1.85*D2``, ``=D2*1.75`` -- a cell reference and a constant,
#: in either order, which is every shape the shipped workbook's wing columns
#: use.  Anything else is reported rather than guessed at: a wing this cannot
#: read is a relationship somebody has to look at before it is thrown away.
_WING_FORMULA = re.compile(
    r"^\s*=?\s*(?:\$?([A-Z]{1,3})\$?(\d+)\s*\*\s*([0-9]*\.?[0-9]+)"
    r"|([0-9]*\.?[0-9]+)\s*\*\s*\$?([A-Z]{1,3})\$?(\d+))\s*$")


def migrate_wing_ratios(workbook, out=None, *, in_place: bool = False) -> dict:
    """Lift the pair sheets' wing formulas onto the ``WING_RATIOS`` tab.

    Run once, on a workbook that is becoming a database.  Until it is run, a
    pair sheet is a small spreadsheet model: ``ST 10D`` is ``=ST 25D * 3.25``
    and ``RR 10D`` is ``=RR 25D * 1.85``, a different multiple per pair and per
    tenor, and the tool cannot see any of it.  Writing a quote into such a
    sheet leaves the wing beside it holding a number computed from the cell
    that has just been replaced.

    So: read every multiplier, write them onto a tab the tool reads, and
    flatten the sheets to numbers.  Nothing is guessed -- a wing whose formula
    is not a constant times the 25-delta cell of its own row is reported and
    left exactly as it is, and the file is written only if every sheet was
    understood or the caller asked for it anyway.

    Returns ``{"written", "backup", "pairs", "ratios", "problems", "notes"}``.
    """
    import io
    import openpyxl
    from .marketdata import SMILE_COLUMNS

    src = Path(workbook)
    if not src.exists():
        raise SessionError(f"workbook not found: {src}")
    dst = Path(out) if out else (
        src if in_place else src.with_name(src.stem + EXPORT_SUFFIX + src.suffix))
    if dst.resolve() == src.resolve() and not in_place:
        raise SessionError(f"{src.name} is the workbook itself; writing into it needs "
                           "in_place (--in-place), or name another file")
    with src.open("rb") as fh:
        blob = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    values = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    cached = _formula_cache(wb, values)

    problems: list[str] = []
    notes: list[str] = []
    by_pair: dict[str, dict[str, dict[str, float | None]]] = {}
    pairs: list[str] = []
    for sheet in list(wb.sheetnames):
        found, _ = _quote_sheet(wb, sheet)
        if found is None:
            continue
        name, ws, col_of, expiry_col, row_of = found
        rows: dict[str, dict[str, float | None]] = {}
        readable = True
        for tenor, r in row_of.items():
            row: dict[str, float | None] = {}
            for wing, target in RATIO_WINGS.items():
                cell = ws.cell(row=r, column=col_of[target])
                text = cell.value
                if isinstance(text, str) and text.startswith("="):
                    pass
                elif _is_formula(text):
                    text = "=" + str(getattr(text, "text", "")).lstrip("=")
                else:
                    row[wing] = None          # typed in its own right
                    continue
                m = _WING_FORMULA.match(str(text))
                base_col = col_of[target.replace("_10", "_25")]
                if m:
                    col, rownum, factor = (m.group(1), m.group(2), m.group(3)) if m.group(1) \
                        else (m.group(5), m.group(6), m.group(4))
                    from openpyxl.utils import get_column_letter
                    if col == get_column_letter(base_col) and int(rownum) == r:
                        row[wing] = float(factor)
                        continue
                readable = False
                problems.append(
                    f"{name} {tenor}: {QUOTE_LABELS[target]} is {str(text)!r}, which is not a "
                    f"constant times the {QUOTE_LABELS[target.replace('_10', '_25')]} of its "
                    f"own row; it was left as it is and no ratio was written for it")
                row[wing] = None
            rows[tenor] = row
        if rows:
            by_pair[name.upper()] = rows
            pairs.append(name)
            if readable:
                n, probs, ns = _write_quote_cells(wb, values, name, {}, rows)
                problems.extend(probs)
                notes.extend(ns)
            else:
                notes.append(f"{name}: left as it is, because a wing on it could not be read")

    if not by_pair:
        raise SessionError(f"{src.name} has no pair sheets with the four quote columns "
                           f"({', '.join(SMILE_COLUMNS)})")
    notes.append(_write_wing_ratios_sheet(wb, by_pair))
    backup = ""
    if dst.exists() and dst.resolve() == src.resolve():
        bak = _backup_path(src)
        with open(bak, "wb") as fh:
            fh.write(blob)
        backup = str(bak)
        notes.append(f"the workbook as it was is kept at {bak.name}")
    _save_workbook(wb, dst, cached)
    found = sum(1 for rows in by_pair.values() for row in rows.values()
                if any(v is not None for v in row.values()))
    return {"written": str(dst), "backup": backup, "pairs": sorted(pairs),
            "ratios": found, "problems": problems, "notes": notes}


def _write_wing_ratios_sheet(wb, by_pair: dict) -> str:
    """Write the ``WING_RATIOS`` tab: pair, tenor, st, rr.

    Whole for the pairs written and merged with whatever is already there for
    the rest, because a multiplier belongs to the pair whose wings it shapes
    and a table half rewritten would leave another pair's rows describing a
    sheet that has since been flattened.  A wing with no multiple is written
    blank, which is how the tab says "quoted in its own right" -- not as a 1,
    which would be a multiple that happens to be one.
    """
    from .marketdata import _norm
    head = ("pair", "tenor", "st", "rr")
    keep: dict[tuple, dict] = {}
    if WING_RATIOS_SHEET in wb.sheetnames:
        old = wb[WING_RATIOS_SHEET]
        cols = {}
        header_row = 0
        for r in range(1, old.max_row + 1):
            names = {_norm(old.cell(row=r, column=c).value or ""): c
                     for c in range(1, old.max_column + 1)}
            if "pair" in names and "tenor" in names:
                cols, header_row = names, r
                break
        for r in range(header_row + 1, old.max_row + 1 if header_row else 1):
            pair = str(old.cell(row=r, column=cols["pair"]).value or "").strip().upper()
            tenor = str(old.cell(row=r, column=cols["tenor"]).value or "").strip().upper()
            if not pair or not tenor or pair.startswith("#"):
                continue
            row = {}
            for wing in RATIO_WINGS:
                c = cols.get(wing)
                v = old.cell(row=r, column=c).value if c else None
                row[wing] = float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) \
                    else None
            keep[(pair, tenor)] = row
        del wb[WING_RATIOS_SHEET]
    for pair, rows in by_pair.items():
        for tenor in [k for (p, k) in list(keep) if p == pair.upper()]:
            keep.pop((pair.upper(), tenor), None)
        for tenor, row in rows.items():
            keep[(pair.upper(), str(tenor).upper())] = row

    ws = wb.create_sheet(WING_RATIOS_SHEET)
    ws.cell(row=1, column=1, value="# how each tenor's 10-delta wings follow its 25-delta "
                                   "ones: ST 10d = ST 25d x st, RR 10d = RR 25d x rr. "
                                   "A blank is no multiple -- that wing is quoted itself.")
    for i, name in enumerate(head, start=1):
        ws.cell(row=2, column=i, value=name)
    r = 3
    for (pair, tenor), row in sorted(keep.items()):
        ws.cell(row=r, column=1, value=pair)
        ws.cell(row=r, column=2, value=tenor)
        for i, wing in enumerate(RATIO_WINGS, start=3):
            if row.get(wing) is not None:
                ws.cell(row=r, column=i, value=float(row[wing]))
        r += 1
    n = sum(1 for row in keep.values() if any(v is not None for v in row.values()))
    return (f"{WING_RATIOS_SHEET}: {len(keep)} row(s), {n} carrying a multiplier")


def _quote_sheet(wb, pair: str):
    """The pair's own quote sheet, its header columns and its tenor rows.

    ``(sheet_name, worksheet, {field: column}, expiry_column, {TENOR: row})``,
    or a reason it could not be found.
    """
    from .marketdata import SMILE_COLUMNS, _norm
    sheet = next((sn for sn in wb.sheetnames if sn.strip().upper() == pair.upper()), None)
    if sheet is None:
        return None, (f"{pair}: the workbook has no {pair!r} sheet, so its quotes were not "
                      f"written; add the sheet or take the pair out of CONFIG")
    ws = wb[sheet]
    header: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip():
            header.setdefault(_norm(v), c)
    missing = [k for k in SMILE_COLUMNS if k not in header]
    if missing:
        return None, (f"{pair}: sheet {sheet!r} has no {', '.join(missing)} column(s), so "
                      f"its quotes were not written")
    col_of = {f: header[label] for label, f in SMILE_COLUMNS.items() if f != "tenor"}
    expiry_col = header["expiry"]
    row_of: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=expiry_col).value
        key = "" if v is None else str(v).strip().upper()
        if key:
            row_of.setdefault(key, r)
    return (sheet, ws, col_of, expiry_col, row_of), ""


def _write_quote_cells(wb, values, pair: str, edits: dict,
                       ratios: dict | None = None) -> tuple[int, list[str], list[str]]:
    """Write the pair's quotes into its own sheet, as numbers, all of them.

    The one place a session writes a *quote* rather than a mark.  Everything
    else this module exports goes into a row the workbook keeps for it
    (``atm 1m``, ``shift rho25``) and leaves the sheet's own numbers alone;
    these are the sheet's own numbers, so an edited quote replaces the cell it
    came from and the workbook is then the database it is being used as.

    **The whole sheet is flattened, not only the cells that changed.**  A pair
    sheet's 10-delta columns were formulas over its 25-delta ones, and writing
    one cell of such a sheet leaves the wing beside it holding a value computed
    from the number that has just been replaced: the file is then internally
    inconsistent, and the next person to open it in Excel gets a wing that
    recalculates itself into something nobody marked.  There is no half of this
    that is safe, so every quote on a sheet this touches is written as the
    number the tool is using, and the relationships that produced them live on
    the ``WING_RATIOS`` tab where the tool can read them.

    ``ratios`` is ``{TENOR: {"st": float | None, "rr": ...}}`` in force for this
    pair.  A wing with a ratio is written as the 25-delta times it.

    A tenor the sheet does not quote is appended, and only with all four of its
    quotes (a ratio counts as supplying one): the reader refuses a half-quoted
    row and finding that out on the way back in is worse than being told here.
    """
    problems: list[str] = []
    notes: list[str] = []
    found, why = _quote_sheet(wb, pair)
    if found is None:
        return 0, [why], notes
    sheet, ws, col_of, expiry_col, row_of = found
    vs = values[sheet] if sheet in values.sheetnames else None
    ratios = ratios or {}

    def cached(row: int, f: str):
        """What the sheet holds for one quote today: the number Excel last
        computed for it, or the number typed there."""
        if vs is None:
            return None
        v = vs.cell(row=row, column=col_of[f]).value
        return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None

    n = 0
    over_formula = False
    tenors = list(row_of) + [str(t).strip().upper() for t in edits if
                             str(t).strip().upper() not in row_of]
    for key in tenors:
        fields = {}
        for raw, vals in edits.items():
            if str(raw).strip().upper() == key:
                fields = vals
        ratio = {}
        for raw, vals in ratios.items():
            if str(raw).strip().upper() == key:
                ratio = vals or {}
        row = row_of.get(key)
        # What every one of the four is worth, before anything is written:
        # the sheet's own number, the typed one over it, and the derived one
        # over that.  Read first and written after, because a derived wing
        # reads the 25-delta cell it is taken from.
        want: dict[str, float | None] = {}
        for f in QUOTE_FIELDS:
            want[f] = fields.get(f, cached(row, f) if row is not None else None)
        for wing, target in (("st", "st_10"), ("rr", "rr_10")):
            factor = ratio.get(wing)
            base = want[target.replace("_10", "_25")]
            if factor is not None and base is not None:
                want[target] = float(base) * float(factor)
        if row is None:
            absent = [QUOTE_LABELS[f] for f in QUOTE_FIELDS if want[f] is None]
            if absent:
                problems.append(
                    f"{pair} {key}: sheet {sheet!r} does not quote this tenor and "
                    f"{', '.join(absent)} {'is' if len(absent) == 1 else 'are'} blank, so no "
                    f"row was added; a quoted tenor is all four of its numbers or none")
                continue
            row = ws.max_row + 1
            ws.cell(row=row, column=expiry_col, value=key)
            row_of[key] = row
            notes.append(f"{pair}: {key} added to sheet {sheet!r} as a newly quoted tenor")
        for f in QUOTE_FIELDS:
            if want[f] is None:
                continue
            cell = ws.cell(row=row, column=col_of[f])
            if _is_formula(cell.value):
                over_formula = True
            cell.value = float(want[f])
            n += 1
    if over_formula:
        notes.append(f"{pair}: sheet {sheet!r} held formulas in its quote columns and is now "
                     f"numbers throughout. A quote written beside a formula that reads it "
                     f"would leave the two disagreeing, so the sheet is flattened whole; the "
                     f"relationship between the wings lives on {WING_RATIOS_SHEET} now")
    return n, problems, notes


def _is_formula(v) -> bool:
    """A cell openpyxl will write back as a formula.

    Not only ``"=C2*3"``: Excel saves the same expression as an *array*
    formula (``{=C2*3}``, the legacy CSE spelling, and what a dynamic-array
    build writes) and openpyxl hands that back as an ``ArrayFormula`` object,
    which is not a string and so was not recognised here.  Every array-formula
    cell therefore lost its cached value on export, and the smile sheets of
    the shipped workbook are entirely array formulas -- the copy came back
    with 126 "blank quote" problems, one per quote it should have carried.
    """
    if isinstance(v, str):
        return v.startswith("=")
    # Imported here for the same reason openpyxl itself is imported inside
    # ``export_workbook``: nothing else in this module needs it.
    from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
    return isinstance(v, (ArrayFormula, DataTableFormula))


def _formula_cache(wb, values) -> dict[int, dict[str, float]]:
    """Every numeric value Excel last computed for a formula cell, by
    worksheet position and coordinate.  Only numbers: the sheets the tool
    reads are numeric, and a string result would need the shared-string
    table rewritten, which is not worth carrying for a label."""
    out: dict[int, dict[str, float]] = {}
    for i, ws in enumerate(wb.worksheets, start=1):
        vs = values[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if _is_formula(cell.value):
                    cv = vs[cell.coordinate].value
                    if isinstance(cv, (int, float)) and not isinstance(cv, bool):
                        out.setdefault(i, {})[cell.coordinate] = float(cv)
    return out


# The formula element is kept whole rather than rebuilt, because it carries
# attributes: an array formula is ``<f t="array" ref="B2">C2*3</f>`` and a
# shared one can be ``<f t="shared" si="0"/>`` with no text at all.  Matching
# only a bare ``<f>`` skipped both, so the substitution silently did nothing.
# The empty value element openpyxl leaves behind is written as ``<v></v>``.
# This matched only the self-closing ``<v/>`` until 2026-09-02, so nothing
# substituted and every array-formula quote in the shipped workbook came back
# blank from an export -- the copy reloaded with one "blank quote" problem per
# cell, which is the whole failure this cache exists to prevent, arriving
# through a spelling.  Both forms are accepted now: which one openpyxl emits is
# openpyxl's business and not something to depend on -- it writes ``<v></v>``
# here, with and without lxml, on 3.1.5.  Two tests already covered this and
# both were red on the working copy, which is the shape of a dependency that
# moved under a check nobody re-ran.
_EMPTY_V = re.compile(
    r'<c r="([A-Z]+[0-9]+)"([^>]*)>(<f[^>]*/>|<f[^>]*>[^<]*</f>)(?:<v\s*/>|<v></v>)</c>')


def _restore_formula_cache(xlsx: bytes, cached: dict[int, dict[str, float]]) -> bytes:
    """Put the cached values back into the file openpyxl wrote.

    openpyxl writes ``<c r="B2"><f>C2*3</f><v /></c>``, and its worksheet
    files are numbered in ``wb.worksheets`` order, so this is one
    substitution per formula cell whose value is known.  A cell the export
    wrote a number into is no longer a formula and is not touched here.
    """
    import io
    import zipfile
    if not cached:
        return xlsx
    src = zipfile.ZipFile(io.BytesIO(xlsx))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as out:
        for item in src.infolist():
            data = src.read(item.filename)
            m = re.fullmatch(r"xl/worksheets/sheet(\d+)\.xml", item.filename)
            if m and int(m.group(1)) in cached:
                values = cached[int(m.group(1))]

                def fill(match):
                    v = values.get(match.group(1))
                    if v is None:
                        return match.group(0)
                    return (f'<c r="{match.group(1)}"{match.group(2)}>{match.group(3)}'
                            f'<v>{v!r}</v></c>')
                data = _EMPTY_V.sub(fill, data.decode("utf-8")).encode("utf-8")
            out.writestr(item, data)
    return buf.getvalue()


def _save_workbook(wb, dst: Path, cached: dict | None = None) -> None:
    """Atomic, like ``write``: a workbook is either the old one or the new
    one, never half of each."""
    import io
    dst.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".volkit-", suffix=".xlsx", dir=str(dst.parent))
    try:
        os.close(fd)
        buf = io.BytesIO()
        wb.save(buf)
        with open(tmp, "wb") as fh:
            fh.write(_restore_formula_cache(buf.getvalue(), cached or {}))
        os.replace(tmp, dst)
    except BaseException:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise
