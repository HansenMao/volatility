"""The workbook's configuration tabs.

The settings a desk maintains by hand -- the defended peg ranges, the kACE
pillars and the width posted at each, the lunar holidays no rule can derive --
used to be a loose CSV each, sitting beside ``vol_marks.xlsx``.  Four files is
four things to copy onto a new machine, four things to keep in step with the
marks they describe, and four things a desk can be missing one of without
finding out until the screen that needed it is empty.  They are **tabs of the
workbook** now, and this module is the one place that reads one.

A configuration tab is an ordinary table with a header row:

    pair     lower   upper   note
    USDHKD   7.75    7.85    HKMA Convertibility Undertakings

Two conventions, both inherited from the files these replace:

* **A row whose first cell starts with ``#`` is a comment**, wherever it
  appears.  The CSVs carried a paragraph each -- what the file was for, and in
  the bands' case which pairs were deliberately *not* listed and why -- and
  that reasoning is worth more inside the workbook a desk actually opens than
  in a file that has been deleted.  A desk can add its own the same way.
* **The header is found, not assumed to be row 1**, so those notes can sit
  above it.  The first row carrying every column the reader requires is the
  header; anything above it is prose.

Blank rows and blank trailing columns are ignored, and a column heading is
matched case- and space-insensitively (``Weak Share`` is ``weak_share``),
because a person maintaining a tab in Excel will capitalise it however reads
best and should not have to know which.

Adding a configuration later is a tab and a parser: name the tab in
:data:`SHEETS` so an error can say what it was for, and call :func:`read_rows`
for it.  Nothing else here changes, and no new file appears beside the exe.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from .paths import app_dir, find_data_file

#: The marks workbook, when nothing names one.  Configuration lives in the
#: same file as the marks it describes -- that is the whole point of the move.
WORKBOOK_FILENAME = "vol_marks.xlsx"

#: Every configuration tab, and what it holds.  One list, so a missing tab can
#: be reported by name and with its purpose, and so ``volkit check`` can say
#: which ones a workbook has without going looking for them.
SHEETS: dict[str, str] = {
    "PEG_BANDS": "managed / pegged trading bands: pair, lower, upper, note",
    "KACE_SPREADS": "the kACE pillars and the ATM width at each: pair, tenor, spread",
    "HOLIDAYS": "holiday dates no rule derives: country, date, remove",
    "CONVENTIONS": "a pair's quoting conventions where they differ from the market's: "
                   "pair, premium (the currency it is paid in), atmf beyond (a tenor, "
                   "never, or always), delta (spot or forward)",
    "RATES": "simple deposit rates for the discount factors a spot delta and a paid "
             "premium need: currency, tenor, rate (% p.a.)",
    "WING_RATIOS": "how each tenor's 10-delta wings follow its 25-delta ones: "
                   "pair, tenor, st, rr",
    "Vega Weights": "how far each tenor moves when the anchor moves one vol point: "
                    "tenor, default, and a column per pair that needs its own",
}

#: The columns each tab is edited with, and which of them decide where its
#: header is.  A tab the screens can edit has to have a fixed shape: the
#: reader matches headings by meaning and will take them in any order, but a
#: writer that guessed at the order would reorder a desk's own columns on
#: every save.  A tab not listed here is read-only -- it can be maintained in
#: Excel like anything else, and the screen says so rather than offering a
#: table it cannot write back.
EDITABLE: dict[str, tuple[tuple[str, ...], tuple[str, ...]]] = {
    "PEG_BANDS": (("pair", "lower", "upper", "note"), ("pair", "lower", "upper")),
    "KACE_SPREADS": (("pair", "tenor", "spread"), ("pair", "tenor")),
    "HOLIDAYS": (("country", "date", "remove"), ("country", "date")),
    "CONVENTIONS": (("pair", "premium", "atmf beyond", "delta"), ("pair",)),
    "RATES": (("currency", "tenor", "rate"), ("currency", "tenor", "rate")),
    "WING_RATIOS": (("pair", "tenor", "st", "rr"), ("pair", "tenor")),
    "Vega Weights": (("tenor", "default", "note"), ("tenor", "default")),
}

#: Tabs whose columns are not the whole list.  ``Vega Weights`` carries a
#: column per pair that has its own curve shape, and which pairs those are is
#: the desk's business rather than this module's -- so the fixed columns above
#: are the ones the tab must have and the ones a screen starts it with, and
#: whatever else is on the sheet (or in what a screen sends back) is carried
#: through.  Everywhere else a tab's columns are exactly what ``EDITABLE``
#: says: a writer that guessed would reorder a desk's own columns on a save.
OPEN_COLUMNS: frozenset[str] = frozenset({"Vega Weights"})

#: Columns an open tab keeps at the right-hand end however many it grows.
TRAILING: frozenset[str] = frozenset({"note"})


class ConfigSheetError(ValueError):
    """A configuration tab cannot be read, and this says why."""


def default_workbook() -> Path:
    """The workbook to read configuration from when nothing names one.

    Beside the exe, or ``files/`` in a source tree; the first that exists, and
    the path it *would* have when neither does, so an error names somewhere a
    person can look.
    """
    found = find_data_file(WORKBOOK_FILENAME, f"files/{WORKBOOK_FILENAME}")
    return found if found is not None else app_dir() / WORKBOOK_FILENAME


@dataclass(frozen=True)
class Row:
    """One data row of a configuration tab, and where it was.

    ``number`` is the row as Excel numbers it, because an error about a
    configuration tab is read by somebody who is about to go and fix it.
    """

    number: int
    cells: dict[str, object]
    sheet: str = ""

    def __contains__(self, key: str) -> bool:
        return normalise(key) in self.cells

    def raw(self, key: str):
        """One cell, by a heading spelled however the caller has it.

        Normalised on the way in because an open tab's columns are named by
        the desk -- a screen asking for ``USDJPY`` and a sheet that stored it
        as ``usdjpy`` are asking about one column.
        """
        return self.cells.get(normalise(key))

    def text(self, key: str) -> str:
        """The cell as trimmed text; ``""`` when it is blank or absent."""
        v = self.raw(key)
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
        if isinstance(v, float) and v.is_integer():
            # Excel hands back 2.0 for a cell showing 2; a tenor or a country
            # code that round-tripped through a number must not become "2.0".
            return str(int(v))
        return str(v).strip()

    def real(self, key: str) -> float | None:
        """The cell as a number, or ``None`` when blank; a bad cell is refused."""
        v = self.raw(key)
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            raise ConfigSheetError(
                f"{self.sheet or 'configuration'} row {self.number}: {key} is {v!r}, "
                f"which is not a number") from None

    def day(self, key: str) -> date | None:
        """The cell as a date, or ``None`` when blank.

        Excel gives a date cell back as a ``datetime`` and a text cell as
        text, and a desk will produce both in the same column -- one typed,
        one pasted.  Both are dates and are read as such.
        """
        v = self.raw(key)
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        text = str(v).strip()
        if not text:
            return None
        from .timeutil import parse_datetime
        try:
            return parse_datetime(text).date()
        except ValueError:
            raise ConfigSheetError(
                f"{self.sheet or 'configuration'} row {self.number}: {key} is {v!r}, "
                f"which is not a date") from None

    def flag(self, key: str) -> bool:
        """``1``/``true``/``yes``/``y`` and the column's own name are true."""
        text = self.text(key).strip().lower()
        return text in {"1", "true", "yes", "y", "on", key}


def normalise(heading) -> str:
    """A column heading as the readers spell it: lower case, underscores."""
    return str(heading).strip().lower().replace(" ", "_").replace("-", "_")


def check_open_columns(sheet: str, columns) -> None:
    """Refuse a column an open tab's own reader would refuse, before writing it.

    An open tab's extra columns are **pairs** -- that is what "open" means
    here.  Written without this, a mistyped heading is accepted, the tab's
    reader refuses it on the next load, and the tab is unreadable until
    somebody opens the workbook in Excel and fixes it by hand.  The screen
    validates the box too; this is the same rule on the route behind it.
    """
    if sheet not in OPEN_COLUMNS:
        return
    fixed = {normalise(c) for c in EDITABLE[sheet][0]}
    for column in columns:
        name = normalise(column)
        if name in fixed or (len(name) == 6 and name.isalpha()):
            continue
        raise ConfigSheetError(
            f"{sheet} has no column {column!r}. Its columns are "
            f"{', '.join(EDITABLE[sheet][0])}, and one per pair -- six letters, like "
            f"USDJPY")


def columns_for(sheet: str, rows=()) -> tuple[str, ...]:
    """The columns one tab is written with, fixed ones first.

    For an ordinary tab this is exactly what :data:`EDITABLE` declares.  For
    an open one it is that list plus every other column the rows carry, in the
    order they first appear -- which is how a pair column added on the screen
    reaches the sheet, and how one already on the sheet survives a save that
    did not mention it.
    """
    fixed = EDITABLE[sheet][0]
    if sheet not in OPEN_COLUMNS:
        return tuple(fixed)
    # ``note`` is written last whatever else the tab grows, because a comment
    # column wedged between the weights and the pairs is a column a desk
    # scrolls past on every row.
    head = [c for c in fixed if c not in TRAILING]
    tail = [c for c in fixed if c in TRAILING]
    seen = {c for c in fixed}
    for row in rows or ():
        for key in row:
            name = normalise(key)
            if not name or name in seen:
                continue
            seen.add(name)
            # A pair is a proper noun and is written as one: the reader
            # matches headings case-insensitively, and a desk reading the tab
            # in Excel should see USDJPY rather than usdjpy.
            head.append(name.upper() if len(name) == 6 and name.isalpha() else name)
    return tuple(head + tail)


def open_workbook(path: str | Path):
    """An openpyxl reader over the workbook, with the file itself already shut.

    The same reasoning as :func:`marketdata.open_workbook`, for the same
    reason: a reader left alive holds the handle until a garbage collection
    nobody schedules, and on Windows that is enough to stop Excel saving the
    very workbook this tool just read.  Reading the bytes first means the file
    is open for as long as it takes to copy it and never between calls.
    """
    import openpyxl

    p = Path(path)
    if not p.exists():
        raise ConfigSheetError(f"no workbook at {p}")
    try:
        with p.open("rb") as fh:
            blob = fh.read()
        return openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except ConfigSheetError:
        raise
    except Exception as exc:  # noqa: BLE001 - openpyxl raises a zoo of these
        raise ConfigSheetError(f"cannot open {p}: {exc}") from exc


def sheet_names(path: str | Path) -> list[str]:
    """Every tab in the workbook, in workbook order."""
    wb = open_workbook(path)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def present(path: str | Path) -> list[str]:
    """Which of the known configuration tabs this workbook actually has."""
    names = sheet_names(path)
    return [s for s in SHEETS if match_sheet(names, s) is not None]


def match_sheet(names, sheet: str) -> str | None:
    """The workbook's own spelling of a tab, or ``None`` if it has not got it.

    Exactly first, then ignoring case, spaces and underscores.  The tabs are
    named by desks in Excel and this one predates the tool that reads it --
    ``Vega Weights`` is what the desk called it -- so a workbook where
    somebody has since typed ``VEGA_WEIGHTS`` is the same tab and is read, and
    a tab written back keeps the name the file already has rather than
    acquiring a second one beside it.
    """
    want = normalise(sheet)
    for name in names:
        if name == sheet:
            return name
    for name in names:
        if normalise(name) == want:
            return name
    return None


def read_rows(path: str | Path, sheet: str, *,
              required: tuple[str, ...] = ()) -> list[Row] | None:
    """The data rows of a configuration tab, or ``None`` when there is no such tab.

    ``required`` names the columns the tab has to carry.  They decide where
    the header is -- the first row that has all of them -- and a tab that has
    none of them anywhere is refused by name rather than read as though its
    first row of prose were a heading.

    ``None`` and an empty list are different answers and both happen: no tab
    at all is a workbook that was never given this configuration, while a tab
    with a header and no rows is a desk that has deliberately emptied it.
    """
    wb = open_workbook(path)
    try:
        found = match_sheet(wb.sheetnames, sheet)
        if found is None:
            return None
        grid = [list(r) for r in wb[found].iter_rows(values_only=True)]
    finally:
        wb.close()

    want = tuple(normalise(c) for c in required)
    header: list[str] | None = None
    out: list[Row] = []
    for n, raw in enumerate(grid, start=1):
        cells = list(raw)
        while cells and (cells[-1] is None or str(cells[-1]).strip() == ""):
            cells.pop()
        if not cells:
            continue
        first = "" if cells[0] is None else str(cells[0]).strip()
        if first.startswith("#"):
            continue
        if header is None:
            names = [normalise(c) if c is not None else "" for c in cells]
            if all(w in names for w in want):
                header = names
            continue
        if not any(c is not None and str(c).strip() != "" for c in cells):
            continue
        row = {}
        for key, value in zip(header, cells):
            if not key:
                continue
            row[key] = value.strip() if isinstance(value, str) else value
            if row[key] == "":
                row[key] = None
        out.append(Row(number=n, cells=row, sheet=sheet))

    if header is None:
        raise ConfigSheetError(
            f"the {sheet!r} tab of {Path(path).name} has no header row naming "
            f"{', '.join(required)}; it holds {SHEETS.get(sheet, 'configuration')}")
    return out


def write_rows(wb, sheet: str, columns, rows, header=None) -> str:
    """Replace one configuration tab's data, keeping the prose above it.

    ``wb`` is an open openpyxl workbook -- the caller saves it, because a tab
    is written as part of a workbook write that has a backup and a staleness
    check around it and never on its own.

    The comment lines above the header are the reasoning a desk wrote down
    about the tab (:mod:`configsheets` reads them as comments precisely so
    they can be kept), so they are carried over verbatim.  Everything from the
    header down is replaced: a tab edited a row at a time by a screen is a tab
    whose rows have no identity worth preserving, and merging them by position
    is how a deleted row becomes an edited one.

    Returns a line for the report.
    """
    # Which columns identify the old header row.  Not ``columns``: an open
    # tab is written with whatever pair columns it has grown, and a header
    # that does not yet carry one of them is still the header -- looked for by
    # every column, a tab gaining its first pair column would find no header,
    # keep the "#" lines from *below* the data as prose and write them above
    # the new one.
    keep: list[str] = []
    at = len(wb.sheetnames)
    found = match_sheet(wb.sheetnames, sheet)
    if found is not None:
        sheet = found                       # keep the name the workbook has
        # ...and its place in the tab bar.  The tab is replaced rather than
        # edited, and one recreated at the end of the workbook is a tab that
        # jumped on somebody every time a setting was saved.
        at = wb.sheetnames.index(sheet)
        old = wb[sheet]
        want = {normalise(c) for c in (header or columns)}
        for r in range(1, old.max_row + 1):
            first = old.cell(row=r, column=1).value
            text = "" if first is None else str(first).strip()
            names = {normalise(old.cell(row=r, column=c).value or "")
                     for c in range(1, old.max_column + 1)}
            if want <= names:
                break                       # the header: everything below is data
            if text.startswith("#"):
                keep.append(text)
        del wb[sheet]
    ws = wb.create_sheet(sheet, at)
    r = 1
    for line in keep:
        ws.cell(row=r, column=1, value=line)
        r += 1
    for i, name in enumerate(columns, start=1):
        ws.cell(row=r, column=i, value=name)
    r += 1
    written = 0
    keys = [normalise(c) for c in columns]
    for row in rows:
        # Matched the way the reader matches: a heading written ``USDJPY``
        # and a row that came back keyed ``usdjpy`` are one column, and
        # looking it up by the spelling in the header would write the whole
        # column blank.
        cells = {normalise(k): v for k, v in row.items()}
        if not any(cells.get(k) not in (None, "") for k in keys):
            continue
        for i, key in enumerate(keys, start=1):
            v = cells.get(key)
            if v is not None and v != "":
                ws.cell(row=r, column=i, value=v)
        r += 1
        written += 1
    return f"{sheet}: {written} row(s)"


def missing(sheet: str, path: str | Path) -> str:
    """The one sentence said when a tab a reader needs is not there."""
    return (f"{Path(path).name} has no {sheet!r} tab. It holds "
            f"{SHEETS.get(sheet, 'configuration')}; add it as a tab of the workbook "
            f"with those column headings.")
