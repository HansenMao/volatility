"""Test suite for volkit.

Uses ``unittest`` rather than pytest so it runs on a bare Python install --
the same reason the web interface is stdlib-only.

    python3 -m unittest discover -s tests -v
"""

from __future__ import annotations

import math
import shutil
import textwrap
import unittest
import inspect as _inspect
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path

import numpy as np
import pandas as pd

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from volkit import black, sabr, smile
from volkit.atm import AtmCurve, BackboneParams
from volkit.black import DeltaConvention
from volkit.book import Book
from volkit.calendars import DEFAULT_CALENDARS, CalendarSet, easter
from volkit.cross import (CorrelationCurve, CrossAtmCurve, dollar_legs,
                          infer_leg_signs)
from volkit import exotics
from volkit.banded import Band, BetaBandSmile, JumpSpec, calibrate_band_smile, load_bands
from volkit.feed import FeedError, MarketFeed, pip_divisor
from volkit import analytics, history, listed, marketmaker, moments, quotes
from volkit.events import EventSchedule
from volkit.knowledge import KnowledgeBank, PairKnowledge, Rule, suggest_rules
from volkit import marketdata
from volkit.marketdata import ExcelSource, MarketDataError
from volkit.numerics import ConvergenceError, fixed_point, integrate_piecewise, solve_scalar
from volkit.pricing import (OptionLeg, StrikeSpec, expiry_datetime, parse_strike, price_strip,
                            quick_vol, resolve_expiry)
from volkit.smile import SmileSlice, fit_svi
from volkit.surface import PARAM_NAMES, SmileMark, VolSurface, fit_param_term_structure
from volkit.timeutil import (Clock, DAYS_IN_YEAR, TenorError, UTC, add_tenor,
                             normalise_tenor, parse_datetime, parse_tenor,
                             tenor_to_years)
from volkit.timeweight import DEFAULT_SESSION_HOURS, TimeWeighting, session_shares

def _source(*parts: str) -> str:
    """A file of this project, read as UTF-8.

    Never in the locale's encoding: that is cp1252 on the Windows box the
    build runs on, ``index.html`` is full of em dashes, and the whole suite
    ended there with "'charmap' codec can't decode byte 0x81".
    """
    return Path(__file__).resolve().parents[1].joinpath(*parts).read_text(encoding="utf-8")


WORKBOOK = Path(__file__).resolve().parents[1] / "files" / "vol_marks.xlsx"
ASOF = Clock(datetime(2024, 2, 28, 12, 0, tzinfo=UTC))


class TestTimeUtil(unittest.TestCase):
    def test_tenor_units(self):
        self.assertAlmostEqual(tenor_to_years("2Y"), 2.0)
        self.assertAlmostEqual(tenor_to_years("3M"), 0.25, places=6)
        self.assertAlmostEqual(tenor_to_years("5D"), 5 / 365.2425)

    def test_unknown_unit_raises(self):
        """Legacy get_years_time returned 1.0 for '1D' -- one year."""
        with self.assertRaises(TenorError):
            tenor_to_years("1X")
        self.assertLess(tenor_to_years("1D"), tenor_to_years("1W"))

    def test_month_end_clamping(self):
        self.assertEqual(add_tenor(date(2024, 1, 31), "1M"), date(2024, 2, 29))
        self.assertEqual(add_tenor(date(2023, 1, 31), "1M"), date(2023, 2, 28))
        self.assertEqual(add_tenor(date(2024, 11, 30), "3M"), date(2025, 2, 28))

    def test_clock_roundtrip(self):
        t = ASOF.years_to(datetime(2025, 2, 28, 12, tzinfo=UTC))
        self.assertAlmostEqual(ASOF.years_to(ASOF.datetime_from_years(t)), t, places=12)

    def test_naive_datetime_treated_as_utc(self):
        self.assertEqual(parse_datetime("2024-02-02 10:00").tzinfo, UTC)

    def test_a_unit_may_be_spelled_out(self):
        """'1wk' is a week and '3mth' three months.

        The unit used to be a single letter, so a desk that wrote its own
        shorthand into the expiry box was told its tenor could not be parsed.
        """
        for text, want in (("1wk", "1W"), ("1 wk", "1W"), ("1-week", "1W"),
                           ("2weeks", "2W"), ("3mth", "3M"), ("3 months", "3M"),
                           ("1mo", "1M"), ("2yr", "2Y"), ("2 years", "2Y"),
                           ("10days", "10D"), ("8d", "8D"), ("o/n", "O/N")):
            with self.subTest(text):
                self.assertEqual(normalise_tenor(text), want)
        self.assertAlmostEqual(tenor_to_years("1wk"), tenor_to_years("1W"))
        self.assertAlmostEqual(tenor_to_years("3mth"), tenor_to_years("3M"))
        # O/N is one day, which is what the code means.
        self.assertEqual(parse_tenor("o/n"), (1.0, "d"))
        with self.assertRaises(TenorError):
            tenor_to_years("1wkk")

    def test_a_date_with_no_year_is_the_next_one_of_it(self):
        """'06 Nov' is the coming sixth of November, not a parse error.

        The year is obvious to whoever typed it, and the reference date is
        the book's clock rather than the machine's, so the same box read
        twice reads the same way.
        """
        today = date(2026, 9, 1)
        for text in ("06 Nov", "06Nov", "6-Nov", "Nov 6", "November 6"):
            with self.subTest(text):
                self.assertEqual(parse_datetime(text, today=today).date(),
                                 date(2026, 11, 6))
        # Already gone this year, so it is next year's.
        self.assertEqual(parse_datetime("31 Aug", today=today).date(), date(2027, 8, 31))
        # Today itself matches: the horizon starts now.
        self.assertEqual(parse_datetime("01 Sep", today=today).date(), today)
        # A time of day survives.
        self.assertEqual(parse_datetime("06 Nov 15:00", today=today),
                         datetime(2026, 11, 6, 15, 0, tzinfo=UTC))
        # 29 February is the one day whose next occurrence is not within a
        # year; answering it beats refusing it.
        self.assertEqual(parse_datetime("29 Feb", today=today).date(), date(2028, 2, 29))
        # A year that is given still wins, and nothing that parsed moves.
        self.assertEqual(parse_datetime("15Sep26", today=today).date(), date(2026, 9, 15))

    def test_a_year_less_date_with_no_reference_says_what_is_missing(self):
        """It is never the wall clock: the clock is injected (§4)."""
        with self.assertRaises(ValueError) as caught:
            parse_datetime("06 Nov")
        self.assertIn("no year", str(caught.exception))
        # A purely numeric year-less date stays ambiguous and is refused:
        # '06/11' is a day and a month in one country and the reverse in
        # another, and there is nothing in it to say which.
        with self.assertRaises(ValueError):
            parse_datetime("06/11", today=date(2026, 9, 1))

    def test_the_clock_is_what_says_which_year(self):
        from volkit.timeutil import Clock
        clock = Clock(datetime(2026, 9, 1, 12, tzinfo=UTC))
        self.assertEqual(clock.coerce_datetime("06 Nov").date(), date(2026, 11, 6))


class TestNumerics(unittest.TestCase):
    def test_bracketed_solve(self):
        self.assertAlmostEqual(solve_scalar(lambda x: x * x - 2, 1.0, lo_bound=0), math.sqrt(2))

    def test_unattainable_target_raises(self):
        with self.assertRaises(ConvergenceError):
            solve_scalar(lambda x: x * x + 1.0, 1.0)

    def test_fixed_point_detects_divergence(self):
        """Legacy loops ran a fixed 10 iterations and returned the last value."""
        self.assertAlmostEqual(fixed_point(math.cos, 1.0), 0.7390851332151607, places=8)
        with self.assertRaises(ConvergenceError):
            fixed_point(lambda x: 2 * x + 1, 1.0, max_iter=30)

    def test_piecewise_integration_is_exact_on_discontinuities(self):
        f = lambda x: np.where(x < 1.0, 1.0, 3.0)
        self.assertAlmostEqual(integrate_piecewise(f, np.array([0.0, 1.0, 2.0])), 4.0, places=12)
        self.assertAlmostEqual(integrate_piecewise(np.sin, np.linspace(0, math.pi, 9)), 2.0, places=10)


class TestBlack(unittest.TestCase):
    def test_strike_delta_roundtrip_all_conventions(self):
        n = 0
        for vol in (0.02, 0.05, 0.12, 0.30):
            for t in (1 / 365, 0.02, 0.5, 2.0):
                for pa in (False, True):
                    for d, call in ((0.25, True), (-0.25, False), (0.10, True), (-0.10, False)):
                        k = black.strike_from_delta(d, 1.0, vol, t, call, pa)
                        # 1e-10 is far tighter than any quote resolution; the
                        # residual is the strike solver's xtol, not model error.
                        self.assertAlmostEqual(float(black.delta(1.0, k, vol, t, call, pa)), d, places=10)
                        n += 1
        self.assertGreater(n, 100)

    def test_dns_strike_is_delta_neutral(self):
        for pa in (False, True):
            for vol in (0.05, 0.30):
                k = black.dns_strike(1.0, vol, 0.5, pa)
                total = (float(black.delta(1.0, k, vol, 0.5, True, pa))
                         + float(black.delta(1.0, k, vol, 0.5, False, pa)))
                self.assertAlmostEqual(total, 0.0, places=14)

    def test_premium_adjusted_call_peak_matches_brute_force(self):
        """The pa call delta is non-monotone; the legacy solver could land
        on the wrong branch."""
        ks = np.linspace(0.01, 8.0, 400_000)
        d = black.delta(1.0, ks, 0.30, 10.0, True, True)
        _, peak = black._pa_call_delta_peak(1.0, 0.30, 10.0)
        self.assertAlmostEqual(peak, float(d.max()), places=5)

    def test_unattainable_premium_adjusted_delta_raises(self):
        with self.assertRaises(ConvergenceError):
            black.strike_from_delta(0.45, 1.0, 0.30, 10.0, True, True)

    def test_put_call_parity(self):
        c = float(black.price(1.0, 1.05, 0.11, 0.5, True))
        p = float(black.price(1.0, 1.05, 0.11, 0.5, False))
        self.assertAlmostEqual(c - p, 1.0 - 1.05, places=13)

    def test_implied_vol_inverts_price(self):
        px = float(black.price(1.0, 1.05, 0.11, 0.5, True))
        self.assertAlmostEqual(black.implied_vol(px, 1.0, 1.05, 0.5, True), 0.11, places=11)

    def test_rejects_bad_inputs(self):
        with self.assertRaises(ValueError):
            black.dns_strike(1.0, -0.1, 0.5)
        with self.assertRaises(ValueError):
            black.strike_from_delta(1.5, 1.0, 0.1, 0.5, True)

    def test_theta_vanna_and_volga_match_a_finite_difference(self):
        """Closed forms, pinned against the price they are derivatives of.

        These are what the exchange-traded positions panel reports as the
        Black-Scholes column, so an algebra slip in one of them would be a
        plausible-looking risk number with nothing to contradict it.
        """
        h = 1e-6
        for F, K, v, t in ((1.09, 1.12, 0.085, 0.37), (1.09, 1.09, 0.075, 0.04),
                           (150.0, 141.0, 0.11, 1.4)):
            for call in (True, False):
                # theta is the derivative with respect to *calendar* time, so
                # against t (time to expiry) it comes back with a sign change.
                fd = -(float(black.price(F, K, v, t + h, call))
                       - float(black.price(F, K, v, t - h, call))) / (2 * h)
                self.assertAlmostEqual(float(black.theta(F, K, v, t)) / fd, 1.0, places=6)
                fd = (float(black.delta(F, K, v + h, t, call))
                      - float(black.delta(F, K, v - h, t, call))) / (2 * h)
                self.assertAlmostEqual(float(black.vanna(F, K, v, t)) / fd, 1.0, places=6)
            fd = (float(black.vega(F, K, v + h, t)) - float(black.vega(F, K, v - h, t))) / (2 * h)
            self.assertAlmostEqual(float(black.volga(F, K, v, t)) / fd, 1.0, places=6)


class TestSabr(unittest.TestCase):
    def test_z_over_x_is_continuous_at_the_money(self):
        z = np.array([-1e-3, -1e-8, 0.0, 1e-8, 1e-3])
        vals = sabr._z_over_x(z, -0.4)
        self.assertTrue(np.all(np.isfinite(vals)))
        self.assertAlmostEqual(float(vals[2]), 1.0, places=14)
        self.assertLess(abs(float(vals[1]) - float(vals[3])), 1e-6)

    def test_atm_closed_form(self):
        p = sabr.SabrParams(0.09, -0.3, 0.8, 1.0)
        self.assertAlmostEqual(sabr.atm_vol(p), float(sabr.lognormal_vol(1.0, p)), places=15)

    def test_calibration_reprices_quotes(self):
        for rr in (-0.025, 0.0, 0.025):
            cal = sabr.calibrate(0.07, rr, 0.01, 0.10, 1.0, False)
            self.assertTrue(cal.converged, cal.message)
            _, cv = sabr.smile_strike_and_vol(cal.params, 0.10, 1.0, True, False)
            _, pv = sabr.smile_strike_and_vol(cal.params, -0.10, 1.0, False, False)
            self.assertAlmostEqual(cv - pv, rr, places=8)

    def test_risk_reversal_sign_convention(self):
        """Positive RR means calls over, which must give a positive rho."""
        self.assertGreater(sabr.calibrate(0.07, 0.025, 0.01, 0.25, 1.0, False).params.rho, 0)
        self.assertLess(sabr.calibrate(0.07, -0.025, 0.01, 0.25, 1.0, False).params.rho, 0)

    def test_smile_strike_converges_or_raises(self):
        p = sabr.SabrParams(0.09, -0.3, 0.8, 1.0)
        k, v = sabr.smile_strike_and_vol(p, 0.25, 1.0, True, False)
        self.assertAlmostEqual(float(black.delta(1.0, k, v, 1.0, True, False)), 0.25, places=9)


class TestSmile(unittest.TestCase):
    def setUp(self):
        self.t = 0.25
        self.atm = 0.0765
        self.c25 = sabr.calibrate(self.atm, -0.0024, 0.0024, 0.25, self.t, True).params
        self.c10 = sabr.calibrate(self.atm, -0.0044, 0.0072, 0.10, self.t, True).params
        self.slice = SmileSlice.build(self.t, self.atm, self.c25, self.c10, True)

    def test_svi_reproduces_its_anchors(self):
        got = np.asarray(self.slice.vol(self.slice.strikes), dtype=float)
        np.testing.assert_allclose(got, self.slice.vols, atol=1e-9)

    def test_svi_is_arbitrage_free(self):
        self.assertTrue(self.slice.svi.arbitrage_free, self.slice.svi.warnings)
        self.assertEqual(self.slice.svi.params.violates(), [])
        g = self.slice.svi.params.durrleman(np.linspace(-1.5, 1.5, 500))
        self.assertGreaterEqual(float(np.min(g)), -1e-8)

    def test_svi_uses_five_parameters_not_twelve(self):
        """The legacy fit had 12 free parameters for 5 points."""
        p = self.slice.svi.params
        self.assertEqual(len({"a", "b", "rho", "m", "sigma"} & set(vars(p))), 5)

    def test_all_interpolators_agree_at_the_anchors(self):
        for method in ("SVI", "VV25", "VV10", "SABR25", "SABR10"):
            sl = SmileSlice.build(self.t, self.atm, self.c25, self.c10, True, method=method)
            self.assertAlmostEqual(float(sl.vol(sl.strikes[2])), self.atm, delta=2e-3)

    def test_vanna_volga_survives_degenerate_strike(self):
        """Legacy getVV divided by d1*d2, which vanishes at two strikes."""
        v = smile.vanna_volga_vol(np.array([1.0, 1.02]), 0.25, 0.97, 1.0, 1.03, 0.08, 0.076, 0.079)
        self.assertTrue(np.all(np.isfinite(v)))


class TestAtmCurve(unittest.TestCase):
    def setUp(self):
        self.curve = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0), ASOF)

    def test_integration_is_exact_on_its_panels(self):
        """Splitting at the hourly/event breakpoints makes each panel smooth,
        so a 5-point rule is already exact -- raising the order changes
        nothing.  The legacy code ran adaptive quad with limit=500 on the
        discontinuous integrand instead."""
        for T in (0.02, 0.25, 1.0):
            self.curve._int_cache.clear()
            self.curve.quad_order = 5
            low = self.curve.integrated_variance(T)
            self.curve._int_cache.clear()
            self.curve.quad_order = 20
            high = self.curve.integrated_variance(T)
            self.curve._int_cache.clear()
            self.curve.quad_order = 5
            self.assertLess(abs(low - high) / high, 1e-14)

    def test_integration_matches_dense_reference(self):
        # A uniform trapezoid rule converges only slowly across the hourly
        # jumps, so its own error sets the tolerance here.
        for T in (0.02, 0.25, 1.0):
            fast = self.curve.integrated_variance(T)
            ts = np.linspace(0, T, 8_000_001)
            ref = float(np.trapezoid(self.curve.instantaneous_variance(ts), ts))
            self.assertLess(abs(fast - ref) / ref, 5e-7)

    def test_term_structure_is_monotone_in_the_right_direction(self):
        v_1w = self.curve.term_vol(tenor_to_years("1w"))
        v_1y = self.curve.term_vol(tenor_to_years("1y"))
        self.assertLess(v_1w, v_1y)  # initial 6.05 below long term 7.65

    def test_neighbour_tenors_outside_range(self):
        """Legacy argmax(tenors > t) returned (last, first) past the end."""
        self.assertEqual(self.curve._neighbour_tenors(5.0), (None, None))
        self.assertEqual(self.curve._neighbour_tenors(1e-6), (None, None))
        self.assertEqual(self.curve._neighbour_tenors(0.3), ("3m", "6m"))

    def test_tenor_overwrite_is_honoured(self):
        """And it is anchored where the tenor actually is on the calendar.

        Read at ``tenor_to_years`` the overwrite is a hair off, because that
        is not where 3M is: a month is 30 or 31 days, not a nominal 30.44.
        ``tenor_years`` is the curve's own reading, and it is the one the
        marks and the pricer both use.
        """
        t3 = self.curve.tenor_years("3m")
        self.curve.overwrite_tenor("3m", 0.09)
        self.assertAlmostEqual(self.curve.term_vol(t3), 0.09, places=8)
        self.assertNotAlmostEqual(self.curve.term_vol(tenor_to_years("3m")), 0.09, places=8)
        self.curve.clear_overwrite("3m")
        self.assertNotAlmostEqual(self.curve.term_vol(t3), 0.09, places=4)

    def test_weekend_volatility_is_damped(self):
        sat = self.curve.daily_vol(datetime(2024, 3, 2, 12, tzinfo=UTC))
        wed = self.curve.daily_vol(datetime(2024, 3, 6, 12, tzinfo=UTC))
        self.assertLess(sat, wed)

    def test_zero_length_window_does_not_divide_by_zero(self):
        self.assertEqual(self.curve.integrated_variance(0.5, 0.5), 0.0)
        with self.assertRaises(ValueError):
            self.curve.integrated_vol(0.5, 0.5)

    def test_backbone_cross_term_carries_time(self):
        """The legacy expression omitted t from the rate-vol cross term."""
        p = BackboneParams(0.06, 0.08, 5.0, rate_vol=0.02, rate_corr=0.5)
        c = AtmCurve("USDJPY", p, ASOF)
        sigma = 0.08 - (0.08 - 0.06) * math.exp(-5.0 * 0.5)
        expected = math.sqrt(sigma**2 + 2 * 0.5 * sigma * 0.02 * 0.5 + (0.02 * 0.5) ** 2)
        self.assertAlmostEqual(float(c.backbone_vol(0.5)), expected, places=14)

    def test_rate_vol_zero_is_unaffected(self):
        p = BackboneParams(0.06, 0.08, 5.0, rate_vol=0.0, rate_corr=0.9)
        c = AtmCurve("USDJPY", p, ASOF)
        sigma = 0.08 - (0.08 - 0.06) * math.exp(-5.0 * 0.5)
        self.assertAlmostEqual(float(c.backbone_vol(0.5)), sigma, places=15)


class TestEvents(unittest.TestCase):
    def test_calibrated_height_reproduces_the_quoted_bump(self):
        """The bump is quoted over the event's own 24 hours, so that is where
        it must be delivered exactly."""
        curve = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0), ASOF,
                         events=EventSchedule())
        when = datetime(2024, 3, 20, 18, 0, tzinfo=UTC)
        ev = curve.events.add(when, 0.030, "FOMC")
        self.assertEqual(curve.calibrate_events(), [])
        self.assertAlmostEqual(curve.achieved_bump(ev), 0.030, places=9)

    def test_event_height_is_stable_across_the_day_roll(self):
        """Under the legacy vol-day reading an event a minute before the 14:00
        roll needed a 12x height and dumped the overflow on the next day."""
        heights = {}
        for minute, mode in ((59, "forward24h"), (1, "forward24h")):
            curve = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0),
                             ASOF, events=EventSchedule(window_mode=mode))
            hour = 13 if minute == 59 else 14
            curve.events.add(datetime(2024, 3, 20, hour, minute, tzinfo=UTC), 0.02, "E")
            curve.calibrate_events()
            heights[minute] = curve.events.events[0].height
        self.assertLess(abs(heights[59] / heights[1] - 1.0), 0.10)

    def test_clustered_events_each_deliver_their_own_bump(self):
        """Solved independently, two nearby events each stop delivering their
        quote once the other is switched on."""
        curve = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0), ASOF,
                         events=EventSchedule())
        a = curve.events.add(datetime(2024, 3, 20, 16, 0, tzinfo=UTC), 0.02, "A")
        b = curve.events.add(datetime(2024, 3, 20, 18, 0, tzinfo=UTC), 0.03, "B")
        curve.calibrate_events()
        self.assertAlmostEqual(curve.achieved_bump(a), 0.02, places=8)
        self.assertAlmostEqual(curve.achieved_bump(b), 0.03, places=8)

    def test_weekend_and_holiday_events_are_flagged(self):
        curve = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0), ASOF,
                         events=EventSchedule())
        curve.events.add(datetime(2024, 3, 23, 16, 0, tzinfo=UTC), 0.02, "SAT")
        problems = curve.calibrate_events()
        self.assertTrue(any("weekly market closure" in p for p in problems), problems)

    def test_event_does_not_leak_into_unrelated_days(self):
        curve = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0), ASOF,
                         events=EventSchedule())
        base = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0), ASOF)
        curve.events.add(datetime(2024, 3, 20, 18, 0, tzinfo=UTC), 0.030, "FOMC")
        curve.calibrate_events()
        far = datetime(2024, 3, 25, 12, tzinfo=UTC)
        self.assertAlmostEqual(curve.daily_vol(far), base.daily_vol(far), places=12)

    def test_two_legs_weights_add_and_the_adjustment_sits_on_top(self):
        """An event is weighted per currency; a pair's bump is its two legs
        superposed plus the pair's own adjustment.  They add: a bump is a
        variance increment over twice the volatility, so two bumps add to
        first order, and a root-sum-square would be the rule for two event
        volatilities, which a bump is not."""
        from volkit.events import EventEntry, leg_weights, pair_bump, superpose
        self.assertEqual(superpose(0.015, 0.003), 0.018)
        self.assertAlmostEqual(pair_bump({"USD": 0.015, "JPY": 0.003}, "USDJPY", 0.002), 0.020)
        # A leg the table does not name weighs nothing; EURUSD sees only USD.
        self.assertAlmostEqual(pair_bump({"USD": 0.015, "JPY": 0.003}, "EURUSD"), 0.015)
        # CNH and CNY are one market for this purpose, unless CNY was named.
        self.assertEqual(leg_weights({"CNH": 0.02}, "USDCNY"), {"USD": 0.0, "CNY": 0.02})
        self.assertEqual(leg_weights({"CNH": 0.02, "CNY": 0.01}, "USDCNY"), {"USD": 0.0, "CNY": 0.01})
        # An entry typed as one number is that number, carried as the
        # adjustment, so ``bump == legs + adjust`` holds for every event.
        one = EventEntry(datetime(2026, 9, 1, tzinfo=UTC), 0.02, "x").resolve("EURUSD")
        self.assertEqual((one.bump, one.adjust, one.weights), (0.02, 0.02, {"EUR": 0.0, "USD": 0.0}))
        parts = EventEntry(datetime(2026, 9, 1, tzinfo=UTC), None, "x",
                           {"USD": 0.015}, 0.002).resolve("USDJPY")
        self.assertAlmostEqual(parts.bump, 0.017)
        self.assertEqual(parts.weights, {"USD": 0.015, "JPY": 0.0})
        # A total that disagrees with its parts is refused, not averaged.
        with self.assertRaises(ValueError):
            EventEntry(datetime(2026, 9, 1, tzinfo=UTC), 0.05, "x", {"USD": 0.015}, 0.002).resolve("USDJPY")

    def test_weighted_event_calibrates_to_the_superposed_bump(self):
        from volkit.events import EventEntry
        curve = AtmCurve("USDJPY", BackboneParams(0.0605, 0.0765, 5.0, 0.007, 50.0), ASOF,
                         events=EventSchedule())
        when = datetime(2024, 3, 20, 18, 0, tzinfo=UTC)
        problems = curve.set_events([EventEntry(when, None, "FOMC", {"USD": 0.015, "JPY": 0.003}, 0.002)])
        self.assertEqual(problems, [])
        ev = curve.events.events[0]
        self.assertAlmostEqual(ev.bump, 0.020)
        self.assertAlmostEqual(ev.adjust, 0.002)
        self.assertAlmostEqual(curve.achieved_bump(ev), 0.020, places=9)
        # The old three-tuple spelling still works and reads as an adjustment.
        curve.set_events([(when, 0.01, "T")])
        self.assertEqual(curve.events.events[0].adjust, 0.01)

    def test_event_rows_in_points_are_read_once(self):
        """The panel and the session file post the parts; a row may carry the
        total too when it agrees, and a bad row is named rather than fatal."""
        from volkit.events import event_entries
        rows = [{"when": "2026-09-16T18:00", "weights": {"USD": 1.5, "JPY": 0.3}, "adjust": 0.2,
                 "label": "FOMC"},
                {"when": "2026-09-16T18:00", "weights": {"USD": 1.5}, "adjust": 0.0, "bump": 1.5},
                {"when": "2026-10-01T12:00", "bump": 0.75},
                {"when": "2026-10-01T12:00", "weights": {"USD": 1.5}, "adjust": 0.0, "bump": 2.0},
                {"when": "", "bump": 1.0}]
        entries, problems = event_entries(rows)
        self.assertEqual(len(entries), 3)
        self.assertAlmostEqual(entries[0].resolve("USDJPY").bump, 0.020)
        self.assertAlmostEqual(entries[2].resolve("USDJPY").adjust, 0.0075)
        self.assertEqual(len(problems), 2)
        self.assertTrue(any("not the weights" in p for p in problems), problems)
        self.assertTrue(any("no date/time" in p for p in problems), problems)

    def test_addon_is_vectorised(self):
        sched = EventSchedule()
        ev = sched.add(datetime(2024, 3, 1, 12, tzinfo=UTC), 0.02, "x")
        ev.height = 0.5
        sched.refresh(ASOF)
        t = np.linspace(0, 0.05, 100)
        self.assertEqual(sched.addon(t).shape, t.shape)


class TestCalendars(unittest.TestCase):
    def test_easter(self):
        self.assertEqual(easter(2024), date(2024, 3, 31))
        self.assertEqual(easter(2025), date(2025, 4, 20))
        self.assertEqual(easter(2027), date(2027, 3, 28))

    def test_holiday_lookup(self):
        c = CalendarSet()
        self.assertTrue(c.is_holiday("USDJPY", date(2024, 7, 4)))
        self.assertIn("JP", c.holiday_countries("USDJPY", date(2024, 4, 29)))
        self.assertFalse(c.is_holiday("EURGBP", date(2024, 7, 4)))

    def test_overrides_add_dates_without_code_changes(self):
        """Legacy left '# manually add Chinese holidays' as a TODO."""
        c = CalendarSet()
        self.assertFalse(c.is_holiday("USDCNH", date(2025, 1, 29)))
        c.add_overrides("CN", ["2025-01-29"])
        self.assertTrue(c.is_holiday("USDCNH", date(2025, 1, 29)))

    def test_spot_and_expiry(self):
        c = CalendarSet()
        self.assertEqual(c.spot_lag("USDCAD"), 1)
        self.assertEqual(c.spot_lag("USDJPY"), 2)
        exp = c.expiry_date("USDJPY", "1M", date(2024, 2, 28))
        self.assertTrue(c.is_business_day("USDJPY", exp))


class TestFxDateConventions(unittest.TestCase):
    """The market's own construction: settlement first, expiry back from it.

    Pinned in detail because every one of these was arrived at by getting it
    wrong first, and because a date that is a day out is a whole day of
    volatility and two business days of swap points.
    """

    def setUp(self):
        self.c = CalendarSet()

    def test_the_settlement_date_is_the_anchor_and_the_expiry_comes_from_it(self):
        d = self.c.fx_dates("EURUSD", "1M", date(2026, 9, 1))
        self.assertEqual(d.spot, date(2026, 9, 3))       # T+2
        self.assertEqual(d.delivery, date(2026, 10, 5))  # 3 Oct is a Saturday
        self.assertEqual(d.expiry, date(2026, 10, 1))    # the spot lag back
        # and the two are consistent both ways round
        self.assertEqual(self.c.delivery_from_expiry("EURUSD", d.expiry), d.delivery)
        self.assertEqual(self.c.expiry_from_delivery("EURUSD", d.delivery), d.expiry)

    def test_a_us_holiday_rules_out_a_value_date_but_does_not_stop_the_count(self):
        """The half of the spot convention that is easiest to conflate.

        Counting US holidays as non-business days for a pair with no dollar in
        it would push EURJPY spot out a day every Thanksgiving, which is not
        what the market does.  The date the count lands on must still be one
        USD can settle on, because every FX trade settles through New York.
        """
        thanksgiving = date(2026, 11, 26)
        self.assertTrue(self.c.is_holiday("USD", thanksgiving))
        self.assertFalse(self.c.is_holiday("EURJPY", thanksgiving))
        # EUR and JPY are both open on the 26th, so it is the second of the two
        # counted days -- and is then rolled off because USD is shut.
        self.assertTrue(self.c.is_business_day("EURJPY", thanksgiving))
        self.assertFalse(self.c.is_settlement_day("EURJPY", thanksgiving))
        self.assertEqual(self.c.spot_date("EURJPY", date(2026, 11, 24)),
                         date(2026, 11, 27))

    def test_the_end_of_month_rule(self):
        """Off a month-end spot, a month tenor settles on a month end.

        Without it a 1M dealt off a 28-Feb spot settles 28-Mar where the
        market settles 31-Mar, and the expiry is then a business day early.
        """
        d = self.c.fx_dates("EURUSD", "1M", date(2026, 2, 25))
        self.assertEqual(d.spot, date(2026, 2, 27))      # the last value date of Feb
        self.assertTrue(self.c.is_month_end("EURUSD", d.spot))
        self.assertEqual(d.delivery, date(2026, 3, 31))  # not 27 Mar
        self.assertIn("end-of-month", d.rule)
        # and it carries down the curve, not just to the first month
        self.assertEqual(self.c.delivery_date("EURUSD", "3M", date(2026, 2, 25)),
                         date(2026, 5, 29))

    def test_a_spot_that_is_not_a_month_end_takes_the_ordinary_roll(self):
        d = self.c.fx_dates("EURUSD", "1M", date(2026, 1, 29))
        self.assertFalse(self.c.is_month_end("EURUSD", d.spot))
        self.assertNotIn("end-of-month", d.rule)
        self.assertEqual(d.delivery, date(2026, 3, 2))

    def test_a_day_tenor_is_business_days_from_the_trade_date(self):
        """And so the short tenors stay distinct.

        Adding calendar days to the spot date and taking the spot lag off the
        end -- the old construction -- collapses them: the two days subtracted
        swallow the weekend the addition just crossed.  Dealt on a Wednesday,
        "1D" and "2D" both came back Thursday.
        """
        wed = date(2026, 9, 2)
        got = [self.c.expiry_date("EURUSD", f"{n}D", wed) for n in (1, 2, 3, 4)]
        # Thu, Fri, then over the weekend *and* over US Labor Day on the 7th
        self.assertTrue(self.c.is_holiday("EURUSD", date(2026, 9, 7)))
        self.assertEqual(got, [date(2026, 9, 3), date(2026, 9, 4),
                               date(2026, 9, 8), date(2026, 9, 9)])
        self.assertEqual(len(set(got)), 4)

    def test_overnight_is_one_business_day_settling_from_its_own_spot(self):
        d = self.c.fx_dates("EURUSD", "O/N", date(2026, 9, 1))
        self.assertEqual(d.expiry, date(2026, 9, 2))
        self.assertEqual(d.delivery, date(2026, 9, 4))
        # "1D" is the same dates under a different name
        one = self.c.fx_dates("EURUSD", "1D", date(2026, 9, 1))
        self.assertEqual((one.expiry, one.delivery), (d.expiry, d.delivery))

    def test_usdcad_settles_t_plus_one_all_the_way_through(self):
        d = self.c.fx_dates("USDCAD", "1M", date(2026, 9, 1))
        self.assertEqual(self.c.spot_lag("USDCAD"), 1)
        self.assertEqual(d.spot, date(2026, 9, 2))
        self.assertEqual(self.c.add_business_days("USDCAD", d.expiry, 1), d.delivery)

    def test_a_pair_s_own_holiday_moves_its_settlement_and_not_another_pair_s(self):
        """3 Nov 2026 is Culture Day in Japan and an ordinary Tuesday elsewhere."""
        self.assertEqual(self.c.delivery_date("EURUSD", "2M", date(2026, 9, 1)),
                         date(2026, 11, 3))
        self.assertEqual(self.c.delivery_date("USDJPY", "2M", date(2026, 9, 1)),
                         date(2026, 11, 4))

    def test_the_short_date_codes_parse(self):
        for text, expect in (("O/N", (1.0, "d")), ("on", (1.0, "d")),
                             ("T/N", (2.0, "d")), ("s/n", (3.0, "d")),
                             ("S/W", (1.0, "w"))):
            self.assertEqual(parse_tenor(text), expect, text)
        self.assertEqual(normalise_tenor("o/n"), "O/N")
        self.assertEqual(normalise_tenor("3m"), "3M")
        self.assertEqual(normalise_tenor("1D"), "1D")   # not folded into O/N

    def test_expiry_years_is_the_calendar_and_not_the_nominal_length(self):
        """The reading that moved every mark; see MIGRATION.md 1.6."""
        clock = Clock(datetime(2026, 9, 1, 12, tzinfo=UTC))
        t = self.c.expiry_years("EURUSD", "1M", clock)
        expected = clock.years_to(datetime(2026, 10, 1, tzinfo=UTC))
        self.assertAlmostEqual(t, expected, places=12)
        self.assertNotAlmostEqual(t, tenor_to_years("1M"), places=4)


class TestTimeWeighting(unittest.TestCase):
    def test_session_shares_sum_to_one(self):
        total = sum(session_shares(DEFAULT_SESSION_HOURS).values())
        np.testing.assert_allclose(total, np.ones(24))

    def test_us_holiday_only_damps_us_hours(self):
        """The legacy 6x24 matrix could not express new calendar combinations."""
        tw = TimeWeighting("USDJPY")
        july4_ny = datetime(2024, 7, 4, 14, tzinfo=UTC)
        july4_tok = datetime(2024, 7, 4, 2, tzinfo=UTC)
        normal_ny = datetime(2024, 7, 2, 14, tzinfo=UTC)
        self.assertLess(tw.weight_at_datetime(july4_ny), tw.weight_at_datetime(normal_ny))
        self.assertAlmostEqual(tw.weight_at_datetime(july4_tok), tw.hourly_weight[2], places=12)

    def test_weekend_uses_the_real_market_close(self):
        tw = TimeWeighting("USDJPY")
        self.assertTrue(tw.is_closed(datetime(2024, 7, 6, 12, tzinfo=UTC)))     # Saturday
        self.assertTrue(tw.is_closed(datetime(2024, 7, 5, 23, tzinfo=UTC)))     # Fri after close
        self.assertFalse(tw.is_closed(datetime(2024, 7, 5, 20, tzinfo=UTC)))    # Fri before close
        self.assertFalse(tw.is_closed(datetime(2024, 7, 7, 23, tzinfo=UTC)))    # Sun after open


class TestCross(unittest.TestCase):
    def test_leg_signs(self):
        self.assertEqual(infer_leg_signs("AUDJPY", "AUDUSD", "USDJPY"), (1, -1))
        self.assertEqual(infer_leg_signs("EURGBP", "EURUSD", "GBPUSD"), (1, 1))

    def test_unrelated_legs_raise(self):
        with self.assertRaises(ValueError):
            infer_leg_signs("AUDJPY", "EURUSD", "GBPCHF")

    def test_correlation_bounds_are_enforced(self):
        with self.assertRaises(ValueError):
            CorrelationCurve(1.4, 0.2)

    def test_cross_vol_exceeds_legs_when_usd_positions_oppose(self):
        """AUDJPY = AUDUSD x USDJPY, so the variances add rather than cancel."""
        book = Book.from_excel(WORKBOOK, ASOF).build(["AUDJPY"])
        t = tenor_to_years("3m")
        cross = book["AUDJPY"].atm.term_vol(t)
        self.assertGreater(cross, book["AUDUSD"].atm.term_vol(t))
        self.assertGreater(cross, book["USDJPY"].atm.term_vol(t))


class TestSurface(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        cls.s = cls.book["USDJPY"]
        cls.expiry = datetime(2024, 5, 28, tzinfo=UTC)

    def test_all_tenors_calibrate(self):
        self.assertTrue(self.s.fits)
        self.assertTrue(all(f.ok for f in self.s.fits), [f.message for f in self.s.fits if not f.ok])

    def test_density_integrates_to_one(self):
        """Legacy getDensity raised NameError and mis-scaled by K^2."""
        g = np.linspace(0.5, 1.8, 601)
        total = float(np.trapezoid([self.s.density(float(k), self.expiry) for k in g], g))
        self.assertAlmostEqual(total, 1.0, places=4)

    def test_density_is_non_negative(self):
        g = np.linspace(0.85, 1.2, 120)
        self.assertGreaterEqual(min(self.s.density(float(k), self.expiry) for k in g), -1e-9)

    def test_delta_strike_matches_its_delta(self):
        for d, call in ((0.25, True), (0.10, False)):
            k, v = self.s.delta_strike(self.expiry, d, call)
            got = float(black.delta(1.0, k, v, self.book.clock.years_to(self.expiry), call, self.s.conv))
            self.assertAlmostEqual(abs(got), d, places=9)

    def test_slice_is_cached(self):
        a = self.s.slice_at(self.expiry)
        b = self.s.slice_at(self.expiry)
        self.assertIs(a, b)

    def test_vectorised_strike_query(self):
        ks = np.linspace(0.95, 1.05, 50)
        self.assertEqual(np.asarray(self.s.vol(ks, self.expiry)).shape, ks.shape)

    def test_past_expiry_raises_rather_than_returning_zero(self):
        """The legacy GUI wrapped this in a bare except and showed 0.0000."""
        with self.assertRaises(ValueError):
            self.s.vol(1.0, datetime(2020, 1, 1, tzinfo=UTC))

    def test_param_term_structure_decay_is_non_negative(self):
        for name, ts in self.s.term.items():
            self.assertGreaterEqual(ts.decay, 0.0, name)

    def test_anchoring_pins_the_quoted_tenors(self):
        self.s.anchor_tenors = True
        self.s._slices.clear()
        try:
            fit = self.s.fits[3]
            self.assertAlmostEqual(self.s.params_at(fit.t)["rho25"], fit.rho25, places=9)
        finally:
            self.s.anchor_tenors = False
            self.s._slices.clear()

    def test_flat_param_curve_does_not_divide_by_zero(self):
        ts = fit_param_term_structure([0.1, 0.2, 0.3], [0.5, 0.5, 0.5])
        self.assertTrue(all(math.isfinite(x) for x in (ts.initial, ts.final, ts.decay)))


class TestMarketData(unittest.TestCase):
    def _workbook(self, config: dict, pairs: list) -> Path:
        """A minimal workbook with the given CONFIG columns, in a temp dir.

        PARAMS carries whatever ``pairs`` names, so a test can say what the
        sheet declares and nothing else.  Every pair also gets a smile sheet,
        because a real workbook has one: a pair CONFIG names with no sheet
        behind it is now a reported problem rather than a silent skip, and a
        fixture that left them out would be asserting on a workbook nobody
        would ship.
        """
        import tempfile
        d = Path(tempfile.mkdtemp())
        path = d / "book.xlsx"
        self.addCleanup(shutil.rmtree, d, True)
        params = pd.DataFrame(
            {p: [8.0, 9.0, 0.0, 0.0, 5.0, 0.0, 50.0] for p in pairs},
            index=["initial", "long term", "ratevol", "addon", "MR",
                   "rate corr", "short decay"],
        )
        for p in pairs:
            if "USD" not in (p[:3], p[3:6]):
                params[p] = [0.4, 0.3, 0.0, 0.0, 4.0, 0.0, 50.0]
        tenors = [t for t in config.get("TENORS", ["1m", "3m"]) if t]
        smile = pd.DataFrame({
            "expiry": tenors,
            "ST 10D": [0.60] * len(tenors),
            "ST 25D": [0.20] * len(tenors),
            "RR 25D": [-0.10] * len(tenors),
            "RR 10D": [-0.19] * len(tenors),
        })
        with pd.ExcelWriter(path) as xw:
            width = max(len(v) for v in config.values())
            padded = {k: list(v) + [None] * (width - len(v)) for k, v in config.items()}
            pd.DataFrame(padded).to_excel(xw, sheet_name="CONFIG", index=False)
            params.to_excel(xw, sheet_name="PARAMS")
            for p in pairs:
                smile.to_excel(xw, sheet_name=p, index=False)
        return path

    def test_a_currency_column_on_the_events_sheet_weights_every_pair_with_that_leg(self):
        """The EVENTS sheet: a column headed USD is the dollar's weight on
        each row and is shared by every pair with a dollar leg; a pair's cell
        on that row is its adjustment on top.  Events used to be dated rows
        on PARAMS, which gave one weight two homes -- the old bug this pins
        is a weight typed on USDJPY and not on EURUSD, and the two pairs
        disagreeing about what the Fed is worth."""
        import tempfile
        d = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, d, True)
        path = d / "book.xlsx"
        idx = ["initial", "long term", "ratevol", "addon", "MR", "rate corr", "short decay"]
        params = pd.DataFrame({
            "USDJPY": [8.0, 9.0, 0, 0, 5.0, 0, 50.0],
            "EURUSD": [8.0, 9.0, 0, 0, 5.0, 0, 50.0],
        }, index=idx)
        # The sheet is typed in Hong Kong time, like every event in this tool.
        events = pd.DataFrame({
            None: ["2026-09-17 06:00", "2026-10-30 19:00"],
            "USD": [1.5, 0.0],
            "JPY": [0.3, 1.5],
            "USDJPY": [0.2, None],
            "EURUSD": [None, None],
        })
        smile = pd.DataFrame({"expiry": ["1m", "3m"], "ST 10D": [0.6, 0.6],
                              "ST 25D": [0.2, 0.2], "RR 25D": [-0.1, -0.1],
                              "RR 10D": [-0.19, -0.19]})
        with pd.ExcelWriter(path) as xw:
            pd.DataFrame({"PAIRS": ["USDJPY", "EURUSD"], "TENORS": ["1m", "3m"]}).to_excel(
                xw, sheet_name="CONFIG", index=False)
            params.to_excel(xw, sheet_name="PARAMS")
            events.to_excel(xw, sheet_name="EVENTS", index=False)
            for name in ("USDJPY", "EURUSD"):
                smile.to_excel(xw, sheet_name=name, index=False)
        data = ExcelSource(path).load()
        self.assertEqual(data.problems, [], data.problems)
        # 06:00 Hong Kong is 22:00 UTC the day before: the sheet's clock is
        # the workbook's, and the model's is UTC.
        uj = {e.when.strftime("%d%b"): e for e in data.events.for_pair("USDJPY")}
        self.assertAlmostEqual(uj["16Sep"].bump, 0.020)          # 1.5 + 0.3 + 0.2
        self.assertEqual(uj["16Sep"].weights, {"USD": 0.015, "JPY": 0.003})
        self.assertAlmostEqual(uj["16Sep"].adjust, 0.002)
        self.assertAlmostEqual(uj["30Oct"].bump, 0.015)          # the JPY leg alone
        eu = {e.when.strftime("%d%b"): e for e in data.events.for_pair("EURUSD")}
        self.assertAlmostEqual(eu["16Sep"].bump, 0.015)          # no JPY leg, no cell
        self.assertAlmostEqual(eu["30Oct"].bump, 0.0)            # nothing on either leg
        # The curve only takes the rows that move it; the panel sees them all.
        touching = data.events.for_pair("EURUSD", touching_only=True)
        self.assertEqual([e.when.strftime("%d%b") for e in touching], ["16Sep"])
        self.assertTrue(any("EVENTS: 2 event(s)" in n for n in data.notes), data.notes)

    def test_a_date_row_left_on_params_is_reported_not_read(self):
        """Events had dated rows on PARAMS once.  One left behind must not be
        read -- two homes for one bump is how a weight comes to mean two
        things -- and must not be ignored either."""
        import tempfile
        d = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, d, True)
        path = d / "book.xlsx"
        idx = ["initial", "long term", "ratevol", "addon", "MR", "rate corr", "short decay",
               "2026-09-16 22:00"]
        params = pd.DataFrame({"USDJPY": [8.0, 9.0, 0, 0, 5.0, 0, 50.0, 0.2]}, index=idx)
        smile = pd.DataFrame({"expiry": ["1m"], "ST 10D": [0.6], "ST 25D": [0.2],
                              "RR 25D": [-0.1], "RR 10D": [-0.19]})
        with pd.ExcelWriter(path) as xw:
            pd.DataFrame({"PAIRS": ["USDJPY"], "TENORS": ["1m"]}).to_excel(
                xw, sheet_name="CONFIG", index=False)
            params.to_excel(xw, sheet_name="PARAMS")
            smile.to_excel(xw, sheet_name="USDJPY", index=False)
        data = ExcelSource(path).load()
        self.assertTrue(any("EVENTS" in p and "is a date" in p for p in data.problems),
                        data.problems)
        self.assertEqual(data.events.rows, [])

    def test_an_events_column_that_is_neither_a_currency_nor_a_pair_is_reported(self):
        """A cell nothing reads is a bump nobody gets."""
        import tempfile
        d = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, d, True)
        path = d / "book.xlsx"
        idx = ["initial", "long term", "ratevol", "addon", "MR", "rate corr", "short decay"]
        params = pd.DataFrame({"USDJPY": [8.0, 9.0, 0, 0, 5.0, 0, 50.0]}, index=idx)
        events = pd.DataFrame({None: ["2026-09-17 06:00"], "USD": [1.5], "USDNOK": [2.0]})
        smile = pd.DataFrame({"expiry": ["1m"], "ST 10D": [0.6], "ST 25D": [0.2],
                              "RR 25D": [-0.1], "RR 10D": [-0.19]})
        with pd.ExcelWriter(path) as xw:
            pd.DataFrame({"PAIRS": ["USDJPY"], "TENORS": ["1m"]}).to_excel(
                xw, sheet_name="CONFIG", index=False)
            params.to_excel(xw, sheet_name="PARAMS")
            events.to_excel(xw, sheet_name="EVENTS", index=False)
            smile.to_excel(xw, sheet_name="USDJPY", index=False)
        data = ExcelSource(path).load()
        self.assertTrue(any("USDNOK" in p for p in data.problems), data.problems)

    def test_real_workbook_loads_without_problems(self):
        data = ExcelSource(WORKBOOK).load()
        self.assertEqual(data.problems, [], data.problems)
        self.assertIn("USDJPY", data.pairs)
        self.assertTrue(data.pairs["AUDJPY"].is_cross)
        self.assertEqual(data.pairs["AUDJPY"].legs, ("AUDUSD", "USDJPY"))

    def test_the_shipped_config_is_two_columns(self):
        """The sheet a desk maintains is pairs and tenors, and nothing else.

        The COR column and the column-per-cross naming its legs are gone: a
        cross has exactly one sensible pair of dollar legs, so it is worked
        out from the name instead of written down thirteen times.
        """
        with marketdata.open_workbook(WORKBOOK) as xls:
            cfg = pd.read_excel(xls, "CONFIG")
        self.assertEqual([str(c).strip().upper() for c in cfg.columns],
                         ["PAIRS", "TENORS"])
        data = ExcelSource(WORKBOOK).load()
        self.assertTrue(data.pairs["EURGBP"].derived)
        self.assertEqual(data.pairs["EURGBP"].legs, ("EURUSD", "GBPUSD"))

    def test_both_layouts_of_the_shipped_workbook_agree(self):
        """The same marks under both CONFIG layouts are the same book.

        ``vol_marks_legacy_format.xlsx`` is the sheet as it was -- BASE, COR
        and a column per cross -- kept because the legacy tool in the repo
        root reads only that.  It is also the strongest guard there is on the
        derivation: the legs it names by hand are the legs the two-column
        sheet works out from the names.
        """
        legacy = WORKBOOK.parent / "vol_marks_legacy_format.xlsx"
        new, old = ExcelSource(WORKBOOK).load(), ExcelSource(legacy).load()
        self.assertEqual(old.problems, [], old.problems)
        self.assertEqual(sorted(new.pairs), sorted(old.pairs))
        for name, spec in new.pairs.items():
            self.assertEqual(spec.is_cross, old.pairs[name].is_cross, name)
            self.assertEqual(spec.legs, old.pairs[name].legs, name)
        self.assertEqual(new.tenor_points, old.tenor_points)

    def test_a_cross_is_broken_into_two_dollar_pairs(self):
        path = self._workbook({"PAIRS": ["EURJPY"], "TENORS": ["1m", "3m"]},
                              ["EURJPY", "EURUSD", "USDJPY"])
        data = ExcelSource(path).load()
        self.assertEqual(data.problems, [], data.problems)
        self.assertTrue(data.pairs["EURJPY"].is_cross)
        self.assertTrue(data.pairs["EURJPY"].derived)
        self.assertEqual(data.pairs["EURJPY"].legs, ("EURUSD", "USDJPY"))
        # The legs nobody listed are there, and say which cross wanted them.
        self.assertEqual(data.pairs["EURUSD"].implied_by, "EURJPY")
        self.assertEqual(data.pairs["USDJPY"].implied_by, "EURJPY")
        self.assertFalse(data.pairs["EURUSD"].is_cross)
        self.assertTrue(any("EURJPY" in n for n in data.notes), data.notes)

    def test_a_derived_cross_is_marked_by_correlation(self):
        """The mechanism is the correlation, not a backbone of its own.

        A cross's initial / long term / MR cells have always meant
        correlation initial / final / decay; deriving the legs must not
        change which of the two a pair gets.
        """
        path = self._workbook({"PAIRS": ["EURJPY"], "TENORS": ["1m", "3m"]},
                              ["EURJPY", "EURUSD", "USDJPY"])
        book = Book.from_excel(path, ASOF).build()
        self.assertIsInstance(book["EURJPY"].atm, CrossAtmCurve)
        # 0.4 -> 0.3 with decay 4: those cells are read as a correlation and
        # are not divided by 100 the way a volatility is.
        self.assertAlmostEqual(float(book["EURJPY"].atm.correlation(0.0)), 0.4)
        self.assertAlmostEqual(float(book["EURJPY"].atm.correlation(50.0)), 0.3)

    def test_a_derived_cross_gets_the_same_legs_the_sheet_used_to_name(self):
        """The orientation is the whole of MIGRATION.md's first entry.

        AUDJPY is AUDUSD and USDJPY -- the dollar in opposite places, so the
        triangle takes +2*rho -- while EURGBP is EURUSD and GBPUSD, the
        dollar in the same place and -2*rho.  A leg written upside down would
        flip that sign silently.
        """
        for pair, legs in (("AUDJPY", ("AUDUSD", "USDJPY")),
                           ("EURGBP", ("EURUSD", "GBPUSD")),
                           ("GBPNZD", ("GBPUSD", "NZDUSD")),
                           ("EURCNH", ("EURUSD", "USDCNH")),
                           ("CNHHKD", ("USDCNH", "USDHKD"))):
            self.assertEqual(dollar_legs(pair), legs)
        self.assertEqual(infer_leg_signs("AUDJPY", *dollar_legs("AUDJPY")), (1, -1))
        self.assertEqual(infer_leg_signs("EURGBP", *dollar_legs("EURGBP")), (1, 1))

    def test_a_pair_config_names_with_no_sheet_is_reported(self):
        """The shipped workbook lost its EURGBP tab to a USDHKD one while
        CONFIG went on naming EURGBP.  Nothing said so: the reader skipped the
        pair in silence, ``volkit check`` reported "no problems found", and
        the first thing to ask that surface for a smile raised "EURGBP: no
        smile term structure; run calibrate() first" -- which names neither
        the workbook nor the tab somebody deleted, and which took the Windows
        build down at the test suite.
        """
        path = self._workbook({"PAIRS": ["USDJPY", "EURUSD"], "TENORS": ["1m", "3m"]},
                              ["USDJPY", "EURUSD"])
        import openpyxl
        wb = openpyxl.load_workbook(path)
        del wb["EURUSD"]
        wb.save(path)
        data = ExcelSource(path).load()
        self.assertEqual(len(data.problems), 1, data.problems)
        self.assertIn("EURUSD", data.problems[0])
        self.assertIn("no 'EURUSD' sheet", data.problems[0])
        self.assertIn("USDJPY", data.marks)
        self.assertNotIn("EURUSD", data.marks)

    def test_a_sheet_whose_rows_cannot_be_read_is_reported(self):
        """The same failure by the other route: the rows are there and unreadable."""
        path = self._workbook({"PAIRS": ["USDJPY"], "TENORS": ["1m", "3m"]}, ["USDJPY"])
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["USDJPY"]
        for r in range(2, ws.max_row + 1):
            for c in range(2, 6):
                ws.cell(row=r, column=c).value = None
        wb.save(path)
        data = ExcelSource(path).load()
        self.assertTrue(any("no readable quotes" in p for p in data.problems), data.problems)

    def test_a_sheet_with_its_columns_and_no_rows_is_a_pair_not_yet_quoted(self):
        """Which is a real state now that a pair is created from the screens.

        It used to be a problem, because the only way to get here was deleting
        rows by hand.  A pair added on the Workbook card arrives exactly like
        this, and a check that goes red on a pair somebody has just made is a
        check people stop reading.  Still *said* -- a pair with no smile is
        worth knowing about -- just not called a fault in the workbook.
        """
        path = self._workbook({"PAIRS": ["USDJPY"], "TENORS": ["1m", "3m"]}, ["USDJPY"])
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb["USDJPY"]
        ws.delete_rows(2, ws.max_row)
        wb.save(path)
        data = ExcelSource(path).load()
        self.assertEqual(data.problems, [])
        self.assertTrue(any("no quotes yet" in n for n in data.notes), data.notes)

    def test_a_pair_asked_for_with_no_quotes_says_so(self):
        """``calibrate_smiles`` used to skip it silently, so the book came
        back looking loaded and refused on the first smile."""
        path = self._workbook({"PAIRS": ["USDJPY", "EURUSD"], "TENORS": ["1m", "3m"]},
                              ["USDJPY", "EURUSD"])
        import openpyxl
        wb = openpyxl.load_workbook(path)
        del wb["EURUSD"]
        wb.save(path)
        book = Book.from_excel(path, ASOF).load_all(["USDJPY", "EURUSD"])
        self.assertTrue(any("EURUSD" in w and "no smile quotes" in w for w in book.warnings),
                        book.warnings)

    def test_a_dollar_pair_has_no_legs_to_derive(self):
        with self.assertRaises(ValueError):
            dollar_legs("USDJPY")
        with self.assertRaises(ValueError):
            dollar_legs("EURUR")

    def test_the_legacy_layout_still_loads_and_its_legs_win(self):
        """BASE / COR / a column per cross is still a workbook we read.

        And a sheet that names legs is not second-guessed: the derived legs
        would be EURUSD and USDJPY, and this one says to go through sterling
        instead.  A leg that is itself a cross is then broken down in its
        turn, which is why the legs are resolved on a work list.
        """
        path = self._workbook(
            {"BASE": ["EURUSD", "USDJPY", "GBPUSD"], "COR": ["EURJPY"],
             "EURJPY": ["EURGBP", "GBPJPY"], "TENORS": ["1m", "3m"]},
            ["EURUSD", "USDJPY", "GBPUSD", "EURJPY", "EURGBP", "GBPJPY"])
        data = ExcelSource(path).load()
        self.assertEqual(data.problems, [], data.problems)
        self.assertTrue(data.pairs["EURJPY"].is_cross)
        self.assertFalse(data.pairs["EURJPY"].derived)
        self.assertEqual(data.pairs["EURJPY"].legs, ("EURGBP", "GBPJPY"))
        # GBPJPY was named by nothing but that column, and is a cross itself.
        self.assertEqual(data.pairs["GBPJPY"].implied_by, "EURJPY")
        self.assertEqual(data.pairs["GBPJPY"].legs, ("GBPUSD", "USDJPY"))

    def test_legs_that_cannot_build_the_cross_are_reported_not_raised(self):
        path = self._workbook(
            {"BASE": ["EURUSD", "USDJPY"], "COR": ["EURJPY"],
             "EURJPY": ["EURUSD", "EURUSD"], "TENORS": ["1m"]},
            ["EURUSD", "USDJPY", "EURJPY"])
        data = ExcelSource(path).load()
        self.assertTrue(any("EURJPY" in p for p in data.problems), data.problems)
        self.assertNotIn("EURJPY", data.pairs)

    def test_a_row_that_is_not_a_pair_is_named_and_the_rest_still_load(self):
        path = self._workbook({"PAIRS": ["EURUSD", "EUR/USD", "USDJPY"],
                               "TENORS": ["1m"]},
                              ["EURUSD", "USDJPY"])
        data = ExcelSource(path).load()
        self.assertIn("EURUSD", data.pairs)
        self.assertIn("USDJPY", data.pairs)
        self.assertTrue(any("EUR/USD" in p for p in data.problems), data.problems)

    def test_a_derivation_reaches_the_page(self):
        """A pair that came out of a convention must not read like one that
        was written down, so the note travels with the state and the page
        shows it -- on the meta line's tooltip, not in the message box, which
        holds errors and warnings only."""
        from volkit.webapp import BookService
        state = BookService(str(WORKBOOK), ASOF).state()
        self.assertTrue(any("EURGBP = EURUSD x GBPUSD" in n for n in state["notes"]),
                        state["notes"])
        page = _source("volkit", "web", "index.html")
        self.assertIn("STATE.notes", page)
        self.assertNotIn("nts.map(x=>`<li", page)

    def test_a_config_with_no_pairs_column_says_what_it_wants(self):
        path = self._workbook({"THINGS": ["EURUSD"], "TENORS": ["1m"]}, ["EURUSD"])
        with self.assertRaises(MarketDataError) as ctx:
            ExcelSource(path).load()
        self.assertIn("PAIRS", str(ctx.exception))

    def test_units_are_converted_once(self):
        data = ExcelSource(WORKBOOK).load()
        self.assertAlmostEqual(data.params["USDJPY"].initial, 0.0605)
        self.assertAlmostEqual(data.marks["USDJPY"][0].st_25, 0.002175)

    def test_cross_correlations_are_not_rescaled(self):
        data = ExcelSource(WORKBOOK).load()
        self.assertAlmostEqual(data.params["AUDJPY"].initial, 0.37)

    def test_missing_file_raises_clearly(self):
        with self.assertRaises(MarketDataError):
            ExcelSource("/nonexistent/nope.xlsx")

    def test_premium_adjustment_follows_the_pair(self):
        data = ExcelSource(WORKBOOK).load()
        self.assertTrue(data.pairs["USDJPY"].resolved_premium_adjusted())
        self.assertFalse(data.pairs["EURUSD"].resolved_premium_adjusted())


class TestBook(unittest.TestCase):
    def test_build_order_puts_legs_first(self):
        book = Book.from_excel(WORKBOOK, ASOF)
        order = book.build_order()
        self.assertLess(order.index("AUDUSD"), order.index("AUDJPY"))
        self.assertLess(order.index("USDJPY"), order.index("AUDJPY"))

    def test_requesting_a_cross_pulls_in_its_legs(self):
        book = Book.from_excel(WORKBOOK, ASOF).build(["AUDJPY"])
        self.assertIn("AUDUSD", book.surfaces)
        self.assertIn("USDJPY", book.surfaces)

    def test_past_events_are_reported_not_silently_used(self):
        book = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])
        self.assertTrue(any("before the valuation time" in w for w in book.warnings))

    def test_tenor_points_available_without_crosses(self):
        """Legacy set self.tenor_points inside the cross loop only."""
        book = Book.from_excel(WORKBOOK, ASOF)
        self.assertTrue(book.data.tenor_points)

    def test_unknown_pair_raises_with_a_useful_message(self):
        """And with the *right* list.

        ``build``/``load_all`` may be narrowed to a few pairs -- ``volkit band
        USDHKD`` narrows them to one -- so when the asked-for pair is not in
        the workbook nothing is built and "available: []" read as an empty
        workbook rather than as a pair the workbook does not carry.  A pair
        the workbook has never heard of is told what the workbook holds; a
        pair it has, but which this book was not asked to build, is told that
        instead.
        """
        book = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])
        with self.assertRaises(KeyError) as ctx:
            book["NOPE"]
        message = str(ctx.exception)
        self.assertIn("is not in", message)
        self.assertIn("USDJPY", message)          # what the workbook does hold
        self.assertIn("EURUSD", message)

        narrowed = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])
        with self.assertRaises(KeyError) as ctx:
            narrowed["EURGBP"]
        message = str(ctx.exception)
        self.assertIn("is in the workbook but is not built", message)
        self.assertIn("USDJPY", message)


class TestPricing(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])

    def test_strike_spec_forms(self):
        self.assertEqual(parse_strike("ATM").kind, "atm")
        self.assertEqual(parse_strike("").kind, "atm")
        self.assertEqual(parse_strike("1.0234").value, 1.0234)
        self.assertEqual((parse_strike("25d").value, parse_strike("25d").is_call), (0.25, True))
        self.assertFalse(parse_strike("10dp").is_call)
        self.assertFalse(parse_strike("-25d").is_call)
        self.assertTrue(parse_strike("25dp").side_explicit)
        self.assertFalse(parse_strike("25d").side_explicit)

    # ---- the marking screen's vol query --------------------------------
    # Two boxes and one number, sharing the pricing screen's strike and
    # expiry vocabulary through `resolve_strike` / `expiry_datetime`.  These
    # pin the sharing: a strike read two ways is a strike that can be read
    # two different ways.

    def test_the_vol_query_reads_the_same_strike_as_a_priced_leg(self):
        """The card and the pricing grid must land on one strike and one vol.

        Both go through `pricing.resolve_strike`; before it there were two
        copies of the same six lines.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        book.feed = MarketFeed.load(FEED)
        for strike in ("ATM", "25d", "10dp", "151.5"):
            q = quick_vol(book, "USDJPY", "1M", strike)
            leg = price_strip(book, [OptionLeg("USDJPY", "1M", strike)])["legs"][0]
            self.assertTrue(leg["ok"], leg.get("error"))
            self.assertAlmostEqual(q["strike"], leg["strike"], places=10, msg=strike)
            self.assertAlmostEqual(q["vol"], leg["vol"], places=10, msg=strike)
            self.assertAlmostEqual(q["forward"], leg["forward"], places=10, msg=strike)

    def test_the_vol_query_takes_its_forward_from_the_feed(self):
        """There is no third box: the level is `Book.market_level_for`'s.

        ``market_level_for`` and not ``market_level``: the forward is the one
        to this option's own **settlement** date, two business days past its
        expiry, which is the date a forward is a price for.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        book.feed = MarketFeed.load(FEED)
        q = quick_vol(book, "USDJPY", "1M", "ATM")
        level = book.market_level_for("USDJPY", date.fromisoformat(q["expiry"]))
        self.assertEqual(q["settle"], level["settle"])
        self.assertTrue(q["scaled"])
        self.assertEqual(q["forward_source"], "feed")
        self.assertAlmostEqual(q["forward"], level["forward"], places=12)
        self.assertAlmostEqual(q["spot"], level["spot"], places=12)

    def test_the_vol_query_answers_in_moneyness_with_no_feed(self):
        """ATM and a delta are moneyness questions and need no level at all.

        Same rule as the smile chart's axis: without a feed it stays in K/F
        and says so, rather than refusing a question it can answer.
        """
        for strike in ("ATM", "25d"):
            q = quick_vol(self.book, "USDJPY", "1M", strike)
            self.assertFalse(q["scaled"], strike)
            self.assertIsNone(q["strike"], strike)
            self.assertIsNone(q["forward"], strike)
            self.assertEqual(q["forward_source"], "none")
            self.assertAlmostEqual(
                q["vol"],
                float(self.book["USDJPY"].vol(q["strike_ratio"],
                                              expiry_datetime(self.book, "USDJPY", "1M"))) * 100,
                places=12)

    def test_an_absolute_strike_with_no_feed_is_refused_by_name(self):
        """The marks are in K/F, so 151.5 cannot be placed against them.

        Read as a ratio it would be a wing nobody asked about, silently.
        """
        with self.assertRaises(ValueError) as ctx:
            quick_vol(self.book, "USDJPY", "1M", "151.5")
        self.assertIn("the feed does not quote USDJPY", str(ctx.exception))

    def test_the_reported_strike_is_the_one_the_vol_was_read_at(self):
        """The card keeps the request in its box and reports the resolution
        under it, so the two must agree: asking again at the strike it named
        is the same read, and the date it named resolves to itself."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        book.feed = MarketFeed.load(FEED)
        asked = quick_vol(book, "USDJPY", "1M", "25d")
        again = quick_vol(book, "USDJPY", "1M", repr(asked["strike"]))
        self.assertAlmostEqual(asked["vol"], again["vol"], places=12)
        self.assertEqual(quick_vol(book, "USDJPY", asked["expiry"], "ATM")["expiry"],
                         asked["expiry"])

    def test_the_strike_box_is_the_only_place_the_wing_is_said(self):
        """A bare `25d` names two strikes and is read on the call, as on the
        pricing screen; `25dp` and `-25d` are how the other one is asked for.

        The card briefly had a wing toggle beside the strike box.  It was a
        second place to say one thing -- and a place that could be set to Call
        against a strike that already said put -- so the strike text is the
        whole of it.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        book.feed = MarketFeed.load(FEED)
        call = quick_vol(book, "USDJPY", "1M", "25d")
        self.assertIs(call["is_call"], True)
        self.assertFalse(call["side_explicit"])
        for text in ("25dp", "-25d"):
            put = quick_vol(book, "USDJPY", "1M", text)
            self.assertIs(put["is_call"], False, text)
            self.assertTrue(put["side_explicit"], text)
            self.assertLess(put["strike"], call["strike"], text)
            self.assertNotAlmostEqual(call["vol"], put["vol"], places=6, msg=text)

    def test_no_side_is_reported_where_there_are_not_two_strikes(self):
        """At the at-the-money and at an absolute strike the volatility is one
        number for the call and the put, so the row names no wing at all."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        book.feed = MarketFeed.load(FEED)
        for text in ("ATM", "151.5"):
            self.assertIsNone(quick_vol(book, "USDJPY", "1M", text)["is_call"], text)

    def test_the_vol_query_reports_the_delta_of_the_strike_it_read(self):
        """The card takes a strike or a delta and answers with both.

        A strike and a delta name one point on the smile; the desk has
        whichever of the two the market gave it, so the card must be able to
        be asked either way and report the other.  The round trip is the
        pin: the delta reported at a 25-delta request is 25, and asking again
        at the strike that came back reports the same delta.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        book.feed = MarketFeed.load(FEED)
        asked = quick_vol(book, "USDJPY", "1M", "25d")
        self.assertAlmostEqual(asked["delta"], 25.0, places=6)
        self.assertIs(asked["delta_is_call"], True)
        back = quick_vol(book, "USDJPY", "1M", repr(asked["strike"]))
        self.assertAlmostEqual(back["delta"], asked["delta"], places=6)
        put = quick_vol(book, "USDJPY", "1M", "10dp")
        self.assertAlmostEqual(put["delta"], -10.0, places=6)
        self.assertIs(put["delta_is_call"], False)

    def test_the_delta_is_reported_under_the_pairs_own_convention(self):
        """Premium adjusted for a USD-base pair, unadjusted otherwise.

        The browser has no business knowing which, so the answer says: a
        premium-adjusted delta-neutral straddle is not at 50 delta and a page
        that assumed it was would report a strike nobody asked about.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])
        book.feed = MarketFeed.load(FEED)
        usd = quick_vol(book, "USDJPY", "1M", "ATM")
        eur = quick_vol(book, "EURUSD", "1M", "ATM")
        self.assertTrue(usd["premium_adjusted"])
        self.assertFalse(eur["premium_adjusted"])
        self.assertLess(usd["delta"], 50.0)
        self.assertGreater(eur["delta"], 50.0)

    def test_the_delta_comes_back_for_a_pair_with_no_feed(self):
        """Delta is a function of moneyness, so it needs no level at all.

        The strike cannot be placed without a feed and comes back as None;
        the delta beside it must not disappear with it, because a delta is
        exactly what a marker asks for when there is no level to hand.
        """
        q = quick_vol(self.book, "USDJPY", "1M", "25d")
        self.assertFalse(q["scaled"])
        self.assertIsNone(q["strike"])
        self.assertAlmostEqual(q["delta"], 25.0, places=6)

    def test_the_vol_query_flags_a_strike_outside_a_managed_band(self):
        """Same rule as a pricing leg: the level the payout depends on is
        checked, and a lognormal wing outside a defended band says so."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDHKD"])
        book.feed = MarketFeed.load(FEED)
        if "USDHKD" not in book.banded_pairs():
            self.skipTest("no USDHKD band on the PEG_BANDS tab")
        band = book["USDHKD"].band
        q = quick_vol(book, "USDHKD", "1M", str(band.upper * 1.02), forward=band.upper)
        self.assertTrue(any("outside the managed band" in w for w in q["warnings"]), q["warnings"])

    def test_bad_strike_spec_raises(self):
        for bad in ("xyz", "60d", "0d"):
            with self.assertRaises(ValueError, msg=bad):
                parse_strike(bad)

    def test_bare_delta_takes_its_wing_from_the_option_type(self):
        """'25d' with type P must resolve the put strike, not the call strike."""
        call = price_strip(self.book, [OptionLeg("USDJPY", "1M", "25d", "C", spot=150.25)])["legs"][0]
        put = price_strip(self.book, [OptionLeg("USDJPY", "1M", "25d", "P", spot=150.25)])["legs"][0]
        self.assertAlmostEqual(call["delta_pct"], 25.0, places=6)
        self.assertAlmostEqual(put["delta_pct"], -25.0, places=6)
        self.assertLess(put["strike"], call["strike"])

    def test_explicit_and_signed_delta_agree(self):
        legs = [OptionLeg("USDJPY", "1M", s, "P", spot=150.25) for s in ("25dp", "-25d")]
        a, b = price_strip(self.book, legs)["legs"]
        self.assertAlmostEqual(a["strike"], b["strike"], places=12)

    def test_tenor_expiry_resolves_on_the_calendar(self):
        d = resolve_expiry(self.book, "USDJPY", "1M")
        self.assertTrue(self.book.calendars.is_business_day("USDJPY", d))
        self.assertEqual(resolve_expiry(self.book, "USDJPY", "2024-05-28"), date(2024, 5, 28))

    def test_forward_points_applied_with_the_pip_divisor(self):
        leg = OptionLeg("USDJPY", "1M", "ATM", spot=150.25, forward_points=-45, pip=100)
        r = price_strip(self.book, [leg])["legs"][0]
        self.assertAlmostEqual(r["forward"], 150.25 - 0.45, places=10)

    def test_one_bad_leg_does_not_break_the_strip(self):
        legs = [
            OptionLeg("USDJPY", "1M", "25d", "C", spot=150.25, label="good"),
            OptionLeg("USDJPY", "not-a-tenor", "25d", "C", spot=150.25, label="bad"),
            OptionLeg("EURUSD", "3M", "ATM", "C", spot=1.0842, label="also good"),
        ]
        out = price_strip(self.book, legs)
        self.assertEqual(out["errors"], 1)
        self.assertTrue(out["legs"][0]["ok"])
        self.assertFalse(out["legs"][1]["ok"])
        self.assertTrue(out["legs"][2]["ok"])
        self.assertIn("not-a-tenor", out["legs"][1]["error"])

    def test_unknown_pair_is_reported_per_leg(self):
        out = price_strip(self.book, [OptionLeg("XXXYYY", "1M", "ATM", spot=1.0)])
        self.assertFalse(out["legs"][0]["ok"])

    def test_put_call_parity_across_two_legs(self):
        legs = [OptionLeg("USDJPY", "3M", "150.0", t, spot=150.25, forward_points=-140)
                for t in ("C", "P")]
        c, p = price_strip(self.book, legs)["legs"]
        self.assertAlmostEqual(c["premium_dom"] - p["premium_dom"],
                               c["forward"] - c["strike"], places=10)

    def test_direction_flips_the_signed_amounts(self):
        base = OptionLeg("USDJPY", "1M", "25d", "C", spot=150.25, notional=10)
        sold = OptionLeg("USDJPY", "1M", "25d", "C", spot=150.25, notional=10, direction=-1)
        b, s = price_strip(self.book, [base, sold])["legs"]
        self.assertAlmostEqual(b["premium_amount"], -s["premium_amount"], places=12)
        self.assertAlmostEqual(b["vega_amount"], -s["vega_amount"], places=12)

    def test_totals_bucket_by_pair(self):
        legs = [
            OptionLeg("USDJPY", "1M", "25d", "C", spot=150.25, notional=10),
            OptionLeg("USDJPY", "1M", "25d", "P", spot=150.25, notional=10, direction=-1),
            OptionLeg("EURUSD", "3M", "ATM", "C", spot=1.0842, notional=20),
        ]
        out = price_strip(self.book, legs)
        self.assertEqual(set(out["totals"]), {"USDJPY", "EURUSD"})
        jpy = [l for l in out["legs"] if l["pair"] == "USDJPY"]
        self.assertAlmostEqual(out["totals"]["USDJPY"]["premium"],
                               sum(l["premium_amount"] for l in jpy), places=10)

    def test_premium_scales_with_notional(self):
        one = price_strip(self.book, [OptionLeg("USDJPY", "1M", "ATM", "C", spot=150.25, notional=1)])
        ten = price_strip(self.book, [OptionLeg("USDJPY", "1M", "ATM", "C", spot=150.25, notional=10)])
        self.assertAlmostEqual(ten["legs"][0]["premium_amount"],
                               10 * one["legs"][0]["premium_amount"], places=10)


class TestSmileShape(unittest.TestCase):
    """Reading a marked smile back as a correlation and a vol of vol.

    ``calibrate`` matches the *market* strangle, because that is what a broker
    quotes.  This one matches the *smile* butterfly, because that is the number
    the analysis screen has off whatever surface is marked -- matching a
    premium condition against a moment would be comparing two different things.
    """

    def test_a_sabr_smile_reads_back_as_the_parameters_it_was_built_from(self):
        for rho, nu, t in ((-0.35, 0.55, 0.25), (0.20, 0.90, 1.0), (-0.60, 1.40, 0.08)):
            with self.subTest(rho=rho, nu=nu, t=t):
                p = sabr.SabrParams(alpha=0.09, rho=rho, volvol=nu, t=t)
                # The at-the-money here is the *delta-neutral straddle*
                # volatility, which is how this book quotes it and what the
                # fit's own at-the-money condition solves alpha against.
                # ``sabr.atm_vol`` is the at-the-forward one, and feeding that
                # in instead moves alpha and with it both parameters.
                # The delta-neutral strike depends on the very volatility it
                # carries, so it is a fixed point; the fit solves alpha at
                # ``dns_strike(f, atm)`` and the test has to hand it an ``atm``
                # consistent with that or the two are a strike apart.
                atm = sabr.atm_vol(p)
                for _ in range(40):
                    nxt = float(sabr.lognormal_vol(
                        black.dns_strike(p.f, atm, t, False), p))
                    if abs(nxt - atm) < 1e-15:
                        atm = nxt
                        break
                    atm = nxt
                _, call = sabr.smile_strike_and_vol(p, 0.25, t, True, False)
                _, put = sabr.smile_strike_and_vol(p, -0.25, t, False, False)
                got = sabr.fit_smile_shape(atm, call - put, 0.5 * (call + put) - atm,
                                           0.25, t, False)
                self.assertTrue(got.converged, got.message)
                self.assertAlmostEqual(got.rho, rho, places=5)
                self.assertAlmostEqual(got.nu, nu, places=5)
                self.assertLess(got.max_error, 1e-7)

    def test_the_correlation_carries_the_sign_of_the_risk_reversal(self):
        """This is the whole reason the number is worth reporting: a risk
        reversal is a price, and rho is what it is a price *of*."""
        t = 0.5
        base = sabr.fit_smile_shape(0.10, 0.0, 0.0030, 0.25, t, False)
        up = sabr.fit_smile_shape(0.10, +0.012, 0.0030, 0.25, t, False)
        down = sabr.fit_smile_shape(0.10, -0.012, 0.0030, 0.25, t, False)
        # Not exactly zero, and it should not be: a zero-correlation SABR smile
        # is symmetric about the *forward*, while the two wings are placed
        # symmetrically about the delta-neutral strike, which is above it.
        self.assertAlmostEqual(base.rho, 0.0, delta=0.05)
        self.assertGreater(up.rho, 0.15)
        self.assertLess(down.rho, -0.15)

    def test_a_wider_butterfly_is_a_higher_vol_of_vol(self):
        t = 0.5
        thin = sabr.fit_smile_shape(0.10, -0.005, 0.0015, 0.25, t, False)
        fat = sabr.fit_smile_shape(0.10, -0.005, 0.0045, 0.25, t, False)
        self.assertGreater(fat.nu, thin.nu * 1.4)

    def test_a_smile_no_sabr_can_reach_says_so_rather_than_returning_the_nearest(self):
        """A steep risk reversal on a flat butterfly is outside the family.

        Returning the nearest parameters in silence would put a number on the
        screen that does not describe the smile it claims to summarise.
        """
        got = sabr.fit_smile_shape(0.10, -0.060, 0.0002, 0.25, 0.5, False)
        self.assertFalse(got.converged)
        self.assertGreater(got.max_error, 1e-6)
        self.assertTrue(got.warnings)
        self.assertIn("nearest", " ".join(got.warnings))

    def test_the_inputs_are_checked(self):
        for bad in ((0.0, 0.0, 0.001, 0.25, 0.5), (0.10, 0.0, 0.001, 0.25, 0.0),
                    (0.10, 0.0, 0.001, 0.0, 0.5), (0.10, 0.0, 0.001, 0.6, 0.5)):
            with self.assertRaises(ValueError):
                sabr.fit_smile_shape(*bad, conv=False)


class TestSabrRobustness(unittest.TestCase):
    def test_alpha_closed_form_matches_hagan(self):
        for rho, nu, t in ((-0.3, 0.8, 1.0), (0.5, 1.5, 2.0), (0.0, 0.5, 0.25)):
            for a in sabr.alpha_roots_at_forward(0.08, rho, nu, t):
                got = float(sabr.lognormal_vol(1.0, sabr.SabrParams(a, rho, nu, t)))
                self.assertAlmostEqual(got, 0.08, places=12)

    def test_multiple_alpha_roots_are_found(self):
        """The at-the-money condition is a cubic and can have two positive roots."""
        roots = sabr.alpha_roots_at_forward(0.08, -0.3, 0.8, 1.0)
        self.assertGreater(len(roots), 1)
        self.assertLess(roots[0], roots[1])

    def test_no_alpha_root_is_reported_not_guessed(self):
        self.assertEqual(sabr.alpha_roots_at_forward(0.08, -0.9, 3.0, 5.0), [])
        K = black.dns_strike(1.0, 0.08, 5.0, False)
        with self.assertRaises(ConvergenceError):
            sabr.alpha_from_atm(0.08, K, -0.9, 3.0, 5.0)

    def test_calibration_is_independent_of_the_starting_point(self):
        """The sweep locates the basin, so no seed can change the answer."""
        base = sabr.calibrate(0.12, -0.055, 0.009, 0.25, 0.25, DeltaConvention(True))
        coarse = sabr.calibrate(0.12, -0.055, 0.009, 0.25, 0.25, DeltaConvention(True),
                                scan=(21, 15))
        self.assertAlmostEqual(base.params.rho, coarse.params.rho, places=6)
        self.assertAlmostEqual(base.params.volvol, coarse.params.volvol, places=6)

    def test_stress_quotes_still_calibrate(self):
        cases = [(0.085, -0.030, 0.0065, 0.25, 1 / 52), (0.085, -0.002, 0.0180, 0.25, 1 / 52),
                 (0.12, -0.055, 0.0090, 0.25, 0.25), (0.28, -0.090, 0.0300, 0.25, 1 / 12),
                 (0.07, -0.010, 0.0004, 0.25, 0.5)]
        for atm, rr, st, d, t in cases:
            cal = sabr.calibrate(atm, rr, st, d, t, DeltaConvention(True))
            self.assertTrue(cal.converged, f"{(atm, rr, st, d, t)}: {cal.message}")
            self.assertLess(cal.max_error, 1e-9)

    def test_extreme_total_vol_raises_rather_than_overflowing(self):
        with self.assertRaises(ValueError):
            black.strike_from_delta(0.25, 1.0, 50.0, 100.0, True, False)

    def test_prior_pulls_the_fit(self):
        free = sabr.calibrate(0.07, -0.02, 0.005, 0.25, 1.0, False)
        prior = sabr.SabrParams(0.07, 0.5, 0.2, 1.0)
        pulled = sabr.calibrate(0.07, -0.02, 0.005, 0.25, 1.0, False,
                                prior=prior, prior_weight=5.0)
        self.assertGreater(pulled.params.rho, free.params.rho)

    def test_book_calibration_is_unique(self):
        """No competing solutions anywhere in the sample workbook."""
        book = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])
        surface = book["USDJPY"]
        for mark in book.data.marks["USDJPY"][:4]:
            t = tenor_to_years(mark.tenor)
            atm = surface.atm.cut_vol(ASOF.datetime_from_years(t), "NY")
            cal = sabr.calibrate(atm, mark.rr_25, mark.st_25, 0.25, t,
                                 surface.conv, max_solutions=3)
            self.assertEqual(cal.alternatives, (), f"{mark.tenor}: {cal.warnings}")


class TestEventTable(unittest.TestCase):
    """The EVENTS sheet in memory: one row per release, weights per currency
    shared across pairs, an adjustment per pair."""

    def setUp(self):
        from volkit.events import EventBook, EventRow
        self.EventBook, self.EventRow = EventBook, EventRow
        self.when = datetime(2026, 9, 16, 22, 0, tzinfo=UTC)
        self.table = EventBook([EventRow(self.when, "FOMC",
                                         weights={"USD": 0.015, "JPY": 0.003},
                                         adjust={"USDJPY": 0.002})])

    def test_the_shipped_workbook_has_its_events_on_the_events_sheet(self):
        data = ExcelSource(WORKBOOK).load()
        self.assertEqual(data.problems, [], data.problems)
        self.assertTrue(data.events.rows)
        self.assertTrue(data.events.currencies())

    def test_a_weight_is_shared_and_an_adjustment_is_not(self):
        """One row, three pairs.  The dollar's weight reaches every pair with
        a dollar leg; only USDJPY's own cell is USDJPY's."""
        uj = self.table.for_pair("USDJPY")[0]
        self.assertAlmostEqual(uj.bump, 0.020)
        self.assertEqual(uj.weights, {"USD": 0.015, "JPY": 0.003})
        self.assertAlmostEqual(uj.adjust, 0.002)
        eu = self.table.for_pair("EURUSD")[0]
        self.assertAlmostEqual(eu.bump, 0.015)
        self.assertAlmostEqual(eu.adjust, 0.0)
        eg = self.table.for_pair("EURGBP")[0]
        self.assertAlmostEqual(eg.bump, 0.0)
        # It is still a row of the sheet: that blank cell is where EURGBP's
        # adjustment would be typed, so the panel sees it and the curve does not.
        self.assertEqual(self.table.for_pair("EURGBP", touching_only=True), [])

    def test_an_alias_leg_takes_the_weight(self):
        """A weight on CNY is the CNH leg's too (``events.CURRENCY_ALIASES``)."""
        self.table.rows[0].weights = {"CNY": 0.004}
        self.assertAlmostEqual(self.table.for_pair("USDCNH")[0].bump, 0.004)

    def test_marking_a_leg_weight_on_one_pair_moves_the_others_and_says_so(self):
        """The panel shows the sheet through one pair's eyes, so a weight
        typed there is the sheet's.  That is the point, and it is never
        hidden: the note names every pair it reached."""
        from volkit.events import EventEntry
        entry = EventEntry(self.when, None, "FOMC", {"USD": 0.02, "JPY": 0.003}, 0.002)
        bad, notes = self.table.set_pair("USDJPY", [entry],
                                         pairs=["USDJPY", "EURUSD", "EURGBP"])
        self.assertEqual(bad, [])
        self.assertTrue(any("EURUSD" in n and "USD" in n for n in notes), notes)
        self.assertNotIn("EURGBP", " ".join(notes))     # no dollar leg
        self.assertAlmostEqual(self.table.for_pair("EURUSD")[0].bump, 0.02)

    def test_the_other_currencies_of_a_row_are_not_cleared_by_a_pair_that_cannot_see_them(self):
        """EURUSD's panel shows EUR and USD.  Applying it must not wipe the
        JPY column -- a screen that never held a number must not zero it."""
        from volkit.events import EventEntry
        entry = EventEntry(self.when, None, "FOMC", {"EUR": 0.0, "USD": 0.015}, 0.0)
        self.table.set_pair("EURUSD", [entry], pairs=["USDJPY", "EURUSD"])
        self.assertAlmostEqual(self.table.rows[0].weights["JPY"], 0.003)
        self.assertAlmostEqual(self.table.rows[0].adjust["USDJPY"], 0.002)

    def test_two_rows_at_one_time_are_refused(self):
        """A row is identified by its time; a second at the same minute would
        overwrite the first rather than add to it."""
        from volkit.events import EventEntry
        bad, _ = self.table.set_pair("USDJPY", [
            EventEntry(self.when, None, "a", {"USD": 0.01}, 0.0),
            EventEntry(self.when, None, "b", {"USD": 0.02}, 0.0)])
        self.assertTrue(bad and "two events" in bad[0], bad)

    def test_set_weights_replaces_the_currencies_and_keeps_every_pair_cell(self):
        problems = self.table.set_weights([
            {"when": self.when, "label": "FOMC", "weights": {"USD": 0.02}}])
        self.assertEqual(problems, [])
        self.assertEqual(self.table.rows[0].weights, {"USD": 0.02})
        self.assertAlmostEqual(self.table.rows[0].adjust["USDJPY"], 0.002)
        self.assertAlmostEqual(self.table.for_pair("USDJPY")[0].bump, 0.022)


class TestCurveMarking(unittest.TestCase):
    def setUp(self):
        self.book = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY", "AUDJPY"])

    def test_backbone_parameters_can_be_remarked(self):
        atm = self.book["USDJPY"].atm
        before = atm.term_vol(1.0)
        self.assertEqual(atm.set_params(long_term_vol=0.09), [])
        self.assertGreater(atm.term_vol(1.0), before)

    def test_bad_parameter_is_rejected_and_curve_unchanged(self):
        atm = self.book["USDJPY"].atm
        before = atm.term_vol(1.0)
        self.assertTrue(atm.set_params(long_term_vol=-1.0))
        self.assertAlmostEqual(atm.term_vol(1.0), before, places=14)

    def test_cross_correlation_can_be_remarked(self):
        atm = self.book["AUDJPY"].atm
        self.assertIsInstance(atm, CrossAtmCurve)
        before = atm.term_vol(1.0)
        self.assertEqual(atm.set_correlation(0.1, 0.1, 1.0), [])
        self.assertNotAlmostEqual(atm.term_vol(1.0), before, places=6)
        self.assertTrue(atm.set_correlation(1.5, 0.1, 1.0))

    def test_the_events_route_takes_parts_and_returns_the_total(self):
        """The panel posts each leg's weight and the adjustment; the rows
        that come back carry both and their total."""
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF)
        when = (ASOF.now + timedelta(days=30)).replace(hour=16, minute=0, second=0, microsecond=0)
        r = service.set_events({"pair": "USDJPY", "events": [
            {"when": when.strftime("%Y-%m-%dT%H:%M"), "weights": {"USD": 1.5, "JPY": 0.3},
             "adjust": 0.2, "label": "FOMC"}]})
        # A dollar weight reaches every dollar pair, so a warning may come
        # back for one of them; none of them is USDJPY's.
        self.assertEqual([p for p in r["problems"] if p.startswith("USDJPY")], [])
        row = next(e for e in r["events"] if e["label"] == "FOMC")
        self.assertAlmostEqual(row["bump"], 2.0)
        self.assertEqual(set(row["weights"]), {"USD", "JPY"})
        self.assertAlmostEqual(row["weights"]["JPY"], 0.3)
        self.assertAlmostEqual(row["adjust"], 0.2)
        self.assertEqual(service.curve({"pair": "USDJPY"})["legs_ccy"], ["USD", "JPY"])
        with self.assertRaises(ValueError):
            service.set_events({"pair": "USDJPY", "events": [
                {"when": when.strftime("%Y-%m-%dT%H:%M"), "weights": {"USD": 1.5}, "adjust": 0,
                 "bump": 3.0}]})

    def test_a_weight_marked_on_one_pair_reaches_the_others(self):
        """The sheet's currency columns are shared.  A dollar weight typed on
        USDJPY's panel is EURUSD's too, and the reply says which pairs moved
        -- the old arrangement kept a copy per pair, so the two could disagree
        about what one release was worth."""
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF)
        service.book.build(["USDJPY", "EURUSD"])
        when = (ASOF.now + timedelta(days=30)).replace(hour=16, minute=0, second=0, microsecond=0)
        r = service.set_events({"pair": "USDJPY", "events": [
            {"when": when.strftime("%Y-%m-%dT%H:%M"), "weights": {"USD": 1.5, "JPY": 0.3},
             "adjust": 0.2, "label": "FOMC"}]})
        self.assertTrue(any("EURUSD" in n for n in r["notes"]), r["notes"])
        eu = next(e for e in service.curve({"pair": "EURUSD"})["events"]
                  if e["label"] == "FOMC")
        self.assertAlmostEqual(eu["bump"], 1.5)     # the dollar leg, no adjustment
        self.assertAlmostEqual(eu["adjust"], 0.0)

    def test_reload_gives_back_the_workbooks_own_rows(self):
        """The Reload button: the EVENTS sheet as the file has it, after a
        session has marked over it.  It reads and applies nothing."""
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF)
        before = service.workbook_events({"pair": "USDJPY"})["events"]
        when = (ASOF.now + timedelta(days=30)).replace(hour=16, minute=0, second=0, microsecond=0)
        service.set_events({"pair": "USDJPY", "events": [
            {"when": when.strftime("%Y-%m-%dT%H:%M"), "weights": {"USD": 1.5}, "adjust": 0,
             "label": "NEW"}]})
        self.assertEqual(service.workbook_events({"pair": "USDJPY"})["events"], before)
        self.assertNotEqual(service.curve({"pair": "USDJPY"})["events"], before)

    def test_the_weights_card_is_a_whole_table_and_keeps_every_pair_cell(self):
        """The optional card is the currency side of the sheet, posted whole.
        Applying it re-solves every pair with those currencies and leaves each
        pair's own adjustment column alone.  A bad cell leaves it untouched."""
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF)
        service.book.build(["USDJPY", "EURUSD"])
        when = (ASOF.now + timedelta(days=30)).replace(hour=16, minute=0, second=0, microsecond=0)
        service.set_events({"pair": "USDJPY", "events": [
            {"when": when.strftime("%Y-%m-%dT%H:%M"), "weights": {"USD": 1.5}, "adjust": 0.4,
             "label": "X"}]})
        table = service.event_weights()
        self.assertIn("USD", table["currencies"])
        row = next(e for e in table["events"] if e["label"] == "X")
        self.assertAlmostEqual(row["weights"]["USD"], 1.5)
        self.assertIn("USDJPY", row["pairs"])

        rows = [{"when": e["when"] + "Z", "label": e["label"],
                 "weights": dict(e["weights"])} for e in table["events"]]
        next(r for r in rows if r["label"] == "X")["weights"]["USD"] = 2.0
        r = service.set_event_weights({"weights": rows})
        self.assertIn("USD", r["currencies"])
        # USDJPY's own adjustment survived the currency-side replacement.
        uj = next(e for e in service.curve({"pair": "USDJPY"})["events"] if e["label"] == "X")
        self.assertAlmostEqual(uj["adjust"], 0.4)
        self.assertAlmostEqual(uj["bump"], 2.4)
        # And EURUSD, which has no cell of its own, moved with the weight.
        eu = next(e for e in service.curve({"pair": "EURUSD"})["events"] if e["label"] == "X")
        self.assertAlmostEqual(eu["bump"], 2.0)

        next(r for r in rows if r["label"] == "X")["weights"]["USD"] = "much"
        with self.assertRaises(ValueError):
            service.set_event_weights({"weights": rows})
        uj = next(e for e in service.curve({"pair": "USDJPY"})["events"] if e["label"] == "X")
        self.assertAlmostEqual(uj["bump"], 2.4)

    def test_events_can_be_set_and_reprice_their_bump(self):
        atm = self.book["USDJPY"].atm
        base = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])["USDJPY"].atm
        # 16:00 UTC sits just after the 14:00 roll, so the spike has its whole
        # volatility day ahead of it and no leakage warning is expected.
        when = (ASOF.now + timedelta(days=30)).replace(hour=16, minute=0, second=0, microsecond=0)
        self.assertEqual(atm.set_events([(when, 0.02, "TEST")]), [])
        self.assertAlmostEqual(atm.achieved_bump(atm.events.events[0]), 0.02, places=9)
        # It still lifts the calendar volatility day, just not by the same amount.
        self.assertGreater(atm.daily_vol(when) - base.daily_vol(when), 0.01)

    def test_event_near_the_day_roll_no_longer_distorts(self):
        """Under the legacy vol-day reading, a release shortly before 14:00 UTC
        was calibrated against a sliver of its own day: the height exploded and
        the overflow landed on the next day.  Forward-24h windows remove it."""
        atm = self.book["USDJPY"].atm
        early = (ASOF.now + timedelta(days=30)).replace(hour=16, minute=0, second=0, microsecond=0)
        late = (ASOF.now + timedelta(days=37)).replace(hour=13, minute=45, second=0, microsecond=0)
        problems = atm.set_events([(early, 0.015, "EARLY"), (late, 0.015, "LATE")])
        self.assertFalse(any("volatility-day roll" in p for p in problems), problems)
        heights = {e.label: e.height for e in atm.events.events}
        self.assertLess(abs(heights["LATE"] / heights["EARLY"] - 1.0), 0.35)
        for ev in atm.events.events:
            self.assertAlmostEqual(atm.achieved_bump(ev), 0.015, places=9)

    def test_legacy_vol_day_mode_still_available(self):
        from volkit.events import EventSchedule as ES
        atm = self.book["USDJPY"].atm
        atm.events = ES(window_mode="vol_day")
        late = (ASOF.now + timedelta(days=30)).replace(hour=13, minute=45, second=0, microsecond=0)
        problems = atm.set_events([(late, 0.015, "LATE")])
        self.assertTrue(any("volatility-day roll" in p for p in problems), problems)

    def test_setting_events_clears_the_previous_schedule(self):
        atm = self.book["USDJPY"].atm
        a = ASOF.now + timedelta(days=20)
        b = ASOF.now + timedelta(days=40)
        atm.set_events([(a, 0.01, "A"), (b, 0.01, "B")])
        self.assertEqual(len(atm.events.events), 2)
        atm.set_events([(b, 0.02, "B")])
        self.assertEqual([e.label for e in atm.events.events], ["B"])


FEED = Path(__file__).resolve().parents[1] / "files" / "market_feed.csv"


class TestExotics(unittest.TestCase):
    def setUp(self):
        self.S, self.vol, self.t, self.F = 100.0, 0.10, 1.0, 100.0
        self.drift = exotics.implied_drift(self.S, self.F, self.t)

    def test_touch_probability_matches_monte_carlo(self):
        """The closed form is the reference; the simulator must agree with it,
        because the bent-barrier case has only the simulator."""
        for B in (110.0, 90.0, 120.0):
            a = exotics.touch_probability(self.S, B, self.vol, self.t, self.drift)
            m, se = exotics._touch_mc(self.S, B, self.vol, self.t, self.drift,
                                      B > self.S, "extend", 0.0, True, 120_000, 192, 3)
            self.assertLess(abs(m - a), 4.0 * se, f"barrier {B}: analytic {a}, mc {m}+/-{se}")

    def test_barrier_at_spot_touches_immediately(self):
        self.assertEqual(exotics.touch_probability(100.0, 100.0, 0.1, 1.0, 0.0), 1.0)

    def test_zero_drift_is_log_symmetric(self):
        up = exotics.touch_probability(100, 110, 0.1, 1, 0.005)
        down = exotics.touch_probability(100, 100 * 100 / 110, 0.1, 1, 0.005)
        self.assertAlmostEqual(up, down, places=12)

    def test_one_touch_plus_no_touch_is_one(self):
        a = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5)
        b = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, is_no_touch=True)
        self.assertAlmostEqual(a.price + b.price, 1.0, places=12)

    def test_extend_costs_more_than_either_bend(self):
        kw = dict(buffer_pct=1.0, conservative=True, paths=40_000, steps=128)
        flat = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, mode="none").price
        ext = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, mode="extend", **kw).price
        bf = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, mode="bend_front", **kw).price
        bb = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, mode="bend_back", **kw).price
        for bent in (bf, bb):
            self.assertGreater(bent, flat)
            self.assertLess(bent, ext)

    def test_overhedge_side_flips_the_shift(self):
        sell = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, mode="extend",
                                 buffer_pct=1.0, conservative=True)
        buy = exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, mode="extend",
                                buffer_pct=1.0, conservative=False)
        self.assertLess(sell.barrier_used, 160.0)     # toward spot: easier to touch
        self.assertGreater(buy.barrier_used, 160.0)
        self.assertGreater(sell.price, buy.price)

    def test_unknown_overhedge_mode_raises(self):
        with self.assertRaises(ValueError):
            exotics.one_touch(150.0, 160.0, 0.09, 0.5, 149.5, mode="wobble")

    def test_digital_fair_value_includes_the_skew_term(self):
        """The digital is -dC/dK through the smile, so a sloping smile moves it
        away from N(d2); using N(d2) as the benchmark made the overhedge cost
        come out negative."""
        skewed = lambda K: 0.06 + 0.02 * (K - 100.0) / 100.0
        r = exotics.european_digital(100.0, 105.0, 1.0, 100.0, skewed, ramp_pct=0.0)
        self.assertNotAlmostEqual(r.fair_value, r.flat_vol_price, places=4)
        self.assertLess(r.fair_value, r.flat_vol_price)   # upward skew lowers a call digital

    def test_digital_ramp_is_monotone_and_never_cheaper(self):
        flat = lambda K: 0.08
        prev = -1.0
        for ramp in (0.0, 0.1, 0.25, 0.5, 1.0, 2.0):
            r = exotics.european_digital(100.0, 105.0, 1.0, 100.0, flat,
                                         ramp_pct=ramp, conservative=True)
            self.assertGreaterEqual(r.overhedge_cost, prev - 1e-12)
            prev = r.overhedge_cost
        self.assertGreater(prev, 0.0)

    def test_digital_ramp_converges_to_fair_value(self):
        """A one-sided ramp approximates the derivative at the ramp midpoint,
        so the gap closes linearly in the ramp width rather than instantly."""
        flat = lambda K: 0.08
        gaps = []
        for ramp in (0.4, 0.2, 0.1, 0.05, 0.025):
            r = exotics.european_digital(100.0, 105.0, 1.0, 100.0, flat, ramp_pct=ramp)
            gaps.append(abs(r.price - r.fair_value))
        for a, b in zip(gaps, gaps[1:]):
            self.assertLess(b, 0.6 * a)      # roughly halves each time
        self.assertLess(gaps[-1], 6e-4)

    def test_call_and_put_digitals_sum_to_one(self):
        flat = lambda K: 0.08
        c = exotics.european_digital(100.0, 105.0, 1.0, 100.0, flat, is_call=True).fair_value
        p = exotics.european_digital(100.0, 105.0, 1.0, 100.0, flat, is_call=False).fair_value
        self.assertAlmostEqual(c + p, 1.0, places=12)


class TestFeed(unittest.TestCase):
    def setUp(self):
        self.feed = MarketFeed.load(FEED)

    def test_sample_feed_loads_clean(self):
        self.assertEqual(self.feed.problems, [], self.feed.problems)
        self.assertIn("USDJPY", self.feed)

    def test_pip_divisor_by_term_currency(self):
        self.assertEqual(pip_divisor("USDJPY"), 100.0)
        self.assertEqual(pip_divisor("EURUSD"), 10000.0)

    def test_interpolation_is_exact_at_the_pillars(self):
        """Every pillar reads back exactly where the curve puts it.

        Where it puts it is its own **delivery date** -- a broker's 1M swap
        points are the points to the 1M value date, and a month is 30 or 31
        days and not a nominal 30.44.  Asking at ``tenor_to_years`` used to be
        the same question and no longer is; it is now a point a day or so away
        from the pillar, which is exactly the gap this placement closes.
        """
        pf = self.feed.pairs["USDJPY"]
        self.assertIsNotNone(pf.times)
        for t, pts in zip(pf.times, pf.points):
            got, _ = pf.forward_points(t)
            self.assertAlmostEqual(got, pts, places=10)

    def test_a_pillar_is_placed_on_its_own_delivery_date(self):
        """The axis is years from the spot date, and a tenor pillar is on it.

        Pinned because it is what makes an option's forward land *on* a
        pillar rather than between two of them: the option is read at its own
        settlement date, and the pillar is quoted to the same date.
        """
        from volkit.calendars import DEFAULT_CALENDARS
        pf = self.feed.pairs["USDJPY"]
        for tenor, t in zip(pf.tenors, pf.times):
            delivery = DEFAULT_CALENDARS.delivery_date("USDJPY", tenor, self.feed.today)
            self.assertAlmostEqual(
                t, (delivery - pf.spot_date).days / DAYS_IN_YEAR, places=12, msg=tenor)
        # 1M is 31 days here and not the 30.44 a nominal year fraction gives
        self.assertNotAlmostEqual(pf.times[pf.tenors.index("1M")],
                                  tenor_to_years("1M"), places=4)

    def test_interpolation_between_pillars_is_bracketed(self):
        pf = self.feed.pairs["USDJPY"]
        lo, _ = pf.forward_points(tenor_to_years("1m"))
        hi, _ = pf.forward_points(tenor_to_years("2m"))
        mid, extrap = pf.forward_points(0.5 * (tenor_to_years("1m") + tenor_to_years("2m")))
        self.assertFalse(extrap)
        self.assertTrue(min(lo, hi) < mid < max(lo, hi))

    def test_points_go_to_zero_at_the_very_front(self):
        pts, _ = self.feed.pairs["USDJPY"].forward_points(0.0)
        self.assertAlmostEqual(pts, 0.0, places=12)

    def test_beyond_the_last_pillar_is_flagged_not_trended(self):
        pf = self.feed.pairs["USDJPY"]
        far, extrap = pf.forward_points(5.0)
        last, _ = pf.forward_points(tenor_to_years("1y"))
        self.assertTrue(extrap)
        self.assertAlmostEqual(far, last, places=10)

    def test_quote_is_json_safe(self):
        import json
        q = self.feed.quote("USDJPY", 0.2)
        self.assertIsInstance(q["extrapolated"], bool)
        json.dumps(q)   # must not need a default= coercion

    def test_missing_file_and_unknown_pair_raise(self):
        with self.assertRaises(FeedError):
            MarketFeed.load("/nonexistent/feed.csv")
        with self.assertRaises(FeedError):
            self.feed.quote("XXXYYY", 0.2)


class TestDatedFeed(unittest.TestCase):
    """A feed line may name a date instead of a tenor.

    Added 2026-08-31.  The front of a bank's own forward file is not on
    standard tenors: it is the overnight and the tom-next, each quoted as one
    *day* of points rather than as points from spot, and neither of them has a
    tenor to be written as.  Reading them needs a spot date, which is why the
    valuation date is threaded in from the caller's clock rather than taken
    from the machine.
    """

    def feed(self, body, **kw):
        import tempfile
        tmp = tempfile.mkdtemp()
        path = Path(tmp) / "dated.csv"
        path.write_text(textwrap.dedent(body).lstrip(), encoding="utf-8")
        self.addCleanup(shutil.rmtree, tmp, True)
        return MarketFeed.load(path, **kw)

    def test_a_dated_pillar_lands_where_the_same_tenor_does(self):
        """One axis for both spellings, so a mixed file has no seam in it.

        A date 28 days after spot and the tenor a desk would call it are the
        same point on the curve, not two.
        """
        f = self.feed("""
            USDJPY,SPOT,150.25
            USDJPY,SPOT DATE,2026-09-01
            USDJPY,2026-09-29,-11.40
        """)
        self.assertEqual(f.problems, [], f.problems)
        pf = f.pairs["USDJPY"]
        self.assertEqual(pf.tenors, ["2026-09-29"])
        self.assertAlmostEqual(pf.forward_points(28 / DAYS_IN_YEAR)[0], -11.40, places=10)
        # and read by the date it was quoted at, which is the same knot
        self.assertAlmostEqual(pf.points_on("2026-09-29")[0], -11.40, places=10)
        self.assertAlmostEqual(pf.forward_on("2026-09-29")[0], 150.25 - 0.1140, places=10)

    def test_a_date_on_or_before_spot_is_one_day_of_points(self):
        """The near side is stated as rates and accumulated back from spot.

        T/N is the day ending on the spot date, so it prices the day *before*
        spot; O/N the one before that.  Cumulative points at the spot date are
        zero by definition, and everything on the near side is negative of the
        sum of the days between.
        """
        f = self.feed("""
            # asof: 2026-08-27
            USDJPY,SPOT,150.25
            USDJPY,SPOT DATE,2026-08-31
            USDJPY,2026-08-31,-0.40
            USDJPY,2026-08-28,-0.10
        """)
        self.assertEqual(f.problems, [], f.problems)
        pf = f.pairs["USDJPY"]
        self.assertEqual([d.isoformat() for d, _ in pf.daily],
                         ["2026-08-28", "2026-08-31"])
        self.assertEqual(pf.tenors, [])          # nothing on the far side
        self.assertAlmostEqual(pf.points_on("2026-08-31")[0], 0.0, places=12)
        # the day ending on spot is the tom-next: it prices spot minus one
        self.assertAlmostEqual(pf.points_on("2026-08-30")[0], +0.40, places=12)
        # 29-Aug and 30-Aug are unquoted, so their rate is interpolated
        # between -0.10 (28th) and -0.40 (31st): -0.20 and -0.30.
        self.assertAlmostEqual(pf.points_on("2026-08-29")[0], 0.40 + 0.30, places=12)
        self.assertAlmostEqual(pf.points_on("2026-08-28")[0], 0.70 + 0.20, places=12)
        self.assertAlmostEqual(pf.points_on("2026-08-27")[0], 0.90 + 0.10, places=12)
        # and a negative swap point means the earlier date is the higher rate
        self.assertGreater(pf.forward_on("2026-08-27")[0], pf.spot)

    def test_the_near_side_interpolates_rates_and_not_the_running_total(self):
        """The rows are rates, so it is the rates that are interpolated.

        A straight line through the *cumulative* knots at the 28th and the
        31st would put the 30th at +0.30.  It is +0.40, because the day being
        skipped over is the expensive one.
        """
        f = self.feed("""
            # asof: 2026-08-27
            EURUSD,SPOT,1.0842
            EURUSD,SPOT DATE,2026-08-31
            EURUSD,2026-08-31,-0.40
            EURUSD,2026-08-28,-0.10
        """)
        pf = f.pairs["EURUSD"]
        straight_line = 0.90 * (31 - 30) / (31 - 28)
        self.assertNotAlmostEqual(pf.points_on("2026-08-30")[0], straight_line, places=6)
        self.assertAlmostEqual(pf.points_on("2026-08-30")[0], 0.40, places=12)

    def test_a_date_already_delivered_is_passed_over_and_said(self):
        """A forward that has already delivered is not a forward.

        Passed over rather than refused: a published file carrying yesterday's
        row is an ordinary thing, and it is a note and not a problem.  What it
        may never be is silently placed, which would put a knot behind the
        valuation date and drag the front of the curve onto it.
        """
        f = self.feed("""
            # asof: 2026-08-27
            USDJPY,SPOT,150.25
            USDJPY,2026-08-26,-0.10
            USDJPY,2026-09-30,-11.40
        """)
        self.assertEqual(f.problems, [], f.problems)
        self.assertTrue(any("2026-08-26" in n and "passed over" in n for n in f.notes), f.notes)
        pf = f.pairs["USDJPY"]
        self.assertEqual(pf.daily, [])
        self.assertEqual(pf.tenors, ["2026-09-30"])

    def test_a_dated_row_with_no_valuation_date_anywhere_is_refused(self):
        """It is never guessed at.

        The spot date places every dated row, and a spot date needs a day to
        count from.  A wall-clock reading here would be the one call to
        ``utcnow`` inside the model, and would move the whole near side of the
        curve on a valuation in the past.
        """
        f = self.feed("""
            USDJPY,SPOT,150.25
            USDJPY,2026-09-30,-11.40
        """)
        self.assertTrue(any("valuation date" in p for p in f.problems), f.problems)
        self.assertEqual(f.pairs["USDJPY"].tenors, [])
        # and given one, the same file reads
        f2 = self.feed("""
            USDJPY,SPOT,150.25
            USDJPY,2026-09-30,-11.40
        """, today=date(2026, 8, 27))
        self.assertEqual(f2.problems, [], f2.problems)
        self.assertEqual(f2.pairs["USDJPY"].tenors, ["2026-09-30"])

    def test_a_stated_spot_date_beats_the_calendar_and_says_so(self):
        """A publisher knows its own holidays; this tool's calendar may not.

        The near side is placed against that date, so a day's disagreement
        moves the tom-next onto the overnight.  Stated, it is used and said;
        derived, that is said too, because the two must not read alike.
        """
        derived = self.feed("""
            # asof: 2026-08-27
            USDJPY,SPOT,150.25
            USDJPY,2026-09-30,-11.40
        """)
        self.assertEqual(derived.pairs["USDJPY"].spot_date, date(2026, 8, 31))
        self.assertTrue(any("business days after" in n for n in derived.notes), derived.notes)

        stated = self.feed("""
            # asof: 2026-08-27
            USDJPY,SPOT,150.25
            USDJPY,SPOT DATE,2026-09-01
            USDJPY,2026-09-30,-11.40
        """)
        self.assertEqual(stated.pairs["USDJPY"].spot_date, date(2026, 9, 1))
        self.assertTrue(any("as the file states it" in n for n in stated.notes), stated.notes)
        # one day of spot date is one day of curve: the same pillar sits a
        # day further out, so every point interpolated inside it moves.
        self.assertNotAlmostEqual(derived.pairs["USDJPY"].forward_points(0.04)[0],
                                  stated.pairs["USDJPY"].forward_points(0.04)[0], places=6)

    def test_the_callers_clock_beats_the_files_asof_and_the_difference_is_said(self):
        """A valuation in the past against this morning's file is ordinary.

        It is also a fact about what is being priced, so it is reported rather
        than absorbed into a spot date nobody can check.
        """
        f = self.feed("""
            # asof: 2026-08-27
            USDJPY,SPOT,150.25
            USDJPY,2026-09-30,-11.40
        """, today=date(2026, 8, 20))
        self.assertEqual(f.today, date(2026, 8, 20))
        self.assertTrue(any("written as of 2026-08-27" in n for n in f.notes), f.notes)

    def test_tenor_and_dated_rows_mix_in_one_curve(self):
        """A file need not choose: the front dated, the back on tenors."""
        f = self.feed("""
            # asof: 2026-08-27
            USDJPY,SPOT,150.25
            USDJPY,SPOT DATE,2026-08-31
            USDJPY,2026-08-31,-0.40
            USDJPY,2026-09-30,-11.40
            USDJPY,3M,-35.0
            USDJPY,1Y,-146.0
        """)
        self.assertEqual(f.problems, [], f.problems)
        pf = f.pairs["USDJPY"]
        self.assertEqual(pf.tenors, ["2026-09-30", "3M", "1Y"])   # sorted by time
        # The 3M pillar reads back exactly where the curve puts it, which is
        # its own delivery date and not a nominal 30.44-day quarter.
        self.assertAlmostEqual(
            pf.forward_points(pf.times[pf.tenors.index("3M")])[0], -35.0, places=10)
        self.assertAlmostEqual(pf.points_on("2026-09-30")[0], -11.40, places=10)
        self.assertAlmostEqual(pf.points_on("2026-08-30")[0], +0.40, places=10)

    def test_a_label_that_is_neither_says_both(self):
        """The tenor is tried first, so nothing a tenor feed reads moves."""
        f = self.feed("""
            USDJPY,SPOT,150.25
            USDJPY,banana,-11.40
        """)
        self.assertEqual(len(f.problems), 1, f.problems)
        self.assertIn("cannot parse tenor", f.problems[0])
        self.assertIn("not a date either", f.problems[0])

    def test_a_tenor_feed_reads_back_at_its_own_pillars(self):
        """The sample feed's every pillar, and its front and back ends.

        The file states no spot date, but it carries an ``asof`` line, and a
        valuation date is all a spot date needs -- so this feed has one and
        every tenor pillar is placed on its own delivery date.  The pillar
        *values* are untouched by any of that; what moved is where they sit,
        and they read back exactly there.
        """
        f = MarketFeed.load(FEED)
        self.assertEqual(f.problems, [], f.problems)
        pf = f.pairs["USDJPY"]
        self.assertEqual(pf.spot_date, DEFAULT_CALENDARS.spot_date("USDJPY", f.today))
        for t, pts in zip(pf.times, pf.points):
            self.assertAlmostEqual(pf.forward_points(t)[0], pts, places=10)
        self.assertAlmostEqual(pf.forward_points(0.0)[0], 0.0, places=12)
        # the front pillar is still scaled toward zero at the spot date
        half = 0.5 * pf.times[0]
        self.assertAlmostEqual(pf.forward_points(half)[0], 0.5 * pf.points[0], places=10)
        self.assertFalse(pf.forward_points(half)[1])
        self.assertTrue(pf.forward_points(5.0)[1])

    def test_the_feed_is_read_against_the_books_clock(self):
        """``load_for`` is the one caller-facing spelling, and it injects it.

        Every screen and every command goes through it, so a dated feed cannot
        be placed one way on a screen and another way in the batch command
        beside it.
        """
        import tempfile
        from volkit.feed import load_for

        class FakeBook:
            clock = Clock(datetime(2026, 8, 27, 10, 0, tzinfo=UTC))
            calendars = DEFAULT_CALENDARS

        tmp = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, tmp, True)
        path = Path(tmp) / "f.csv"
        path.write_text("USDJPY,SPOT,150.25\nUSDJPY,2026-09-30,-11.40\n", encoding="utf-8")
        f = load_for(FakeBook(), path)
        self.assertEqual(f.today, date(2026, 8, 27))
        self.assertEqual(f.pairs["USDJPY"].spot_date, date(2026, 8, 31))


class TestImpliedCrossFeed(unittest.TestCase):
    """A cross the file does not quote is implied from its two legs.

    Two spot rates and two swap points are all an implied cross rate has ever
    been, and a file that quotes EURUSD and USDJPY *is* quoting EURJPY.  The
    arithmetic lives in `feed.compose_level` and nowhere else: `Book`
    contributes only the workbook's opinion about which legs a cross has, so
    there is one place for the triangle's signs to be written (§5 item 1 is
    what a second place costs).
    """

    def feed(self, body, **kw):
        import tempfile
        tmp = tempfile.mkdtemp()
        path = Path(tmp) / "cross.csv"
        path.write_text(textwrap.dedent(body).lstrip(), encoding="utf-8")
        self.addCleanup(shutil.rmtree, tmp, True)
        return MarketFeed.load(path, **kw)

    def test_the_feed_itself_implies_a_cross_it_does_not_quote(self):
        """It used to refuse by name, and only ``Book`` knew the triangle.

        Anything holding a feed and not a book -- the feed status route's own
        quote box among them -- got ``no feed for 'EURJPY'`` off a file that
        was quoting both of its legs.
        """
        f = MarketFeed.load(FEED)
        q = f.quote("EURJPY", 0.25)
        a, b = f.quote("EURUSD", 0.25), f.quote("USDJPY", 0.25)
        self.assertTrue(q["derived"])
        self.assertEqual(q["via"], "EURUSD and USDJPY")
        self.assertAlmostEqual(q["spot"], a["spot"] * b["spot"], places=12)
        self.assertAlmostEqual(q["forward"], a["forward"] * b["forward"], places=12)
        # the points are the cross's own, in the cross's own pips, and are
        # never the legs' points added
        self.assertEqual(q["pip"], 100.0)
        self.assertAlmostEqual(q["points"], (q["forward"] - q["spot"]) * 100.0, places=10)
        self.assertNotAlmostEqual(q["points"], a["points"] + b["points"], places=3)

    def test_a_divided_cross_is_divided(self):
        """EURGBP is EURUSD over GBPUSD, and the sign comes from one place."""
        f = MarketFeed.load(FEED)
        q = f.quote("EURGBP", 0.25)
        a, b = f.quote("EURUSD", 0.25), f.quote("GBPUSD", 0.25)
        self.assertEqual(q["via"], "EURUSD and GBPUSD")
        self.assertAlmostEqual(q["spot"], a["spot"] / b["spot"], places=12)
        self.assertAlmostEqual(q["forward"], a["forward"] / b["forward"], places=12)

    def test_a_cross_nobody_declared_still_has_its_dollar_legs(self):
        """GBPJPY is not in the sample workbook, and the market quotes it.

        The legs of a cross are a fact about its name, not a decision -- which
        is why `cross.dollar_legs` exists and why CONFIG stopped asking for
        them.  Refusing a level for want of a row in a spreadsheet that has
        nothing to do with the feed was the same refusal-by-name in a
        different place.
        """
        f = MarketFeed.load(FEED)
        self.assertNotIn("GBPJPY", f)
        q = f.quote("GBPJPY", 0.25)
        self.assertEqual(q["via"], "GBPUSD and USDJPY")
        self.assertAlmostEqual(q["forward"],
                               f.quote("GBPUSD", 0.25)["forward"]
                               * f.quote("USDJPY", 0.25)["forward"], places=12)

    def test_half_a_triangle_is_still_a_refusal_and_names_the_missing_leg(self):
        """A guessed leg is a level nobody published wearing a published one."""
        f = MarketFeed.load(FEED)
        with self.assertRaises(FeedError) as caught:
            f.quote("GBPNZD", 0.25)
        self.assertIn("NZDUSD", str(caught.exception))
        self.assertIn("GBPUSD and NZDUSD", str(caught.exception))

    def test_the_workbook_wins_when_it_names_a_crosss_legs(self):
        """A sheet that says something is not second-guessed by a convention."""
        f = MarketFeed.load(FEED)
        named = {"EURJPY": ("EURUSD", "USDJPY")}
        q = f.quote("EURJPY", 0.25, lambda p: named.get(p))
        self.assertEqual(q["via"], "EURUSD and USDJPY")
        # and a pair the caller has no opinion about falls through to the
        # dollar legs rather than being refused
        self.assertEqual(f.quote("AUDJPY", 0.25, lambda p: named.get(p))["via"],
                         "AUDUSD and USDJPY")

    def test_the_book_still_gets_exactly_the_numbers_it_did(self):
        """The composition moved into the feed; not one figure moved with it."""
        book = Book.from_excel(WORKBOOK, ASOF)
        book.feed = MarketFeed.load(FEED)
        for pair, via in [("EURJPY", "EURUSD and USDJPY"),
                          ("EURGBP", "EURUSD and GBPUSD"),
                          ("AUDJPY", "AUDUSD and USDJPY")]:
            level = book.market_level(pair, 0.25)
            legs = via.split(" and ")
            a = book.market_level(legs[0], 0.25)
            b = book.market_level(legs[1], 0.25)
            self.assertTrue(level["feed"])
            self.assertTrue(level["derived"])
            self.assertEqual(level["via"], via)
            expect = (a["forward"] * b["forward"] if pair != "EURGBP"
                      else a["forward"] / b["forward"])
            self.assertAlmostEqual(level["forward"], expect, places=12)
        # and a dollar pair the file quotes is untouched by any of it
        self.assertFalse(book.market_level("EURUSD", 0.25)["derived"])

    def test_an_implied_cross_reaches_the_near_side_too(self):
        """The overnight of a cross is the two legs' overnights, composed.

        Read **on the date** rather than at a time, so each leg is placed
        against its own spot date: a cross of a T+1 pair and a T+2 pair has
        two different dates at one ``t``, and the tom-next is a day wide.
        """
        f = self.feed("""
            # asof: 2026-08-27
            EURUSD,SPOT,1.0842
            EURUSD,SPOT DATE,2026-08-31
            EURUSD,2026-08-28,0.03
            EURUSD,2026-08-31,0.12
            EURUSD,2026-09-30,5.8
            USDJPY,SPOT,150.25
            USDJPY,SPOT DATE,2026-08-31
            USDJPY,2026-08-28,-0.06
            USDJPY,2026-08-31,-0.24
            USDJPY,2026-09-30,-11.4
        """)
        self.assertEqual(f.problems, [], f.problems)
        for when in ["2026-08-27", "2026-08-30", "2026-08-31", "2026-09-30"]:
            q = f.quote_on("EURJPY", when)
            a, b = f.quote_on("EURUSD", when), f.quote_on("USDJPY", when)
            self.assertTrue(q["derived"], when)
            self.assertAlmostEqual(q["forward"], a["forward"] * b["forward"],
                                   places=10, msg=when)
        # the cross's points are zero on its own spot date, by definition,
        # and the near side is on the other side of it
        self.assertAlmostEqual(f.quote_on("EURJPY", "2026-08-31")["points"], 0.0, places=10)
        self.assertGreater(f.quote_on("EURJPY", "2026-08-30")["points"], 0.0)
        self.assertLess(f.quote_on("EURJPY", "2026-09-30")["points"], 0.0)

    def test_a_dated_read_of_a_cross_needs_both_legs_to_have_a_spot_date(self):
        """A leg with no valuation date behind it has no dates on it, and says so.

        A file that states a spot date, or one loaded against a clock, has one
        for every pair -- the spot date is derived from the valuation date, and
        that is what lets a *tenor* pillar be placed on its own delivery date.
        With neither there is nothing to derive it from, and a dated read is
        refused by name rather than answered from a guessed origin.
        """
        f = self.feed("""
            EURUSD,SPOT,1.0842
            EURUSD,1M,5.8
            USDJPY,SPOT,150.25
            USDJPY,1M,-11.4
        """)                            # no asof line, and loaded with no clock
        self.assertIsNone(f.today)
        self.assertIsNone(f.pairs["USDJPY"].spot_date)
        with self.assertRaises(FeedError) as caught:
            f.quote_on("EURJPY", "2026-09-30")
        self.assertIn("spot date", str(caught.exception))


class TestExoticLegs(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        cls.book.feed = MarketFeed.load(FEED)

    def _one(self, **kw):
        return price_strip(self.book, [OptionLeg("USDJPY", "3M", **kw)])["legs"][0]

    def test_feed_fills_spot_and_forward(self):
        r = self._one(strike="ATM")
        self.assertTrue(r["feed_used"])
        self.assertAlmostEqual(r["spot"], 150.25, places=10)
        self.assertNotAlmostEqual(r["forward"], r["spot"], places=6)

    def test_explicit_spot_overrides_the_feed(self):
        r = self._one(strike="ATM", spot=160.0, forward_points=0.0, pip=100.0)
        self.assertFalse(r["feed_used"])
        self.assertAlmostEqual(r["spot"], 160.0, places=10)

    def test_touch_leg_needs_a_barrier(self):
        r = self._one(product="one_touch")
        self.assertFalse(r["ok"])
        self.assertIn("barrier", r["error"])

    def test_one_touch_and_no_touch_sum_to_one(self):
        a = self._one(product="one_touch", barrier="158")
        b = self._one(product="no_touch", barrier="158")
        self.assertAlmostEqual(a["premium_dom"] + b["premium_dom"], 1.0, places=10)
        self.assertAlmostEqual(a["delta_pct"], -b["delta_pct"], places=8)

    def test_overhedge_raises_both_price_and_risk(self):
        flat = self._one(product="one_touch", barrier="158")
        buff = self._one(product="one_touch", barrier="158", overhedge="extend",
                         buffer_pct=0.5, conservative=True)
        self.assertGreater(buff["premium_dom"], flat["premium_dom"])
        self.assertGreater(abs(buff["delta_pct"]), abs(flat["delta_pct"]))
        self.assertGreater(buff["overhedge_cost"], 0.0)

    def test_bent_barrier_reports_monte_carlo_error(self):
        r = self._one(product="one_touch", barrier="158", overhedge="bend_front",
                      buffer_pct=0.5)
        self.assertIn("monte carlo", r["pricing_method"])
        self.assertGreater(r["mc_error"], 0.0)

    def test_digital_ramp_raises_price_and_delta(self):
        flat = self._one(product="digital", strike="155", option_type="C")
        ramp = self._one(product="digital", strike="155", option_type="C", ramp_pct=1.0)
        self.assertGreater(ramp["premium_dom"], flat["premium_dom"])
        self.assertGreater(ramp["delta_pct"], flat["delta_pct"])

    def test_unknown_product_is_reported_per_leg(self):
        r = self._one(product="rainbow")
        self.assertFalse(r["ok"])
        self.assertIn("rainbow", r["error"])


class TestImpliedRRFly(unittest.TestCase):
    def test_anchoring_makes_the_surface_reproduce_its_quotes(self):
        """The marking check the implied-vs-quoted table displays."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        surface = book["USDJPY"]
        mark = {m.tenor.upper(): m for m in book.data.marks["USDJPY"]}["3M"]
        expiry = ASOF.datetime_from_years(tenor_to_years("3m"))
        loose = abs(surface.risk_reversal(expiry, 0.25) - mark.rr_25)
        surface.anchor_tenors = True
        surface._slices.clear()
        tight = abs(surface.risk_reversal(expiry, 0.25) - mark.rr_25)
        self.assertLess(tight, loose)
        self.assertLess(tight, 1e-4)


class TestBandedSmile(unittest.TestCase):
    def setUp(self):
        self.band = Band("USDHKD", 7.75, 7.85)
        self.F, self.t, self.atm = 7.8020, 0.25, 0.004424

    def _fit(self, t=None, **kw):
        return calibrate_band_smile(self.band, self.F, t or self.t, self.atm,
                                    conv=DeltaConvention(True), **kw)

    def test_band_validation(self):
        for lo, hi in ((8.0, 7.0), (0.0, 0.0)):
            with self.assertRaises(ValueError):
                Band("X", lo, hi)

    def test_forward_outside_the_band_is_rejected(self):
        with self.assertRaises(ValueError):
            calibrate_band_smile(self.band, 7.90, self.t, self.atm)

    def test_atm_and_forward_are_matched_exactly(self):
        smile, rep = self._fit()
        self.assertTrue(rep["converged"], rep)
        self.assertAlmostEqual(smile.mean, self.F, places=9)
        self.assertLess(abs(rep["atm_residual_vol"]), 1e-8)

    def test_breach_probability_is_positive_and_real(self):
        """The peg can break; that probability belongs in the price.  It is an
        output of the hazard, not something forced to zero."""
        _, rep = self._fit(jump=JumpSpec(hazard=0.02))
        self.assertGreater(rep["prob_outside_band"], 0.0)
        self.assertLess(rep["prob_outside_band"], rep["prob_broken"])

    def test_breach_probability_compounds_with_the_hazard(self):
        """A hazard rate composes across the term structure; a per-tenor
        probability would not."""
        probs = []
        for t in (1 / 12, 0.25, 0.5, 1.0):
            _, rep = self._fit(t=t, jump=JumpSpec(hazard=0.02))
            probs.append(rep["prob_broken"])
            self.assertAlmostEqual(rep["prob_broken"], 1 - math.exp(-0.02 * t), places=12)
        self.assertTrue(all(a < b for a, b in zip(probs, probs[1:])))

    def test_analytic_breach_matches_monte_carlo(self):
        smile, rep = self._fit(jump=JumpSpec(hazard=0.05))
        rng = np.random.default_rng(4)
        hold, pw, ps = smile.weights
        mw, ms = smile.break_levels
        n = 400_000
        outside = 0.0
        for p, level, vol in ((pw, mw, smile.jump.weak_vol), (ps, ms, smile.jump.strong_vol)):
            sq = vol * math.sqrt(self.t)
            S = level * np.exp(-0.5 * sq * sq + sq * rng.standard_normal(n))
            outside += p * float(((S < self.band.lower) | (S > self.band.upper)).mean())
        self.assertAlmostEqual(rep["prob_outside_band"], outside, places=4)

    def test_options_outside_the_band_carry_jump_value(self):
        smile, _ = self._fit(jump=JumpSpec(hazard=0.02))
        prices = [float(smile.call_price(K)) for K in (7.86, 7.90, 8.00, 8.50)]
        self.assertTrue(all(p > 0 for p in prices), prices)
        self.assertTrue(all(a > b for a, b in zip(prices, prices[1:])))

    def test_expected_devaluation_shifts_the_peg_intact_distribution(self):
        """With an asymmetric break the forward constraint pushes the in-band
        mean to the strong side; that is a real effect, not an artefact."""
        _, rep = self._fit(jump=JumpSpec(hazard=0.01, weak_share=1.0, weak_jump=0.05))
        self.assertLess(rep["in_band_mean_shift"], 0.0)
        _, flat = self._fit(jump=JumpSpec(hazard=0.0))
        self.assertAlmostEqual(flat["in_band_mean_shift"], 0.0, places=9)

    def test_beta_can_represent_a_u_shape(self):
        """A logit-normal or lognormal cannot; the realised peg distribution is
        U-shaped because the authority defends the edges."""
        u = BetaBandSmile(self.band, 0.6, 0.6, self.t, self.F, JumpSpec(hazard=0.0))
        self.assertTrue(u.u_shaped)
        d = u.density(np.array([7.755, 7.80, 7.845]))
        self.assertGreater(d[0], d[1])
        self.assertGreater(d[2], d[1])

    def test_call_put_parity_holds_against_the_model_forward(self):
        smile, _ = self._fit(jump=JumpSpec(hazard=0.02))
        for K in (7.60, 7.76, 7.80, 7.84, 8.10):
            c, p = float(smile.call_price(K)), float(smile.put_price(K))
            self.assertAlmostEqual(c - p, smile.mean - K, places=12)

    def test_implied_vol_reprices_the_model_everywhere(self):
        smile, _ = self._fit(jump=JumpSpec(hazard=0.02))
        for K in (7.70, 7.77, 7.80, 7.83, 7.95):
            v = smile.implied_vol(K)
            self.assertTrue(np.isfinite(v), K)
            is_call = K >= smile.mean
            ref = float(smile.call_price(K)) if is_call else float(smile.put_price(K))
            self.assertAlmostEqual(float(black.price(smile.mean, K, v, self.t, is_call)),
                                   ref, places=10)

    def test_density_integrates_to_one_including_the_tails(self):
        smile, _ = self._fit(jump=JumpSpec(hazard=0.05))
        g = np.linspace(4.0, 16.0, 300_001)
        self.assertAlmostEqual(float(np.trapezoid(smile.density(g), g)), 1.0, places=4)

    def test_atm_floor_from_the_hazard_is_diagnosed(self):
        """Break risk alone sets a floor under the ATM volatility.  A quote
        below it is not a solver failure but a statement about the marks."""
        with self.assertRaises(ConvergenceError) as ctx:
            self._fit(jump=JumpSpec(hazard=0.05, weak_share=1.0, weak_jump=0.08))
        self.assertIn("inconsistent", str(ctx.exception))
        self.assertIn("at least", str(ctx.exception))

    def test_forward_constraint_failure_is_diagnosed_separately(self):
        """Enough expected devaluation and the peg-intact mean would have to sit
        outside the band for the forward to match at all."""
        with self.assertRaises(ValueError) as ctx:
            self._fit(t=5.0, jump=JumpSpec(hazard=0.05))
        self.assertIn("outside the band", str(ctx.exception))

    def test_hazard_inversion_responds_to_the_assumed_jump(self):
        """Backing the hazard out of the wings must actually move with the
        assumption, and say so when it cannot be done."""
        got = []
        for wj, sv in ((0.03, 0.06), (0.06, 0.10), (0.12, 0.18)):
            _, rep = self._fit(risk_reversal=0.0022, strangle=0.0014, solve_hazard=True,
                               jump=JumpSpec(hazard=0.02, weak_jump=wj,
                                             strong_jump=wj * 0.7, weak_vol=sv,
                                             strong_vol=sv * 0.8))
            got.append(rep["hazard"])
        self.assertTrue(all(a > b for a, b in zip(got, got[1:])), got)
        self.assertGreater(got[0] / got[-1], 2.0)

    def test_jump_spec_validation(self):
        for kw in ({"hazard": -1.0}, {"weak_share": 1.5}, {"weak_vol": 0.0}):
            with self.assertRaises(ValueError):
                JumpSpec(**kw)

    def test_the_peg_bands_tab_loads_and_rejects_degenerate_rows(self):
        bands = load_bands(Path(__file__).resolve().parents[1] / "files" / "vol_marks.xlsx")
        self.assertIn("USDHKD", bands)
        self.assertEqual(bands["USDHKD"].lower, 7.75)

    def test_a_workbook_with_no_peg_bands_tab_says_so_by_name(self):
        """The tab is asked for by name, so its absence is named, not shrugged at."""
        legacy = Path(__file__).resolve().parents[1] / "files" / "vol_marks_legacy_format.xlsx"
        with self.assertRaises(ValueError) as ctx:
            load_bands(legacy)
        self.assertIn("PEG_BANDS", str(ctx.exception))


class TestConfigurationTabs(unittest.TestCase):
    """The settings that used to be a CSV each, now tabs of the workbook.

    Three things are pinned: the reader's own conventions (a header found
    below prose, a '#' row skipped wherever it sits, an absent tab answered
    with None rather than an empty table), that the shipped workbook actually
    carries all three tabs, and that the holidays it lists reach the book
    without reaching every *other* book in the process.
    """

    def _workbook(self, rows, sheet="PEG_BANDS"):
        import tempfile
        import openpyxl
        d = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, d, True)
        wb = openpyxl.Workbook()
        wb.active.title = sheet
        for row in rows:
            wb.active.append(row)
        path = d / "marks.xlsx"
        wb.save(path)
        return path

    def test_an_absent_tab_is_none_and_an_empty_one_is_an_empty_list(self):
        from volkit import configsheets
        path = self._workbook([["pair", "lower", "upper"]])
        self.assertIsNone(configsheets.read_rows(path, "HOLIDAYS",
                                                 required=("country", "date")))
        self.assertEqual(configsheets.read_rows(path, "PEG_BANDS",
                                                required=("pair", "lower", "upper")), [])

    def test_the_header_is_found_below_prose_and_comments_are_skipped(self):
        from volkit import configsheets
        path = self._workbook([
            ["# what this tab is for"],
            ["# and a second line of it"],
            ["Pair", "Lower", "Upper", "Note"],
            ["USDHKD", 7.75, 7.85, "the peg"],
            ["# deliberately not listed: USDCNY"],
            ["USDXXX", 1.0, 2.0, None],
        ])
        rows = configsheets.read_rows(path, "PEG_BANDS",
                                      required=("pair", "lower", "upper"))
        self.assertEqual([r.text("pair") for r in rows], ["USDHKD", "USDXXX"])
        # The row number is the one Excel shows, so an error can be looked up.
        self.assertEqual([r.number for r in rows], [4, 6])
        self.assertEqual(rows[0].real("lower"), 7.75)
        self.assertEqual(rows[0].text("note"), "the peg")
        self.assertEqual(rows[1].text("note"), "")

    def test_a_tab_with_no_header_at_all_is_refused_by_name(self):
        from volkit import configsheets
        path = self._workbook([["# nothing but prose"], ["# and more of it"]])
        with self.assertRaises(configsheets.ConfigSheetError) as ctx:
            configsheets.read_rows(path, "PEG_BANDS", required=("pair", "lower", "upper"))
        self.assertIn("PEG_BANDS", str(ctx.exception))

    def test_the_shipped_workbook_carries_every_configuration_tab(self):
        """Every tab the tool needs -- which is not quite every tab it reads.

        ``WING_RATIOS`` is the exception on purpose: a workbook that has not
        been through ``volkit migrate-wings`` quotes all four wings itself,
        which is what every workbook did before the tab existed, and a
        configuration a workbook is allowed not to have must not be asserted
        into existence by a test.  The rest are not optional -- a missing
        ``PEG_BANDS`` is a managed pair with no band and a screen that says
        nothing about it.
        """
        from volkit import configsheets
        optional = {"WING_RATIOS"}
        needed = [s for s in configsheets.SHEETS if s not in optional]
        self.assertEqual([s for s in configsheets.present(WORKBOOK) if s not in optional],
                         needed)

    def test_the_holidays_tab_reaches_the_book_and_no_further(self):
        """A lunar holiday belongs to the workbook that lists it.

        The overrides were loadable but never loaded before this: a Chinese
        New Year in the file moved no expiry.  They are the book's calendars
        now -- and a *copy*, because a book that added dates to the shared set
        would change the expiry of every book loaded after it.
        """
        from volkit.calendars import DEFAULT_CALENDARS
        book = Book.from_excel(WORKBOOK, ASOF)
        cny = date(2026, 2, 17)
        self.assertTrue(book.calendars.is_holiday("CNH", cny))
        self.assertIsNot(book.calendars, DEFAULT_CALENDARS)
        self.assertEqual(DEFAULT_CALENDARS.overrides, {})

    def test_a_workbook_with_no_holidays_tab_keeps_the_shared_calendars(self):
        from volkit.calendars import DEFAULT_CALENDARS
        legacy = WORKBOOK.parent / "vol_marks_legacy_format.xlsx"
        self.assertIs(Book.from_excel(legacy, ASOF).calendars, DEFAULT_CALENDARS)


class TestBreakRegimeFit(unittest.TestCase):
    """Stage B of the band calibration: the break regime from the wings.

    The body is exact from the ATM and the forward, as it always was; what is
    new is that the hazard and the share of breaks (and any other break
    parameter somebody frees) come out of both wings at both deltas by least
    squares, per tenor or across the whole curve, with the residual of every
    quote reported and the identifiability *measured* at the answer.
    """

    BAND = Band("USDHKD", 7.75, 7.85)
    F = 7.8020
    TRUTH = JumpSpec(hazard=0.03, weak_share=0.7)
    START = JumpSpec(hazard=0.01, weak_share=0.85)

    def quotes(self, t, atm, deltas=(0.25, 0.10), truth=None):
        """Wings a surface under ``truth`` would quote: the model's own implied
        volatility at the model's own delta strikes, iterated to a fixed point."""
        from volkit.banded import WingQuote
        conv = DeltaConvention(True)
        sm, _ = calibrate_band_smile(self.BAND, self.F, t, atm, jump=truth or self.TRUTH, conv=conv)
        out = []
        for d in deltas:
            vc = vp = atm
            for _ in range(40):
                kc = black.strike_from_delta(d, self.F, vc, t, True, conv)
                kp = black.strike_from_delta(-d, self.F, vp, t, False, conv)
                vc, vp = sm.implied_vol(kc), sm.implied_vol(kp)
            out.append(WingQuote(d, vc - vp, 0.5 * (vc + vp) - atm))
        return tuple(out)

    def test_solve_hazard_is_the_one_parameter_one_instrument_case(self):
        """The wrapper: hazard alone against the strangle premium, bracketed,
        and the number it gives is the number it gave before the refit."""
        _, rep = calibrate_band_smile(self.BAND, self.F, 0.25, 0.004424, risk_reversal=0.0022,
                                      strangle=0.0014, solve_hazard=True,
                                      conv=DeltaConvention(True))
        # The number the pre-refit solver gave for this case, to the last digit.
        self.assertAlmostEqual(rep["hazard"], 0.05006201744360903, places=12)
        self.assertEqual(rep["fit"]["method"], "bracketed solve")
        self.assertEqual(rep["fit"]["free"], ["hazard"])
        self.assertTrue(rep["fit"]["converged"])
        self.assertEqual(len(rep["fit"]["residuals"]), 1)
        self.assertLess(abs(rep["fit"]["residuals"][0]["residual"]), 1e-10)

    def test_a_planted_regime_is_recovered_from_one_tenors_wings(self):
        from volkit.banded import calibrate_band_wings
        t, atm = 0.25, 0.0044
        _, rep = calibrate_band_wings(self.BAND, self.F, t, atm, self.quotes(t, atm),
                                      conv=DeltaConvention(True), jump=self.START)
        fit = rep["fit"]
        self.assertAlmostEqual(fit["fitted"]["hazard"], 0.03, places=6)
        self.assertAlmostEqual(fit["fitted"]["weak_share"], 0.7, places=5)
        self.assertLess(fit["rmse"], 1e-7)
        self.assertTrue(fit["converged"], fit["notes"])
        self.assertEqual(fit["n_quotes"], 4)
        self.assertEqual(fit["held"]["weak_jump"], self.START.weak_jump)
        # Every quote reports its own residual, and the strikes their placement.
        self.assertEqual({r["instrument"] for r in fit["residuals"]}, {"rr", "fly"})
        self.assertEqual({w["delta"] for w in rep["wings"]}, {0.25, 0.10})
        self.assertIn("call_in_band", rep["wings"][0])
        # The ATM and the forward are still exact: the body was profiled out.
        self.assertTrue(rep["converged"])

    def test_the_term_structure_shares_one_regime_and_reports_each_tenors_own(self):
        from volkit.banded import TenorQuotes, calibrate_band_term_structure
        curve = (("1M", 1 / 12, 0.0035), ("3M", 0.25, 0.0044), ("6M", 0.5, 0.0055),
                 ("1Y", 1.0, 0.0075))
        tenors = [TenorQuotes(t, self.F, atm, self.quotes(t, atm), name) for name, t, atm in curve]
        out = calibrate_band_term_structure(self.BAND, tenors, conv=DeltaConvention(True),
                                            jump=self.START)
        self.assertAlmostEqual(out["jump"].hazard, 0.03, places=6)
        self.assertAlmostEqual(out["jump"].weak_share, 0.7, places=5)
        self.assertEqual(out["fit"]["n_quotes"], 16)
        self.assertEqual(out["n_tenors"], 4)
        own = [h["hazard"] for h in out["hazard_by_tenor"]]
        self.assertEqual(len(own), 4)
        for h in own:
            self.assertAlmostEqual(h, 0.03, places=5)
        self.assertEqual(out["hazard_slope_note"], "")
        for row in out["rows"]:
            self.assertTrue(row["used"], row)
            self.assertTrue(row["converged"])

    def test_a_tenor_no_hazard_can_fit_sits_out_with_its_reason(self):
        """A band too narrow for a long tenor's ATM would have made the
        shared hazard ceiling zero and taken every other tenor down with it."""
        from volkit.banded import TenorQuotes, calibrate_band_term_structure
        band = Band("X", 0.97, 1.03)
        conv = DeltaConvention(True)
        good = TenorQuotes(0.25, 1.0, 0.03, self.quotes_for(band, 1.0, 0.25, 0.03, conv), "3M")
        bad = TenorQuotes(2.0, 1.0, 0.20, self.quotes_for(band, 1.0, 0.25, 0.03, conv), "2Y")
        out = calibrate_band_term_structure(band, [good, bad], conv=conv, jump=self.START)
        self.assertEqual(out["n_tenors"], 1)
        self.assertTrue(out["rows"][0]["used"])
        self.assertFalse(out["rows"][1]["used"])
        self.assertIn("band", out["rows"][1]["message"])

    def quotes_for(self, band, F, t, atm, conv):
        from volkit.banded import WingQuote
        sm, _ = calibrate_band_smile(band, F, t, atm, jump=self.TRUTH, conv=conv)
        out = []
        for d in (0.25, 0.10):
            vc = vp = atm
            for _ in range(40):
                kc = black.strike_from_delta(d, F, vc, t, True, conv)
                kp = black.strike_from_delta(-d, F, vp, t, False, conv)
                vc, vp = sm.implied_vol(kc), sm.implied_vol(kp)
            out.append(WingQuote(d, vc - vp, 0.5 * (vc + vp) - atm))
        return tuple(out)

    def test_identifiability_is_measured_not_assumed(self):
        """Free more than the wings can see and the Jacobian says so: the
        condition number blows up, the near-degenerate pair is named, and a
        parameter the quotes do not move is marked as not informed."""
        from volkit.banded import DEGENERATE_CONDITION, calibrate_band_wings
        t, atm = 0.25, 0.003
        quotes = self.quotes(t, atm, truth=JumpSpec(hazard=0.02, weak_share=0.7))
        _, rep = calibrate_band_wings(self.BAND, self.F, t, atm, quotes,
                                      conv=DeltaConvention(True),
                                      jump=JumpSpec(hazard=0.01, weak_share=0.7),
                                      free=("hazard", "weak_share", "weak_vol", "strong_vol"))
        fit = rep["fit"]
        cond = fit["condition"]
        self.assertTrue(cond != cond or cond > DEGENERATE_CONDITION, cond)
        self.assertIsNotNone(fit["degenerate"])
        self.assertTrue(any("degenerate" in n for n in fit["notes"]))
        self.assertFalse(all(v["informed"] for v in fit["sensitivity"].values()))
        # And the hazard against the jump size, from strikes inside the band,
        # is *not* degenerate -- the forward constraint moves the body with
        # the jump, and that is visible from inside.  Measured, not argued.
        _, rep = calibrate_band_wings(self.BAND, self.F, t, atm, quotes,
                                      conv=DeltaConvention(True),
                                      jump=JumpSpec(hazard=0.01, weak_share=0.7),
                                      free=("hazard", "weak_jump"))
        self.assertLess(rep["fit"]["condition"], DEGENERATE_CONDITION)
        self.assertIsNone(rep["fit"]["degenerate"])

    def test_more_parameters_than_quotes_is_said(self):
        from volkit.banded import calibrate_band_wings
        t, atm = 0.25, 0.0044
        _, rep = calibrate_band_wings(self.BAND, self.F, t, atm, self.quotes(t, atm, (0.25,)),
                                      conv=DeltaConvention(True), jump=self.START,
                                      free=("hazard", "weak_share", "weak_jump"))
        self.assertTrue(any("underdetermined" in n for n in rep["fit"]["notes"]))

    def test_a_free_name_that_is_not_a_break_parameter_is_refused(self):
        from volkit.banded import WingQuote, calibrate_band_wings
        t, atm = 0.25, 0.0044
        for free in (("hazard", "beta"), (), ("hazard", "hazard")):
            with self.assertRaises(ValueError):
                calibrate_band_wings(self.BAND, self.F, t, atm, self.quotes(t, atm),
                                     conv=DeltaConvention(True), free=free)
        with self.assertRaises(ValueError):
            WingQuote(0.25, 0.001, 0.001, fit=("strangle",))

    def test_the_surface_level_fit_proposes_and_marks_nothing(self):
        """The card's *Fit from the wings*: the proposal is in the card's own
        units, the surface's treatment is untouched, and a tenor the band
        cannot hold keeps its row and its reason."""
        from volkit.banded import BandTreatment, fit_band_treatment
        from volkit.feed import MarketFeed
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        book.feed = MarketFeed.load(FEED)
        surface = book["EURUSD"]
        surface.band = Band("EURUSD", 1.05, 1.12, "synthetic, for the test only")
        surface.forward_lookup = lambda t: book.forward_at("EURUSD", t)
        before = surface.band_treatment
        out = fit_band_treatment(surface, ["1M", "3M", "1Y"], cut="NY",
                                 treatment=BandTreatment(mode="mixture", jump=JumpSpec(weak_jump=0.05)))
        self.assertIs(surface.band_treatment, before)
        self.assertEqual(out["free"], ["hazard", "weak_share"])
        prop = out["proposal"]
        self.assertEqual(prop["weak_jump"], 5.0)              # held, in percent
        self.assertEqual(prop["mode"], "mixture")
        self.assertFalse(prop["solve_hazard"])
        self.assertGreater(prop["hazard"], 0.0)
        self.assertEqual([r["tenor"] for r in out["rows"]], ["1M", "3M", "1Y"])
        self.assertFalse(out["rows"][2]["used"])
        self.assertIn("band", out["rows"][2]["message"])
        self.assertEqual(len(out["hazard_by_tenor"]), 2)
        # The proposal round-trips through the same reader Apply uses.
        again = BandTreatment.from_request(prop)
        self.assertAlmostEqual(again.jump.hazard, out["jump"].hazard if "jump" in out
                               else prop["hazard"] / 100.0)

    def test_the_cli_prints_the_same_proposal(self):
        import argparse
        import io
        from contextlib import redirect_stdout
        from volkit.cli import _print_band_fit
        from volkit.feed import MarketFeed
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        book.feed = MarketFeed.load(FEED)
        surface = book["EURUSD"]
        surface.band = Band("EURUSD", 1.05, 1.12, "synthetic, for the test only")
        surface.forward_lookup = lambda t: book.forward_at("EURUSD", t)
        buf = io.StringIO()
        with redirect_stdout(buf):
            rc = _print_band_fit(argparse.Namespace(fit="hazard,weak_share", cut="NY"),
                                 surface, ["1M", "3M", "1Y"])
        text = buf.getvalue()
        self.assertEqual(rc, 1)                                # the 1Y row failed, in place
        self.assertIn("free: hazard, weak_share", text)
        self.assertIn("own hazard", text)
        self.assertIn("nothing was marked", text)
        self.assertIn("1Y", text)


class TestBandGuard(unittest.TestCase):
    def test_surface_flags_a_strike_outside_the_band(self):
        book = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])
        surface = book["USDJPY"]
        self.assertEqual(surface.band_check(7.90, 7.80), [])   # no band set
        surface.band = Band("USDHKD", 7.75, 7.85)
        self.assertEqual(surface.band_check(7.80, 7.80), [])
        warn = surface.band_check(7.90, 7.80)
        self.assertTrue(warn)
        self.assertIn("managed band", warn[0])

    def test_the_treatment_decides_whether_the_warning_is_said_at_all(self):
        """'off' is a marking, not an oversight, and a BAND price already has
        the peg in it."""
        from volkit.banded import BandTreatment
        book = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])
        surface = book["USDJPY"]
        surface.band = Band("USDHKD", 7.75, 7.85)
        self.assertTrue(surface.band_check(7.90, 7.80))
        surface.set_band_treatment(BandTreatment(mode="off"))
        self.assertEqual(surface.band_check(7.90, 7.80), [])
        surface.set_band_treatment(BandTreatment(mode="mixture"))
        self.assertEqual(surface.band_check(7.90, 7.80, method="BAND"), [])
        # Marked as a mixture but priced lognormally: that is worth saying.
        self.assertIn("switch the method to BAND", surface.band_check(7.90, 7.80, "SVI")[0])

    def test_a_leg_is_checked_at_the_level_its_payout_depends_on(self):
        """A vanilla struck outside a managed band used to go unflagged.

        The check only ever looked at ``leg.barrier``, so the one product a
        band matters most for -- an option struck out at the edge -- was the
        one product nothing was said about, while a barrier left behind on a
        leg that had since been switched to a vanilla was checked instead.
        The level checked is now the one the payout actually depends on.
        """
        from volkit.pricing import OptionLeg, price_leg
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        book["EURUSD"].band = Band("EURUSD", 1.05, 1.12, "synthetic, for the test only")

        def warnings(*, method="SVI", **kw):
            leg = OptionLeg(pair="EURUSD", expiry="2024-05-28", spot=1.0842,
                            method=method, cut="TK", **kw)
            out = price_leg(book, leg)
            self.assertTrue(out.ok, out.error)
            return [w for w in out.warnings if "managed band" in w]

        self.assertTrue(warnings(strike="1.3000"))          # outside: said
        self.assertEqual(warnings(strike="1.0900"), [])     # inside: nothing
        # A barrier left on a leg whose product no longer uses one is not read.
        self.assertEqual(warnings(strike="1.0900", barrier="1.3000"), [])
        # The touch products are checked at their barrier, as before.
        self.assertTrue(warnings(product="one_touch", barrier="1.3000"))
        # A leg already priced with BAND is not told to switch to it.
        book["EURUSD"].forward_lookup = lambda t: 1.0859825226390687
        self.assertEqual(warnings(strike="1.3000", method="BAND"), [])

    def test_the_band_edges_can_be_overridden_on_the_screen(self):
        from volkit.banded import BandTreatment
        book = Book.from_excel(WORKBOOK, ASOF).build(["USDJPY"])
        surface = book["USDJPY"]
        surface.band = Band("USDHKD", 7.75, 7.85)
        surface.set_band_treatment(BandTreatment(upper=7.90))
        self.assertEqual(surface.band_check(7.88, 7.80), [])
        self.assertIn("7.9", surface.band_check(7.95, 7.80)[0])


class TestBandInterpolation(unittest.TestCase):
    """The BAND method: the regime mixture priced through the surface.

    The plumbing this pins is the piece the band model was missing.  A band is
    an absolute price range and the surface works in strike over forward, so
    a slice read in moneyness has to divide the band by an outright forward --
    and refuse, rather than guess a level, when there is no feed to divide by.
    """

    BAND = Band("EURUSD", 1.05, 1.12, "synthetic, for the test only")
    EXPIRY = datetime(2024, 5, 28, tzinfo=UTC)

    def surface(self, *, feed=True):
        from volkit.feed import MarketFeed
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        if feed:
            book.feed = MarketFeed.load(FEED)
        surface = book["EURUSD"]
        surface.band = self.BAND
        surface.forward_lookup = lambda t: book.forward_at("EURUSD", t)
        return book, surface

    def test_the_band_is_divided_into_the_surfaces_own_space(self):
        _, surface = self.surface()
        band = surface.band_for_slice(0.25, 1.0)
        self.assertAlmostEqual(band.lower * 1.08602, 1.05, places=4)
        self.assertAlmostEqual(band.upper * 1.08602, 1.12, places=4)
        # A slice built at the outright forward is already in the band's space.
        self.assertEqual(surface.band_for_slice(0.25, 1.086).lower, 1.05)

    def test_without_a_feed_it_refuses_rather_than_guessing_a_level(self):
        _, surface = self.surface(feed=False)
        with self.assertRaises(ValueError) as ctx:
            surface.vol(1.0, self.EXPIRY, "BAND", "NY")
        self.assertIn("moneyness", str(ctx.exception))
        self.assertIn("feed", str(ctx.exception))

    def test_the_wings_collapse_outside_the_band(self):
        """The whole point: a lognormal wing pays for a level the peg forbids."""
        _, surface = self.surface()
        inside = float(surface.vol(1.0, self.EXPIRY, "BAND", "NY"))
        outside = float(surface.vol(1.03, self.EXPIRY, "BAND", "NY"))
        lognormal = float(surface.vol(1.03, self.EXPIRY, "SVI", "NY"))
        self.assertGreater(inside, outside)
        self.assertLess(outside, lognormal * 0.75)

    def test_the_treatment_is_part_of_the_cache_key(self):
        """Two hazards are two smiles; a cache that could not tell them apart
        would serve the first answer for the rest of the session."""
        from volkit.banded import BandTreatment, JumpSpec
        _, surface = self.surface()
        low = float(surface.vol(1.02, self.EXPIRY, "BAND", "NY"))
        surface.set_band_treatment(BandTreatment(mode="mixture",
                                                 jump=JumpSpec(hazard=0.40)))
        high = float(surface.vol(1.02, self.EXPIRY, "BAND", "NY"))
        self.assertNotAlmostEqual(low, high, places=6)
        self.assertGreater(high, low)          # more break risk, more wing value

    def test_the_feed_level_is_part_of_the_cache_key_as_well(self):
        """The old bug: it was not.

        The feed is a publication and is re-read all morning (the auto-reload
        switch exists for exactly that), and a band is placed against whatever
        it then says -- so two spots are two smiles, the same way two hazards
        are.  Nothing in the key moved when the feed did, so the marking
        screen's band card printed the *republished* forward in its own column
        beside probabilities still calibrated against the old one.
        """
        _, surface = self.surface()
        level = {"f": 1.086}
        surface.forward_lookup = lambda t: level["f"]
        before = float(surface.vol(1.01, self.EXPIRY, "BAND", "NY"))
        level["f"] = 1.09              # the market is republished higher
        after = float(surface.vol(1.01, self.EXPIRY, "BAND", "NY"))
        self.assertNotAlmostEqual(before, after, places=6)
        # and it is still a cache: the same level gives the same slice back.
        level["f"] = 1.086
        self.assertAlmostEqual(
            float(surface.vol(1.01, self.EXPIRY, "BAND", "NY")), before, places=12)

    def test_a_blend_is_between_the_two_and_says_it_is_not_a_model(self):
        from volkit.banded import BandTreatment
        _, surface = self.surface()
        band = float(surface.vol(1.02, self.EXPIRY, "BAND", "NY"))
        logn = float(surface.vol(1.02, self.EXPIRY, "SVI", "NY"))
        warnings = surface.set_band_treatment(BandTreatment(mode="mixture", blend=0.5))
        mixed = float(surface.vol(1.02, self.EXPIRY, "BAND", "NY"))
        self.assertAlmostEqual(mixed, 0.5 * band + 0.5 * logn, places=10)
        self.assertTrue(any("arbitrage free" in w for w in warnings))

    def test_a_delta_strike_is_found_where_the_fixed_point_will_not_contract(self):
        """v -> vol(K(v)) contracts only while the smile is gentle.

        A band smile is not: its wings fall away where the peg's support runs
        out, so the 10 delta strikes came back as "did not converge" and took
        the whole smile table with them. Delta is still monotone in strike, so
        the bracketed solve one level down finds them.
        """
        from volkit.numerics import fixed_point
        _, surface = self.surface()
        sl = surface.slice_at(self.EXPIRY, "BAND", "NY")
        with self.assertRaises(ConvergenceError):        # the primary path alone
            fixed_point(lambda v: float(sl.vol(black.strike_from_delta(
                -0.10, sl.forward, v, sl.t, False, sl.conv))),
                sl.atm_vol, tol=1e-11, max_iter=80, what="10d put")
        table = {r["label"]: r["strike"] for r in
                 surface.smile_table(self.EXPIRY, method="BAND", cut="NY")}
        self.assertEqual(len(table), 5)
        self.assertLess(table["10d put"], table["25d put"])
        self.assertLess(table["25d put"], table["ATM"])
        self.assertLess(table["ATM"], table["25d call"])
        self.assertLess(table["25d call"], table["10d call"])
        # And the wings sit inside the band, which is the whole point of it.
        self.assertGreater(table["10d put"], sl.band.lower)
        self.assertLess(table["10d call"], sl.band.upper)

    def test_a_delta_the_smile_never_reaches_says_so(self):
        """Not "did not converge": the peg's support ran out.

        With the hazard marked at zero -- a peg that cannot break, which §6 of
        CLAUDE.md says is not a thing to believe, but is a thing to be able to
        ask for -- the distribution has compact support and there is no 10
        delta call at all. That is a statement about the marks.
        """
        from volkit.banded import BandTreatment, JumpSpec
        _, surface = self.surface()
        surface.set_band_treatment(BandTreatment(mode="mixture",
                                                 jump=JumpSpec(hazard=0.0)))
        sl = surface.slice_at(self.EXPIRY, "BAND", "NY")
        with self.assertRaises(ConvergenceError) as ctx:
            sl.strike_from_delta(0.10, True)
        self.assertIn("never reaches", str(ctx.exception))
        self.assertIn("peg", str(ctx.exception))

    def test_a_pair_with_no_band_is_told_which_pairs_have_one(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        with self.assertRaises(ValueError) as ctx:
            book["EURUSD"].vol(1.0, self.EXPIRY, "BAND", "NY")
        self.assertIn("PEG_BANDS", str(ctx.exception))

    def test_the_panel_reports_breach_probability_and_keeps_a_failed_row(self):
        from volkit.banded import band_panel
        _, surface = self.surface()
        panel = band_panel(surface, ["3M", "1Y"], cut="NY")
        self.assertTrue(panel["has_band"])
        good = panel["rows"][0]
        self.assertEqual(good["message"], "")
        self.assertGreater(good["prob_outside_band"], 0.0)
        self.assertLess(good["prob_outside_band"], good["prob_broken"])
        # 1Y cannot be fitted inside a 6% band, and the row says so in place.
        self.assertEqual(len(panel["rows"]), 2)
        self.assertIn("band", panel["rows"][1]["message"])

    def test_the_calibration_names_the_bound_that_was_missed(self):
        """Reporting only the break-risk floor named one cause for both
        failures, and sent a marker to lower a hazard that was not the
        problem."""
        from volkit.banded import calibrate_band_smile
        with self.assertRaises(ConvergenceError) as narrow:
            calibrate_band_smile(Band("X", 0.99, 1.01), 1.0, 1.0, 0.20)
        self.assertIn("wide", str(narrow.exception))
        with self.assertRaises(ConvergenceError) as jumpy:
            calibrate_band_smile(Band("X", 0.90, 1.10), 1.0, 1.0, 0.005,
                                 jump=JumpSpec(hazard=0.5, weak_jump=0.005,
                                               strong_jump=0.005, weak_vol=0.30,
                                               strong_vol=0.30))
        self.assertIn("at least", str(jumpy.exception))


class TestBandTreatmentRequest(unittest.TestCase):
    """Percentages at the edge, decimals in the middle -- converted once."""

    def test_the_screens_units_are_converted_exactly_once(self):
        from volkit.banded import BandTreatment
        t = BandTreatment.from_request({"mode": "mixture", "hazard": "3", "weak_jump": 8,
                                        "blend": "80", "delta": 10, "lower": "7.74"})
        self.assertAlmostEqual(t.jump.hazard, 0.03)
        self.assertAlmostEqual(t.jump.weak_jump, 0.08)
        self.assertAlmostEqual(t.blend, 0.80)
        self.assertAlmostEqual(t.delta, 0.10)
        self.assertEqual(t.lower, 7.74)
        self.assertEqual(t.to_request()["hazard"], 3.0)

    def test_a_blank_field_keeps_the_default_rather_than_becoming_zero(self):
        """A hazard silently set to zero is a peg that cannot break."""
        from volkit.banded import BandTreatment, JumpSpec
        t = BandTreatment.from_request({"hazard": "", "weak_vol": None})
        self.assertEqual(t.jump.hazard, JumpSpec().hazard)
        self.assertEqual(t.jump.weak_vol, JumpSpec().weak_vol)

    def test_nonsense_is_refused_by_name(self):
        from volkit.banded import BandTreatment
        with self.assertRaises(ValueError) as ctx:
            BandTreatment.from_request({"hazard": "soon"})
        self.assertIn("hazard", str(ctx.exception))
        with self.assertRaises(ValueError):
            BandTreatment.from_request({"mode": "ignore"})
        with self.assertRaises(ValueError):
            BandTreatment.from_request({"blend": "140"})


class TestListedOptions(unittest.TestCase):
    """Exchange traded options: a table of strikes, not three broker quotes."""

    EXPIRY = datetime(2024, 6, 14, 19, 0, tzinfo=UTC)

    # -- the parser ------------------------------------------------------
    def test_parser_reports_the_columns_it_chose(self):
        """The legacy tool's silence about its inputs is what this replaces."""
        t = listed.parse_quote_table("Strike\tIV\n1.05\t8.20\n1.08\t7.50\n1.10\t7.65\n")
        self.assertEqual(t.delimiter, "tab")
        self.assertEqual(t.strike_column, 1)
        self.assertEqual(t.vol_unit, "percent")
        self.assertAlmostEqual(t.quotes[0].vol, 0.0820)

    def test_parser_takes_the_mid_of_bid_and_ask(self):
        t = listed.parse_quote_table("Strike,Bid,Ask\n1.05,8.10,8.30\n1.08,7.40,7.60\n1.10,7.55,7.75\n")
        self.assertEqual(t.delimiter, "comma")
        self.assertAlmostEqual(t.quotes[0].vol, 0.0820)
        self.assertTrue(any("mid" in n for n in t.notes))

    def test_decimal_is_something_a_person_says(self):
        """A table is read in volatility points as written (§4).  A table of
        decimals is loaded with vol_unit='decimal'; it used to be inferred
        from the level, which read a managed pair's own table 100x too big."""
        dec = listed.parse_quote_table("1.05 0.0820\n1.08 0.0750\n1.10 0.0765\n",
                                       vol_unit="decimal")
        pct = listed.parse_quote_table("1.05 8.20\n1.08 7.50\n1.10 7.65\n")
        self.assertEqual(dec.vol_unit, "decimal")
        self.assertEqual(pct.vol_unit, "percent")
        for a, b in zip(dec.quotes, pct.quotes):
            self.assertAlmostEqual(a.vol, b.vol, places=12)

    def test_a_table_below_one_is_read_as_written_and_says_so(self):
        """0.82 is eight tenths of a volatility point, which is what a managed
        pair's listed table looks like.  Read as written, and noted."""
        t = listed.parse_quote_table("1.05 0.82\n1.08 0.75\n1.10 0.76\n")
        self.assertEqual(t.vol_unit, "percent")
        self.assertAlmostEqual(t.quotes[0].vol, 0.0082)
        self.assertTrue(any("as written" in n for n in t.notes))

    def test_a_table_straddling_one_is_read_and_not_refused(self):
        """0.95 beside 8.20 used to be refused as ambiguous.  Both are points."""
        t = listed.parse_quote_table("1.05\t0.95\n1.08\t8.20\n1.10\t7.70\n")
        self.assertAlmostEqual(t.quotes[0].vol, 0.0095)
        self.assertAlmostEqual(t.quotes[1].vol, 0.0820)

    def test_every_unusable_line_is_returned_with_a_reason(self):
        t = listed.parse_quote_table(
            "Strike\tIV\n1.05\t8.2\nn/a\t7.0\n1.08\t\n1.10\t7.65\n1.12\t7.9\n")
        self.assertEqual(len(t.quotes), 3)
        self.assertEqual([n for n, _, _ in t.skipped], [3, 4])

    def test_explicit_columns_override_the_headers(self):
        t = listed.parse_quote_table("Strike\tDelta\tIV\n1.05\t0.30\t8.2\n"
                                     "1.08\t0.50\t7.5\n1.10\t0.62\t7.65\n",
                                     vol_column=3)
        self.assertAlmostEqual(t.quotes[0].vol, 0.082)

    def test_duplicate_strikes_keep_the_out_of_the_money_quote(self):
        """An in-the-money option has little time value, so its implied vol is noise."""
        t = listed.parse_quote_table(
            "K\tIV\tType\n1.05\t8.2\tC\n1.05\t8.4\tP\n1.08\t7.5\tC\n1.10\t7.65\tC\n")
        kept, notes = listed.dedupe(t.quotes, forward=1.08)
        self.assertEqual(len(kept), 3)
        self.assertAlmostEqual(kept[0].vol, 0.084)      # 1.05 is below the forward: the put
        self.assertTrue(notes)

    # -- the fit ---------------------------------------------------------
    def _known(self):
        t, f = 0.5, 1.0850
        p = sabr.SabrParams(alpha=0.082, rho=-0.31, volvol=0.65, t=t, beta=1.0, f=f)
        ks = f * np.array([0.85, 0.90, 0.95, 1.00, 1.02, 1.05, 1.10, 1.15, 1.22])
        return p, ks, np.asarray(sabr.lognormal_vol(ks, p))

    def test_the_fit_recovers_parameters_it_was_generated_from(self):
        p, ks, vs = self._known()
        fit = listed.fit_sabr(ks, vs, p.t, p.f)
        self.assertAlmostEqual(fit.params.alpha, p.alpha, places=6)
        self.assertAlmostEqual(fit.params.rho, p.rho, places=5)
        self.assertAlmostEqual(fit.params.volvol, p.volvol, places=5)
        self.assertLess(fit.rmse, 1e-10)
        self.assertTrue(fit.converged)

    def test_the_answer_does_not_depend_on_the_sweep_resolution(self):
        """The box is swept before anything is polished, so no start point matters."""
        p, ks, vs = self._known()
        a = listed.fit_sabr(ks, vs, p.t, p.f, scan=(9, 7))
        b = listed.fit_sabr(ks, vs, p.t, p.f, scan=(21, 15))
        self.assertAlmostEqual(a.params.rho, b.params.rho, places=6)
        self.assertAlmostEqual(a.params.volvol, b.params.volvol, places=6)

    def test_the_order_of_the_rows_does_not_change_the_fit(self):
        p, ks, vs = self._known()
        order = np.array([4, 0, 8, 2, 6, 1, 7, 3, 5])
        a = listed.fit_sabr(ks, vs, p.t, p.f)
        b = listed.fit_sabr(ks[order], vs[order], p.t, p.f)
        self.assertAlmostEqual(a.params.rho, b.params.rho, places=8)

    def test_three_quotes_is_an_interpolation_and_says_so(self):
        """Zero residuals from three points are not evidence of anything."""
        p, ks, vs = self._known()
        fit = listed.fit_sabr(ks[[0, 3, 7]], vs[[0, 3, 7]], p.t, p.f)
        self.assertLess(fit.rmse, 1e-10)
        self.assertEqual(fit.degrees_of_freedom, 0)
        self.assertTrue(any("exact interpolation" in w for w in fit.warnings))

    def test_fewer_than_three_strikes_is_rejected(self):
        p, ks, vs = self._known()
        with self.assertRaises(ValueError):
            listed.fit_sabr(ks[:2], vs[:2], p.t, p.f)
        with self.assertRaises(ValueError):
            listed.fit_sabr([1.0, 1.0, 1.0], [0.08, 0.08, 0.08], p.t, p.f)


    # -- parameters given rather than fitted ------------------------------
    def test_a_held_parameter_comes_back_exactly_as_it_was_given(self):
        """A number somebody typed must not come back rounded.

        Alpha travels through a logarithm inside the optimiser, and
        exp(log(a)) is not always a again."""
        p, ks, vs = self._known()
        fit = listed.fit_sabr(ks, vs, p.t, p.f, fixed={"alpha": 0.0925})
        self.assertEqual(fit.params.alpha, 0.0925)
        self.assertEqual(fit.fixed, ("alpha",))
        self.assertEqual(fit.free, ("rho", "volvol"))
        self.assertEqual(fit.degrees_of_freedom, len(ks) - 2)

    def test_holding_a_parameter_fits_the_others_around_it(self):
        p, ks, vs = self._known()
        free = listed.fit_sabr(ks, vs, p.t, p.f)
        held = listed.fit_sabr(ks, vs, p.t, p.f, fixed={"rho": -0.10})
        self.assertEqual(held.params.rho, -0.10)
        # The other two moved to make the best of it, and the fit is worse
        # than the free one -- which is the whole point of reporting it.
        self.assertNotAlmostEqual(held.params.alpha, free.params.alpha, places=4)
        self.assertGreater(held.rmse, free.rmse)
        self.assertTrue(any("given, not fitted" in w for w in held.warnings))

    def test_holding_all_three_fits_nothing_and_says_so(self):
        """Priced against the curve you typed: a legitimate thing to ask for,
        and it must not report a convergence it never attempted."""
        p, ks, vs = self._known()
        fit = listed.fit_sabr(ks, vs, p.t, p.f,
                              fixed={"alpha": p.alpha, "rho": p.rho, "volvol": p.volvol})
        self.assertEqual((fit.params.alpha, fit.params.rho, fit.params.volvol),
                         (p.alpha, p.rho, p.volvol))
        self.assertEqual(fit.free, ())
        self.assertEqual(fit.degrees_of_freedom, len(ks))
        self.assertIn("no fit", fit.message)
        self.assertLess(fit.rmse, 1e-12)          # these quotes came off that curve
        self.assertTrue(any("nothing was fitted at all" in w for w in fit.warnings))

    def test_a_blank_override_is_not_a_zero(self):
        """The screen sends an empty box for every parameter it is not
        holding; reading that as a number would mark a curve nobody asked
        for."""
        p, ks, vs = self._known()
        free = listed.fit_sabr(ks, vs, p.t, p.f)
        blank = listed.fit_sabr(ks, vs, p.t, p.f,
                                fixed={"alpha": None, "rho": "", "volvol": None})
        self.assertEqual(blank.fixed, ())
        self.assertAlmostEqual(blank.params.rho, free.params.rho, places=9)

    def test_a_held_parameter_that_is_not_a_parameter_is_refused(self):
        p, ks, vs = self._known()
        for bad in ({"rho": 1.4}, {"alpha": -0.01}, {"volvol": 0.0}, {"beta": 0.5}):
            with self.assertRaises(ValueError):
                listed.fit_sabr(ks, vs, p.t, p.f, fixed=bad)

    def test_holding_two_leaves_two_quotes_enough(self):
        """The three-quote rule is about free parameters, not about SABR."""
        p, ks, vs = self._known()
        fit = listed.fit_sabr(ks[:2], vs[:2], p.t, p.f,
                              fixed={"rho": p.rho, "volvol": p.volvol})
        self.assertAlmostEqual(fit.params.alpha, p.alpha, places=6)
        with self.assertRaises(ValueError):
            listed.fit_sabr(ks[:1], vs[:1], p.t, p.f, fixed={"rho": p.rho})

    def test_the_panel_carries_the_overrides_and_reports_them(self):
        """What the browser sends is what the fit holds, and the answer says
        which numbers were typed -- one that was is otherwise indistinguishable
        from one the market implied."""
        panel = listed.panel_from_request({
            "underlying": "CUSTOM", "expiry": "2024-06-14 19:00", "forward": 1.085,
            "text": "1.00 9.10\n1.05 8.40\n1.085 8.20\n1.12 8.35\n1.16 8.90\n",
            "rho": "-0.20", "alpha": "", "volvol": None,
        })
        self.assertEqual((panel.rho, panel.alpha, panel.volvol), (-0.20, None, None))
        out = panel.run(None, clock=ASOF)
        self.assertEqual(out["fit"]["rho"], -0.20)
        self.assertEqual(out["fit"]["fixed"], ["rho"])
        self.assertEqual(out["fit"]["free"], ["alpha", "volvol"])
        with self.assertRaises(ValueError):
            listed.panel_from_request({"underlying": "CUSTOM", "forward": 1.0,
                                       "expiry": "2024-06-14", "text": "1.0 8.0",
                                       "rho": "steep"})

    def test_vega_weighting_favours_the_money(self):
        p, ks, vs = self._known()
        fit = listed.fit_sabr(ks, vs, p.t, p.f, weighting="vega")
        near = min(range(len(ks)), key=lambda i: abs(ks[i] - p.f))
        self.assertEqual(near, int(np.argmax(fit.weights)))
        flat = listed.fit_sabr(ks, vs, p.t, p.f, weighting="equal")
        self.assertTrue(np.allclose(flat.weights, 1.0))

    def test_strikes_that_miss_the_forward_are_flagged(self):
        p, ks, vs = self._known()
        keep = ks > p.f
        fit = listed.fit_sabr(ks[keep], vs[keep], p.t, p.f)
        self.assertTrue(any("do not bracket the forward" in w for w in fit.warnings))

    def test_negative_density_is_reported_rather_than_hidden(self):
        """Hagan's expansion arbitrages in the wings.  Marking the risk beats clipping it."""
        p = sabr.SabrParams(alpha=0.30, rho=-0.90, volvol=3.0, t=3.0, beta=1.0, f=1.0)
        msgs = listed.arbitrage_warnings(p, 0.4, 2.5)
        self.assertTrue(msgs and "negative risk-neutral density" in msgs[0])
        calm = sabr.SabrParams(alpha=0.09, rho=-0.10, volvol=0.35, t=0.25, beta=1.0, f=1.0)
        self.assertEqual(listed.arbitrage_warnings(calm, 0.9, 1.1), [])

    def test_the_arbitrage_check_does_not_depend_on_the_contract_units(self):
        """Second differences of a yen future's prices are tiny in raw units;
        checking in moneyness stops rounding being read as arbitrage."""
        for f in (0.0068, 1.0, 145.0, 4200.0):
            calm = sabr.SabrParams(alpha=0.09 * f, rho=-0.10, volvol=0.35,
                                   t=0.25, beta=1.0, f=f)
            self.assertEqual(listed.arbitrage_warnings(calm, f * 0.9, f * 1.1), [],
                             msg=f"false positive at forward {f}")

    # -- the mapping onto an FX pair -------------------------------------
    def test_inverted_strikes_round_trip(self):
        u = listed.resolve_underlying("6J")
        self.assertTrue(u.invert)
        self.assertAlmostEqual(u.to_fx(0.006850), 1.0 / 0.006850, places=9)
        self.assertAlmostEqual(u.from_fx(u.to_fx(0.006850)), 0.006850, places=12)

    def test_scale_is_applied_before_the_inversion(self):
        u = listed.resolve_underlying("6J", scale=1e-6)
        self.assertAlmostEqual(u.to_fx(6850.0), 1.0 / 0.006850, places=9)

    def test_a_non_inverted_contract_is_left_alone(self):
        u = listed.resolve_underlying("6E")
        self.assertFalse(u.invert)
        self.assertEqual(u.pair, "EURUSD")
        self.assertAlmostEqual(u.to_fx(1.0850), 1.0850)

    def test_overrides_beat_the_registry(self):
        u = listed.resolve_underlying("CUSTOM", pair="EURGBP", invert=True, scale=2.0)
        self.assertEqual((u.pair, u.invert, u.scale), ("EURGBP", True, 2.0))

    def test_a_code_this_build_does_not_know_is_taken_as_typed(self):
        """The contract was a dropdown, and every contract missing from it had
        to be entered as CUSTOM -- at which point two of them on one screen
        cannot be told apart, and a position line naming one is refused as
        ambiguous with no way to settle it.  A typed code is now a contract
        with the name that was typed."""
        u = listed.resolve_underlying("M6E SEP26", pair="EURUSD", contract_size=12_500)
        self.assertFalse(u.known)
        self.assertEqual(u.code, "M6E SEP26")
        self.assertEqual((u.pair, u.invert, u.scale), ("EURUSD", False, 1.0))
        self.assertEqual(u.contract_size, 12_500)
        # Nothing is inferred from the name: a code that merely looks like the
        # euro brings no pair, no direction and no size of its own.
        bare = listed.resolve_underlying("M6E SEP26")
        self.assertIsNone(bare.pair)
        self.assertEqual(bare.contract_size, 0.0)
        # And the ones in the registry are still known, with everything on.
        self.assertTrue(listed.resolve_underlying("6J").known)

    def test_a_typed_code_says_it_was_typed_rather_than_looked_up(self):
        """Otherwise a typo (6R for 6E) is indistinguishable from a contract
        that genuinely has no mapping, which is a comparison quietly gone."""
        r = listed.panel_from_request({
            "underlying": "6R", "expiry": "2024-06-14 19:00", "forward": 1.085,
            "text": "1.06 7.9\n1.085 7.42\n1.11 7.6"}).run(clock=ASOF)
        self.assertFalse(r["underlying"]["known"])
        self.assertEqual(r["underlying"]["code"], "6R")
        self.assertTrue(any("taken as typed" in n for n in r["notes"]))

    def test_a_mis_pasted_cell_is_still_refused_as_a_contract_code(self):
        """Free text is not "anything at all": a code has a shape, and a
        pasted quote row landing in the box must not become a contract."""
        for bad in ("1.0800\t7.42", "x" * 40, "6E, 1.0800"):
            with self.assertRaises(ValueError):
                listed.resolve_underlying(bad)

    def test_the_settlement_currency_is_derived_from_the_pair_and_the_direction(self):
        """The premium comes out of Black-76 in the currency the *listed
        strike axis* is quoted in.  Every CME contract works out as USD, which
        is the whole reason the money columns have always added up across
        them; a typed contract need not."""
        self.assertEqual(listed.resolve_underlying("6E").premium_ccy, "USD")
        self.assertEqual(listed.resolve_underlying("6J").premium_ccy, "USD")
        self.assertEqual(listed.resolve_underlying("6C").premium_ccy, "USD")
        self.assertEqual(
            listed.resolve_underlying("XYZ", pair="EURGBP").premium_ccy, "GBP")
        # No pair, no answer -- the premium is still a number, in a currency
        # this tool was never told.
        self.assertEqual(listed.resolve_underlying("CUSTOM").premium_ccy, "")

    # -- against the marked surface --------------------------------------
    def _panel_from_the_book(self, code, pair, fx_forward, book):
        """Quote a listed table off the book's own smile, then read it back."""
        surface = book[pair]
        u = listed.resolve_underlying(code)
        listed_fwd = u.from_fx(fx_forward)
        ks = np.linspace(listed_fwd * 0.90, listed_fwd * 1.12, 9)
        fx_ks = np.asarray(u.to_fx(ks), dtype=float)
        vols = np.asarray(surface.vol(fx_ks / fx_forward, self.EXPIRY, "SVI", "NY"))
        text = "Strike\tIV\n" + "\n".join(f"{k:.10f}\t{v * 100:.8f}"
                                          for k, v in zip(ks, vols))
        return listed.panel_from_request({
            "underlying": code, "expiry": "2024-06-14 19:00", "forward": listed_fwd,
            "text": text, "cut": "NY", "method": "SVI"})

    def test_an_inverted_table_maps_back_onto_the_same_marks(self):
        """The whole inversion, end to end: a USDJPY smile quoted as a yen future
        and read back must give the marks it started from.  A sign or a
        reciprocal in the wrong place shows up here as a wing-sized error."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        out = self._panel_from_the_book("6J", "USDJPY", 150.0, book).run(book)
        for row in out["rows"]:
            self.assertAlmostEqual(row["book_vol"], row["market_vol"], places=6)
        self.assertEqual(out["comparison"]["pair"], "USDJPY")
        self.assertAlmostEqual(out["comparison"]["forward_fx"], 150.0, places=9)

    def test_an_upright_table_maps_back_onto_the_same_marks(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        out = self._panel_from_the_book("6E", "EURUSD", 1.0850, book).run(book)
        for row in out["rows"]:
            self.assertAlmostEqual(row["book_vol"], row["market_vol"], places=6)

    def test_the_book_delta_strikes_come_back_on_the_listed_axis(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        out = self._panel_from_the_book("6J", "USDJPY", 150.0, book).run(book)
        u = listed.resolve_underlying("6J")
        for a in out["comparison"]["anchors"]:
            self.assertAlmostEqual(u.to_fx(a["listed_strike"]), a["fx_strike"], places=6)
        labels = [a["label"] for a in out["comparison"]["anchors"]]
        self.assertIn("ATM", labels)
        self.assertIn("25d call", labels)

    def test_the_panel_uses_the_books_clock(self):
        """Nothing in the model may read the wall clock; the fit must use the same
        365.2425-day year as the mark it is compared with."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        out = self._panel_from_the_book("6J", "USDJPY", 150.0, book).run(book)
        self.assertAlmostEqual(out["years"], ASOF.years_to(self.EXPIRY), places=12)
        self.assertEqual(out["valuation"][:16], ASOF.now.isoformat()[:16])

    def test_a_past_expiry_is_rejected_with_both_times(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        panel = self._panel_from_the_book("6J", "USDJPY", 150.0, book)
        panel.expiry = "2023-01-01 19:00"
        with self.assertRaises(ValueError) as cm:
            panel.run(book)
        self.assertIn("not in the future", str(cm.exception))

    def test_an_unmapped_contract_fits_but_does_not_compare(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        out = listed.panel_from_request({
            "underlying": "CUSTOM", "expiry": "2024-06-14 19:00", "forward": 4200.0,
            "text": "Strike\tIV\n3800\t18.4\n4000\t16.9\n4200\t15.8\n4400\t15.4\n4600\t15.6\n",
        }).run(book, clock=ASOF)
        self.assertIsNone(out["comparison"])
        self.assertTrue(out["fit"]["converged"])

    def test_a_pair_outside_the_book_is_named_in_the_error(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        panel = listed.panel_from_request({
            "underlying": "CUSTOM", "pair": "USDNOK", "expiry": "2024-06-14 19:00",
            "forward": 10.5, "text": "K\tIV\n10.0\t9.4\n10.5\t9.0\n11.0\t9.2\n"})
        with self.assertRaises(ValueError) as cm:
            panel.run(book)
        self.assertIn("USDNOK", str(cm.exception))

    def test_a_browser_datetime_is_accepted(self):
        """An HTML datetime-local field emits a 'T'; parse_datetime wants a space."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        panel = self._panel_from_the_book("6J", "USDJPY", 150.0, book)
        panel.expiry = "2024-06-14T19:00"
        self.assertAlmostEqual(panel.run(book)["years"], ASOF.years_to(self.EXPIRY), places=12)


HISTORY = Path(__file__).resolve().parents[1] / "files" / "history_sample.xlsx"
FEED = Path(__file__).resolve().parents[1] / "files" / "market_feed.csv"


def _flat_distribution(vol, t, n=1601, span=6.0):
    """A lognormal with no smile, built the same way a real one is."""
    sd = vol * math.sqrt(t)
    x = np.linspace(-span * sd, span * sd, n)
    k = np.exp(x)
    c = np.asarray(black.price(1.0, k, vol, t, True), dtype=float)
    cdf = 1.0 + np.gradient(c, x, edge_order=2) / k
    return moments.Distribution(x=x, pdf=np.gradient(cdf, x, edge_order=2), cdf=cdf, t=t)


class TestMoments(unittest.TestCase):
    """Reading a smile as a distribution, and combining two of them."""

    def test_a_flat_smile_is_read_back_as_a_lognormal(self):
        d = _flat_distribution(0.10, 0.5)
        m = d.moments()
        self.assertAlmostEqual(m.annualised_vol(0.5), 0.10, places=5)
        # Two numerical differentiations of a call curve leave a little grid
        # dust; a few times 1e-5 of skew is three orders below anything a
        # butterfly would show.
        self.assertAlmostEqual(m.skew, 0.0, places=4)
        self.assertAlmostEqual(m.excess_kurtosis, 0.0, places=3)
        self.assertAlmostEqual(d.mgf(1.0), 1.0, places=6)      # it is a martingale

    def test_the_copula_reproduces_the_exact_variance_triangle(self):
        """Two smile-less legs have an answer in closed form; the grid must find it."""
        t = 0.5
        for va, vb, rho, ca, cb in ((0.10, 0.08, 0.30, 1, 1), (0.10, 0.08, -0.40, 1, -1),
                                    (0.12, 0.12, 0.60, 1, -1), (0.20, 0.15, 0.80, 1, 1)):
            comb = moments.combine(_flat_distribution(va, t), _flat_distribution(vb, t),
                                   (ca, cb), rho, DeltaConvention(False))
            table = comb.table()
            exact = math.sqrt(va * va + vb * vb + 2 * ca * cb * rho * va * vb)
            # 2e-5 in volatility is two ten-thousandths of a vol point.
            self.assertAlmostEqual(table["atm"], exact, delta=2e-5,
                                   msg=f"rho={rho} coefficients=({ca},{cb})")
            self.assertAlmostEqual(table["rr25"], 0.0, places=4)
            self.assertAlmostEqual(table["fly25"], 0.0, places=4)

    def test_the_forward_shift_is_the_triangles_own_convexity(self):
        """The product of two martingales is not one, and an inverted leg is not
        one at all.  Both are known from the legs, so what is left over after
        subtracting them is grid error and must be tiny."""
        t = 0.5
        for va, vb, rho, ca, cb in ((0.10, 0.08, 0.30, 1, 1), (0.12, 0.12, -0.90, 1, -1),
                                    (0.20, 0.15, 0.80, 1, 1)):
            comb = moments.combine(_flat_distribution(va, t), _flat_distribution(vb, t),
                                   (ca, cb), rho, DeltaConvention(False))
            self.assertLess(abs(comb.shift - comb.convexity), 1e-4,
                            msg=f"rho={rho} coefficients=({ca},{cb})")
            self.assertEqual(comb.warnings, ())

    def test_triangle_coefficients_are_not_just_the_variance_signs(self):
        """The variance triangle only ever needs the *product* of the two signs.

        An odd cumulant needs them one at a time, and getting the product right
        while getting the individual signs wrong leaves the at-the-money
        correct and flips the risk reversal -- which is the cross-pair error
        this project already had to fix once (MIGRATION.md 1.1).
        """
        self.assertEqual(moments.triangle_coefficients("AUDJPY", "AUDUSD", "USDJPY"), (1, 1))
        self.assertEqual(moments.triangle_coefficients("EURGBP", "EURUSD", "GBPUSD"), (1, -1))
        self.assertEqual(moments.triangle_coefficients("GBPNZD", "GBPUSD", "NZDUSD"), (1, -1))

    def test_the_coefficients_agree_with_the_curve_builders_signs(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all()
        crosses = [(n, s.legs) for n, s in book.data.pairs.items() if s.is_cross]
        self.assertTrue(crosses)
        for name, legs in crosses:
            ca, cb = moments.triangle_coefficients(name, *legs)
            sa, sb = infer_leg_signs(name, *legs)
            self.assertEqual(ca * cb, -sa * sb, msg=name)

    def test_legs_that_share_no_currency_are_refused(self):
        with self.assertRaises(ValueError):
            moments.triangle_coefficients("EURJPY", "AUDUSD", "GBPCAD")

    def test_a_marked_smile_reads_back_as_a_usable_distribution(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        expiry = ASOF.datetime_from_years(0.5)
        d = moments.distribution_from_surface(book["EURUSD"], expiry, method="SVI", cut="NY")
        self.assertGreater(d.captured, 0.99)
        self.assertLess(abs(d.forward_error), 1e-3)
        self.assertEqual(d.warnings, ())
        # A smile with a positive butterfly is fat-tailed: that is what a
        # butterfly *is*, so the density must say so.
        self.assertGreater(d.moments().excess_kurtosis, 0.0)

    def test_the_machinery_reproduces_a_leg_it_is_given_alone(self):
        """The noise floor.  A cross difference smaller than this means nothing."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        surface = book["EURUSD"]
        expiry = ASOF.datetime_from_years(0.25)
        d = moments.distribution_from_surface(surface, expiry, method="SVI", cut="NY")
        table = surface.smile_table(expiry, method="SVI", cut="NY")
        by = {r["label"]: r["vol"] for r in table}
        ref = {"atm": by["ATM"], "rr25": by["25d call"] - by["25d put"],
               "fly25": 0.5 * (by["25d call"] + by["25d put"]) - by["ATM"]}
        err = moments.reconstruction_error(d, surface.conv, ref)
        for key, value in err.items():
            self.assertLess(abs(value), 2e-4, msg=f"{key} came back {value * 100:+.4f} vol points")

    def test_two_different_expiries_cannot_be_combined(self):
        with self.assertRaises(ValueError):
            moments.combine(_flat_distribution(0.1, 0.5), _flat_distribution(0.1, 0.25),
                            (1, 1), 0.3, DeltaConvention(False))


class TestHistory(unittest.TestCase):
    """The historical workbook, and what the market actually did."""

    def test_headers_are_read_by_meaning_not_by_position(self):
        cases = {
            "Fwd 1M": ("forward", "1M", None),
            "1M swap points": ("points", "1M", None),
            "3m atm vol": ("atm", "3M", None),
            "RR25 1M": ("rr", "1M", 25),
            "1M 25d rr": ("rr", "1M", 25),
            "10RR 3M": ("rr", "3M", 10),
            "1M 10d fly": ("bf", "1M", 10),
            "implied vol 6m": ("atm", "6M", None),
        }
        for header, (field_name, tenor, delta) in cases.items():
            got = history.parse_header(header)
            self.assertIsNotNone(got, header)
            self.assertEqual((got.field, got.tenor, got.delta), (field_name, tenor, delta), header)

    def test_a_ten_delta_wing_is_not_a_ten_day_tenor(self):
        """'RR 10d 1M' has two tokens that parse as tenors.

        Deciding in token order files the whole column under a 10-day maturity
        that does not exist, and the ten-delta series then goes missing without
        a word.
        """
        got = history.parse_header("RR 10d 1M")
        self.assertEqual((got.field, got.tenor, got.delta), ("rr", "1M", 10))
        got = history.parse_header("BF 10d 6M")
        self.assertEqual((got.field, got.tenor, got.delta), ("bf", "6M", 10))

    def test_a_header_naming_nothing_is_reported_not_guessed(self):
        self.assertIsNone(history.parse_header("Trader note"))
        self.assertIsNone(history.parse_header("3M"))          # a tenor with no field

    def test_the_volatility_unit_is_decided_per_sheet_not_per_column(self):
        """A 25 delta risk reversal of -0.89 vol points is below 1 in magnitude.

        Sniffing each column on its own reads that as a decimal and returns it
        a hundred times too large, while the at-the-money column beside it is
        read correctly -- so the error shows up only in the skew.
        """
        h = history.load_history(HISTORY, ["EURUSD", "USDJPY", "EURJPY", "GBPUSD"])
        for pair in ("EURUSD", "USDJPY", "EURJPY"):
            atm = h[pair].series("atm", "3M")[-1]
            rr = h[pair].series("rr", "3M", 25)[-1]
            self.assertTrue(0.01 < atm < 0.50, f"{pair} ATM came back as {atm}")
            self.assertLess(abs(rr), 0.05, f"{pair} risk reversal came back as {rr}")

    def test_a_low_at_the_money_is_still_read_as_points(self):
        """A pegged pair marks its at-the-money below one volatility point.

        The reader used to call any sheet whose at-the-money sat under 1.0 a
        sheet of decimals, so USDHKD's 0.35 came back as 0.35 *decimal* and
        the monitor showed it at 35 vol points.  What the sheet says is what
        the number is: 0.35 points, and the risk reversal and butterfly beside
        it on the same scale.  A genuinely decimal sheet is something the
        caller says, with ``vol_unit='decimal'``.
        """
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pegged.xlsx"
            pd.DataFrame({
                "Date": pd.to_datetime(["2024-05-01", "2024-05-02", "2024-05-03"]),
                "Spot": [7.81, 7.812, 7.815],
                "3M ATM": [0.35, 0.36, 0.34],
                "3M 25d RR": [0.12, 0.13, 0.11],
                "3M 25d BF": [0.08, 0.08, 0.09],
            }).to_excel(path, sheet_name="USDHKD", index=False)
            h = history.load_history(path, ["USDHKD"])
            hist = h["USDHKD"]
            self.assertAlmostEqual(hist.series("atm", "3M")[-1], 0.0034)
            self.assertAlmostEqual(hist.series("rr", "3M", 25)[-1], 0.0011)
            self.assertAlmostEqual(hist.series("bf", "3M", 25)[-1], 0.0009)
            # It is the one reading somebody might have meant the other way,
            # so it is said once rather than guessed at in silence.
            self.assertTrue(any("read as written" in p for p in h.problems), h.problems)
            # And the other reading is still available, by name.
            dec = history.load_history(path, ["USDHKD"], vol_unit="decimal")
            self.assertAlmostEqual(dec["USDHKD"].series("atm", "3M")[-1], 0.34)

    def test_the_monitor_shows_a_low_at_the_money_as_it_is_written(self):
        """The same number, at the edge a person reads it at (§4).

        curves is decimals throughout and the page multiplies by 100, so a
        0.35 point at-the-money read as a decimal reached the monitor tile at
        35.00 -- a hundred times the mark, on the screen a desk opens first.
        """
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "pegged.xlsx"
            pd.DataFrame({
                "Date": pd.to_datetime(["2024-05-01", "2024-05-02"]),
                "Spot": [7.81, 7.812],
                "3M ATM": [0.35, 0.36],
                "3M 25d RR": [0.12, 0.13],
            }).to_excel(path, sheet_name="USDHKD", index=False)
            h = history.load_history(path, ["USDHKD"])
            from volkit import curves
            curve = curves.history_curve(h["USDHKD"])
            point = curve.at("3M")
            self.assertAlmostEqual(point.values["atm"] * 100.0, 0.36)
            self.assertAlmostEqual(point.values["rr25"] * 100.0, 0.13)

    def test_forcing_the_unit_matches_what_auto_detected(self):
        a = history.load_history(HISTORY, ["EURUSD"])
        b = history.load_history(HISTORY, ["EURUSD"], vol_unit="percent")
        self.assertTrue(np.allclose(a["EURUSD"].series("atm", "3M"),
                                    b["EURUSD"].series("atm", "3M")))
        c = history.load_history(HISTORY, ["EURUSD"], vol_unit="decimal")
        self.assertTrue(np.allclose(c["EURUSD"].series("atm", "3M"),
                                    100.0 * a["EURUSD"].series("atm", "3M")))

    def test_unreadable_columns_are_named(self):
        h = history.load_history(HISTORY, ["EURUSD", "USDJPY", "EURJPY", "GBPUSD"])
        self.assertTrue(any("Trader note" in p for p in h.problems))

    def test_forward_points_are_turned_into_outrights(self):
        """USDJPY's sheet quotes swap points; EURUSD's quotes outrights."""
        h = history.load_history(HISTORY, ["EURUSD", "USDJPY", "EURJPY", "GBPUSD"])
        jpy = h["USDJPY"]
        self.assertIn("3M", jpy.forwards)
        self.assertGreater(jpy.forwards["3M"][-1], 50.0)       # an outright, not points
        self.assertNotEqual(jpy.forwards["3M"][-1], jpy.spot[-1])

    def test_a_sheet_that_is_not_a_pair_is_skipped_with_a_reason(self):
        h = history.load_history(HISTORY, ["EURUSD", "USDJPY", "EURJPY", "GBPUSD"])
        self.assertEqual(sorted(h.pairs), ["EURJPY", "EURUSD", "GBPUSD", "USDJPY"])
        with self.assertRaises(history.HistoryError):
            h["AUDNZD"]

    def test_volatility_time_matches_the_models_own_integration(self):
        """Realized and implied must be measured on the same clock.

        A flat-backbone curve integrates to ``sigma * sqrt(voltime / t)``; if
        the window measure here did not agree with the one inside the model,
        every realized-against-implied number on the screen would be biased by
        the difference.
        """
        from volkit.timeweight import TimeWeighting
        curve = AtmCurve(pair="EURUSD",
                         params=BackboneParams(initial_vol=0.10, long_term_vol=0.10,
                                               mean_reversion=1.0),
                         clock=ASOF, weighting=TimeWeighting("EURUSD"))
        t = 0.5
        vt = history.volatility_time(
            "EURUSD", ASOF.now, ASOF.now + timedelta(days=t * 365.2425))
        self.assertAlmostEqual(curve.integrated_vol(t), 0.10 * math.sqrt(vt / t), places=4)

    def test_a_calendar_year_is_less_than_a_year_of_volatility_time(self):
        vt = history.volatility_time("EURUSD", datetime(2023, 2, 28, tzinfo=UTC),
                                     datetime(2024, 2, 28, tzinfo=UTC))
        self.assertLess(vt, 0.9)          # weekends and holidays are nearly free
        self.assertGreater(vt, 0.6)

    def test_the_three_annualisations_are_all_reported_and_differ(self):
        h = history.load_history(HISTORY, ["EURUSD"])
        r = history.realized(h["EURUSD"], 365)
        self.assertGreater(r.vol, r.vol_calendar)      # weighted time is shorter
        self.assertGreater(r.vol, r.vol_count)
        self.assertEqual(r.annualisation, "weighted")
        # The ratio between two of them is arithmetic, not an estimate.
        self.assertAlmostEqual(r.vol / r.vol_count,
                               math.sqrt((r.observations / 252.0) / r.vol_time), places=9)

    def test_the_count_annualisation_recovers_the_walk_it_was_simulated_at(self):
        """The sample is a lognormal walk with one step per business day."""
        h = history.load_history(HISTORY, ["EURUSD"])
        r = history.realized(h["EURUSD"], 730)
        self.assertAlmostEqual(r.vol_count, 0.0524, delta=0.006)

    def test_too_short_a_window_is_refused_with_the_count(self):
        h = history.load_history(HISTORY, ["EURUSD"])
        with self.assertRaises(history.HistoryError) as cm:
            history.realized(h["EURUSD"], 7)
        self.assertIn("observation", str(cm.exception))

    def test_skew_and_kurtosis_carry_their_standard_errors(self):
        h = history.load_history(HISTORY, ["EURUSD"])
        r = history.realized(h["EURUSD"], 365)
        self.assertAlmostEqual(r.skew_se, math.sqrt(6.0 / r.observations), delta=0.02)
        self.assertAlmostEqual(r.kurtosis_se, math.sqrt(24.0 / r.observations), delta=0.06)

    def test_shape_projects_onto_a_horizon_by_the_independence_rule(self):
        """Skewness falls as 1/sqrt(n) and excess kurtosis as 1/n in the steps."""
        h = history.load_history(HISTORY, ["EURUSD"])
        r = history.realized(h["EURUSD"], 365)
        t = 0.25
        n = t / (r.vol_time / r.observations)
        self.assertAlmostEqual(r.scaled_skew(t), r.skew / math.sqrt(n), places=12)
        self.assertAlmostEqual(r.scaled_excess_kurtosis(t), r.excess_kurtosis / n, places=12)
        self.assertLess(abs(r.scaled_skew(t)), abs(r.skew))

    # -- the swap points are part of what was realized --------------------

    def _synthetic(self, pair="EURUSD", n=400, vol=0.10, carry=0.05, points_vol=0.0, seed=7):
        """A sheet with a known spot walk and a known carry curve.

        ``carry`` is the annualised continuous carry the forward is built at;
        ``points_vol`` is how much that carry itself wobbles day to day.  With
        ``points_vol`` zero the forward's only extra motion is the *decay* of
        the points, which is a known slide and not a risk.
        """
        rng = np.random.default_rng(seed)
        dt = 1.0 / 252.0
        dates = [date(2022, 1, 3) + timedelta(days=int(i)) for i in range(n)]
        steps = rng.normal(0.0, vol * math.sqrt(dt), n - 1)
        spot = 1.10 * np.exp(np.concatenate(([0.0], np.cumsum(steps))))
        c = carry + (rng.normal(0.0, points_vol, n) if points_vol else 0.0)
        h = history.PairHistory(pair=pair, dates=dates, spot=spot)
        for tenor, tau in (("1M", 1.0 / 12.0), ("3M", 0.25), ("1Y", 1.0)):
            h.forwards[tenor] = spot * np.exp(c * tau)
        return h

    def test_a_pure_carry_decay_is_not_counted_as_volatility(self):
        """The points *decaying* by one day of carry is a known slide.

        Leaving it in the sum of squares books the carry itself as
        volatility, which is exactly backwards for the pairs the forward basis
        exists for -- so a forward built on a perfectly constant 5% carry must
        realize what spot realized, to the last digit.
        """
        h = self._synthetic(carry=0.05, points_vol=0.0)
        r = history.realized(h, 365, basis="forward", basis_tenor="1Y")
        self.assertEqual(r.basis, "forward")
        self.assertAlmostEqual(r.vol, r.vol_spot, places=12)
        self.assertAlmostEqual(r.carry_rate, 0.05, places=9)
        self.assertLess(r.points_vol, 1e-12)

    def test_the_swap_points_moving_is_realized_volatility(self):
        """A forward whose carry wobbles realizes more than spot did.

        The tenor multiplies it: a one-year forward carries a whole year of
        the carry's move, a one-month forward a twelfth of it.
        """
        h = self._synthetic(carry=0.05, points_vol=0.01)
        spot_only = history.realized(h, 365, basis="spot", basis_tenor="1Y")
        year = history.realized(h, 365, basis="forward", basis_tenor="1Y")
        month = history.realized(h, 365, basis="forward", basis_tenor="1M")
        self.assertGreater(year.vol, spot_only.vol)
        self.assertGreater(year.points_vol, month.points_vol * 5.0)
        # Independent by construction, so the variances add.
        self.assertAlmostEqual(year.vol ** 2,
                               year.vol_spot ** 2 + year.points_vol ** 2,
                               delta=0.05 * year.vol ** 2)

    def test_a_tenor_the_sheet_does_not_quote_is_interpolated_not_dropped(self):
        """Falling back to spot on the misses put two different measurements
        in one column, so the term structure of realized volatility grew steps
        at whichever tenors the sheet happened to quote."""
        h = self._synthetic(carry=0.05, points_vol=0.01)
        self.assertNotIn("6M", h.forwards)
        r = history.realized(h, 365, basis="auto", basis_tenor="6M")
        self.assertEqual(r.basis, "forward")
        self.assertTrue(any("interpolated" in w for w in r.warnings))
        three, one = (history.realized(h, 365, basis="forward", basis_tenor=t).points_vol
                      for t in ("3M", "1Y"))
        self.assertGreater(r.points_vol, three)
        self.assertLess(r.points_vol, one)

    def test_a_sheet_with_no_points_at_all_falls_back_and_says_so(self):
        h = self._synthetic()
        h.forwards.clear()
        r = history.realized(h, 365, basis="auto", basis_tenor="3M")
        self.assertEqual(r.basis, "spot")
        self.assertTrue(any("realized on spot" in w for w in r.warnings))
        with self.assertRaises(history.HistoryError):
            history.realized(h, 365, basis="forward", basis_tenor="3M")

    # -- what the volatility itself did -----------------------------------

    def test_vol_dynamics_recovers_the_correlation_and_vol_of_vol_it_was_built_with(self):
        """rho and nu are the two numbers a SABR smile is made of, and under
        beta = 1 they are directly measurable off the quoted at-the-money."""
        rng = np.random.default_rng(11)
        n, dt = 1500, 1.0 / 252.0
        rho, nu = -0.55, 0.60
        z1 = rng.normal(size=n - 1)
        z2 = rho * z1 + math.sqrt(1.0 - rho * rho) * rng.normal(size=n - 1)
        spot = 1.10 * np.exp(np.concatenate(([0.0], np.cumsum(0.10 * math.sqrt(dt) * z1))))
        vol = 0.10 * np.exp(np.concatenate(([0.0], np.cumsum(nu * math.sqrt(dt) * z2))))
        h = history.PairHistory(
            pair="EURUSD",
            dates=[date(2020, 1, 1) + timedelta(days=int(i)) for i in range(n)],
            spot=spot, atm={"3M": vol})
        d = history.vol_dynamics(h, 5000, "3M")
        self.assertEqual(d.source, "quoted")
        self.assertAlmostEqual(d.rho, rho, delta=3.0 * d.rho_se)
        # nu is per unit of *volatility time*, not per 252 business days, for
        # the same reason the realized volatility is: it has to be comparable
        # with the nu a marked smile implies.  The series above was built on a
        # flat 1/252 step, so recovering it means converting onto the model's
        # own clock first -- a calendar year holds about 0.78 years of it.
        per_step = d.vol_time / d.observations
        self.assertAlmostEqual(d.nu, nu * math.sqrt(dt / per_step), delta=0.08 * d.nu)
        self.assertLess(per_step, dt)

    def test_vol_dynamics_falls_back_to_a_rolling_volatility_and_warns(self):
        """A rolling average moves less than the thing it averages, so the
        fallback is a floor and has to say so."""
        rng = np.random.default_rng(3)
        n = 400
        spot = 1.10 * np.exp(np.cumsum(rng.normal(0.0, 0.006, n)))
        h = history.PairHistory(
            pair="EURUSD",
            dates=[date(2021, 1, 1) + timedelta(days=int(i)) for i in range(n)],
            spot=spot)
        d = history.vol_dynamics(h, 5000, "3M")
        self.assertEqual(d.source, "rolling")
        self.assertTrue(any("floor" in w for w in d.warnings))

    def test_a_tenor_the_sheet_does_not_quote_uses_the_nearest_one_by_name(self):
        """Interpolating a volatility column would be interpolating something
        whose *changes* are the measurement; the nearest real column is used."""
        h = self._synthetic()
        h.atm["3M"] = np.full(len(h.dates), 0.10) * np.exp(
            np.linspace(0.0, 0.3, len(h.dates)))
        d = history.vol_dynamics(h, 5000, "2M")
        self.assertEqual(d.tenor, "3M")
        self.assertTrue(any("3M column instead" in w for w in d.warnings))

    def test_a_percentile_locates_todays_mark_in_its_own_history(self):
        h = history.load_history(HISTORY, ["EURUSD"])
        series = h["EURUSD"].series("atm", "3M")
        lo, hi = float(np.min(series)), float(np.max(series))
        self.assertEqual(history.implied_stats(h["EURUSD"], 3650, "atm", "3M",
                                               current=lo - 0.01).percentile, 0.0)
        self.assertEqual(history.implied_stats(h["EURUSD"], 3650, "atm", "3M",
                                               current=hi + 0.01).percentile, 100.0)


class TestAnalysis(unittest.TestCase):
    """Carry and roll, fair value, and the cross triangle."""

    @classmethod
    def setUpClass(cls):
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all()
        cls.book.feed = MarketFeed.load(FEED)
        cls.history = history.load_history(HISTORY, cls.book.pairs)

    def test_every_advertised_target_resolves_to_legs(self):
        for key in analytics.TARGETS:
            legs = analytics._target_legs(key)
            self.assertTrue(legs, key)
            self.assertAlmostEqual(sum(w for w, _, _ in legs),
                                   0.0 if key.startswith("rr") else
                                   (0.0 if key.startswith("fly") else 1.0), places=12, msg=key)
        with self.assertRaises(ValueError):
            analytics._target_legs("nonsense")

    def test_the_roll_splits_exactly_into_term_and_smile(self):
        rows = [r for r in analytics.carry_table(self.book, "EURUSD", horizon_days=30, cut="NY")
                if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertAlmostEqual(r.roll, r.roll_term + r.roll_smile, places=15, msg=r.tenor)

    def test_without_a_forward_feed_the_smile_slide_is_zero_and_says_so(self):
        """USDCNY has marks but no feed, so the strike can only be held in moneyness.

        Not a cross: a cross the feed quotes both legs of has a forward built
        from the triangle (``TestCrossLevelsFromTheLegs``), so EURJPY -- which
        this used to be written on -- is no longer a pair with no feed.
        """
        rows = [r for r in analytics.carry_table(self.book, "USDCNY", horizon_days=30, cut="NY")
                if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertEqual(r.roll_smile, 0.0)
            self.assertTrue(any("forward feed" in w for w in r.warnings), r.tenor)

    def test_a_forward_curve_makes_the_smile_slide_bite(self):
        rows = {r.tenor: r for r in analytics.carry_table(
            self.book, "EURUSD", horizon_days=30, target="25dp", cut="NY") if r.expiry}
        self.assertTrue(any(abs(r.roll_smile) > 1e-6 for r in rows.values()))

    def test_a_tenor_shorter_than_the_horizon_is_reported_not_dropped(self):
        rows = analytics.carry_table(self.book, "EURUSD", horizon_days=30, cut="NY")
        self.assertEqual([r.tenor for r in rows], list(self.book.data.tenor_points))
        short = [r for r in rows if not r.expiry]
        self.assertTrue(short)
        for r in short:
            self.assertTrue(r.warnings and "horizon" in r.warnings[0])

    def test_the_carry_delta_is_the_smile_delta_not_the_black_scholes_one(self):
        """The bug this was written for.

        The carry table reported a Black-Scholes delta -- the sensitivity with
        the volatility *held fixed* as the forward moves -- for a table whose
        entire subject is a fixed strike sliding under a moving forward.  The
        volatility that strike is marked at moves with it, and on a skewed
        surface the difference is several delta.
        """
        moved = 0
        for pair in ("EURUSD", "USDJPY"):
            for target in ("25dc", "25dp", "10dp"):
                rows = [r for r in analytics.carry_table(
                    self.book, pair, horizon_days=7, target=target, cut="NY") if r.expiry]
                self.assertTrue(rows)
                for r in rows:
                    self.assertIsNotNone(r.smile_delta, f"{pair} {target} {r.tenor}")
                    self.assertAlmostEqual(r.skew_delta, r.smile_delta - r.delta, places=15)
                    if abs(r.skew_delta) > 0.01:
                        moved += 1
        self.assertGreater(moved, 20, "the skew moved no delta by a whole point")

    def test_the_smile_delta_is_black_scholes_plus_vega_times_the_skew_slope(self):
        """``dV/dF = dV/dF|sigma + vega * dsigma/dF``, and nothing else.

        Pinned by finite difference so a change to the smile delta that broke
        this identity could not pass as a refinement.
        """
        for pair in ("EURUSD", "USDJPY"):
            surface = self.book[pair]
            rows = [r for r in analytics.carry_table(
                self.book, pair, horizon_days=7, target="25dp", cut="NY") if r.expiry]
            for r in rows[:4]:
                expiry = self.book.clock.datetime_from_years(r.t)
                f, k = r.forward, r.strike
                bs = float(black.delta(f, k, r.level, r.t, False))
                vega = float(black.vega(f, k, r.level, r.t))
                eps = f * 1e-4
                slope = (float(surface.vol(k / (f + eps), expiry, "SVI", "NY"))
                         - float(surface.vol(k / (f - eps), expiry, "SVI", "NY"))) / (2 * eps)
                self.assertAlmostEqual(r.delta, bs, places=12, msg=f"{pair} {r.tenor}")
                # Both sides are central differences at different bumps, so
                # they agree to the second-order term and not beyond it; a
                # short tenor has the most curvature and the widest gap.
                self.assertAlmostEqual(r.smile_delta, bs + vega * slope, delta=1e-3,
                                       msg=f"{pair} {r.tenor}")

    def test_the_smile_delta_values_the_whole_of_what_the_forward_move_does(self):
        """And the Black-Scholes delta values only half the story.

        The forward reaches the position twice: through the price at a fixed
        volatility (``carry_pnl``) and through the mark, because the strike's
        moneyness changed (``vega * roll_smile``).  The smile delta is the
        first-order coefficient of *both*; the Black-Scholes delta is the
        coefficient of the first alone.  Measured over a one-day horizon,
        where the second order terms have not had room to matter.
        """
        checked = 0
        for pair in ("EURUSD", "USDJPY"):
            for target in ("25dc", "25dp", "10dp"):
                rows = [r for r in analytics.carry_table(
                    self.book, pair, horizon_days=1, target=target, cut="NY") if r.expiry]
                for r in rows:
                    # A one-day roll is only "small" against a tenor with room
                    # in it: a week rolled by a day is a sixth of its life and
                    # the second-order terms are no longer negligible.
                    move = r.forward_rolled - r.forward
                    whole = r.vega * r.roll_smile + r.carry_pnl
                    if r.t < 1.0 / 12.0 or abs(whole) < 1e-9 or abs(move) < 1e-12:
                        continue
                    where = f"{pair} {target} {r.tenor}"
                    self.assertAlmostEqual(r.smile_delta * move / whole, 1.0, delta=0.02,
                                           msg=where)
                    self.assertAlmostEqual(r.delta * move / r.carry_pnl, 1.0, delta=0.02,
                                           msg=where)
                    # And the point: the Black-Scholes delta is not a reading
                    # of the whole move at all, by a margin far outside the
                    # tolerance above.
                    self.assertGreater(abs(r.delta * move - whole),
                                       5.0 * abs(r.smile_delta * move - whole), where)
                    checked += 1
        self.assertGreater(checked, 20)

    def test_the_at_the_money_straddle_is_delta_neutral_only_in_black_scholes(self):
        """A straddle is long vega, and on a skewed surface the volatility
        moves with the forward, so the delta-neutral strike is delta neutral
        in one column and not in the other.  Reporting the Black-Scholes zero
        alone said the position had no forward exposure when it had several
        delta of it."""
        # EURUSD quotes an unadjusted delta, so its delta-neutral straddle
        # really is Black-Scholes delta neutral and the whole of the smile
        # delta is the skew.
        rows = [r for r in analytics.carry_table(
            self.book, "EURUSD", horizon_days=7, target="atm", cut="NY") if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertAlmostEqual(r.delta, 0.0, places=9, msg=r.tenor)
            self.assertAlmostEqual(r.skew_delta, r.smile_delta, places=12, msg=r.tenor)
        self.assertGreater(max(abs(r.smile_delta) for r in rows), 0.0)

        # USDJPY quotes a premium-adjusted one, so the delta-neutral strike is
        # neutral in *that* convention and carries a little unadjusted delta
        # already.  The skew is still much the larger part of what it runs.
        rows = [r for r in analytics.carry_table(
            self.book, "USDJPY", horizon_days=7, target="atm", cut="NY") if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertLess(abs(r.delta), 0.05, r.tenor)
        biggest = max(rows, key=lambda r: abs(r.smile_delta))
        self.assertGreater(abs(biggest.smile_delta), 0.05)
        self.assertGreater(abs(biggest.skew_delta), 2.0 * abs(biggest.delta))

    def test_the_carry_deltas_are_term_currency_not_the_quoted_convention(self):
        """A premium-adjusted delta is a hedge ratio in the *other* currency.

        Multiplying it by a move in the forward does not give the money the
        position made, and money is what this table reports, so both columns
        are dV/dF and the surface's own convention is deliberately not used.
        """
        surface = self.book["USDJPY"]
        self.assertTrue(bool(surface.conv), "USDJPY should be premium adjusted here")
        r = [x for x in analytics.carry_table(
            self.book, "USDJPY", horizon_days=7, target="25dc", cut="NY") if x.expiry][-1]
        quoted = float(black.delta(r.forward, r.strike, r.level, r.t, True, surface.conv))
        self.assertAlmostEqual(quoted, 0.25, places=6)          # the strike is a 25 delta one
        self.assertNotAlmostEqual(r.delta, quoted, places=3)    # and this column is not that
        self.assertAlmostEqual(
            r.delta, float(black.delta(r.forward, r.strike, r.level, r.t, True)), places=12)
        # The override exists for exactly this caller and changes nothing else.
        expiry = self.book.clock.datetime_from_years(r.t)
        own = surface.smile_delta(r.forward, r.strike, expiry, True, "SVI", "NY")
        term = surface.smile_delta(r.forward, r.strike, expiry, True, "SVI", "NY", conv=False)
        self.assertAlmostEqual(term, r.smile_delta, places=12)
        self.assertNotAlmostEqual(own, term, places=3)

    def test_the_forward_carry_is_delta_hedged_before_it_pays_for_a_break_even(self):
        """A break-even volatility is a property of the strike, not of the side.

        ``carry_pnl`` is the whole revaluation at the rolled forward and is
        the right number for a spot-hedged *position*.  It is the wrong one
        for a break-even: put-call parity puts the entire difference between
        the call's and the put's revaluation at one strike into the
        first-order term ``delta * (F2 - F1)``, so read unhedged the same
        strike is "rich" as a call and "cheap" as a put by a quarter of the
        forward move.  ``carry_hedged`` takes that term out and what is left
        is the gamma over the move -- which, being the convexity of a long
        option, cannot be negative whichever side it is written as.  That is
        the bug the relative-value grid's carry signal had: the score flipped
        sign across the strike axis with the option's direction.
        """
        for pair in ("EURUSD", "USDJPY", "EURJPY"):
            for target in ("atm", "25dp", "25dc", "10dp", "10dc"):
                rows = [r for r in analytics.carry_table(
                    self.book, pair, horizon_days=7, target=target, cut="NY")
                    if r.expiry and r.carry_pnl is not None]
                self.assertTrue(rows, f"{pair} {target}")
                for r in rows:
                    where = f"{pair} {target} {r.tenor}"
                    self.assertAlmostEqual(
                        r.carry_hedged,
                        r.carry_pnl - r.delta * (r.forward_rolled - r.forward),
                        places=15, msg=where)
                    # Convexity: V(F2) - V(F1) - delta * (F2 - F1) >= 0.
                    self.assertGreaterEqual(r.carry_hedged, -1e-15, msg=where)

    def test_a_call_and_a_put_at_one_strike_carry_the_same_break_even(self):
        """The exact identity behind the fix, at one strike rather than two.

        The carry columns are read at a 25 delta *call* strike and a 25 delta
        *put* strike, which are two different strikes, so the grid alone
        cannot show that the direction has stopped mattering.  Here both sides
        are priced at one strike: the raw revaluations differ by the whole
        forward move, and the hedged ones are the same number.
        """
        r = [x for x in analytics.carry_table(
            self.book, "USDJPY", horizon_days=7, target="25dc", cut="NY") if x.expiry][-1]
        f1, f2, k, vol, t = r.forward, r.forward_rolled, r.strike, r.level, r.t
        pnl, hedged = {}, {}
        for call in (True, False):
            pnl[call] = float(black.price(f2, k, vol, t, call) - black.price(f1, k, vol, t, call))
            hedged[call] = pnl[call] - float(black.delta(f1, k, vol, t, call)) * (f2 - f1)
        self.assertAlmostEqual(pnl[True] - pnl[False], f2 - f1, places=12)
        self.assertGreater(abs(pnl[True] - pnl[False]), 0.9 * abs(f2 - f1))
        # A yen strike is ~150, so the cancellation leaves a few times 1e-14
        # of dust on a number of order 1e-5; the identity is exact in exact
        # arithmetic and this is the last bit of a double, not a difference.
        self.assertAlmostEqual(hedged[True], hedged[False], delta=1e-12)
        self.assertAlmostEqual(hedged[True], r.carry_hedged, delta=1e-12)

    def test_the_fair_value_carry_is_the_gamma_and_not_the_residual_delta(self):
        """What the delta hedge changed on the fair-value card, and where.

        The at-the-money straddle is delta neutral in the pair's **own**
        quoted convention, so on a pair quoting an unadjusted delta the hedge
        is exactly zero and nothing moves.  On a premium-adjusted pair the
        delta-neutral strike is neutral in *that* convention and its ``dV/dF``
        is not quite zero, so the old ``carry_value`` carried a residual first
        order term.  The tell is that it grew with the tenor -- a delta term
        is linear in the forward move -- while a real gamma term is flat.
        """
        eur = analytics.fair_value_table(self.book, "EURUSD", None, horizon_days=7, cut="NY")
        self.assertFalse(bool(self.book["EURUSD"].conv))
        for r in eur:
            if r.carry_pnl is not None:
                self.assertAlmostEqual(r.carry_hedged, r.carry_pnl, places=15, msg=r.tenor)
        jpy = [r for r in analytics.fair_value_table(
            self.book, "USDJPY", None, horizon_days=7, cut="NY") if r.carry_pnl is not None]
        self.assertTrue(bool(self.book["USDJPY"].conv))
        self.assertTrue(jpy)
        for r in jpy:
            self.assertLess(abs(r.carry_hedged), 0.5 * abs(r.carry_pnl), msg=r.tenor)
        # Flat, not growing: the residual delta was linear in the forward move
        # and so scaled with the tenor, and the gamma over one horizon's move
        # does not.  Unhedged this ratio was better than twenty.
        values = [abs(r.carry_value) for r in jpy]
        unhedged = [abs(r.carry_value * r.carry_pnl / r.carry_hedged) for r in jpy]
        self.assertLess(max(values), 2.0 * min(values))
        self.assertGreater(max(unhedged), 10.0 * min(unhedged))

    def test_the_fair_value_roll_is_the_atm_roll_whatever_is_displayed(self):
        """Feeding an at-the-money implied and a risk-reversal roll into one
        break-even mixes two different positions."""
        a = analytics.fair_value_table(self.book, "EURUSD", None, horizon_days=30, cut="NY")
        b = analytics.fair_value_table(self.book, "EURUSD", None, horizon_days=30, cut="NY")
        carry_rr = {r.tenor: r for r in analytics.carry_table(
            self.book, "EURUSD", horizon_days=30, target="rr25", cut="NY")}
        carry_atm = {r.tenor: r for r in analytics.carry_table(
            self.book, "EURUSD", horizon_days=30, target="atm", cut="NY")}
        self.assertEqual([r.roll for r in a], [r.roll for r in b])
        for r in a:
            self.assertAlmostEqual(r.roll, carry_atm[r.tenor].roll, places=15)
            self.assertNotAlmostEqual(r.roll, carry_rr[r.tenor].roll, places=6)

    def test_richness_is_implied_less_realized_less_the_roll_and_the_carry(self):
        """The break-even gained a third term when the forward curve's
        price-side carry was added; ``carry_value`` is it."""
        rows = analytics.fair_value_table(self.book, "EURUSD", self.history["EURUSD"],
                                          horizon_days=30, cut="NY")
        priced = [r for r in rows if r.fair is not None]
        self.assertTrue(priced)
        for r in priced:
            self.assertAlmostEqual(r.richness,
                                   r.implied - r.realized - r.roll_value - r.carry_value,
                                   places=15)
            self.assertAlmostEqual(r.roll_value, r.roll * r.roll_multiplier, places=15)

    def test_a_tenor_with_too_little_history_keeps_its_row(self):
        rows = analytics.realized_table(self.book, "EURUSD", self.history["EURUSD"], cut="NY")
        self.assertEqual([r.tenor for r in rows], list(self.book.data.tenor_points))
        blank = [r for r in rows if r.observations == 0]
        self.assertTrue(blank)
        self.assertTrue(blank[0].warnings)

    def test_matching_the_window_to_the_tenor_is_the_default(self):
        rows = analytics.realized_table(self.book, "EURUSD", self.history["EURUSD"], cut="NY")
        for r in rows:
            self.assertAlmostEqual(r.window_days, r.t * 365.2425, places=6)
        fixed = analytics.realized_table(self.book, "EURUSD", self.history["EURUSD"],
                                         lookback_days=90, cut="NY")
        self.assertTrue(all(r.window_days == 90.0 for r in fixed))

    def test_the_variance_triangle_reproduces_the_marked_cross(self):
        """The book builds a cross from exactly this expression, so its own
        at-the-money mark must sit on it bar the cross's own add-on."""
        rows = analytics.triangle_table(self.book, "EURJPY", cut="NY", tenors=["3m", "1y"],
                                        with_noise=False)
        for r in rows:
            self.assertLess(abs(r.variance_triangle_atm - r.marked["atm"]), 0.0015, msg=r.tenor)
            self.assertAlmostEqual(r.implied_correlation, r.rho, delta=0.02, msg=r.tenor)

    def test_the_distribution_triangle_sits_above_the_variance_one(self):
        """It carries the legs' whole densities, whose variance is larger than
        their at-the-money volatilities by the convexity of their own smiles."""
        r = analytics.triangle_table(self.book, "EURJPY", cut="NY", tenors=["3m"],
                                     with_noise=False)[0]
        self.assertGreater(r.smile_convexity, 0.0)
        self.assertLess(r.smile_convexity, 0.01)

    def test_the_triangle_reports_its_own_noise_floor(self):
        r = analytics.triangle_table(self.book, "EURJPY", cut="NY", tenors=["3m"])[0]
        self.assertTrue(r.noise)
        for key, value in r.noise.items():
            self.assertLess(value, 5e-4, msg=key)

    def test_the_triangle_coefficients_reach_the_row(self):
        r = analytics.triangle_table(self.book, "EURGBP", cut="NY", tenors=["3m"],
                                     with_noise=False)[0]
        self.assertEqual(r.coefficients, (1, -1))

    def test_a_cross_loaded_on_its_own_still_has_a_triangle(self):
        """``load_all`` builds a cross's legs but deliberately does not fit
        their smiles -- nothing else needs them.  The triangle does, and it is
        the only thing that does, so it must arrange them itself rather than
        returning an empty table."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURGBP"])
        self.assertFalse(book["EURUSD"].fits)          # not fitted by the load
        rows = analytics.triangle_table(book, "EURGBP", cut="NY", tenors=["3m"],
                                        with_noise=False)
        self.assertEqual(len(rows), 1)
        self.assertIn("rr25", rows[0].triangle)
        self.assertTrue(book["EURUSD"].fits)           # fitted on demand

    def test_a_tenor_the_triangle_cannot_build_keeps_its_row(self):
        """An earlier cut dropped the tenor and its reason together, so when
        every tenor failed the table came back silently empty."""
        def boom(*a, **kw):
            raise ValueError("no density here")

        original = moments.distribution_from_surface
        moments.distribution_from_surface = boom
        try:
            rows = analytics.triangle_table(self.book, "EURJPY", cut="NY", tenors=["3m"])
        finally:
            moments.distribution_from_surface = original
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].marked, {})
        self.assertEqual(rows[0].coefficients, (1, 1))
        self.assertTrue(any("no density here" in w for w in rows[0].warnings))

    def test_a_pair_that_is_not_a_cross_has_no_triangle(self):
        with self.assertRaises(ValueError) as cm:
            analytics.triangle_table(self.book, "EURUSD", cut="NY")
        self.assertIn("not a cross", str(cm.exception))

    # -- the forward curve's own carry -----------------------------------

    def test_the_at_the_money_row_carries_no_delta(self):
        """The at-the-money is a straddle, and a straddle at the delta-neutral
        strike has no delta.

        Reading that leg as the call alone -- which is how ``_target_legs``
        marks it, because it only ever needed a *strike* -- handed the
        at-the-money row half a unit of forward carry that nobody is running.
        """
        rows = [r for r in analytics.carry_table(self.book, "EURUSD", horizon_days=30,
                                                 target="atm", cut="NY") if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertAlmostEqual(r.delta, 0.0, places=9, msg=r.tenor)
            # ... so its carry is only the gamma over the forward's move.
            self.assertLess(abs(r.carry_pnl), 1e-5, r.tenor)

    def test_a_directional_target_earns_the_forwards_roll_down(self):
        """A 25 delta call is long the forward, and the forward rolls down."""
        rows = [r for r in analytics.carry_table(self.book, "EURUSD", horizon_days=30,
                                                 target="25dc", cut="NY") if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertAlmostEqual(r.delta, 0.25, delta=0.02, msg=r.tenor)
            # Full revaluation, so it is delta times the move plus the gamma
            # over it -- close to the linear reading but not equal to it.
            linear = r.delta * (r.forward_rolled - r.forward)
            self.assertAlmostEqual(r.carry_pnl, linear, delta=0.05 * abs(linear), msg=r.tenor)
            self.assertAlmostEqual(r.carry_vols, r.carry_pnl / r.vega, places=12, msg=r.tenor)
            self.assertAlmostEqual(r.total_pnl, r.pnl + r.carry_pnl, places=15, msg=r.tenor)
            # The rate the swap points are quoting, not a price difference.
            self.assertAlmostEqual(
                r.carry_rate,
                (r.forward_rolled - r.forward) / (r.forward * (30.0 / 365.2425)), places=12)

    def test_without_a_feed_the_carry_is_unavailable_not_zero(self):
        """f1 and f2 are both 1.0 without a feed, so every carry figure would
        come out an exact zero -- a silent zero dressed as a measurement.

        USDCNY rather than EURJPY: the feed carries neither, but it carries
        both of EURJPY's legs and so builds its forward from the triangle.
        """
        rows = [r for r in analytics.carry_table(self.book, "USDCNY", horizon_days=30,
                                                 target="25dc", cut="NY") if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertIsNone(r.carry_pnl, r.tenor)
            self.assertIsNone(r.delta, r.tenor)
            self.assertIsNone(r.total_pnl, r.tenor)
            self.assertTrue(math.isnan(r.carry_rate), r.tenor)

    def test_a_risk_reversal_declines_to_state_its_carry_in_vols(self):
        """A risk reversal has almost no net vega, so dividing its premium
        carry by that vega is a division by nearly nothing."""
        rows = [r for r in analytics.carry_table(self.book, "EURUSD", horizon_days=30,
                                                 target="rr25", cut="NY") if r.expiry]
        self.assertTrue(rows)
        for r in rows:
            self.assertIsNotNone(r.carry_pnl, r.tenor)
            self.assertIsNone(r.carry_vols, r.tenor)
            self.assertIsNone(r.total_pnl, r.tenor)      # a combination has no single P&L

    def test_fair_value_is_realized_plus_the_roll_and_the_carry(self):
        rows = analytics.fair_value_table(self.book, "EURUSD", self.history["EURUSD"],
                                          horizon_days=30, cut="NY")
        priced = [r for r in rows if r.fair is not None]
        self.assertTrue(priced)
        for r in priced:
            self.assertAlmostEqual(r.fair, r.realized + r.roll_value + r.carry_value, places=15)
            self.assertAlmostEqual(r.richness, r.implied - r.fair, places=15)
            # The straddle is delta neutral, so the price-side carry is second
            # order; the first-order forward effect is in the smile slide.
            self.assertLess(abs(r.carry_value), 0.001, r.tenor)

    # -- what "realized" is measured on ----------------------------------

    def test_realized_defaults_to_the_forward_and_says_so(self):
        rows = analytics.realized_table(self.book, "EURUSD", self.history["EURUSD"])
        live = [r for r in rows if r.observations]
        self.assertTrue(live)
        for r in live:
            self.assertEqual(r.realized_basis, "forward", r.tenor)
            self.assertIsNotNone(r.points_vol, r.tenor)
            # The spot-only figure is kept beside it rather than replaced.
            self.assertTrue(math.isfinite(r.realized_spot), r.tenor)
        spot = [r for r in analytics.realized_table(self.book, "EURUSD",
                                                    self.history["EURUSD"],
                                                    realized_basis="spot") if r.observations]
        self.assertTrue(all(r.realized_basis == "spot" for r in spot))
        self.assertAlmostEqual(spot[-1].realized, live[-1].realized_spot, places=12)

    # -- the wings as a SABR shape ----------------------------------------

    def test_the_marked_wings_read_back_as_a_correlation_and_a_vol_of_vol(self):
        rows = analytics.realized_table(self.book, "USDJPY", self.history["USDJPY"],
                                        with_sabr=True)
        live = [r for r in rows if r.implied_rho is not None]
        self.assertTrue(live)
        for r in live:
            # USDJPY's risk reversal is marked for the downside, so the
            # correlation that produces it is negative.
            self.assertLess(r.marked_rr, 0.0, r.tenor)
            self.assertLess(r.implied_rho, 0.0, r.tenor)
            self.assertGreater(r.implied_nu, 0.0, r.tenor)
            self.assertLess(r.implied_shape_error, 1e-6, r.tenor)
            if r.realized_rho is not None:
                self.assertAlmostEqual(r.rho_difference, r.implied_rho - r.realized_rho, places=15)
                self.assertAlmostEqual(r.nu_difference, r.implied_nu - r.realized_nu, places=15)

    def test_the_sabr_shape_is_off_unless_it_is_asked_for(self):
        rows = analytics.realized_table(self.book, "USDJPY", self.history["USDJPY"])
        self.assertTrue(all(r.implied_rho is None and r.realized_rho is None for r in rows))

    def test_the_dynamics_are_measured_over_their_own_window_not_the_lookback(self):
        """The bug: both difference columns blank at every tenor.

        ``(rho, nu)`` used to be measured over whatever realized lookback the
        screen was set to.  They need more paired observations than a realized
        volatility needs returns, so on a three-week lookback every tenor's
        realized figure came back and every tenor's measured pair was missing
        -- and with it both ``diff`` columns, which is the whole point of the
        card.  The window is a property of the measurement, not of the tenor
        being forecast, and it is never shorter than the lookback.
        """
        short = analytics.realized_table(self.book, "USDJPY", self.history["USDJPY"],
                                         lookback_days=21, with_sabr=True)
        self.assertTrue(short)
        for r in short:
            self.assertEqual(r.window_days, 21.0, r.tenor)
            self.assertGreaterEqual(r.dynamics_days, history.DYNAMICS_DAYS, r.tenor)
            self.assertIsNotNone(r.realized_rho, r.tenor)
            self.assertIsNotNone(r.rho_difference, r.tenor)
            self.assertIsNotNone(r.nu_difference, r.tenor)
        # A lookback longer than the dynamics window is not shortened to it.
        long = analytics.realized_table(self.book, "USDJPY", self.history["USDJPY"],
                                        lookback_days=500, with_sabr=True)
        self.assertTrue(all(r.dynamics_days == 500.0 for r in long))

    def test_a_tenor_with_no_realized_window_still_carries_its_sabr_shape(self):
        """A one-week window can never hold a week of returns.

        The realized statistics for that row are unavailable and say so; the
        wings as a SABR shape do not depend on them -- the marked half needs
        no history at all -- so losing the column group with them left the
        card's first row permanently blank.
        """
        rows = analytics.realized_table(self.book, "USDJPY", self.history["USDJPY"],
                                        with_sabr=True)
        short = [r for r in rows if not r.observations]
        self.assertTrue(short, "no tenor came up short, so this pins nothing")
        for r in short:
            self.assertIsNotNone(r.implied_rho, r.tenor)
            self.assertIsNotNone(r.realized_rho, r.tenor)
            self.assertIsNotNone(r.rho_difference, r.tenor)


class TestAnalysisApi(unittest.TestCase):
    def test_the_payload_carries_no_number_a_browser_cannot_parse(self):
        """Python's json writes NaN, which JSON.parse refuses.

        One unavailable cell would take the whole response down in the
        browser, so non-finite floats become null on the way out.
        """
        import json as _json
        from volkit.webapp import BookService, _finite
        service = BookService(str(WORKBOOK), ASOF, feed_path=str(FEED),
                              history_path=str(HISTORY))
        payload = service.analysis({"pair": "EURUSD", "cut": "NY", "horizon_days": "30",
                                    "lookback_days": "match", "noise": "0"})
        text = _json.dumps(_finite(payload), default=str)
        self.assertNotIn("NaN", text)
        self.assertNotIn("Infinity", text)
        self.assertEqual(_json.loads(text)["pair"], "EURUSD")

    def test_sections_fail_independently(self):
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF)      # no feed, no history
        payload = service.analysis({"pair": "EURUSD", "cut": "NY"})
        self.assertTrue(payload["carry"])               # still rolls, in moneyness
        self.assertIsNone(payload["realized"])
        self.assertIn("history", payload["unavailable"])
        self.assertIn("triangle", payload["unavailable"])


class TestRelativeValue(unittest.TestCase):
    """Scoring the expiry / strike surface.

    The grid is not a new model: every signal is one of the comparisons the
    Analysis screen already makes, read at a strike instead of at the
    at-the-money.  What is pinned here is that it stays that way -- the
    at-the-money column has to reproduce the fair-value table exactly, the
    three additive signals have to add to the richness, and a signal a cell
    does not have has to be renormalised away rather than counted as a zero.
    """

    @classmethod
    def setUpClass(cls):
        from volkit import relvalue
        cls.relvalue = relvalue
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all()
        cls.book.feed = MarketFeed.load(FEED)
        cls.history = history.load_history(HISTORY, cls.book.pairs)
        cls.grid = relvalue.relative_value(cls.book, "EURUSD", cls.history["EURUSD"],
                                           horizon_days=7, cut="NY")

    def test_the_at_the_money_column_reproduces_the_fair_value_table(self):
        """Two ways of computing one number is how they drift apart.

        The grid extends ``fair_value_table``'s break-even from the
        at-the-money to a strike; at the at-the-money itself it has to be the
        same arithmetic, to the last bit, or the screen is showing a richness
        in one card and a different richness in the next.
        """
        fair = {r.tenor: r for r in analytics.fair_value_table(
            self.book, "EURUSD", self.history["EURUSD"], horizon_days=7, cut="NY")}
        checked = 0
        for row in self.grid.rows:
            cell = [c for c in row.cells if c.column == "atm"]
            got = fair.get(row.tenor)
            if not cell or cell[0].richness is None or got is None or got.richness is None:
                continue
            self.assertAlmostEqual(cell[0].richness, got.richness, places=12, msg=row.tenor)
            checked += 1
        self.assertGreater(checked, 3, "no tenor was actually compared")

    def test_the_three_additive_signals_add_to_the_richness(self):
        """``level + shape + carry`` is ``implied(K) - fair(K)`` and nothing else.

        The other two answer different questions -- where the mark sits in its
        own history, and what the legs of a cross imply -- and adding them in
        would make the volatility-point column mean nothing.
        """
        seen = 0
        for row in self.grid.rows:
            for cell in row.cells:
                by = {s.name: s for s in cell.signals}
                parts = [by[n].value for n in self.relvalue.ADDITIVE]
                if any(p is None for p in parts):
                    self.assertIsNone(cell.richness, f"{row.tenor} {cell.column}")
                    continue
                self.assertAlmostEqual(cell.richness, sum(parts), places=15)
                seen += 1
        self.assertGreater(seen, 10)

    def test_the_at_the_money_carries_no_shape_by_statement(self):
        """Zero because the at-the-money *is* the level, not because two
        near-equal numbers happened to cancel."""
        for row in self.grid.rows:
            for cell in row.cells:
                if cell.column != "atm":
                    continue
                shape = [s for s in cell.signals if s.name == "shape"][0]
                if shape.value is None:
                    continue
                self.assertEqual(shape.value, 0.0, row.tenor)
                self.assertIn("level", shape.message)

    def test_the_at_the_money_shape_is_shown_and_not_averaged_in(self):
        """A statement is not a measurement.

        The at-the-money's shape is zero by construction, and averaging that
        zero into the score pulled every at-the-money cell a fifth of the way
        to the middle -- the same failure the module refuses when a signal is
        *missing*, arriving through the one signal that is present.  It is
        still reported, with its value and its reason, because "zero" and "not
        measured" are different answers.
        """
        checked = 0
        for row in self.grid.rows:
            for cell in row.cells:
                if cell.column != "atm" or cell.score is None:
                    continue
                shape = [s for s in cell.signals if s.name == "shape"][0]
                if shape.value is None:
                    continue
                self.assertEqual(shape.value, 0.0, row.tenor)
                self.assertFalse(shape.used, row.tenor)
                self.assertNotIn("shape", cell.used, row.tenor)
                rest = [s for s in cell.signals if s.used]
                self.assertAlmostEqual(
                    cell.score,
                    sum(s.weight * s.value for s in rest) / sum(s.weight for s in rest),
                    places=12, msg=row.tenor)
                checked += 1
        self.assertGreater(checked, 2)

    def test_the_shape_signal_survives_a_short_realized_lookback(self):
        """The bug: shape zero at the at-the-money and unavailable everywhere else.

        The comparison smile's ``(rho, nu)`` were measured over the realized
        lookback, which needs *more* paired observations than a realized
        volatility needs returns.  Set the lookback to three weeks and the
        level signal went on working while the shape signal was blank at every
        strike of every tenor -- which reads as a signal that does not work
        rather than as a window that is too short.  They come off the history
        window now, for the same reason the scale does.
        """
        grid = self.relvalue.relative_value(
            self.book, "EURUSD", self.history["EURUSD"], horizon_days=7, cut="NY",
            lookback_days=21)
        wings = 0
        for row in grid.rows:
            if row.realized_rho is None:
                continue
            self.assertGreaterEqual(row.dynamics_days, history.DYNAMICS_DAYS, row.tenor)
            for cell in row.cells:
                if cell.column == "atm":
                    continue
                shape = [s for s in cell.signals if s.name == "shape"][0]
                self.assertIsNotNone(shape.value, f"{row.tenor} {cell.column}")
                wings += 1
        self.assertGreater(wings, 10, "no wing was actually scored on its shape")

    def test_the_score_is_the_weighted_mean_of_the_signals_it_used(self):
        """In **volatility points**, and of no others.

        The score was the weighted mean of the *z-scores* until 2026-08-31 and
        the desk asked for the points: how unusual a difference is is a
        statistic about a series, and how much you are being paid is the
        number the mark is moved by.  A missing signal is still renormalised
        away rather than counted as a zero, which would drag every score
        toward the middle.
        """
        scored = 0
        for row in self.grid.rows:
            for cell in row.cells:
                if cell.score is None:
                    continue
                used = [s for s in cell.signals if s.used]
                self.assertEqual(sorted(s.name for s in used), sorted(cell.used))
                total = sum(s.weight for s in used)
                self.assertAlmostEqual(
                    cell.score, sum(s.weight * s.value for s in used) / total, places=12)
                self.assertAlmostEqual(cell.confidence,
                                       total / sum(self.grid.weights.values()), places=12)
                scored += 1
        self.assertGreater(scored, 10)

    def test_the_score_is_on_the_same_footing_as_the_richness_beside_it(self):
        """One unit across the whole card.

        The three additive signals sum to the richness, so a cell scored on
        exactly those three at equal-enough weights lands within their own
        range.  What is pinned is the weaker and more useful thing: the score
        is never outside the span of the values it averaged, which a weighted
        mean cannot be and a weighted mean of z-scores plainly could.
        """
        checked = 0
        for row in self.grid.rows:
            for cell in row.cells:
                if cell.score is None:
                    continue
                vals = [s.value for s in cell.signals if s.used]
                self.assertGreaterEqual(cell.score, min(vals) - 1e-12, row.tenor)
                self.assertLessEqual(cell.score, max(vals) + 1e-12, row.tenor)
                checked += 1
        self.assertGreater(checked, 10)

    def test_a_signal_carries_its_z_wherever_there_is_a_scale_and_says_why_not(self):
        """The z is reported beside the value and scored on by nothing.

        It used to be the score, so it was present exactly where a signal was
        used.  Now it is the reading of *how unusual* the value is, so it
        follows the scale instead: present wherever the history can measure
        one, absent where it cannot, used or not.
        """
        for row in self.grid.rows:
            for cell in row.cells:
                for sig in cell.signals:
                    if sig.value is not None and cell.scale is not None:
                        self.assertIsNotNone(sig.z, f"{row.tenor} {cell.column} {sig.name}")
                        self.assertAlmostEqual(sig.z, sig.value / cell.scale, places=12)
                    else:
                        self.assertIsNone(sig.z)
                    if not sig.used:
                        self.assertTrue(sig.message or cell.message,
                                        f"{row.tenor} {cell.column} {sig.name} is silent")

    def test_the_carry_signal_does_not_carry_the_option_s_own_direction(self):
        """The bug this was written for: the score flipped sign across a row.

        The carry signal used to be built on ``carry_pnl``, the whole
        revaluation of the column's option at the rolled forward.  At a strike
        with any delta on it that is dominated by ``delta * (F2 - F1)`` -- a
        directional number with nothing to say about a volatility -- and by
        put-call parity it is equal and *opposite* for the call columns and
        the put columns.  So one row of the grid was pushed rich on one side
        and cheap on the other, the composite changed sign somewhere between
        them, and the at-the-money column barely showed it because a
        delta-neutral straddle has almost no first-order term.  On the sample
        marks a USDJPY one-year 25 delta put scored ``+13.8`` against the call's
        ``-0.46``, and 30 basis points of forward carry was the whole of it.

        Delta hedged, what is left is the gamma over the move: the same
        positive number whichever side the strike is written as, so the signal
        can no longer split a row by direction.
        """
        h = 7 / 365.2425
        rows = {col.name: {r.tenor: r for r in analytics.carry_table(
            self.book, "USDJPY", horizon_days=7, target=col.target, cut="NY")}
            for col in self.relvalue.COLUMNS}
        split = 0
        for tenor in self.book.data.tenor_points:
            here = [rows[c][tenor] for c in rows if rows[c][tenor].carry_pnl is not None]
            if len(here) < len(self.relvalue.COLUMNS):
                continue
            # The forward's half of the signal, column by column.  Hedged it is
            # a gamma and is one sign right across the row, so it cannot be
            # what separates the wings; unhedged it took the sign of each
            # column's own delta, and did.
            hedged = [r.carry_hedged * (r.t / h) / r.vega for r in here]
            raw = [r.carry_pnl * (r.t / h) / r.vega for r in here]
            self.assertEqual(len({v >= 0 for v in hedged}), 1, msg=tenor)
            if len({v >= 0 for v in raw}) > 1:
                split += 1
                self.assertGreater(max(raw) - min(raw), 5.0 * (max(hedged) - min(hedged)),
                                   msg=tenor)
        self.assertGreater(split, 3, "the old reading split the row at most tenors")

        # And on the grid itself: the whole carry signal, roll included, now
        # spans less than a volatility point across a row.
        grid = self.relvalue.relative_value(self.book, "USDJPY", self.history["USDJPY"],
                                            horizon_days=7, cut="NY")
        checked = 0
        for row in grid.rows:
            values = [c.signal["carry"].value for c in row.cells]
            if any(v is None for v in values):
                continue
            self.assertLess(max(values) - min(values), 0.01, msg=row.tenor)
            checked += 1
        self.assertGreater(checked, 4)

    def test_the_scale_is_measured_over_its_own_window_not_the_realized_lookback(self):
        """The bug this was written for.

        Scoring divided every signal by the standard deviation of the *same*
        window the realized volatility was measured over, which is matched to
        each tenor.  A month of a one-month at-the-money is a handful of
        observations of a smooth series, and an ordinary half point of
        richness came out at thirty standard deviations.  How much a
        volatility moves is a slower measurement and gets its own window.
        """
        short = self.relvalue.relative_value(
            self.book, "EURUSD", self.history["EURUSD"], horizon_days=7, cut="NY",
            history_days=40)
        long = self.relvalue.relative_value(
            self.book, "EURUSD", self.history["EURUSD"], horizon_days=7, cut="NY",
            history_days=500)
        pairs = 0
        for a, b in zip(short.rows, long.rows):
            for ca, cb in zip(a.cells, b.cells):
                if ca.scale is None or cb.scale is None:
                    continue
                self.assertNotAlmostEqual(ca.scale, cb.scale, places=6)
                # Same volatility points either way: only the denominator moved.
                if ca.richness is not None:
                    self.assertAlmostEqual(ca.richness, cb.richness, places=15)
                pairs += 1
        self.assertGreater(pairs, 10)
        self.assertNotEqual(short.history_days, long.history_days)

    def test_a_wing_the_sheet_does_not_quote_borrows_a_scale_and_says_so(self):
        """A substituted denominator is still a substitution.

        A z-score is only as meaningful as what it was divided by, so the cell
        names the series the scale came from rather than quietly using another
        one.
        """
        borrowed = [c for r in self.grid.rows for c in r.cells
                    if c.scale is not None and c.scale_source not in ("", c.column)]
        for cell in borrowed:
            self.assertEqual(cell.scale_source, "atm")
            self.assertIsNone(cell.history_mean, "a borrowed scale is not a history")

    def test_without_a_history_what_can_be_measured_is_still_scored(self):
        """The bug the change to volatility points fixed.

        A score in standard deviations needed a scale, and a scale needed the
        historical sheet -- so a pair the sheet does not quote scored nothing
        at all, in every cell, while its carry had been measured perfectly
        well.  In volatility points the scale is context and not the
        denominator: the ``history`` signal goes (it *is* the history) and so
        does every z, and what is left is scored on its own points.
        """
        grid = self.relvalue.relative_value(self.book, "EURUSD", None, horizon_days=7, cut="NY")
        self.assertIn("history", grid.unavailable)
        self.assertIsNotNone(grid.summary["mean_score"])
        measured = 0
        for row in grid.rows:
            for cell in row.cells:
                self.assertIsNone(cell.scale)
                for sig in cell.signals:
                    self.assertIsNone(sig.z, f"{row.tenor} {cell.column} {sig.name}")
                hist = [s for s in cell.signals if s.name == "history"][0]
                self.assertFalse(hist.used)
                self.assertIsNone(hist.value)
                carry = [s for s in cell.signals if s.name == "carry"][0]
                if carry.value is not None:
                    self.assertTrue(carry.used, f"{row.tenor} {cell.column}")
                    self.assertIn("carry", cell.used)
                    self.assertIsNotNone(cell.score)
                    measured += 1
        self.assertGreater(measured, 10)

    def test_the_row_keeps_the_whole_realized_measurement_not_one_field_of_it(self):
        """The bug this was written for.

        The grid kept ``stats.vol`` and threw away the decomposition
        ``history.realized`` had already measured, so a cell scored rich on
        ``level`` gave no way to say whether the richness was genuine forward
        variance or a level comparison against a thin estimate.
        """
        checked = 0
        for row in self.grid.rows:
            if row.realized is None:
                continue
            stats = history.realized(self.history["EURUSD"], row.window_days,
                                     annualisation="weighted", basis="auto",
                                     basis_tenor=row.tenor)
            self.assertAlmostEqual(row.realized_spot, stats.vol_spot, places=15)
            self.assertAlmostEqual(row.realized_forward, stats.vol_forward, places=15)
            self.assertAlmostEqual(row.points_vol, stats.points_vol, places=15)
            self.assertAlmostEqual(row.points_correlation, stats.points_correlation, places=15)
            self.assertAlmostEqual(row.realized_carry_rate, stats.carry_rate, places=15)
            checked += 1
        self.assertGreater(checked, 3)

    def test_the_vol_support_number_is_the_ratio_and_not_the_level_of_the_points(self):
        """A large carry says nothing on its own about whether the forward is
        more volatile than spot.  What answers that is the ratio of the two
        realized volatilities, so the ratio is what the row carries."""
        rows = [r for r in self.grid.rows if r.forward_vol_ratio is not None]
        self.assertTrue(rows)
        for row in rows:
            self.assertAlmostEqual(row.forward_vol_ratio,
                                   row.realized_forward / row.realized_spot, places=12)
            # The sample's swap points barely move, so the forward and spot
            # measurements are the same variance -- which is the reading, and
            # it is nothing like the -1.8%/yr level of the carry beside it.
            self.assertAlmostEqual(row.forward_vol_ratio, 1.0, places=2)
            self.assertLess(row.realized_carry_rate, -0.01)

    def test_no_row_carries_a_non_finite_number_out_of_the_decomposition(self):
        """``history.realized`` uses nan for 'not measured on this basis' in
        some fields and None in others; either one reaching JSON would take
        the whole grid down in the browser."""
        for row in self.grid.rows:
            for name in ("realized_spot", "realized_forward", "points_vol",
                         "points_correlation", "realized_carry_rate", "forward_vol_ratio"):
                value = getattr(row, name)
                if value is not None:
                    self.assertTrue(math.isfinite(value), f"{row.tenor} {name}")

    def test_the_triangle_is_read_at_the_strike_the_cross_is_marked_at(self):
        """``triangle_table`` compares an at-the-money, a risk reversal and a
        butterfly; a call at that delta is ``atm + fly + rr/2``, so the same
        combination of the differences is the difference at its strike."""
        rows = {r.tenor: r for r in analytics.triangle_table(
            self.book, "EURJPY", cut="NY")}
        grid = self.relvalue.relative_value(self.book, "EURJPY", self.history["EURJPY"],
                                            horizon_days=7, cut="NY")
        checked = 0
        for row in grid.rows:
            tri = rows.get(row.tenor)
            if tri is None:
                continue
            for cell in row.cells:
                sig = [s for s in cell.signals if s.name == "triangle"][0]
                if sig.value is None:
                    continue
                d = tri.difference
                if cell.column == "atm":
                    want = d["atm"]
                else:
                    tag = f"{int(round(cell.delta * 100))}"
                    want = d["atm"] + d[f"fly{tag}"] + (0.5 if cell.is_call else -0.5) * d[f"rr{tag}"]
                self.assertAlmostEqual(sig.value, want, places=15,
                                       msg=f"{row.tenor} {cell.column}")
                checked += 1
        self.assertGreater(checked, 10)

    def test_a_difference_inside_the_triangle_noise_floor_is_shown_but_not_scored(self):
        """That section's own rule: a difference smaller than what the
        machinery gets wrong on the legs alone is not a difference.  Scoring
        it anyway would put the reconstruction error into the answer."""
        real = analytics.triangle_table(self.book, "EURJPY", cut="NY")
        loud = [analytics.TriangleRow(
            tenor=r.tenor, t=r.t, rho=r.rho, coefficients=r.coefficients, marked=r.marked,
            triangle=r.triangle, difference=r.difference,
            noise={k: 10.0 for k in r.difference}, variance_triangle_atm=r.variance_triangle_atm,
            smile_convexity=r.smile_convexity, leg_atm=r.leg_atm,
            implied_correlation=r.implied_correlation) for r in real]
        old = self.relvalue.triangle_table
        self.relvalue.triangle_table = lambda *a, **k: loud
        try:
            grid = self.relvalue.relative_value(self.book, "EURJPY", self.history["EURJPY"],
                                                horizon_days=7, cut="NY")
        finally:
            self.relvalue.triangle_table = old
        seen = 0
        for row in grid.rows:
            for cell in row.cells:
                sig = [s for s in cell.signals if s.name == "triangle"][0]
                if sig.value is None:
                    continue
                self.assertFalse(sig.used, f"{row.tenor} {cell.column}")
                self.assertIn("noise floor", sig.message)
                self.assertNotIn("triangle", cell.used)
                seen += 1
        self.assertGreater(seen, 10)

    def test_a_pair_that_is_not_a_cross_has_no_triangle_and_is_not_charged_for_it(self):
        for row in self.grid.rows:
            for cell in row.cells:
                sig = [s for s in cell.signals if s.name == "triangle"][0]
                self.assertIsNone(sig.value)
                self.assertIn("not a cross", sig.message)
        self.assertIn("not a cross", self.grid.unavailable["triangle"])

    def test_a_tenor_that_cannot_be_rolled_keeps_its_row_and_carries_the_reason(self):
        """Dropping it makes a short grid look like a complete one."""
        grid = self.relvalue.relative_value(self.book, "EURUSD", self.history["EURUSD"],
                                            horizon_days=90, cut="NY")
        self.assertEqual([r.tenor for r in grid.rows], list(self.book.data.tenor_points))
        blocked = [(r, c) for r in grid.rows for c in r.cells
                   if [s for s in c.signals if s.name == "carry"][0].value is None]
        self.assertTrue(blocked)
        for _, cell in blocked:
            carry = [s for s in cell.signals if s.name == "carry"][0]
            self.assertTrue(carry.message)

    def test_the_carry_horizon_and_the_dominance_threshold_are_one_statement(self):
        """``T = 0.64 * sigma**2 / c**2`` is exactly where ``z`` reaches 0.8.

        The factor is the threshold squared, so it is derived from it rather
        than written down twice: two constants that could drift apart would
        put the annotation and the horizon on opposite sides of the line.
        """
        rv = self.relvalue
        self.assertAlmostEqual(rv.CARRY_HORIZON_FACTOR, rv.CARRY_DOMINANT_Z ** 2, places=15)
        sigma, carry, spot = 0.10, 0.06, 1.25
        at_a_year = rv._regime(spot, spot * math.exp(carry * 1.0), 1.0, sigma, None)
        horizon = at_a_year["carry_horizon_days"] / 365.2425
        at_the_horizon = rv._regime(spot, spot * math.exp(carry * horizon), horizon,
                                    sigma, None)
        self.assertAlmostEqual(at_the_horizon["regime_z"], rv.CARRY_DOMINANT_Z, places=12)
        # The claim is that the horizon is where the two cross, so it is
        # tested either side of it rather than exactly on it: the boundary
        # itself lands within a rounding error of the threshold.
        for scale, dominant in ((1.02, True), (0.98, False)):
            t = horizon * scale
            side = rv._regime(spot, spot * math.exp(carry * t), t, sigma, None)
            self.assertEqual(side["carry_dominant"], dominant, f"at {scale:g} of the horizon")
        # And the sign of the carry does not decide the regime: a discount and
        # a premium of the same size are the same distance travelled.
        mirrored = rv._regime(spot, spot * math.exp(-carry * horizon), horizon, sigma, None)
        self.assertAlmostEqual(mirrored["regime_z"], at_the_horizon["regime_z"], places=12)

    def test_a_carry_to_volatility_ratio_alone_would_flag_a_free_float(self):
        """The bug in the obvious version of this test.

        Read as "carry over volatility" alone, USDJPY on a five point rate
        differential and ten volatility points scores 0.53 -- right beside
        USDCNH's 0.50 -- and USDJPY is not managed in any sense.  What
        separates them is the second condition: a managed float's realized
        volatility is *low in absolute terms*, which is the suppressed
        diffusion itself rather than a consequence of it.
        """
        rv = self.relvalue

        class Row:
            def __init__(self, ratio, vol):
                self.carry_to_vol, self.realized = ratio, vol

        def verdict(ratio, vol):
            return rv.suppressed_diffusion([Row(ratio, vol)] * 3)

        # The ratio alone does not separate these two.
        self.assertGreater(verdict(0.53, 0.095)["carry_to_vol"],
                           verdict(0.50, 0.050)["carry_to_vol"])
        self.assertFalse(verdict(0.53, 0.095)["managed"], "USDJPY is not a managed float")
        self.assertTrue(verdict(0.50, 0.050)["managed"], "USDCNH has the shape")
        # A high-carry, high-volatility pair is deliberately outside it: its
        # diffusion is not suppressed, it is merely expensive.
        self.assertFalse(verdict(1.40, 0.250)["managed"])
        # And an ordinary free float fails the first condition instead.
        self.assertFalse(verdict(0.25, 0.060)["managed"])
        self.assertFalse(verdict(None, None)["managed"])

    def test_a_carry_dominated_tenor_warns_and_marks_the_signal_it_dominates(self):
        """Said where the signal is read, not only in the row's warnings."""
        import tempfile
        from volkit.feed import MarketFeed
        rv = self.relvalue
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "steep.csv"
            path.write_text(
                "# a deliberately steep curve: EURUSD at a 7%/yr discount\n"
                "EURUSD,SPOT,1.08000\n"
                "EURUSD,1M,-60.0\nEURUSD,3M,-180.0\n"
                "EURUSD,6M,-360.0\nEURUSD,1Y,-720.0\n", encoding="utf-8")
            old = self.book.feed
            self.book.feed = MarketFeed.load(path)
            self.addCleanup(setattr, self.book, "feed", old)
            grid = rv.relative_value(self.book, "EURUSD", self.history["EURUSD"],
                                     horizon_days=7, cut="NY")
        hot = [r for r in grid.rows if r.carry_dominant]
        self.assertTrue(hot, "a 7%/yr carry against a 6% mark should dominate by one year")
        for row in hot:
            self.assertGreaterEqual(row.regime_z, rv.CARRY_DOMINANT_Z)
            self.assertTrue(any("carry trade" in w for w in row.warnings), row.tenor)
            for cell in row.cells:
                carry = [s for s in cell.signals if s.name == "carry"][0]
                if carry.value is not None:
                    self.assertIn("carry dominated", carry.message)
        # The weight is not tapered on the strength of it: the score is still
        # the weighted mean of what it used, and the weights are the desk's.
        for row in hot:
            for cell in row.cells:
                used = [s for s in cell.signals if s.used]
                if not used:
                    continue
                total = sum(s.weight for s in used)
                self.assertAlmostEqual(
                    cell.score, sum(s.weight * s.value for s in used) / total, places=12)
                for sig in used:
                    self.assertEqual(sig.weight, grid.weights[sig.name])

    def test_the_forward_comes_from_the_feed_before_the_carry_table(self):
        """A tenor the horizon cannot roll still has a forward and a regime.

        Reading it off the carry table alone left every tenor shorter than the
        horizon with no forward, and therefore no absolute strikes, on a pair
        whose forward the feed was quoting perfectly well.
        """
        grid = self.relvalue.relative_value(self.book, "EURUSD", self.history["EURUSD"],
                                            horizon_days=90, cut="NY")
        unrollable = [r for r in grid.rows
                      if all([s for s in c.signals if s.name == "carry"][0].value is None
                             for c in r.cells) and r.cells]
        self.assertTrue(unrollable, "no tenor was blocked by a 90-day horizon")
        for row in unrollable:
            self.assertIsNotNone(row.forward, row.tenor)
            self.assertIsNotNone(row.spot, row.tenor)
            self.assertIsNotNone(row.regime_z, row.tenor)
            for cell in row.cells:
                self.assertIsNotNone(cell.strike, f"{row.tenor} {cell.column}")

    def test_a_shared_signal_is_one_number_for_the_row_and_is_marked_as_one(self):
        """``level`` is a statement about the level, and a level is one number
        per expiry.  Printed in five cells with nothing to tie them together,
        one at-the-money mispricing reads as five confirmations."""
        rv = self.relvalue
        spread = 0
        for row in self.grid.rows:
            if not row.cells:
                continue
            for name in rv.SHARED:
                values = [[s for s in c.signals if s.name == name][0].value
                          for c in row.cells]
                self.assertEqual(len(set(values)), 1, f"{row.tenor} {name} is not shared")
                zs = {[s for s in c.signals if s.name == name][0].z for c in row.cells}
                if len(zs) > 1:
                    spread += 1        # one number, but each cell's own scale
            for cell in row.cells:
                for sig in cell.signals:
                    self.assertEqual(sig.shared, sig.name in rv.SHARED, sig.name)
        self.assertGreater(spread, 2, "a shared value should still take each cell's scale")

    def test_the_state_response_names_the_shared_signal_for_the_page(self):
        """The page marks it without knowing which signal it happens to be."""
        from volkit.webapp import BookService
        from volkit.relvalue import SHARED
        state = BookService(str(WORKBOOK), ASOF).state()
        signals = state["analysis"]["signals"]
        self.assertEqual({s["key"] for s in signals if s["shared"]}, set(SHARED))
        self.assertTrue(all("weight" in s for s in signals))

    def test_a_weight_that_is_not_a_signal_is_refused_not_ignored(self):
        with self.assertRaises(self.relvalue.RelativeValueError) as ctx:
            self.relvalue.resolve_weights({"gut_feel": 1.0})
        self.assertIn("gut_feel", str(ctx.exception))
        with self.assertRaises(self.relvalue.RelativeValueError):
            self.relvalue.resolve_weights({"level": "quite a lot"})
        with self.assertRaises(self.relvalue.RelativeValueError):
            self.relvalue.resolve_weights({"level": -1})
        with self.assertRaises(self.relvalue.RelativeValueError):
            self.relvalue.resolve_weights({k: 0 for k in self.relvalue.WEIGHTS})

    def test_a_reweighting_moves_the_score_and_nothing_else(self):
        heavy = self.relvalue.relative_value(
            self.book, "EURUSD", self.history["EURUSD"], horizon_days=7, cut="NY",
            weights={"carry": 5.0})
        moved = 0
        for a, b in zip(self.grid.rows, heavy.rows):
            for ca, cb in zip(a.cells, b.cells):
                self.assertAlmostEqual(ca.implied, cb.implied, places=15)
                if ca.richness is not None:
                    self.assertAlmostEqual(ca.richness, cb.richness, places=15)
                if ca.score is not None and "carry" in ca.used:
                    if abs(ca.score - cb.score) > 1e-9:
                        moved += 1
        self.assertGreater(moved, 5, "reweighting the carry changed no score")

    def test_the_panel_is_the_only_reader_of_the_request(self):
        panel = self.relvalue.panel_from_request({
            "pair": "EURUSD", "cut": "NY", "method": "SVI", "horizon_days": "7",
            "lookback_days": "match", "history_days": "250", "annualisation": "weighted",
            "realized_basis": "auto", "triangle": "0", "weights": {"carry": "0.4"}})
        self.assertEqual(panel.pair, "EURUSD")
        self.assertIsNone(panel.lookback_days)
        self.assertFalse(panel.with_triangle)
        self.assertEqual(panel.weights["carry"], 0.4)
        for bad in ({"pair": ""}, {"pair": "EURUSD", "history_days": "soon"},
                    {"pair": "EURUSD", "lookback_days": "-3"},
                    {"pair": "EURUSD", "weights": "level=1"}):
            with self.assertRaises(self.relvalue.RelativeValueError):
                self.relvalue.panel_from_request(bad)

    def test_the_command_line_and_the_screen_run_the_same_panel(self):
        """The CLI builds the panel the browser posts, so a cell quoted off
        the screen can be reproduced in a batch job."""
        from volkit.webapp import BookService
        request = {"pair": "EURUSD", "cut": "NY", "method": "SVI", "horizon_days": "7",
                   "lookback_days": "match", "history_days": "250", "triangle": "1"}
        service = BookService(str(WORKBOOK), ASOF, feed_path=str(FEED),
                              history_path=str(HISTORY))
        service.load_history({"path": str(HISTORY)})
        served = service.relative_value(request)
        direct = self.relvalue.panel_from_request(request).run(service.book, service.history)
        self.assertEqual(served["summary"]["headline"], direct.summary["headline"])
        self.assertEqual(served["pair"], "EURUSD")

    def test_the_payload_carries_no_number_a_browser_cannot_parse(self):
        """Python's json writes NaN and JSON.parse refuses it, so one
        unscored cell would take the whole grid down in the browser."""
        import json as _json
        from volkit.webapp import BookService, _finite
        service = BookService(str(WORKBOOK), ASOF, feed_path=str(FEED),
                              history_path=str(HISTORY))
        service.load_history({"path": str(HISTORY)})
        text = _json.dumps(_finite(service.relative_value(
            {"pair": "EURJPY", "cut": "NY", "horizon_days": "7"})), default=str)
        self.assertNotIn("NaN", text)
        self.assertNotIn("Infinity", text)
        self.assertEqual(_json.loads(text)["pair"], "EURJPY")

    def test_the_route_belongs_to_the_analysis_screen(self):
        from volkit import screens
        owner = {r: s.name for s in screens.SCREENS for r in s.routes}
        self.assertEqual(owner["/api/relvalue"], "analysis")


class TestTimestampReading(unittest.TestCase):
    """One place reads a timestamp, and it reads what the tool writes.

    The listed panel, the events route and the session loader had each
    patched ISO 8601 into shape for themselves, and had done it differently:
    only one of them understood a trailing Z or an offset, so a string that
    parsed on one screen failed on the next.  Worse, that one *dropped* the
    offset and then stamped the result UTC, which reads 19:00+09:00 as 19:00Z.
    """

    def test_iso_8601_is_read_the_way_the_tool_writes_it(self):
        from volkit.timeutil import UTC, parse_datetime
        want = datetime(2024, 2, 28, 12, 0, tzinfo=UTC)
        for text in ("2024-02-28 12:00", "2024-02-28T12:00", "2024-02-28T12:00:00",
                     "2024-02-28T12:00:00Z", "2024-02-28T12:00:00+00:00"):
            with self.subTest(text):
                self.assertEqual(parse_datetime(text), want)
        # The tabular formats are tried first and are untouched.
        self.assertEqual(parse_datetime("2/28/2024 12:00"), want)
        self.assertEqual(parse_datetime("28-Feb-24"),
                         datetime(2024, 2, 28, tzinfo=UTC))
        # And what it prints, it reads.
        self.assertEqual(parse_datetime(want.isoformat()), want)

    def test_an_offset_is_converted_and_not_thrown_away(self):
        from volkit.timeutil import UTC, parse_datetime
        self.assertEqual(parse_datetime("2026-09-11T19:00+09:00"),
                         datetime(2026, 9, 11, 10, 0, tzinfo=UTC))

    def test_nonsense_still_says_so(self):
        from volkit.timeutil import parse_datetime
        with self.assertRaises(ValueError) as caught:
            parse_datetime("next tuesday")
        self.assertIn("next tuesday", str(caught.exception))

    def test_the_listed_panel_takes_its_expiry_straight_from_the_browser(self):
        from volkit.listed import _normalise_expiry
        self.assertEqual(_normalise_expiry(" 2026-09-11T19:00Z "), "2026-09-11T19:00Z")
        with self.assertRaises(ValueError):
            _normalise_expiry("   ")


class TestSmileStrikeScale(unittest.TestCase):
    """The strike axis reads in levels when the feed can supply one.

    The smile is fitted in moneyness whatever happens -- that is the space the
    surface works in -- so this is a *scale* on the way out and never a change
    to a number.  Without a feed there is no honest level to name and it stays
    in K/F rather than inventing one.
    """

    def test_the_smile_carries_the_feed_level_and_no_vol_moves(self):
        from volkit.webapp import BookService
        with_feed = BookService(str(WORKBOOK), ASOF, feed_path=str(FEED))
        without = BookService(str(WORKBOOK), ASOF)
        q = {"pair": "EURUSD", "expiry": "2024-05-28", "method": "SVI", "cut": "TK"}
        a, b = with_feed.smile(dict(q)), without.smile(dict(q))

        self.assertTrue(a["feed"])
        self.assertGreater(a["forward"], 0.0)
        self.assertGreater(a["spot"], 0.0)
        self.assertFalse(b["feed"])
        self.assertIsNone(b["forward"])
        self.assertIsNone(b["spot"])

        # The strikes come back in moneyness either way: the page multiplies.
        self.assertEqual([r["k"] for r in a["curve"]], [r["k"] for r in b["curve"]])
        self.assertEqual([r["v"] for r in a["curve"]], [r["v"] for r in b["curve"]])
        self.assertEqual(a["atm"], b["atm"])

    def test_the_level_is_the_one_the_band_model_would_place_against(self):
        """One lookup for both, so a strike a chart names and a band edge the
        model places can never come from different forwards."""
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF, feed_path=str(FEED))
        payload = service.smile({"pair": "USDJPY", "expiry": "2024-05-28",
                                 "method": "SVI", "cut": "TK"})
        book = service.book
        expiry = date(2024, 5, 28)
        # One lookup, and it is the settlement-date one: the chart's axis
        # scale, the band model's placement and `market_level_for` are the
        # same number, and the payload says which date it is a price to.
        level = book.market_level_for("USDJPY", expiry)
        self.assertEqual(payload["forward"], level["forward"])
        self.assertEqual(payload["settle"], level["settle"])
        self.assertEqual(payload["settle"],
                         book.settlement_date("USDJPY", expiry).isoformat())
        self.assertEqual(payload["forward"],
                         book.forward_at("USDJPY", payload["t"], expiry=expiry))
        # and it is *not* the plain time reading, which drops the spot lag
        self.assertNotEqual(payload["forward"],
                            book.market_level("USDJPY", payload["t"])["forward"])
        # A pair the feed does not cover says so rather than guessing a level.
        self.assertFalse(service.book.market_level("XXXYYY", 0.25)["feed"])
        self.assertIsNone(service.book.forward_at("XXXYYY", 0.25))


class TestCrossLevelsFromTheLegs(unittest.TestCase):
    """A cross the feed quotes only through its legs still has a level.

    The bug: the feed publishes EURUSD and USDJPY and therefore publishes
    EURJPY, but every level lookup asked the feed for the pair by name and
    refused.  On the market-maker screen that made a loaded feed invisible --
    a quote written against an absolute strike came back "there is no forward
    feed for EURJPY" while the pricing screen was quoting both its legs off
    the same file.
    """

    def book(self, pairs=("EURJPY", "EURGBP", "GBPNZD")):
        from volkit.feed import MarketFeed
        book = Book.from_excel(WORKBOOK, ASOF).load_all(list(pairs))
        book.feed = MarketFeed.load(FEED)
        return book

    def test_a_cross_is_the_product_of_its_legs(self):
        book = self.book()
        t = 0.25
        level = book.market_level("EURJPY", t)
        self.assertTrue(level["feed"])
        self.assertTrue(level["derived"])
        self.assertEqual(level["via"], "EURUSD and USDJPY")
        legs = (book.feed.quote("EURUSD", t), book.feed.quote("USDJPY", t))
        self.assertAlmostEqual(level["forward"], legs[0]["forward"] * legs[1]["forward"], places=12)
        self.assertAlmostEqual(level["spot"], legs[0]["spot"] * legs[1]["spot"], places=12)
        # The points are the cross's own, in the cross's own pips, and never
        # the legs' points added together.
        self.assertAlmostEqual(level["spot"] + level["points"] / level["pip"],
                               level["forward"], places=12)
        self.assertEqual(level["pip"], 100.0)

    def test_a_cross_of_two_same_side_legs_divides(self):
        """EURGBP is EURUSD / GBPUSD, not EURUSD * GBPUSD."""
        book = self.book()
        t = 0.25
        level = book.market_level("EURGBP", t)
        a, b = book.feed.quote("EURUSD", t), book.feed.quote("GBPUSD", t)
        self.assertAlmostEqual(level["forward"], a["forward"] / b["forward"], places=12)
        self.assertEqual(level["via"], "EURUSD and GBPUSD")

    def test_a_pair_the_feed_quotes_itself_is_not_derived(self):
        book = self.book(["EURUSD"])
        level = book.market_level("EURUSD", 0.25)
        self.assertTrue(level["feed"])
        self.assertFalse(level["derived"])
        self.assertEqual(level["via"], "")

    def test_a_leg_the_feed_does_not_carry_is_still_a_refusal(self):
        """Half a triangle is not a level, and is refused rather than guessed."""
        book = self.book()
        self.assertNotIn("NZDUSD", book.feed.pairs)
        level = book.market_level("GBPNZD", 0.25)
        self.assertFalse(level["feed"])
        self.assertIsNone(level["forward"])
        self.assertIsNone(book.forward_at("GBPNZD", 0.25))

    def test_the_market_maker_prices_an_absolute_strike_on_a_cross(self):
        """The bug, on the screen it was found on."""
        from volkit import marketmaker as mm
        book = self.book(["EURJPY"])
        panel = mm.panel_from_request({
            "pair": "EURJPY", "cut": "NY", "target_source": "quotes",
            "fit_curve": False, "tune_wings": False,
            "text": "1M ATM 8.2/8.6\n1M, 162.00, 8.4/8.8\n3M, 162.00, 8.4/8.8\n"})
        rows = panel.run(book)["market"]["rows"]
        self.assertEqual(len(rows), 3)
        for row in rows:
            self.assertIsNotNone(row["model_before"], row["raw"])
            self.assertEqual([w for w in row["warnings"] if "forward feed" in w], [])

    def test_the_derivation_is_said_once_and_not_once_a_tenor(self):
        from volkit import marketmaker as mm
        book = self.book(["EURJPY"])
        panel = mm.panel_from_request({
            "pair": "EURJPY", "cut": "NY", "target_source": "quotes",
            "fit_curve": False, "tune_wings": False,
            "text": "1M, 162.00, 8.4/8.8\n2M, 162.00, 8.4/8.8\n3M, 162.00, 8.4/8.8\n"})
        notes = [n for n in panel.run(book)["market"]["notes"] if "triangle" in n]
        self.assertEqual(len(notes), 1, notes)
        self.assertIn("EURUSD and USDJPY", notes[0])

    def test_every_screen_reads_the_one_lookup(self):
        """``analytics._forward_at`` is the same number as ``market_level``."""
        from volkit.analytics import _forward_at
        book = self.book(["EURJPY"])
        fwd, real, note = _forward_at(book, "EURJPY", 0.25)
        self.assertTrue(real)
        self.assertEqual(fwd, book.market_level("EURJPY", 0.25)["forward"])
        self.assertIn("triangle", note)


class TestCurveComparison(unittest.TestCase):
    """Several curves side by side, and the same curve on other dates."""

    def book(self, pairs=("EURUSD",)):
        return Book.from_excel(WORKBOOK, ASOF).load_all(list(pairs))

    def history(self, book):
        return history.load_history(HISTORY, book.pairs)

    def test_the_surface_and_the_quotes_differ_by_the_fit_residual(self):
        """The same comparison the marking screen's implied-vs-quoted table
        makes, extended across curves: small, and not zero."""
        from volkit import curves
        book = self.book()
        panel = curves.ComparePanel(curves=(curves.CurveRequest("surface", "EURUSD"),
                                            curves.CurveRequest("marks", "EURUSD")))
        r = panel.run(book)
        self.assertEqual(r["base"], 0)
        self.assertTrue(r["curves"][0]["is_base"])
        diffs = [p["diffs"]["atm"] for p in r["curves"][1]["points"]
                 if p["diffs"]["atm"] is not None]
        self.assertTrue(diffs)
        self.assertLess(max(abs(d) for d in diffs), 0.01)
        self.assertTrue(any(abs(d) > 1e-9 for d in diffs))

    def test_the_tenor_axis_is_the_union_and_a_gap_is_a_gap(self):
        """A curve that does not quote a tenor leaves a blank, which is not the
        same thing as the row being missing."""
        from volkit import curves
        book = self.book()
        r = curves.ComparePanel(curves=(
            curves.CurveRequest("marks", "EURUSD"),
            curves.CurveRequest("history", "EURUSD", "latest"))).run(book, self.history(book))
        self.assertIn("2Y", r["tenors"])            # in the workbook, not in the sheet
        self.assertIn("1M", r["tenors"])
        hist = r["curves"][1]
        self.assertIsNone(hist_point(hist, "2Y"))
        self.assertIsNotNone(hist_point(hist, "1M"))
        self.assertTrue(any("not every curve quotes every tenor" in n for n in r["notes"]))

    def test_a_date_snaps_back_to_the_last_row_on_or_before_it(self):
        """Forward-snapping would compare a Friday mark against Monday's."""
        from volkit import curves
        book = self.book()
        hist = self.history(book)["EURUSD"]
        when, note = curves.resolve_history_date(hist, "2024-02-25")   # a Sunday
        self.assertLessEqual(when, date(2024, 2, 25))
        self.assertIn("nearest row on or before", note)
        self.assertEqual(curves.resolve_history_date(hist, "")[0], hist.dates[-1])
        back, _ = curves.resolve_history_date(hist, "-30d")
        self.assertLess(back, hist.dates[-1])
        with self.assertRaises(curves.CurveError):
            curves.resolve_history_date(hist, "the other day")

    def test_a_curve_that_cannot_be_built_keeps_its_place(self):
        """Dropping it makes a short comparison look complete."""
        from volkit import curves
        book = self.book()
        r = curves.ComparePanel(curves=(
            curves.CurveRequest("surface", "EURUSD"),
            curves.CurveRequest("history", "USDHKD"))).run(book, self.history(book))
        self.assertEqual(len(r["curves"]), 2)
        self.assertFalse(r["curves"][1]["ok"])
        self.assertIn("no sheet for USDHKD", r["curves"][1]["message"])

    def test_the_base_falls_back_when_the_one_chosen_could_not_be_built(self):
        from volkit import curves
        book = self.book()
        r = curves.ComparePanel(curves=(curves.CurveRequest("history", "USDHKD"),
                                        curves.CurveRequest("surface", "EURUSD")),
                                base=0).run(book, self.history(book))
        self.assertEqual(r["base"], 1)
        self.assertTrue(r["curves"][1]["is_base"])

    def test_a_pasted_curve_is_read_in_points_as_written(self):
        """The level is not evidence of the unit (§4): a managed pair's whole
        curve sits below 1.0, and reading that as decimals put it on the
        monitor at a hundred times its mark."""
        from volkit import curves
        c = curves.parse_pasted_curve("1M 8.20 -0.35 0.22\n3M 8.45")
        self.assertTrue(c.ok)
        self.assertAlmostEqual(c.at("1M").values["atm"], 0.0820)
        self.assertAlmostEqual(c.at("1M").values["rr25"], -0.0035)
        self.assertIsNone(c.at("3M").values["rr25"])
        low = curves.parse_pasted_curve("1M 0.35 -0.02")
        self.assertTrue(low.ok)
        self.assertAlmostEqual(low.at("1M").values["atm"], 0.0035)
        self.assertIn("volatility points", low.source)

    def test_a_pasted_curve_that_straddles_one_is_read_and_not_refused(self):
        from volkit import curves
        c = curves.parse_pasted_curve("1M 8.20\n3M 0.35")
        self.assertTrue(c.ok)
        self.assertAlmostEqual(c.at("1M").values["atm"], 0.0820)
        self.assertAlmostEqual(c.at("3M").values["atm"], 0.0035)
        bad = curves.parse_pasted_curve("1M\n3M 8.4")
        self.assertFalse(bad.ok)
        self.assertIn("line 1", bad.message)

    def test_a_command_line_spec_reads_the_same_panel(self):
        from volkit import curves
        self.assertEqual(curves.parse_spec("history:-30d:eurusd", "USDJPY"),
                         curves.CurveRequest("history", "EURUSD", "-30d"))
        self.assertEqual(curves.parse_spec("marks", "USDJPY").pair, "USDJPY")
        # A date on a source that has none would silently be ignored.
        with self.assertRaises(curves.CurveError):
            curves.parse_spec("surface:2024-01-15", "EURUSD")
        with self.assertRaises(curves.CurveError):
            curves.parse_spec("yesterday", "EURUSD")

    def test_the_endpoint_is_a_pure_function_of_its_request(self):
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF, history_path=str(HISTORY))
        payload = {"curves": [{"kind": "surface", "pair": "EURUSD"},
                              {"kind": "history", "pair": "EURUSD", "date": "-60d"}],
                   "cut": "NY", "method": "SVI", "field": "rr25", "base": 0}
        first = service.compare_curves(payload)
        self.assertEqual(first, service.compare_curves(payload))
        self.assertEqual(first["field"], "rr25")


def hist_point(curve, tenor):
    return next((p for p in curve["points"] if p["tenor"].upper() == tenor.upper()), None)


class TestMarketMakerApi(unittest.TestCase):
    """The endpoints, and the one piece of server state on the whole tool."""

    RUN = ("1M ATM 6.05/6.35 in 100mm vega\n"
           "3M ATM 6.25/6.55\n"
           "6M ATM 6.60/6.90\n"
           "1Y atm 7.00/7.30\n"
           "1M 25d rr -0.30/-0.10\n")

    def service(self, tmp):
        from volkit.webapp import BookService
        return BookService(str(WORKBOOK), ASOF, bank_path=str(Path(tmp) / "bank.json"))

    def test_the_payload_carries_no_number_a_browser_cannot_parse(self):
        """Python's json writes NaN, which JSON.parse refuses."""
        import json as _json, tempfile
        from volkit.webapp import _finite
        with tempfile.TemporaryDirectory() as tmp:
            payload = self.service(tmp).mm_fit(
                {"pair": "EURUSD", "text": self.RUN, "target_source": "quotes",
                 "fallback_spread": "0.3", "tune_wings": False})
        text = _json.dumps(_finite(payload), default=str)
        self.assertNotIn("NaN", text)
        self.assertNotIn("Infinity", text)
        self.assertEqual(_json.loads(text)["pair"], "EURUSD")

    def test_the_bank_is_the_only_state_the_server_keeps(self):
        """The panel is posted whole every time, like the listed screen; the
        bank is a file the desk owns, so it lives on the server."""
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            service = self.service(tmp)
            saved = service.mm_save_bank({
                "pair": "EURUSD",
                "rules": [{"kind": "spread", "value": "0.28", "instrument": "atm"},
                          {"kind": "note", "text": "wider into the ECB"}]})
            self.assertTrue(saved["ok"], saved.get("problems"))
            self.assertTrue(Path(saved["written"]).exists())
            # The bank is the quote route's, not the fit's: a width is a
            # property of what we show, and the fit shows nothing.
            out = service.mm_quote({"pair": "EURUSD", "request_text": "1M ATM\n"})
            atm = next(r for r in out["sheet"]["rows"] if r["instrument"] == "atm")
            self.assertAlmostEqual(atm["width"], 0.28)
            # A note is advice, kept apart from the reader's own notes so it
            # cannot get buried: it exists to be read.
            self.assertIn("wider into the ECB", atm["advice"])
            # And it survives a fresh service reading the same file.
            self.assertEqual(len(self.service(tmp).bank.for_pair("EURUSD").rules), 2)

    def test_a_bad_rule_set_is_rejected_without_touching_the_file(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            service = self.service(tmp)
            service.mm_save_bank({"pair": "EURUSD",
                                  "rules": [{"kind": "spread", "value": "0.28"}]})
            bad = service.mm_save_bank({"pair": "EURUSD",
                                        "rules": [{"kind": "spread", "value": "-3"}]})
            self.assertFalse(bad["ok"])
            self.assertTrue(bad["problems"])
            self.assertEqual(len(self.service(tmp).bank.for_pair("EURUSD").rules), 1)

    def test_learning_proposes_and_does_not_save(self):
        """A paste that happens to hold one wide quote must not be able to
        rewrite the desk's ladder without somebody looking at it."""
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            service = self.service(tmp)
            got = service.mm_learn({"pair": "EURUSD", "text": self.RUN,
                                    "target_source": "quotes"})
            self.assertTrue(got["rules"])
            self.assertTrue(all(r["kind"] == "spread" for r in got["rules"]))
            self.assertEqual(service.bank.for_pair("EURUSD").rules, [])

    def test_the_state_endpoint_tells_the_browser_what_it_may_choose_from(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            state = self.service(tmp).state()["marketmaker"]
        for key in ("target_sources", "backbone_knobs", "cross_knobs", "smile_params",
                    "fly_conventions", "vol_units", "rule_kinds", "rule_instruments",
                    "size_bases", "bank"):
            self.assertIn(key, state)
        self.assertIn("overwrites", state["target_sources"])
        self.assertIn("rho25", state["smile_params"])


class TestWebAssets(unittest.TestCase):
    def test_front_end_javascript_parses(self):
        """Guards against shipping a page that dies on load."""
        try:
            import esprima
        except ImportError:
            self.skipTest("esprima not installed")
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        # esprima tops out at ES2017; downlevel the two newer operators used.
        probe = _re.sub(r"\?\.", ".", js.replace("??", " || "))
        esprima.parseScript(probe)

    def test_the_settlement_date_is_an_input_row_and_is_not_repeated_below(self):
        """It moved from Results to Inputs when it turned out to be a box.

        The rule it moved under is the one it used to be an example of: a
        Results row may not repeat an input box.  Left in both places it
        would be one date typed in one place and shown in another, which is
        the disagreement that rule exists to prevent.
        """
        html = _source("volkit", "web", "index.html")
        ins = html.split("const IN=[")[1].split("];")[0]
        outs = html.split("const OUT=[")[1].split("];")[0]
        self.assertIn("['settle','Settlement','text']", ins)
        # Under Expiry, because that is the date it is built from.
        self.assertLess(ins.index("'expiry'"), ins.index("'settle'"))
        self.assertNotIn("['settle',", outs)
        # And the one market fact that is *not* a box stays where it was.
        self.assertIn("['market_source','Market'", outs)

    def test_the_screen_tells_the_server_where_each_of_its_boxes_came_from(self):
        """The page owns this panel, so provenance is the page's to state.

        It fills spot, the swap and the settlement date and then posts what
        is in the boxes, so nothing downstream can tell a level it filled
        from one somebody typed.  The three source flags are the page saying
        which is which; without them the Market row read `typed` for every
        leg, including after `Refresh spot` had just put them all back on the
        feed.
        """
        html = _source("volkit", "web", "index.html")
        # Priced legs are posted whole, so the flags travel with them.
        self.assertIn("post('/api/price',{legs:LEGS})", html)
        # The two routes that post a cut-down leg have to name them.
        for fn in ("async function resolveLegs(force){", "async function refreshFeed(){"):
            body = html.split(fn)[1].split("\n}")[0]
            for field in ("pair:L.pair", "expiry:L.expiry",
                          "settle:L.settle", "settlesrc:L.settlesrc"):
                self.assertIn(field, body, fn)
        # And the server reads all three names.
        py = _source("volkit", "webapp.py")
        for name in ('row.get("settlesrc")', 'row.get("spotsrc")', 'row.get("fwdsrc")'):
            self.assertIn(name, py)

    def test_a_table_body_given_as_one_string_is_not_joined(self):
        """`rows.join is not a function`, on the band card's Fit from the wings.

        `tblHtml` takes the list of <tr> strings and joins it, and every caller
        passed a list -- except the band fit, which pre-joined both of its
        tables and handed the result over as one string. The whole fit had
        already been computed and came back to the desk as an error message
        that read as though the calibration itself had failed. The guard is in
        `tblHtml` rather than at the two call sites, because a table body is
        naturally either shape and there is nothing to catch a third caller.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        body = html.split("function tblHtml(head,rows,cls){")[1].split("\n}")[0]
        self.assertIn("Array.isArray(rows)", body)
        # And the row argument may not be joined on the way in.
        fit = html.split("async function fitBand(){")[1].split("\n}")[0]
        for call in _re.findall(r"tblHtml\(([^;]*?)\)\+'</div>'", fit, flags=_re.S):
            self.assertNotIn(".join(", call.rsplit(",", 1)[-1])

    def test_the_panel_roots_are_siblings_and_the_markup_closes(self):
        """A missing </div> put one panel inside another.

        The marking panel was never closed, so anything added after it landed
        *inside* it and was hidden whenever the marking tab was not showing --
        a whole tab that silently renders nothing. Browsers repair the markup
        on their own, which is exactly why it went unnoticed, so the balance
        is checked here instead.
        """
        from html.parser import HTMLParser
        import re as _re
        html = _source("volkit", "web", "index.html")
        body = _re.sub(r"<script>.*?</script>", "", html, flags=_re.S)
        body = _re.sub(r"<style>.*?</style>", "", body, flags=_re.S)
        void = {"meta", "input", "br", "hr", "img", "link", "source", "col",
                "area", "base", "embed", "param", "track", "wbr"}

        class Scan(HTMLParser):
            def __init__(self):
                super().__init__()
                self.stack, self.bad, self.depth = [], [], {}

            def handle_starttag(self, tag, attrs):
                if tag in void:
                    return
                got = dict(attrs).get("id")
                if got:
                    self.depth[got] = len(self.stack)
                self.stack.append(tag)

            def handle_endtag(self, tag):
                if tag in void:
                    return
                if not self.stack:
                    self.bad.append(f"stray </{tag}>")
                elif self.stack[-1] != tag:
                    self.bad.append(f"</{tag}> closes <{self.stack[-1]}>")
                    self.stack.pop()
                else:
                    self.stack.pop()

        scan = Scan()
        scan.feed(body)
        self.assertEqual(scan.bad, [])
        self.assertEqual(scan.stack, [])
        roots = [scan.depth[p] for p in ("p-pricing", "p-marking", "p-listed", "p-analysis",
                                         "p-mm", "p-monitor")]
        self.assertEqual(len(set(roots)), 1, "the panel roots are not siblings")

    def test_the_nav_shows_the_tabs_in_the_order_screens_declares_them(self):
        """`screens.SCREENS` is the one declaration of what a build has.

        The page hides a tab the build left out by name, so the two lists have
        to agree; and the order is the desk's, not the file's -- Monitor sits
        behind Vol marking because that is the pair of screens a morning
        starts on.  Two orders that drifted apart would put a build's tabs in
        one order and a trimmed build's in another.
        """
        import re as _re
        from volkit import screens
        html = _source("volkit", "web", "index.html")
        nav = html.split('<div class="nav" id="nav">')[1].split("</div>")[0]
        self.assertEqual(_re.findall(r'data-p="([a-z]+)"', nav), list(screens.ALL))
        self.assertEqual([s.label for s in screens.SCREENS],
                         _re.findall(r'data-p="[a-z]+"[^>]*>([^<]+)<', nav))
        # And the page's own panel map, which decides which tab opens when the
        # first screen is not in the build.
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const PANELROOT={")[1].split("};")[0]
        self.assertEqual(_re.findall(r"([a-z]+):'#", block), list(screens.ALL))
        for screen in screens.SCREENS:
            self.assertIn(f"{screen.name}:'#{screen.panel}'", block.replace("\n", " "))

    def test_the_pricing_grid_only_offers_fields_the_product_uses(self):
        """A box that is filled in and then ignored is a silent zero.

        The rows are declared with the products they belong to; a product
        renamed on one side and not the other would hide a field that is
        needed, or show one that is not, with nothing to say so.
        """
        import re as _re
        from volkit.pricing import PRODUCTS
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        decl = (js.split("const VANILLA=")[1].split(";")[0]
                + js.split("const IN=[")[1].split("];")[0]
                + js.split("const OUT=[")[1].split("];")[0])
        # Every list whose first entry is a product is a relevance list, and
        # then all of its entries have to be products.
        named: set[str] = set()
        for group in _re.findall(r"\[((?:'[a-z_]+'\s*,?\s*)+)\]", decl):
            items = _re.findall(r"'([a-z_]+)'", group)
            if items[0] not in PRODUCTS:
                continue
            for name in items:
                self.assertIn(name, PRODUCTS, f"{name!r} is not one of pricing.PRODUCTS")
            named.update(items)
        self.assertTrue(named, "no product relevance lists found in the grid declarations")
        # Every product owns at least one row, or a relevance list has gone
        # stale against a product nobody can price.
        for product in PRODUCTS:
            self.assertIn(product, named, f"no grid row mentions {product!r}")

    def test_every_field_a_pricing_leg_sends_is_one_the_server_reads(self):
        """The same guard the listed and market-maker panels have.

        A leg is owned by the browser and posted whole, so its fields *are*
        the payload; one the server has never heard of is a box that can be
        filled in and is then ignored.  The exceptions are declared here and
        never reach the pricer as themselves.  ``spotsrc`` / ``fwdsrc`` are
        screen state -- which market boxes are still showing the feed's
        numbers, which somebody has typed over, and which of the swap and the
        outright the leg is holding.  ``strikeask`` is the same thing said
        about the strike box: what was asked for before the marks solved it
        into the number now sitting there.  ``swap`` is the outright written the
        other way: the browser converts it where it is typed, exactly as
        every other edge of this tool converts volatility points into
        decimals once, and posts the outright it leaves in the box.  The
        server must not start reading any of them -- the leg it is sent is
        already the answer.
        """
        import re as _re
        from volkit import webapp as _webapp
        BROWSER_SIDE = {"spotsrc", "fwdsrc", "swap", "strikeask"}
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        body = js.split("function defaultLeg(")[1].split("\n}")[0]
        block = body.split("return{")[-1].split("}")[0]     # the object literal itself
        sent = set(_re.findall(r"([A-Za-z]+):", block))
        self.assertIn("forward", sent, "the leg has no outright forward box")
        reader = _inspect.getsource(_webapp.BookService.price)
        for key in sorted(sent - BROWSER_SIDE):
            with self.subTest(key):
                self.assertTrue(f'"{key}"' in reader,
                                f"the grid sends {key!r} and BookService.price never reads it")
        # And the rows the grid draws are fields the leg actually has: a row
        # keyed on something `defaultLeg` does not make is a box that starts
        # blank on every new leg and is read as nothing.
        rows = set(_re.findall(r"\['([a-z]+)','[^']+','(?:text|pair|cut|type|method|side|product|overhedge)'",
                               js.split("const IN=[")[1].split("];")[0]))
        self.assertTrue(rows)
        self.assertEqual(rows - sent, set())
        # The market is three boxes -- spot, the swap and the outright -- and
        # the leg still never sends `points`, which is the name the server
        # reads: the outright is what the screen shows and what is priced,
        # and a stored `points` of zero would pin every forward to spot.
        self.assertLessEqual({"spot", "swap", "forward"}, rows)
        self.assertNotIn("points", sent)

    def test_the_pricing_results_repeat_no_input_box(self):
        """The old screen showed the expiry, spot, the forward, the strike
        and the option type twice: once as a box you fill in and once as an
        answer beneath it.

        They are the same numbers.  What the pricer resolves is written back
        into the boxes -- a tenor becomes the one standard date, `ATM` or
        `25d` becomes the strike it solved to, `Auto` becomes `C` or `P` --
        and it is those that are priced, so a second copy under *Results* is
        one number in two places on one screen and two places for it to
        disagree.
        """
        import re as _re
        from volkit.pricing import PRODUCTS
        js = _source("volkit", "web", "index.html").split("<script>")[1]
        out = js.split("const OUT=[")[1].split("];")[0]
        keys = {k for k in _re.findall(r"\['([a-z_]+)','", out) if k not in PRODUCTS}
        self.assertTrue(keys)
        # `is_call` is the option type and `strike` doubles as the barrier;
        # both have a box of their own above.
        for key in ("expiry", "spot", "forward", "points", "swap", "strike",
                    "is_call", "barrier"):
            self.assertNotIn(key, keys, f"the results still repeat the {key!r} box")

    def test_refresh_spot_puts_every_market_box_back_on_the_feed(self):
        """The old screen had two buttons: `Refresh spot`, which refilled only
        the boxes the feed was already filling, and `Fill legs`, which also
        wrote over a level somebody had typed.  A desk pressing the one named
        after the thing it wanted -- the published spot -- kept its stale
        hand-marked level and was told the feed had been re-read.

        So there is one button, and it hands *every* market box back: a leg
        holding a typed spot or a typed forward is put back on the feed and
        the count is reported, because a hand-marked level that has just
        become the file's must not change in silence.
        """
        js = _source("volkit", "web", "index.html").split("<script>")[1]
        html = _source("volkit", "web", "index.html")
        self.assertNotIn("feedfill", html, "the second feed button is still there")
        body = js.split("async function refreshFeed(")[1].split("\n}")[0]
        # The refill is unconditional -- there is no longer a flag deciding
        # whether a typed box is taken back.
        self.assertIn("applyLegRows(r.legs,true)", body.replace(" ", ""))
        self.assertIn("put back on the feed", body)
        self.assertIn("$('#feedrefresh').onclick=()=>refreshFeed()", js)

    def test_the_vol_query_asks_in_a_strike_or_a_delta_but_never_both(self):
        """Two boxes for one point on the smile, and only one of them is the
        request.

        A box that can be filled in and is then ignored is a silent zero with
        a cursor in it, so typing in one clears the other and the resolution
        goes into the *placeholder* of the box that was left empty -- greyed
        out, so it cannot be mistaken for something typed and cannot be posted
        back as though it had been.
        """
        page = _source("volkit", "web", "index.html")
        js = page.split("<script>")[1]
        self.assertIn('id="vqstrike"', page)
        self.assertIn('id="vqdelta"', page)
        # Each clears the other as it is typed, not at the run.
        self.assertIn("$('#vqstrike').oninput", js)
        self.assertIn("$('#vqdelta').oninput", js)
        # One request goes to the server, and the delta box is the one that
        # gains the `d` the server's grammar wants.
        ask = js.split("function vqAsk(){")[1].split("\n}")[0]
        self.assertIn("vqDeltaText", ask)
        self.assertIn("strike:vqAsk()", js.replace(" ", ""))
        # The resolution lands in the placeholders, never in the values.
        hints = js.split("function vqHints(r){")[1].split("\n}")[0]
        self.assertIn("placeholder", hints)
        self.assertNotIn(".value=", hints)
        # ...and the answer the server sends carries both readings.
        vol = _source("volkit", "pricing.py").split("def quick_vol(")[1].split("\n@dataclass")[0]
        self.assertIn('"delta"', vol)
        self.assertIn('"delta_is_call"', vol)

    def test_the_screens_offer_only_the_cuts_this_desk_marks_on(self):
        """The model knows four cuts; the selectors offer two.

        Filtered in one place rather than at each of the six selectors, and
        filtered in the *page* rather than in `atm.CUTS`, because the command
        line still answers for any of them and the model is not what the desk
        chose to stop looking at.
        """
        page = _source("volkit", "web", "index.html")
        js = page.split("<script>")[1]
        self.assertIn("const SHOWN_CUTS=['TK','NY']", js)
        # Nothing reaches a selector except through the filter: the whole page
        # names `STATE.cuts` once, inside it.
        self.assertEqual(js.count("STATE.cuts"), 1)
        self.assertIn("const all=STATE.cuts||[];",
                      js.split("function cutList(){")[1].split("\n}")[0])
        for sel in ("#mcut", "#ancut", "#mocut", "#mmcut"):
            self.assertIn("fillSel('%s',cutList()," % sel, js)
        # The model itself still has all four: this is a screen preference.
        from volkit.atm import CUTS
        self.assertEqual(sorted(CUTS), ["HK", "LDN", "NY", "TK"])

    def test_the_marking_screen_opens_on_the_pair_it_marks_first(self):
        """A preference about one screen, so it lives on that screen.

        Every other selector still takes the workbook's own first pair, and
        `fillSel` keeps whatever is already chosen -- so this is the value on
        a cold load and a reload does not move the screen.
        """
        page = _source("volkit", "web", "index.html")
        js = page.split("<script>")[1]
        self.assertIn("const MARKING_PAIR='USDCNH'", js)
        self.assertIn("fillSel('#mpair',STATE.pairs||[],markingPair())", js)
        # A workbook with no CNH pair falls back to the first, as before.
        self.assertIn("all.indexOf(MARKING_PAIR)>=0", js)

    def test_the_atm_table_is_readable_with_the_overwrite_column_shut(self):
        """Two columns sit beside each other; three fill the card.

        Stretched across the card a two-column table puts the tenor at one
        edge and its volatility at the other, which is a row read across an
        inch of nothing.  And the tenor a morning is read against is
        highlighted where it stands, rather than moved or pinned out of the
        term structure it belongs to.
        """
        page = _source("volkit", "web", "index.html")
        js = page.split("<script>")[1]
        paint = js.split("function paintMarks(){")[1].split("\nasync function loadMarks")[0]
        self.assertIn("$('#matm').className='mark'+((ow||qcols)?'':' tight')", paint)
        self.assertIn("table.mark.tight{width:auto", page)
        self.assertIn("isKeyTenor(r.tenor)?' class=\"key\"':''", paint)
        self.assertIn("table.mark tr.key td", page)
        self.assertIn("const KEY_TENOR='1M'", js)

    def test_a_volatility_is_shown_to_two_decimals_everywhere(self):
        """One place decides, so two tables cannot disagree about a number.

        The desk quotes volatility to a hundredth of a point; four decimals
        is two columns of noise to scan past.  What is *typed* keeps its full
        precision -- an overwrite box holds the mark as it was marked, not as
        it is displayed -- so nothing is rounded by being looked at.
        """
        page = _source("volkit", "web", "index.html")
        js = page.split("<script>")[1]
        self.assertIn("const VOLDP=2", js)
        self.assertIn("const vnum=v=>num(v,VOLDP)", js)
        self.assertIn("const vsgn=v=>sgn(v,VOLDP)", js)
        # The screens a marker reads first go through it.
        paint = js.split("function paintMarks(){")[1].split("\nasync function loadMarks")[0]
        self.assertIn("vnum(r.cut)", paint)
        self.assertIn("['vol','Vol %',v=>vnum(v),'big']", js)
        self.assertIn("+vnum(r.vol)+", js)          # the vol query's one number
        self.assertIn("const anPct=(v,d=VOLDP)=>", js)   # the analysis screen
        self.assertIn("const anSgn=(v,d=VOLDP)=>", js)
        # The overwrite the marker types is not rounded to what is displayed.
        self.assertIn("(r.overwrite*100).toFixed(4)", paint)

    def test_the_marking_tables_hide_nothing_that_has_been_marked(self):
        """The ATM table lost its curve column and its overwrite column is a
        disclosure; the whole smile-parameter card is another.  Both are shut
        by default, which is the point -- and both would be a way to lose
        sight of a mark somebody made, which is the failure this project
        exists to remove.  So the shut state must still count what is
        overwritten, and the ATM row must still say which tenor it was.
        """
        page = _source("volkit", "web", "index.html")
        js = page.split("<script>")[1]
        paint = js.split("function paintMarks(){")[1].split("\nasync function loadMarks")[0]
        # The curve column is gone from the header the painter writes.
        self.assertNotIn("Curve %", paint)
        self.assertIn("cut %", paint)
        # Shut by default: neither disclosure is ticked in the markup, and
        # both bodies carry `hide`.
        for box in ('id="matmowshow"', 'id="msmileshow"'):
            i = page.index(box)
            tag = page[page.rindex("<input", 0, i):page.index(">", i)]
            self.assertNotIn("checked", tag, f"{box} is ticked in the markup")
        self.assertIn('<div class="row hide" id="matmowrow"', page)
        self.assertIn('<div id="msmilebody" class="hide">', page)
        # ...and shut, each still says what has been marked.
        self.assertIn("matmowcount", paint)
        self.assertIn("owdot", paint)
        self.assertIn("msmilenote", paint)
        self.assertIn("overwritten", paint)
        # A marked term structure replaces the fitted curve at *every* expiry,
        # so a shut card that counted only per-tenor overwrites would hide the
        # broadest mark on the screen.
        self.assertIn("curve", paint)
        self.assertIn("marked", paint)

    def test_a_marked_term_structure_is_posted_as_a_whole_row(self):
        """Three coefficients are one curve.

        Sent one at a time, two of every three requests would ask the server
        to hold a shape nobody typed -- a rho of the old initial and the new
        final -- and the middle one could be refused as out of domain while
        the row the marker sees is perfectly sensible.  So the browser reads
        the row and posts it whole, and the coefficients it names are the
        ones the route reads.
        """
        import re as _re
        from volkit.surface import TERM_COEFFS
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        decl = js.split("const TERMC=[")[1].split("];")[0]
        self.assertEqual(_re.findall(r"'([a-z]+)'", decl), list(TERM_COEFFS))
        body = js.split("function termBody(param){")[1].split("\n}")[0]
        # Every coefficient is read off the row, and a blank box falls back to
        # the fitted value it is showing rather than to zero.
        self.assertIn("TERMC.map", body)
        self.assertIn("placeholder", body)
        apply_ = js.split("async function applyMark(el){")[1].split("\n}")[0]
        self.assertIn("kind:'smile_term'", apply_.replace(" ", ""))
        self.assertIn("kind:'clear_smile_term'", apply_.replace(" ", ""))

    def test_every_class_the_script_looks_up_is_one_it_emits(self):
        """The panel shell and the painter that fills it are different functions.

        Nothing else would catch a renamed class between the two -- the page
        would simply render a panel with no chart and no error.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        for name in set(_re.findall(r"querySelector\('\.([A-Za-z0-9_-]+)'\)", js)):
            self.assertIn(f'class="{name}"', js, f".{name} is looked up but never emitted")

    def test_applying_a_band_reads_the_form_before_it_overwrites_it(self):
        """The old bug: `applyBand` wrote its spinner into `#bandbody` -- which
        is the div holding the treatment fields -- and only then called
        `bandPayload()`, which read `$('#bandmode').value` off a node that had
        just been removed.  Every Apply on the managed-band card died with
        "Cannot read properties of null (reading 'value')" before a request
        was ever made.  A panel's payload is read first, and the failure is
        reported *beside* the form rather than over it: a hazard with a typo
        in it is the ordinary way to get here, and the field the typo is in
        has to stay on screen to be corrected.
        """
        import re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        # the fields really do live inside #bandbody, which is what makes the
        # order matter: renderBand paints the whole form into it.
        painter = js.split("function renderBand(){")[1].split("\n}")[0]
        self.assertIn("id=\"bandmode\"", painter)
        self.assertIn("$('#bandbody').innerHTML=f", painter)

        body = js.split("async function applyBand(){")[1].split("\n}")[0]
        body = re.sub(r"/\*.*?\*/", "", body, flags=re.S)   # the comment names both
        read = body.index("bandPayload()")
        for target in ("$('#banderr')", "$('#bandstatus')", "$('#bandbody')"):
            if target in body:
                self.assertLess(
                    read, body.index(target),
                    f"applyBand writes to {target} before it reads the form")
        self.assertNotIn(
            "$('#bandbody').innerHTML", body,
            "a failed apply must not take the treatment fields off the screen")

    def test_the_listed_panel_fields_are_all_understood_by_the_server(self):
        """A field the browser sends and the server ignores is a setting that
        silently does nothing, which is the failure mode this project exists
        to remove."""
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const EF=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("forward", fields)
        src = _source("volkit", "listed.py")
        handler = src.split("def panel_from_request")[1]
        for f in fields:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")

    def test_the_contract_box_is_free_text_with_the_known_codes_offered(self):
        """The old shape: a <select> built from listed.UNDERLYINGS, so a
        contract missing from that table could only be entered as CUSTOM --
        and two CUSTOM panels on one screen cannot be told apart, which made a
        position line naming either one refused as ambiguous with nothing left
        to settle it.  The box is now an input; the known codes are offered
        through a datalist rather than imposed, and the datalist it names has
        to be one the markup actually holds.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const EF=[")[1].split("];")[0]
        kind = dict(_re.findall(r"\['([a-z_]+)','[^']*','([a-z]+)'", block))
        self.assertEqual(kind.get("underlying"), "code")
        field = js.split("function efield(")[1].split("\nfunction ")[0]
        branch = field.split("if(kind==='code'){")[1].split("}else")[0]
        self.assertIn("<input ", branch)
        self.assertNotIn("<select", branch)
        listname = _re.search(r'list="([a-zA-Z0-9_-]+)"', branch).group(1)
        self.assertIn(f'<datalist id="{listname}">', html)
        # And it is filled from the server's own list, so a code this build
        # knows how to map is one keystroke away.
        self.assertIn(f"$('#{listname}').innerHTML", js)

    def test_the_positions_panel_fields_are_all_understood_by_the_server(self):
        """Same guard as the listed fit panel, for the same reason.

        The positions panel posts its own settings alongside the panels; a
        setting the server never reads would silently do nothing.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const GF=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("vol_bump", fields)
        src = _source("volkit", "listed.py")
        handler = src.split("def positions_from_request")[1]
        for f in fields | {"text", "panels"}:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")
        # And every greek column the table paints is one listed.py declares,
        # so a column cannot reach the screen without a unit beside it.
        from volkit.listed import GREEK_FIELDS
        cols = set(_re.findall(r"\['([a-z_0-9]+)'", js.split("const GCOLS=[")[1].split("];")[0]))
        self.assertEqual(cols, {k for k, _ in GREEK_FIELDS})

    def test_the_market_maker_fields_are_all_understood_by_the_server(self):
        """Same guard as the listed panel, for the same reason.

        A field the browser sends and the server ignores is a setting that
        silently does nothing, which is the failure mode this project exists
        to remove.

        Two lists and two readers, because the screen is two stages: the fit
        panel and the quote panel post different payloads to different routes.
        Checking them against one reader would let a field the fit sends and
        only the quote reads pass, and that field would sit on the fit's own
        toolbar doing nothing.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        src = _source("volkit", "marketmaker.py")

        fit = set(_re.findall(r"\['([a-z_]+)'", js.split("const MF=[")[1].split("];")[0]))
        self.assertIn("text", fit)
        self.assertIn("target_source", fit)
        handler = src.split("def panel_from_request")[1].split("def quote_panel_from_request")[0]
        common = src.split("def _common")[1].split("def panel_from_request")[0]
        for f in fit | {"free", "smile_free", "fit_curve", "tune_wings"}:
            self.assertIn(f'"{f}"', handler + common, f"the fit reader never reads {f!r}")

        quote = set(_re.findall(r"\['([a-z_]+)'", js.split("const MQF=[")[1].split("];")[0]))
        self.assertIn("request_text", quote)
        self.assertIn("fallback_spread", quote)
        handler = src.split("def quote_panel_from_request")[1]
        for f in quote | {"marks"}:
            self.assertIn(f'"{f}"', handler + common, f"the quote reader never reads {f!r}")

    def test_the_desk_agent_fields_are_all_understood_by_the_server(self):
        """Same guard again, for the agent card inside the market-maker tab.

        The card posts its own payload rather than the panel's, so it needs
        its own list checked against its own reader; sharing the market
        maker's would let a field the agent sends go unread by either.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const AF=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("text", fields)
        self.assertIn("half_life", fields)
        handler = _source("volkit", "agent.py").split("def panel_from_request")[1]
        for f in fields - {"counterparty"}:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")
        # The one field the panel reader does not take: it says who showed the
        # market, which only matters when the run is filed to the archive.
        filer = _source("volkit", "webapp.py").split("def mm_agent_file")[1]
        self.assertIn('"counterparty"', filer)

    def test_the_ask_card_fields_are_all_understood_by_the_server(self):
        """The third agent's card posts its own list, pinned against its own reader."""
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const AK=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("text", fields)
        self.assertIn("half_life", fields)
        handler = _source("volkit", "ask.py").split("def panel_from_request")[1]
        for f in fields | {"transcript"}:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")
        # The card belongs to the market-maker screen with the other two agents.
        from volkit import screens
        self.assertIn("/api/mm/ask", screens.BY_NAME["mm"].routes)

    def test_the_ask_route_answers_without_the_paste_and_writes_nothing(self):
        """A question is answered off the archive alone, and the files are untouched."""
        import tempfile
        from pathlib import Path as _P
        from volkit import archive as _arch
        from volkit.webapp import BookService
        folder = _P(tempfile.mkdtemp())
        arc = _arch.Archive.load(folder / "arc.jsonl")
        for i in range(3):
            arc.add(_arch.Observation(kind="quote", pair="EURUSD", at=ASOF.now.isoformat(),
                                      instrument="atm", tenor="1M", bid=8.2, ask=8.6,
                                      counterparty=f"b{i}"))
        arc.flush()
        before = (folder / "arc.jsonl").read_bytes()
        service = BookService(str(WORKBOOK), ASOF, archive_path=str(folder / "arc.jsonl"),
                              journal_path=str(folder / "j.jsonl"))
        out = service.mm_ask({"pair": "EURUSD", "text": "how wide is the 1M atm this week",
                              "half_life": "5", "min_effective": "2", "lookback_days": "90",
                              "include_model_read": True, "narrate": False, "transcript": []})
        self.assertTrue(out["ok"], out)
        self.assertTrue(any("0.400 wide" in f["text"] for f in out["facts"]), out["facts"])
        self.assertIn("model_note", out)
        # A follow-up posted with the transcript keeps the topic and the pair.
        again = service.mm_ask({"pair": "EURUSD", "text": "and the 3M?", "narrate": False,
                                "transcript": [{"q": "how wide is the 1M atm this week",
                                                "a": {"ok": True}}]})
        self.assertEqual(again["question"]["topics"], ["widths"])
        self.assertEqual(again["question"]["tenor"], "3M")
        self.assertEqual(again["turns"], 1)
        # The surface is read in points beside the archive's points.
        marked = service.mm_ask({"pair": "EURUSD", "text": "where is the surface marked in 1M",
                                 "narrate": False, "transcript": []})
        line = next(f for f in marked["facts"] if "ATM " in f["text"] and f["source"] == "surface")
        self.assertGreater(float(line["text"].split("ATM ")[1].split(",")[0]), 1.0, line)
        self.assertEqual((folder / "arc.jsonl").read_bytes(), before)
        self.assertFalse((folder / "j.jsonl").exists())
        with self.assertRaises(Exception):
            service.mm_ask({"pair": "EURUSD", "text": "   "})

    def test_the_agent_card_never_names_a_folder_the_browser_chose(self):
        """A path a page can post is a path anything reaching the page can read.

        The folders the ingest route scans come from the command line and are
        held on the service; the browser chooses *when*, not *where*.
        """
        src = _source("volkit", "webapp.py")
        handler = src.split("def mm_agent_ingest")[1].split("def _agent_model")[0]
        self.assertIn("self.agent_chats", handler)
        self.assertIn("self.agent_sdr", handler)
        for named in ("payload.get(\"chats\")", "payload.get(\"sdr\")",
                      "payload.get(\"folders\")"):
            self.assertNotIn(named, handler)

    def test_a_folder_scan_does_not_hold_the_book_lock(self):
        """A minute of reading is a minute the pricing screen does not answer.

        Reading a folder can take one -- a large dissemination file, or a
        language model working through prose the grammar refused -- so the
        archive has a lock of its own and the book's is borrowed only long
        enough to read the pair list.
        """
        src = _source("volkit", "webapp.py")
        handler = src.split("def mm_agent_ingest")[1].split("def _agent_model")[0]
        self.assertIn("self._archive_lock", handler)
        before_scan = handler.split("ingest_mod.scan")[0]
        # the last lock taken before the scan must be the archive's
        self.assertGreater(before_scan.rindex("self._archive_lock"),
                           before_scan.rindex("self._lock"),
                           "the scan runs under the book's lock")

    def test_the_comparison_panel_fields_are_all_understood_by_the_server(self):
        """Same guard as the listed and market-maker panels.

        A field the browser sends and the server ignores is a setting that
        silently does nothing.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const CF=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("kind", fields)
        self.assertIn("date", fields)
        src = _source("volkit", "curves.py")
        handler = src.split("def panel_from_request")[1]
        for f in fields | {"cut", "method", "field", "base"}:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")

    def test_the_monitor_panel_fields_are_all_understood_by_the_server(self):
        """Same guard as the listed, market-maker and comparison panels.

        A field the browser sends and the server ignores is a setting that
        silently does nothing.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const MOF=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("was_kind", fields)
        self.assertIn("was_date", fields)
        self.assertIn("now_kind", fields)
        src = _source("volkit", "monitor.py")
        handler = src.split("def tile_from_request")[1].split("def panel_from_request")[0]
        for f in fields:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")
        panel = src.split("def panel_from_request")[1]
        for f in ("cut", "method", "field", "tiles", "big"):
            self.assertIn(f'"{f}"', panel, f"the server never reads {f!r}")

    def test_the_relative_value_panel_fields_are_all_understood_by_the_server(self):
        """Same guard as the listed, market-maker, comparison and monitor panels.

        The relative-value grid is posted whole like the rest of them, and a
        field the page sends that the scorer never reads would be a setting
        that appears to do something and does not.
        """
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const RVF=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("history_days", fields)
        self.assertIn("weights", fields)
        src = _source("volkit", "relvalue.py")
        handler = src.split("def panel_from_request")[1]
        for f in fields:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")
        # And every weight box the panel paints is a signal the scorer
        # declares: the boxes are built from the server's own list, so a
        # weight cannot reach the screen that `resolve_weights` would refuse.
        from volkit.relvalue import SIGNALS, WEIGHTS
        self.assertEqual([n for n, _ in SIGNALS], list(WEIGHTS))
        self.assertIn("rvw-", js, "the weight boxes are not built from the server's list")

    def test_the_band_card_fields_are_all_understood_by_the_server(self):
        """The band treatment is marked on the screen and read in one place."""
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        block = js.split("const BFIELDS=[")[1].split("];")[0]
        fields = set(_re.findall(r"\['([a-z_]+)'", block))
        self.assertIn("hazard", fields)
        self.assertIn("blend", fields)
        src = _source("volkit", "banded.py")
        handler = src.split("def from_request")[1]
        for f in fields | {"mode", "solve_hazard"}:
            self.assertIn(f'"{f}"', handler, f"the server never reads {f!r}")
        # The card's *Fit from the wings* posts the same fields plus `free`,
        # and the fit route reads that one itself.
        self.assertIn("body.free=", js)
        fit = _source("volkit", "webapp.py").split("def fit_band")[1].split("def set_band")[0]
        self.assertIn('"free"', fit)
        from volkit import screens
        owner = {r: s.name for s in screens.SCREENS for r in s.routes}
        self.assertEqual(owner["/api/band/fit"], "marking")

    def test_every_element_id_referenced_by_the_script_exists(self):
        import re as _re
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        ids = set(_re.findall(r'id="([^"]+)"', html))
        refs = set(_re.findall(r"\$\('#([a-zA-Z0-9_-]+)'\)", js))
        self.assertEqual(refs - ids - {"c1"}, set())


class TestSmileTermStructureMarks(unittest.TestCase):
    """The three coefficients behind each smile parameter, marked by hand.

    Every smile parameter already had a term structure -- ``final - (final -
    initial) * exp(-decay t)``, fitted across the quoted tenors -- but it was
    computed, shipped on ``/api/term``, and shown nowhere, so the only handles
    on a wing's *shape* across expiries were a per-tenor overwrite and the
    market maker's curve-wide shift.  A marked curve is kept beside the fitted
    one rather than written into it, which is what these pin.
    """

    def surface(self):
        return Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])["EURUSD"]

    def test_a_marked_curve_replaces_the_fitted_one_and_clearing_gives_it_back(self):
        s = self.surface()
        t = s.tenor_years("3M")
        before = s.params_at(t)
        self.assertEqual(s.set_param_term("rho10", -0.2, -0.05, 1.5), [])
        after = s.params_at(t)
        self.assertAlmostEqual(after["rho10"], -0.05 - (-0.05 - -0.2) * math.exp(-1.5 * t))
        # ...and only that parameter moved.
        for name in ("slog10", "slog25", "rho25"):
            self.assertAlmostEqual(after[name], before[name], places=12)
        s.clear_param_terms("rho10")
        self.assertAlmostEqual(s.params_at(t)["rho10"], before["rho10"], places=12)

    def test_a_refit_does_not_discard_a_marked_curve(self):
        """Why ``term_marks`` is a second dict and not an assignment into
        ``term``: the cross triangle recalibrates a leg in place, and a
        workbook export recalibrates every pair, so a mark written into the
        fitted dict would vanish somewhere no screen would ever show.
        """
        s = self.surface()
        s.set_param_term("rho10", -0.2, -0.05, 1.5)
        t = s.tenor_years("3M")
        marked = s.params_at(t)["rho10"]
        s.calibrate()
        self.assertAlmostEqual(s.params_at(t)["rho10"], marked, places=12)
        self.assertIsNotNone(s.term.get("rho10"), "the fit is still underneath it")

    def test_a_bad_coefficient_is_refused_whole(self):
        """Half a curve is not a curve.  A rejected set leaves the parameter
        exactly as it was, and the message quotes the number that was typed.
        """
        s = self.surface()
        t = s.tenor_years("3M")
        before = s.params_at(t)["rho10"]
        for args, wanted in ((("rho10", -1.0, -0.05, 1.5), "strictly inside"),
                             (("rho10", -0.2, -0.05, -1.0), "must not be negative"),
                             (("slog25", 0.0, 0.6, 1.0), "must be positive"),
                             (("slog25", float("nan"), 0.6, 1.0), "finite")):
            problems = s.set_param_term(*args)
            self.assertTrue(problems, args)
            self.assertIn(wanted, problems[0])
            self.assertEqual(s.term_marks, {}, args)
        self.assertAlmostEqual(s.params_at(t)["rho10"], before, places=12)

    def test_the_shift_and_the_marked_curve_compose(self):
        """A wing shift is additive on whatever curve is in force, so it has
        to read the marked one -- a shift measured against a curve the surface
        is no longer using is a shift the market maker cannot reason about.
        """
        s = self.surface()
        t = s.tenor_years("3M")
        s.set_param_term("rho10", -0.2, -0.05, 1.5)
        plain = s.params_at(t)["rho10"]
        s.set_param_shifts({"rho10": 0.03})
        self.assertAlmostEqual(s.params_at(t)["rho10"], plain + 0.03, places=12)
        # And a shift that would push the marked curve out of the domain is
        # reported rather than silently clamped.
        s.set_param_shifts({"rho10": 1.5})
        self.assertTrue([w for w in s.shift_warnings() if "rho10" in w])

    def test_the_route_marks_a_curve_and_the_grid_reports_it(self):
        from volkit.webapp import BookService
        svc = BookService(str(WORKBOOK), clock=ASOF)
        svc.reload()
        rows = {r["param"]: r for r in svc.marks({"pair": "EURUSD", "cut": "NY"})["term"]}
        self.assertEqual(sorted(rows), ["rho10", "rho25", "slog10", "slog25"])
        self.assertIsNone(rows["rho10"]["marked"])
        self.assertIn("decay", rows["rho10"]["fitted"])
        svc.overwrite({"pair": "EURUSD", "kind": "smile_term", "param": "rho10",
                       "initial": "-0.2", "final": "-0.05", "decay": "1.5"})
        rows = {r["param"]: r for r in svc.marks({"pair": "EURUSD", "cut": "NY"})["term"]}
        self.assertEqual(rows["rho10"]["marked"],
                         {"initial": -0.2, "final": -0.05, "decay": 1.5})
        # An empty box is named rather than read as a zero.
        with self.assertRaises(ValueError) as caught:
            svc.overwrite({"pair": "EURUSD", "kind": "smile_term", "param": "rho10",
                           "initial": "", "final": "-0.05", "decay": "1.5"})
        self.assertIn("initial", str(caught.exception))
        svc.overwrite({"pair": "EURUSD", "kind": "clear_smile_term"})
        self.assertTrue(all(r["marked"] is None
                            for r in svc.marks({"pair": "EURUSD", "cut": "NY"})["term"]))

    def test_the_workbook_row_grammar_reads_what_the_export_writes(self):
        from volkit.marketdata import overlay_label
        self.assertEqual(overlay_label("term rho10 decay"), ("term", "rho10", "decay"))
        self.assertEqual(overlay_label("TERM  Rho25  Initial"), ("term", "rho25", "initial"))
        # A row the tool would not read must not sit there looking read.
        for bad in ("term rho10", "term rho10 slope", "term bogus final"):
            self.assertIsNone(overlay_label(bad), bad)


class TestSessionFile(unittest.TestCase):
    """Saving the marks a session made, and putting them back.

    The workbook is never written to (a standing decision), so a morning's
    marking only survives in this file.  What is pinned here is that the round
    trip is exact, that the file is in the units the screen shows, and that a
    file with a bad number in it still restores everything else and says what
    it could not.
    """

    def marked_book(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])
        surface = book["USDJPY"]
        surface.atm.overwrite_tenor("1m", 0.0925)
        surface.overwrite_param("slog25", "3M", 0.61)
        surface.set_param_term("rho10", -0.2, -0.05, 1.5)
        surface.set_param_shifts({"rho25": 0.05})
        surface.anchor_tenors = True
        surface.atm.set_params(long_term_vol=0.081)
        return book

    def test_a_session_round_trips_onto_a_fresh_book(self):
        from volkit import session
        marked = self.marked_book()
        doc = session.capture(marked)
        fresh = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])
        expiry = ASOF.datetime_from_years(0.25)
        self.assertNotAlmostEqual(float(marked["USDJPY"].vol(1.0, expiry)),
                                  float(fresh["USDJPY"].vol(1.0, expiry)), places=6)
        out = session.apply_document(fresh, doc)
        self.assertEqual(out["problems"], [])
        self.assertIn("USDJPY", out["applied"])
        self.assertEqual(float(marked["USDJPY"].vol(1.0, expiry)),
                         float(fresh["USDJPY"].vol(1.0, expiry)))
        self.assertEqual(fresh["USDJPY"].param_shifts, {"rho25": 0.05})
        self.assertTrue(fresh["USDJPY"].anchor_tenors)

    def test_the_file_is_in_the_units_the_screen_shows(self):
        """Volatility points at the edges, decimals in the middle (§4).

        A file written in decimals and read back in points is the bank-width
        bug again: a 0.28 market read as 28 points.
        """
        from volkit import session
        doc = session.capture(self.marked_book(), ["USDJPY"])
        block = doc["pairs"]["USDJPY"]
        self.assertAlmostEqual(block["atm_overwrites"]["1m"], 9.25)
        self.assertAlmostEqual(block["curve"]["long_term_vol"], 8.1)
        # A smile parameter is not a volatility and is not scaled.
        self.assertAlmostEqual(block["smile_overwrites"]["slog25"]["3M"], 0.61)

    def test_the_screen_and_the_file_read_the_curve_the_same_way(self):
        """One conversion, shared, so the two cannot come to disagree."""
        from volkit import session
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF)
        shown = service.curve({"pair": "USDJPY"})["params"]
        self.assertEqual(shown, session.curve_params(service.book["USDJPY"].atm))

    def test_a_bad_value_does_not_take_the_rest_of_the_file_down(self):
        from volkit import session
        doc = session.capture(self.marked_book())
        doc["pairs"]["USDJPY"]["atm_overwrites"]["1m"] = "not a number"
        doc["pairs"]["ZZZFAKE"] = {"curve": {}}
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])
        out = session.apply_document(book, doc)
        self.assertIn("EURUSD", out["applied"])
        self.assertTrue(any("ZZZFAKE" in x for x in out["problems"]))
        self.assertTrue(any("not a number" in x for x in out["problems"]))
        # The rest of USDJPY still went on.
        self.assertEqual(book["USDJPY"].param_shifts, {"rho25": 0.05})

    def test_event_weights_and_adjustment_survive_the_round_trip(self):
        """The file holds the whole event table -- a row per release, weights
        per currency and an adjustment per pair -- and each pair block keeps
        its resolved schedule for the record."""
        from volkit import session
        from volkit.events import EventEntry
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        when = (ASOF.now + timedelta(days=13)).replace(hour=16, minute=0, second=0, microsecond=0)
        # A mark goes onto the table, which is what reaches the curves; a
        # currency weight is shared and has nowhere else to live.
        book.events.set_pair("USDJPY", [
            EventEntry(when, None, "FOMC", {"USD": 0.015, "JPY": 0.003}, 0.002)],
            pairs=book.data.pairs)
        book.apply_events()
        doc = session.capture(book, ["USDJPY"])
        row = doc["pairs"]["USDJPY"]["events"][0]
        self.assertAlmostEqual(row["bump"], 2.0)
        self.assertAlmostEqual(row["weights"]["JPY"], 0.3)
        self.assertAlmostEqual(row["adjust"], 0.2)
        table = doc["event_table"][0]
        self.assertAlmostEqual(table["weights"]["USD"], 1.5)
        self.assertAlmostEqual(table["adjust"]["USDJPY"], 0.2)

        fresh = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        out = session.apply_document(fresh, doc)
        self.assertEqual(out["problems"], [])
        ev = fresh["USDJPY"].atm.events.events[0]
        self.assertAlmostEqual(ev.bump, 0.020)
        self.assertAlmostEqual(ev.weights["JPY"], 0.003)
        self.assertAlmostEqual(ev.adjust, 0.002)

    def test_a_file_from_before_the_event_table_is_rebuilt_from_its_pairs(self):
        """An older session file spread the same events across its pairs.  A
        currency weight was shared even then, so the table is unioned back
        out of them rather than the events being dropped."""
        from volkit import session
        fresh = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        when = (ASOF.now + timedelta(days=13)).replace(hour=16, minute=0, second=0, microsecond=0)
        old = {"pairs": {"USDJPY": {"events": [
            {"when": when.strftime("%Y-%m-%dT%H:%M"), "bump": 1.25, "label": "OLD"}]}}}
        out = session.apply_document(fresh, old)
        self.assertEqual(out["problems"], [])
        self.assertTrue(any("rebuilt from its pairs" in n for n in out["notes"]), out["notes"])
        ev = fresh["USDJPY"].atm.events.events[0]
        self.assertAlmostEqual(ev.bump, 0.0125)
        self.assertAlmostEqual(ev.adjust, 0.0125)

    def test_two_pairs_disagreeing_about_a_currency_weight_are_reported(self):
        """In an older file the same weight sat in two pair blocks and could
        drift apart.  Rebuilding the table names the disagreement rather than
        averaging it -- §4, a total that disagrees with its parts is refused."""
        from volkit import session
        fresh = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])
        when = (ASOF.now + timedelta(days=13)).replace(hour=16, minute=0, second=0, microsecond=0)
        stamp = when.strftime("%Y-%m-%dT%H:%M")
        old = {"pairs": {
            "USDJPY": {"events": [{"when": stamp, "weights": {"USD": 1.5, "JPY": 0.0},
                                   "adjust": 0.0, "label": "FOMC"}]},
            "EURUSD": {"events": [{"when": stamp, "weights": {"USD": 2.0, "EUR": 0.0},
                                   "adjust": 0.0, "label": "FOMC"}]}}}
        out = session.apply_document(fresh, old)
        self.assertTrue(any("disagrees" in p and "USD" in p for p in out["problems"]),
                        out["problems"])

    def test_the_event_table_is_saved_with_the_session(self):
        from volkit import session
        from volkit.events import EventEntry
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])
        when = (ASOF.now + timedelta(days=13)).replace(hour=16, minute=0, second=0, microsecond=0)
        book.events.set_pair("USDJPY", [EventEntry(when, None, "FOMC", {"USD": 0.015}, 0.0)],
                             pairs=book.data.pairs)
        book.apply_events()
        doc = session.capture(book, ["USDJPY"])
        self.assertEqual(len(doc["event_table"]), 1)
        fresh = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY", "EURUSD"])
        out = session.apply_document(fresh, doc)
        self.assertEqual(out["problems"], [])
        # Saved for one pair, restored for the book: the weight is the
        # dollar's, so EURUSD takes it too.
        self.assertAlmostEqual(fresh["EURUSD"].atm.events.events[0].bump, 0.015)
        # A file from before the table existed leaves the book's alone.
        older = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        before = len(older.events.rows)
        session.apply_document(older, {"pairs": {}})
        self.assertEqual(len(older.events.rows), before)

    def test_events_are_replaced_and_not_merged(self):
        """A saved table is the whole table.

        Merging would double every release that appears in both the workbook
        and the file, which nothing downstream could tell from a real bump.
        """
        from volkit import session
        from volkit.events import EventEntry
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        when = ASOF.now + timedelta(days=10)
        book.events.set_pair("EURUSD", [EventEntry(when, None, "TEST", {}, 0.004)],
                             pairs=book.data.pairs)
        book.apply_events()
        doc = session.capture(book, ["EURUSD"])
        self.assertEqual(len(doc["event_table"]), 1)
        fresh = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        session.apply_document(fresh, doc)
        self.assertEqual(len(fresh["EURUSD"].atm.events.events), 1)
        # Applying it twice must not stack the same release up.
        session.apply_document(fresh, doc)
        self.assertEqual(len(fresh["EURUSD"].atm.events.events), 1)

    def test_the_file_is_written_atomically_and_read_back(self):
        import tempfile
        from volkit import session
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sub" / "marks.json"
            written = session.save(self.marked_book(), path, note="morning")
            self.assertEqual(Path(written), path)
            doc = session.load(path)
            self.assertEqual(doc["note"], "morning")
            self.assertIn("USDJPY", doc["pairs"])
        with self.assertRaises(session.SessionError):
            session.load(Path(tmp) / "gone.json")

    def test_a_pair_the_workbook_does_not_build_is_reported_not_skipped(self):
        from volkit import session
        doc = {"pairs": {"NOTAPAIR": {"curve": {}}}, "version": 1}
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        out = session.apply_document(book, doc)
        self.assertEqual(out["applied"], [])
        self.assertTrue(any("NOTAPAIR" in x for x in out["problems"]))

    def test_the_service_saves_and_restores_over_the_api(self):
        import tempfile
        from volkit.webapp import BookService
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "marks.json")
            service = BookService(str(WORKBOOK), ASOF)
            service.overwrite({"pair": "USDJPY", "kind": "atm", "tenor": "1m", "value": "9.25"})
            saved = service.session_save({"path": path})
            self.assertTrue(saved["ok"])
            service.reload()                       # back to the workbook's own marks
            self.assertEqual(service.book["USDJPY"].atm.tenor_overwrites, {})
            out = service.session_load({"path": path})
            self.assertTrue(out["ok"], out["problems"])
            self.assertAlmostEqual(service.book["USDJPY"].atm.tenor_overwrites["1m"], 0.0925)


class TestQuoteMarking(unittest.TestCase):
    """Typing a quote over the one the pair's sheet holds.

    The sheet is where quotes come from, not where they have to come from:
    the marking screen edits the four numbers a tenor is fitted from, the
    session carries them, and writing the session back puts them in the
    sheet's own cell.  What is pinned here is the layering -- ``marks`` stays
    the workbook's, the edit sits beside it, and the fit uses the two
    together -- because an edit written into ``marks`` would be silently lost
    the next time the book handed the surface its quotes.
    """

    def surface(self):
        return Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])["USDJPY"]

    def test_a_typed_quote_is_fitted_and_the_sheets_number_is_kept(self):
        s = self.surface()
        was = {m.tenor.upper(): m.rr_25 for m in s.marks}["3M"]
        before = [f.rho25 for f in s.fits]
        s.overwrite_quote("3M", "rr_25", -0.009)
        s.calibrate()
        self.assertNotEqual(before, [f.rho25 for f in s.fits])
        # The sheet's own quote is untouched, so clearing the box gives it back
        # without a reload -- and the screen can show it as the placeholder.
        self.assertEqual({m.tenor.upper(): m.rr_25 for m in s.marks}["3M"], was)
        row = {r["tenor"]: r for r in s.quote_rows()}["3M"]
        self.assertAlmostEqual(row["rr_25"], -0.009)
        self.assertAlmostEqual(row["rr_25_sheet"], was)
        self.assertTrue(row["marked"])
        s.clear_quote_overwrite("3M", "rr_25")
        s.calibrate()
        self.assertEqual(before, [f.rho25 for f in s.fits])
        self.assertFalse({r["tenor"]: r for r in s.quote_rows()}["3M"]["marked"])

    def test_a_quote_typed_back_onto_the_sheets_own_number_is_not_a_mark(self):
        """A dot that says "changed" on a row nobody changed is noise, and
        the screen reads ``marked`` rather than "is there an entry"."""
        s = self.surface()
        was = {m.tenor.upper(): m.rr_25 for m in s.marks}["3M"]
        s.overwrite_quote("3M", "rr_25", was)
        self.assertFalse({r["tenor"]: r for r in s.quote_rows()}["3M"]["marked"])

    def test_a_tenor_the_sheet_does_not_quote_needs_all_four(self):
        s = self.surface()
        fitted = {f.tenor.upper() for f in s.fits}
        self.assertNotIn("4M", fitted)
        s.overwrite_quote("4M", "rr_25", -0.005)
        s.warnings.clear()
        s.calibrate()
        self.assertNotIn("4M", {f.tenor.upper() for f in s.fits})
        self.assertTrue(any("all four" in w for w in s.warnings), s.warnings)
        # The row still exists on the screen that is creating it, saying what
        # it is: quoted by nobody, fitted by nothing.
        row = {r["tenor"]: r for r in s.quote_rows()}["4M"]
        self.assertFalse(row["quoted"])
        self.assertFalse(row["fitted"])
        for name, v in (("st_25", 0.002), ("rr_10", -0.009), ("st_10", 0.0065)):
            s.overwrite_quote("4M", name, v)
        s.calibrate()
        self.assertIn("4M", {f.tenor.upper() for f in s.fits})
        self.assertTrue({r["tenor"]: r for r in s.quote_rows()}["4M"]["fitted"])

    def test_a_strangle_is_refused_where_the_reader_would_refuse_it(self):
        """The same two checks the workbook reader makes on the cell this
        replaces, so a bad number is caught in the box it was typed in and
        not as a convergence failure three calls later."""
        s = self.surface()
        with self.assertRaises(ValueError) as cm:
            s.overwrite_quote("1M", "st_25", -0.01)
        self.assertIn("positive", str(cm.exception))
        with self.assertRaises(ValueError):
            s.overwrite_quote("1M", "fly_25", 0.01)

    def test_the_screen_shows_the_quotes_on_the_tenor_row(self):
        from volkit.webapp import BookService
        from volkit.surface import QUOTE_FIELDS
        service = BookService(str(WORKBOOK), ASOF)
        rows = {r["tenor"].upper(): r for r in service.marks({"pair": "USDJPY"})["atm"]}
        self.assertEqual(sorted(rows["3M"]["quotes"]), sorted(QUOTE_FIELDS))
        self.assertTrue(rows["3M"]["quoted"])
        # In points, like every volatility the screen shows.
        self.assertAlmostEqual(rows["3M"]["quotes"]["rr_25"],
                               rows["3M"]["quotes_sheet"]["rr_25"])
        self.assertGreater(abs(rows["3M"]["quotes"]["rr_25"]), 0.05)
        out = service.overwrite({"pair": "USDJPY", "kind": "quote", "tenor": "3M",
                                 "field": "rr_25", "value": -0.9})
        self.assertEqual(out["problems"], [])
        again = {r["tenor"].upper(): r
                 for r in service.marks({"pair": "USDJPY"})["atm"]}["3M"]
        self.assertAlmostEqual(again["quotes"]["rr_25"], -0.9)
        self.assertTrue(again["quotes_marked"])
        # A tenor the workbook does not quote appears once it is typed into,
        # even though the book prices no such point.
        service.overwrite({"pair": "USDJPY", "kind": "quote", "tenor": "4M",
                           "field": "rr_25", "value": -0.5})
        self.assertIn("4M", {r["tenor"].upper()
                             for r in service.marks({"pair": "USDJPY"})["atm"]})

    def test_a_session_carries_the_quotes_in_points_and_refits_on_the_way_back(self):
        from volkit import session
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        s = book["USDJPY"]
        s.overwrite_quote("3M", "rr_25", -0.009)
        s.calibrate()
        block = session.capture(book, ["USDJPY"])["pairs"]["USDJPY"]
        self.assertAlmostEqual(block["quote_overwrites"]["3M"]["rr_25"], -0.9)
        fresh = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        out = session.apply_document(fresh, session.capture(book, ["USDJPY"]))
        self.assertEqual(out["problems"], [])
        # Restored *and refitted*: a quote is an input to the fit, so a
        # surface that took one back without refitting would hold the typed
        # number and the old smile at the same time.
        self.assertEqual([f.rho25 for f in s.fits], [f.rho25 for f in fresh["USDJPY"].fits])


class TestWingRatios(unittest.TestCase):
    """The 10-delta wings as a multiple of the 25-delta ones.

    They were formulas in the pair sheets -- ``ST 10D = ST 25D * 3.25`` -- so
    the tool could not see them and a quote written beside one left the wing
    holding a number computed from the cell that had just been replaced.  The
    multiples are data now.  What is pinned is that the derivation is the last
    word on a wing it governs, that typing that wing takes it off the ratio
    rather than fighting it, and that the migration off the formulas does not
    move a single mark.
    """

    def workbook(self, migrate=True) -> Path:
        import shutil
        import tempfile
        from volkit import session
        d = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, d, True)
        wb = d / "vol_marks.xlsx"
        shutil.copy(WORKBOOK, wb)
        if migrate:
            session.migrate_wing_ratios(wb, in_place=True)
        return wb

    def test_the_migration_reads_every_multiple_and_moves_no_mark(self):
        from volkit import session
        from volkit.surface import load_wing_ratios
        wb = self.workbook(migrate=False)
        out = session.migrate_wing_ratios(wb, in_place=True)
        self.assertEqual(out["problems"], [])
        self.assertGreater(out["ratios"], 100)
        ratios = load_wing_ratios(wb)
        # Per pair *and* per tenor: one number for the pair would have been a
        # different workbook from the one the desk has.
        self.assertAlmostEqual(ratios["USDJPY"]["3M"].rr, 1.85)
        self.assertAlmostEqual(ratios["USDJPY"]["1Y"].rr, 1.875)
        self.assertAlmostEqual(ratios["USDHKD"]["3M"].st, 3.4)

        pairs = ["USDJPY", "EURUSD", "USDHKD", "EURJPY"]
        was = Book.from_excel(WORKBOOK, ASOF).load_all(pairs)
        now = Book.from_excel(wb, ASOF).load_all(pairs)
        self.assertEqual(now.data.problems, [])
        for pair in pairs:
            for t in (0.02, 0.25, 1.0, 2.0):
                expiry = ASOF.datetime_from_years(t)
                for k in (0.9, 1.0, 1.1):
                    self.assertAlmostEqual(float(was[pair].vol(k, expiry)),
                                           float(now[pair].vol(k, expiry)), places=8,
                                           msg=(pair, t, k))

    def test_a_quoted_25_delta_moves_the_wing_it_derives(self):
        """The whole point. Before the ratios were data, writing the 25-delta
        left the 10-delta at a value taken from the number just replaced."""
        s = Book.from_excel(self.workbook(), ASOF).load_all(["USDJPY"])["USDJPY"]
        before = {m.tenor.upper(): m for m in s.quoted_marks()}["3M"]
        self.assertAlmostEqual(before.rr_10 / before.rr_25, 1.85)
        s.overwrite_quote("3M", "rr_25", -0.009)
        after = {m.tenor.upper(): m for m in s.quoted_marks()}["3M"]
        self.assertAlmostEqual(after.rr_25, -0.009)
        self.assertAlmostEqual(after.rr_10, -0.009 * 1.85)

    def test_typing_the_wing_takes_it_off_the_ratio_and_clearing_puts_it_back(self):
        """Otherwise the ratio, applied last, wins over the box just typed in
        and the number goes back to what it was on the way out of the field."""
        s = Book.from_excel(self.workbook(), ASOF).load_all(["USDJPY"])["USDJPY"]
        s.overwrite_quote("3M", "rr_10", -0.02)
        self.assertIsNone(s.effective_ratio("3M").rr)
        self.assertAlmostEqual({m.tenor.upper(): m for m in s.quoted_marks()}["3M"].rr_10,
                               -0.02)
        # ...and the strangle beside it is untouched: one wing, one decision.
        self.assertAlmostEqual(s.effective_ratio("3M").st, 3.0)
        s.clear_quote_overwrite("3M", "rr_10")
        self.assertAlmostEqual(s.effective_ratio("3M").rr, 1.85)

    def test_a_ratio_and_the_quotes_it_derives_survive_the_session(self):
        from volkit import session
        wb = self.workbook()
        book = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
        s = book["USDJPY"]
        s.overwrite_ratio("3M", "st", 4.0)
        s.overwrite_quote("6M", "rr_10", -0.02)      # takes 6M off its rr ratio
        s.calibrate()
        block = session.capture(book, ["USDJPY"])["pairs"]["USDJPY"]
        self.assertAlmostEqual(block["wing_ratios"]["3M"]["st"], 4.0)
        # A wing taken off its ratio is written as a null and *kept*: "no
        # multiple here" and "no opinion" are different answers.
        self.assertIsNone(block["wing_ratios"]["6M"]["rr"])
        fresh = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
        out = session.apply_document(fresh, session.capture(book, ["USDJPY"]))
        self.assertEqual(out["problems"], [])
        self.assertAlmostEqual(fresh["USDJPY"].effective_ratio("3M").st, 4.0)
        self.assertIsNone(fresh["USDJPY"].effective_ratio("6M").rr)
        self.assertEqual([f.rho25 for f in s.fits], [f.rho25 for f in fresh["USDJPY"].fits])

    def test_writing_a_quote_leaves_the_sheet_all_numbers_and_consistent(self):
        """A sheet half formula and half number is a workbook that changes
        itself the next time Excel opens it."""
        from volkit import session
        wb = self.workbook(migrate=False)
        book = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
        book["USDJPY"].overwrite_quote("3M", "rr_25", -0.009)
        book["USDJPY"].calibrate()
        session.export_workbook(session.capture(book, ["USDJPY"]), wb, in_place=True)
        marks = {m.tenor.upper(): m
                 for m in Book.from_excel(wb, ASOF).load_all(["USDJPY"])["USDJPY"].marks}
        self.assertAlmostEqual(marks["3M"].rr_25, -0.009)
        import openpyxl
        ws = openpyxl.load_workbook(wb)["USDJPY"]
        self.assertFalse([c.coordinate for row in ws.iter_rows() for c in row
                          if session._is_formula(c.value)])

    def test_a_new_tenor_needs_only_the_25_delta_where_a_ratio_derives_the_wing(self):
        s = Book.from_excel(self.workbook(), ASOF).load_all(["USDJPY"])["USDJPY"]
        s.overwrite_ratio("4M", "st", 3.0)
        s.overwrite_ratio("4M", "rr", 1.85)
        s.overwrite_quote("4M", "st_25", 0.0025)
        s.overwrite_quote("4M", "rr_25", -0.006)
        s.warnings.clear()
        s.calibrate()
        self.assertIn("4M", {f.tenor.upper() for f in s.fits})
        self.assertEqual(s.warnings, [])


class TestWorkbookAsDatabase(unittest.TestCase):
    """The workbook written by the screens rather than opened by a person.

    A store has obligations a book of record does not: a write must not lose
    somebody else's, a clear must actually clear, and the copies it keeps must
    not grow without end.
    """

    def workbook(self) -> Path:
        import shutil
        import tempfile
        d = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, d, True)
        wb = d / "vol_marks.xlsx"
        shutil.copy(WORKBOOK, wb)
        return wb

    def marked(self, wb):
        from volkit import session
        book = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
        book["USDJPY"].atm.overwrite_tenor("1m", 0.0925)
        return session.capture(book, ["USDJPY"])

    def test_a_write_over_a_workbook_that_moved_is_refused(self):
        """Two volkits, or somebody saving from Excel. Last writer wins is how
        a store loses a morning."""
        from volkit import session
        wb = self.workbook()
        stamp = session.workbook_stamp(wb)
        doc = self.marked(wb)
        session.export_workbook(doc, wb, in_place=True, expect=stamp)
        with self.assertRaises(session.SessionError) as cm:
            session.export_workbook(doc, wb, in_place=True, expect=stamp)
        self.assertIn("changed since", str(cm.exception))
        # A copy takes nothing from anybody and is never refused.
        out = session.export_workbook(doc, wb, wb.with_name("copy.xlsx"), expect=stamp)
        self.assertEqual(out["problems"], [])
        # And a person who has looked can say so.
        forced = session.export_workbook(doc, wb, in_place=True, expect=stamp, force=True)
        self.assertTrue(forced["stale"])

    def test_a_cleared_overwrite_is_cleared_in_the_workbook(self):
        """``ws.cell(row, column, value=None)`` does not blank a cell --
        openpyxl assigns only when the value is not None -- so every clear on
        this path was a no-op and an overwrite taken off on the screen came
        back on the next load."""
        from volkit import session
        wb = self.workbook()
        book = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
        book["USDJPY"].atm.overwrite_tenor("1m", 0.0925)
        book["USDJPY"].anchor_tenors = True
        session.export_workbook(session.capture(book, ["USDJPY"]), wb, in_place=True)
        back = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
        self.assertEqual(back["USDJPY"].atm.tenor_overwrites, {"1m": 0.0925})
        self.assertTrue(back["USDJPY"].anchor_tenors)

        back["USDJPY"].atm.clear_overwrite("1m")
        back["USDJPY"].anchor_tenors = False
        session.export_workbook(session.capture(back, ["USDJPY"]), wb, in_place=True,
                                force=True)
        after = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
        self.assertEqual(after["USDJPY"].atm.tenor_overwrites, {})
        self.assertFalse(after["USDJPY"].anchor_tenors)

    def test_the_backups_are_pruned_to_the_most_recent(self):
        from volkit import session
        wb = self.workbook()
        doc = self.marked(wb)
        for _ in range(4):
            session.export_workbook(doc, wb, in_place=True, force=True)
        kept = sorted(wb.parent.glob(f"{wb.stem}{session.BACKUP_INFIX}*{wb.suffix}"))
        self.assertEqual(len(kept), 4)
        self.assertEqual(session.prune_backups(wb, keep=2) and len(
            list(wb.parent.glob(f"{wb.stem}{session.BACKUP_INFIX}*{wb.suffix}"))), 2)

    def test_a_configuration_tab_is_written_and_read_back(self):
        from volkit import session
        wb = self.workbook()
        rows = [{"pair": "USDHKD", "lower": 7.75, "upper": 7.85, "note": "HKMA"},
                {"pair": "USDTRY", "lower": 30.0, "upper": 45.0, "note": "made up"}]
        session.write_config_tabs(wb, {"PEG_BANDS": rows})
        book = Book.from_excel(wb, ASOF)
        self.assertEqual(sorted(book.bands), ["USDHKD", "USDTRY"])
        # A tab this does not write is refused by name rather than written
        # somewhere it would not be read.
        with self.assertRaises(session.SessionError):
            session.write_config_tabs(wb, {"PARAMS": []})

    def test_a_pair_is_added_and_removed_through_the_workbook(self):
        from volkit import session
        wb = self.workbook()
        out = session.add_pair(wb, "USDSGD", atm=6.5, quotes={
            "3M": {"rr_25": -0.40, "st_25": 0.24, "rr_10": -0.74, "st_10": 0.72}})
        self.assertEqual(out["quoted"], 1)
        book = Book.from_excel(wb, ASOF).load_all(["USDSGD"])
        self.assertEqual(book.data.problems, [])
        self.assertIn("USDSGD", book.pairs)
        self.assertEqual([f.tenor for f in book["USDSGD"].fits], ["3M"])

        # A cross is its correlation, not a volatility, and says so.
        with self.assertRaises(session.SessionError) as cm:
            session.add_pair(wb, "SGDJPY", atm=6.5)
        self.assertIn("correlation", str(cm.exception))

        # Removing takes it out of CONFIG and leaves its work where it is, so
        # adding it back finds what it had.
        session.remove_pair(wb, "USDSGD")
        self.assertNotIn("USDSGD", Book.from_excel(wb, ASOF).data.pairs)
        session.add_pair(wb, "USDSGD", atm=6.5)
        again = Book.from_excel(wb, ASOF).load_all(["USDSGD"])
        self.assertEqual([m.tenor for m in again["USDSGD"].marks], ["3M"])

    def test_a_pair_with_its_columns_and_no_quotes_is_said_not_refused(self):
        """A pair created on the screen and not yet marked is a real state
        now; a check that goes red on it is a check people stop reading."""
        from volkit import session
        wb = self.workbook()
        session.add_pair(wb, "USDSGD", atm=6.5)
        data = Book.from_excel(wb, ASOF).data
        self.assertEqual(data.problems, [])
        self.assertTrue(any("no quotes yet" in n for n in data.notes), data.notes)


class TestSessionIntoWorkbook(unittest.TestCase):
    """A session file written into a workbook's own cells.

    The one deliberate exception to "nothing writes to the workbook", so
    what is pinned is what makes it safe: it writes a copy unless told
    otherwise, the original's bytes do not move, and the copy loads as the
    session it came from -- every kind of mark, to the last digit, through
    the ordinary reader.  Cells the tool would not read back would be the
    silent zero this project exists to remove.
    """

    PAIRS = ["USDJPY", "EURUSD", "EURJPY"]

    def marked_session(self, tmp: Path):
        from volkit import session
        from volkit.banded import BandTreatment
        from volkit.events import EventEntry
        book = Book.from_excel(WORKBOOK, ASOF).load_all(self.PAIRS)
        s = book["USDJPY"]
        s.atm.overwrite_tenor("1m", 0.0925)
        s.overwrite_param("slog25", "3M", 0.61)
        s.set_param_term("rho10", -0.2, -0.05, 1.5)
        s.set_param_shifts({"rho25": 0.05})
        s.anchor_tenors = True
        s.atm.set_params(long_term_vol=0.081)
        # A Tuesday release, with weights on both legs and an adjustment.
        # It goes on the book's event table, which is where every event lives:
        # the USD weight is EURUSD's too, and the 0.2 adjustment is USDJPY's
        # alone.
        when = (ASOF.now + timedelta(days=13)).replace(hour=13, minute=30)
        book.events.set_pair("USDJPY", [
            EventEntry(when, None, "NFP", {"USD": 0.004, "JPY": 0.001}, 0.002)],
            pairs=book.data.pairs)
        book.apply_events()
        s.set_band_treatment(BandTreatment.from_request({"mode": "warn", "hazard": 3}))
        book["EURJPY"].atm.correlation  # a cross: its curve is a correlation
        path = tmp / "marks.json"
        session.save(book, path)
        return book, session.load(path)

    def test_the_copy_loads_as_the_session_and_the_original_does_not_move(self):
        import hashlib
        import shutil
        import tempfile
        from volkit import session
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wb = tmp / "vol_marks.xlsx"
            shutil.copy(WORKBOOK, wb)
            before = hashlib.md5(wb.read_bytes()).hexdigest()
            book, doc = self.marked_session(tmp)
            out = session.export_workbook(doc, wb)
            self.assertEqual(out["problems"], [])
            self.assertEqual(Path(out["written"]), tmp / "vol_marks_marked.xlsx")
            self.assertEqual(hashlib.md5(wb.read_bytes()).hexdigest(), before)

            # The reference is the session put on a fresh book *and the smiles
            # recalibrated against it*, which is what a workbook load does.
            ref = Book.from_excel(WORKBOOK, ASOF).load_all(self.PAIRS)
            session.apply_document(ref, doc)
            ref.calibrate_smiles()
            copy = Book.from_excel(out["written"], ASOF).load_all(self.PAIRS)
            self.assertEqual(copy.data.problems, [])
            self.assertFalse([w for w in copy.warnings if "session mark" in w], copy.warnings)
            for pair in self.PAIRS:
                for t in (0.02, 0.08, 0.25, 1.0):
                    expiry = ASOF.datetime_from_years(t)
                    for k in (0.97, 1.0, 1.03):
                        self.assertAlmostEqual(float(ref[pair].vol(k, expiry)),
                                               float(copy[pair].vol(k, expiry)), places=9,
                                               msg=(pair, t, k))
            # Every kind of mark came back through the reader, not just the vols.
            s = copy["USDJPY"]
            self.assertAlmostEqual(s.atm.tenor_overwrites["1m"], 0.0925)
            self.assertEqual(s.param_overwrites, {"slog25": {"3M": 0.61}})
            self.assertEqual(s.param_shifts, {"rho25": 0.05})
            self.assertEqual(sorted(s.term_marks), ["rho10"])
            self.assertAlmostEqual(s.term_marks["rho10"].initial, -0.2)
            self.assertAlmostEqual(s.term_marks["rho10"].final, -0.05)
            self.assertAlmostEqual(s.term_marks["rho10"].decay, 1.5)
            self.assertTrue(s.anchor_tenors)
            self.assertAlmostEqual(s.band_treatment.jump.hazard, 0.03)
            ev = s.atm.events.events[-1]
            self.assertAlmostEqual(ev.bump, 0.007)
            self.assertEqual(ev.weights, {"USD": 0.004, "JPY": 0.001})
            self.assertAlmostEqual(ev.adjust, 0.002)
            # ...and the sheet's formulas are still formulas, with their values.
            # Asked as ``startswith("=")`` this missed the shipped workbook's
            # own spelling: Excel saves ``=C2*3`` as an *array* formula and
            # openpyxl returns an ``ArrayFormula`` object, not a string.  The
            # export dropped the cached value of every such cell and the copy
            # came back with 126 blank quotes, while this line read a repr.
            import openpyxl
            ws = openpyxl.load_workbook(out["written"])["USDJPY"]
            self.assertTrue(session._is_formula(ws["B2"].value), ws["B2"].value)
            self.assertIsNotNone(openpyxl.load_workbook(out["written"], data_only=True)
                                 ["USDJPY"]["B2"].value)

    def test_an_array_formula_keeps_its_cached_value_through_a_copy(self):
        """Excel writes ``=C2*3`` as an array formula ({=C2*3}, the CSE
        spelling), openpyxl hands it back as an ``ArrayFormula`` object and
        writes it as ``<f t="array" ref="B2">``.  Both halves of the cache
        restore missed it -- ``_formula_cache`` tested ``startswith("=")`` on
        a non-string, and the substitution matched only a bare ``<f>`` -- so
        the copy came back with every quote blank.  The shipped workbook's
        smile sheets are array formulas throughout, which is how this reached
        the Windows build as 126 "blank quote" problems.
        """
        import io
        import tempfile
        import openpyxl
        from openpyxl.worksheet.formula import ArrayFormula
        from volkit import session

        d = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, d, True)
        path = d / "arrays.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "USDJPY"
        ws["A1"], ws["B1"] = "expiry", "ST 10D"
        ws["C2"] = 0.2175
        ws["B2"] = ArrayFormula("B2", "=C2*3")
        wb.save(path)

        blob = path.read_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(blob))
        vals = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
        # Nothing has computed it yet, so seed the value the way Excel would.
        vals["USDJPY"]["B2"] = 0.6525
        cached = session._formula_cache(wb, vals)
        self.assertEqual(cached, {1: {"B2": 0.6525}})

        buf = io.BytesIO()
        wb.save(buf)
        out = d / "copy.xlsx"
        out.write_bytes(session._restore_formula_cache(buf.getvalue(), cached))
        # Still a formula, and now readable without opening Excel first.
        kept = openpyxl.load_workbook(out)["USDJPY"]["B2"].value
        self.assertTrue(session._is_formula(kept), kept)
        self.assertAlmostEqual(
            openpyxl.load_workbook(out, data_only=True)["USDJPY"]["B2"].value, 0.6525)

    def test_a_re_quoted_tenor_goes_into_the_sheets_own_cell(self):
        """The one mark that is not written to a row of its own.

        A quote is the sheet's own number, so it replaces the sheet's own
        cell and the pair tab is then the database it is being used as.  The
        copy has to load with the typed quote in it and with every quote
        nobody touched exactly as it was -- including the array formulas the
        shipped workbook's smile sheets are made of, which is the failure
        mode this write is one cell away from at all times.
        """
        import shutil
        import tempfile
        from volkit import session
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wb = tmp / "vol_marks.xlsx"
            shutil.copy(WORKBOOK, wb)
            book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
            s = book["USDJPY"]
            untouched = {m.tenor.upper(): (m.st_10, m.rr_10) for m in s.marks}
            s.overwrite_quote("3M", "rr_25", -0.009)
            for name, v in (("rr_25", -0.005), ("st_25", 0.002),
                            ("rr_10", -0.009), ("st_10", 0.0065)):
                s.overwrite_quote("4M", name, v)
            s.calibrate()
            doc = session.capture(book, ["USDJPY"])
            out = session.export_workbook(doc, wb)
            self.assertEqual(out["problems"], [])
            self.assertTrue(any("newly quoted tenor" in n for n in out["notes"]), out["notes"])

            copy = Book.from_excel(out["written"], ASOF).load_all(["USDJPY"])
            self.assertEqual(copy.data.problems, [], copy.data.problems)
            marks = {m.tenor.upper(): m for m in copy["USDJPY"].marks}
            self.assertAlmostEqual(marks["3M"].rr_25, -0.009)
            self.assertAlmostEqual(marks["4M"].st_10, 0.0065)
            # Everything nobody typed into is the number it was, formulas
            # included -- 1W's strangle is an array formula on this workbook.
            for tenor, (st10, rr10) in untouched.items():
                self.assertAlmostEqual(marks[tenor].st_10, st10, msg=tenor)
                self.assertAlmostEqual(marks[tenor].rr_10, rr10, msg=tenor)
            # And the copy is the session: the workbook's quotes now say what
            # the session's overwrites said, so the two fit the same smile.
            for t in (0.08, 0.25, 1.0):
                expiry = ASOF.datetime_from_years(t)
                for k in (0.97, 1.0, 1.03):
                    self.assertAlmostEqual(float(s.vol(k, expiry)),
                                           float(copy["USDJPY"].vol(k, expiry)),
                                           places=9, msg=(t, k))

    def test_a_half_typed_new_tenor_is_refused_rather_than_half_written(self):
        """A row the reader would call a blank quote is not written at all,
        and the export says so where it happened."""
        import shutil
        import tempfile
        from volkit import session
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wb = tmp / "vol_marks.xlsx"
            shutil.copy(WORKBOOK, wb)
            book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
            book["USDJPY"].overwrite_quote("4M", "rr_25", -0.005)
            doc = session.capture(book, ["USDJPY"])
            out = session.export_workbook(doc, wb)
            self.assertTrue(any("all four" in p for p in out["problems"]), out["problems"])
            copy = Book.from_excel(out["written"], ASOF).load_all(["USDJPY"])
            self.assertEqual(copy.data.problems, [], copy.data.problems)
            self.assertNotIn("4M", {m.tenor.upper() for m in copy["USDJPY"].marks})

    def test_the_events_sheet_is_written_whole_and_a_weight_reaches_every_pair(self):
        """USD 0.4 on the NFP row belongs to every pair with a dollar in it,
        so the sheet is written once from the file's one table rather than
        pair by pair.  The old export wrote a pair's view of a shared weight
        and then had to cancel it in every other pair's column; a table
        written whole has nothing to cancel."""
        import shutil
        import tempfile
        from volkit import session
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wb = tmp / "vol_marks.xlsx"
            shutil.copy(WORKBOOK, wb)
            _, doc = self.marked_session(tmp)
            out = session.export_workbook(doc, wb)
            self.assertEqual(out["problems"], [])
            self.assertTrue(any("EVENTS" in n and "written whole" in n for n in out["notes"]),
                            out["notes"])
            copy = Book.from_excel(out["written"], ASOF).load_all(["EURUSD", "USDJPY"])
            self.assertEqual(copy.data.problems, [], copy.data.problems)
            # One row, and both pairs read it: EURUSD takes the dollar leg
            # alone, USDJPY takes both legs and its own cell.
            eu = [e for e in copy["EURUSD"].atm.events.events if e.when > ASOF.now]
            self.assertEqual([round(e.bump, 12) for e in eu], [0.004])
            self.assertAlmostEqual(eu[0].adjust, 0.0)
            uj = copy["USDJPY"].atm.events.events[-1]
            # A named release keeps its name through the sheet.
            self.assertEqual(uj.label, "NFP")
            self.assertEqual(uj.weights, {"USD": 0.004, "JPY": 0.001})
            self.assertAlmostEqual(uj.bump, 0.007)
            self.assertAlmostEqual(uj.adjust, 0.002)

    def test_writing_over_the_workbook_itself_needs_in_place(self):
        import hashlib
        import shutil
        import tempfile
        from volkit import session
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wb = tmp / "vol_marks.xlsx"
            shutil.copy(WORKBOOK, wb)
            before = hashlib.md5(wb.read_bytes()).hexdigest()
            _, doc = self.marked_session(tmp)
            with self.assertRaises(session.SessionError):
                session.export_workbook(doc, wb, wb)
            out = session.export_workbook(doc, wb, wb, in_place=True)
            self.assertEqual(Path(out["written"]), wb)
            # The bytes it replaced are kept beside it: openpyxl does not
            # carry images or charts through a round trip, so the backup is
            # the only way back from an export.
            self.assertTrue(out["backup"], out["notes"])
            bak = Path(out["backup"])
            self.assertTrue(bak.exists())
            self.assertEqual(hashlib.md5(bak.read_bytes()).hexdigest(), before)
            self.assertTrue(bak.name.startswith("vol_marks.bak-"))
            self.assertTrue(any("kept at" in n for n in out["notes"]), out["notes"])
            copy = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
            self.assertAlmostEqual(copy["USDJPY"].atm.tenor_overwrites["1m"], 0.0925)
            # A second export onto the same file reuses its rows rather than
            # adding another 'atm 1m' under the first.  It also names no
            # output: in_place *is* the destination -- naming none and asking
            # for in place used to write the ``_marked`` copy and report it
            # as having gone into the workbook.
            again = session.export_workbook(doc, wb, in_place=True)
            self.assertEqual(Path(again["written"]), wb)
            self.assertFalse((tmp / "vol_marks_marked.xlsx").exists())
            import openpyxl
            labels = [r[0] for r in openpyxl.load_workbook(wb)["PARAMS"]
                      .iter_rows(min_row=2, values_only=True) if r[0] is not None]
            self.assertEqual(labels.count("atm 1m"), 1)
            self.assertEqual(labels.count("slog25 3m"), 1)
            self.assertEqual(labels.count("shift rho25"), 1)
            self.assertEqual(labels.count("anchor"), 1)

    def test_a_pair_the_workbook_has_no_column_for_is_reported_not_added(self):
        import shutil
        import tempfile
        from volkit import session
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wb = tmp / "vol_marks.xlsx"
            shutil.copy(WORKBOOK, wb)
            doc = {"pairs": {"USDTRY": {"curve": {"initial_vol": 20.0}}}, "version": 1}
            out = session.export_workbook(doc, wb)
            self.assertEqual(out["written"], "")
            self.assertTrue(any("USDTRY" in p and "no PARAMS column" in p
                                for p in out["problems"]), out["problems"])
            self.assertFalse((tmp / "vol_marks_marked.xlsx").exists())

    def test_the_reader_reads_the_rows_the_export_writes(self):
        """The vocabulary, pinned on its own: a label the reader does not
        know is still reported, and a blank row is not a label."""
        from volkit.marketdata import overlay_label
        self.assertEqual(overlay_label("atm 1m"), ("atm", "1m"))
        self.assertEqual(overlay_label("ATM 1M"), ("atm", "1m"))
        self.assertEqual(overlay_label("slog25 3m"), ("smile", "slog25", "3m"))
        self.assertEqual(overlay_label("shift rho25"), ("shift", "rho25"))
        self.assertEqual(overlay_label("Anchor"), ("anchor",))
        self.assertIsNone(overlay_label("shift nothing"))
        self.assertIsNone(overlay_label("atm soon"))
        self.assertIsNone(overlay_label("initial"))

    def test_the_cli_writes_a_copy_and_the_route_writes_the_workbook_itself(self):
        import hashlib
        import io
        import shutil
        import tempfile
        from contextlib import redirect_stdout
        from volkit import cli
        from volkit.webapp import BookService
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wb = tmp / "vol_marks.xlsx"
            shutil.copy(WORKBOOK, wb)
            before = hashlib.md5(wb.read_bytes()).hexdigest()
            _, doc = self.marked_session(tmp)
            buf = io.StringIO()
            with redirect_stdout(buf):
                rc = cli.main(["-w", str(wb), "session", str(tmp / "marks.json"),
                               "--to-workbook", "--pair", "USDJPY"])
            self.assertEqual(rc, 0, buf.getvalue())
            self.assertIn("vol_marks_marked.xlsx", buf.getvalue())
            self.assertIn("is unchanged", buf.getvalue())
            self.assertEqual(hashlib.md5(wb.read_bytes()).hexdigest(), before)

            service = BookService(str(wb), ASOF)
            # A named output is still a copy, and leaves the workbook alone.
            out = service.session_export({"path": str(tmp / "marks.json"),
                                          "out": str(tmp / "via_route.xlsx")})
            self.assertTrue(out["ok"], out["problems"])
            self.assertEqual(Path(out["written"]), tmp / "via_route.xlsx")
            self.assertFalse(out["backup"])
            self.assertEqual(hashlib.md5(wb.read_bytes()).hexdigest(), before)
            # Naming a *named* output that is the workbook is still refused:
            # `out` means a copy there, and the in-place write is the one the
            # button asks for by naming nothing.
            with self.assertRaises(Exception):
                service.session_export({"path": str(tmp / "marks.json"), "out": str(wb)})

            # With no output named the route writes the loaded workbook, and
            # the marks it wrote read back through the ordinary reader.
            out = service.session_export({"path": str(tmp / "marks.json")})
            self.assertTrue(out["ok"], out["problems"])
            self.assertEqual(Path(out["written"]), wb)
            self.assertTrue(out["in_place"])
            self.assertNotEqual(hashlib.md5(wb.read_bytes()).hexdigest(), before)
            self.assertEqual(hashlib.md5(Path(out["backup"]).read_bytes()).hexdigest(),
                             before)
            back = Book.from_excel(wb, ASOF).load_all(["USDJPY"])
            self.assertAlmostEqual(back["USDJPY"].atm.tenor_overwrites["1m"], 0.0925)


class TestMonitorScreen(unittest.TestCase):
    """Small panels: what has moved between two points in time."""

    def book(self, pairs=("EURUSD",)):
        return Book.from_excel(WORKBOOK, ASOF).load_all(list(pairs))

    def history(self, book):
        return history.load_history(HISTORY, book.pairs)

    def test_a_tile_is_the_difference_between_its_two_ends(self):
        from volkit import monitor
        book = self.book()
        hist = self.history(book)
        panel = monitor.MonitorPanel(tiles=(monitor.Tile(pair="EURUSD"),))
        r = panel.run(book, hist)
        tile = r["tiles"][0]
        self.assertTrue(tile["ok"], tile["message"])
        row = next(x for x in tile["rows"] if x["tenor"] == "1M")
        self.assertAlmostEqual(row["change"]["atm"], row["now"]["atm"] - row["was"]["atm"])

    def test_a_broken_end_leaves_the_levels_standing(self):
        """The failure this project exists to remove is the empty panel.

        A tile whose earlier end has no sheet still shows what it could read
        and carries the reason it has no change.
        """
        from volkit import monitor
        book = self.book(["EURUSD", "GBPNZD"])
        panel = monitor.MonitorPanel(tiles=(monitor.Tile(pair="GBPNZD"),))
        tile = panel.run(book, self.history(book))["tiles"][0]
        self.assertTrue(tile["ok"])
        self.assertTrue(any("could not be built" in n for n in tile["notes"]))
        self.assertTrue(any(r["now"]["atm"] is not None for r in tile["rows"]))
        self.assertTrue(all(r["change"]["atm"] is None for r in tile["rows"]))

    def test_two_dated_ends_on_the_same_row_say_so(self):
        """A column of zeros otherwise reads as a quiet market."""
        from volkit import monitor
        book = self.book()
        hist = self.history(book)
        tile = monitor.Tile(pair="EURUSD", was_kind="history", was_date="latest",
                            now_kind="history", now_date="latest")
        got = monitor.run_tile(tile, book, hist)
        self.assertTrue(any("same row" in n for n in got.notes))

    def test_a_tenor_one_end_does_not_quote_is_blank_not_absent(self):
        from volkit import monitor
        book = self.book()
        tile = monitor.run_tile(monitor.Tile(pair="EURUSD"), book, self.history(book))
        tenors = [r["tenor"] for r in tile.rows]
        self.assertEqual(tenors, sorted(tenors, key=tenor_to_years))
        blank = [r for r in tile.rows if r["was"]["atm"] is None]
        self.assertTrue(blank, "the sample history quotes fewer tenors than the book")
        self.assertTrue(all(r["change"]["atm"] is None for r in blank))

    def test_a_date_on_a_source_that_has_none_is_refused(self):
        """It would be typed, ignored, and read back as if it were honoured."""
        from volkit import monitor
        from volkit.curves import CurveError
        with self.assertRaises(CurveError):
            monitor.parse_spec("EURUSD:surface@-1w")
        spec = monitor.parse_spec("EURJPY:history@-1m:history@latest")
        self.assertEqual((spec.was_kind, spec.was_date), ("history", "-1m"))
        self.assertEqual((spec.now_kind, spec.now_date), ("history", "latest"))

    def test_a_tile_that_throws_keeps_its_place(self):
        from volkit import monitor
        book = self.book()
        panel = monitor.MonitorPanel(tiles=(monitor.Tile(pair="NOTAPAIR"),
                                            monitor.Tile(pair="EURUSD", was_kind="marks",
                                                         was_date="", now_kind="surface")))
        r = panel.run(book, None)
        self.assertEqual(len(r["tiles"]), 2)
        self.assertFalse(r["tiles"][0]["ok"])
        self.assertTrue(r["tiles"][1]["ok"])

    def test_a_paste_cannot_be_a_tile_end(self):
        """A tile is rebuilt on every refresh; a paste cannot be."""
        from volkit import monitor
        from volkit.curves import CurveError
        with self.assertRaises(CurveError):
            monitor.Tile(pair="EURUSD", was_kind="paste")

    def test_a_big_move_is_graded_against_the_threshold_it_was_given(self):
        """The eye has to find the handful that matter in a few hundred cells.

        The grade is the model's, not the browser's, so the screen and
        ``volkit monitor`` mark the same cells.  Two tiers: at the threshold
        and at twice it.
        """
        from volkit import monitor
        book = self.book()
        big = 0.0025  # a quarter of a volatility point, in decimals
        tile = monitor.run_tile(monitor.Tile(pair="EURUSD"), book, self.history(book),
                                big=big)
        seen = 0
        for row in tile.rows:
            for f, change in row["change"].items():
                grade = row["grade"][f]
                if change is None:
                    self.assertEqual(grade, 0)
                    continue
                seen += 1
                want = 2 if abs(change) >= 2 * big else (1 if abs(change) >= big else 0)
                self.assertEqual(grade, want, f"{row['tenor']} {f} = {change}")
        self.assertTrue(seen, "no change was graded at all")
        self.assertEqual(tile.moved,
                         sum(1 for r in tile.rows for g in r["grade"].values() if g))
        self.assertEqual(tile.moved_hard,
                         sum(1 for r in tile.rows for g in r["grade"].values() if g > 1))

    def test_every_field_is_graded_not_only_the_highlighted_one(self):
        """What has moved may not be what was being watched.

        The screen highlights one column; a big move in any of the five is
        still a big move, so the grading is across the board.
        """
        from volkit import monitor
        from volkit.curves import CURVE_FIELDS
        book = self.book()
        tile = monitor.run_tile(monitor.Tile(pair="EURUSD"), book, self.history(book),
                                big=1e-9)
        row = next(r for r in tile.rows if all(v is not None for v in r["change"].values()))
        self.assertEqual(sorted(row["grade"]), sorted(CURVE_FIELDS))
        self.assertTrue(all(row["grade"][f] for f in CURVE_FIELDS),
                        "a threshold of nothing should grade every change that exists")

    def test_the_big_move_threshold_is_typed_in_volatility_points(self):
        """Volatility points at the edge, decimals in the middle -- converted once."""
        from volkit import monitor
        panel = monitor.panel_from_request({"tiles": [{"pair": "EURUSD"}], "big": 0.5})
        self.assertAlmostEqual(panel.big, 0.005)
        # And an empty box is the declared default, not a silent zero that
        # would leave a screen with no marks on it and no reason why.
        for empty in ({"tiles": []}, {"tiles": [], "big": ""}, {"tiles": [], "big": None}):
            self.assertAlmostEqual(monitor.panel_from_request(empty).big,
                                   monitor.DEFAULT_BIG_MOVE / 100.0)

    def test_a_threshold_that_cannot_be_compared_against_is_refused(self):
        """Nothing fails silently: a bad threshold says so rather than grading nothing."""
        from volkit import monitor
        from volkit.curves import CurveError
        with self.assertRaises(CurveError):
            monitor.panel_from_request({"tiles": [], "big": "wide"})
        with self.assertRaises(CurveError):
            monitor.MonitorPanel(big=-1.0)
        with self.assertRaises(CurveError):
            monitor.MonitorPanel(big=float("nan"))
        # Zero is the one way to turn the marking off, and it grades nothing
        # rather than grading everything.
        self.assertEqual(monitor.move_grade(9.9, 0.0), 0)

    def test_the_panel_reports_what_it_marked(self):
        """A tile scrolled past still says how much has moved in it."""
        from volkit import monitor
        book = self.book()
        panel = monitor.MonitorPanel(tiles=(monitor.Tile(pair="EURUSD"),), big=1e-9)
        r = panel.run(book, self.history(book))
        self.assertAlmostEqual(r["big"], 1e-7)  # volatility points at the edge
        self.assertEqual(r["moved"], sum(t["moved"] for t in r["tiles"]))
        self.assertTrue(r["moved"] > 0)
        quiet = monitor.MonitorPanel(tiles=(monitor.Tile(pair="EURUSD"),), big=0.0)
        self.assertEqual(quiet.run(book, self.history(book))["moved"], 0)


class TestFeedRefresh(unittest.TestCase):
    """Picking up spot that has just been published."""

    def service(self, feed):
        from volkit.webapp import BookService
        return BookService(str(WORKBOOK), ASOF, feed_path=str(feed))

    def test_a_rewritten_feed_reads_as_stale_until_it_is_refreshed(self):
        import shutil, tempfile, os, time
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "feed.csv"
            shutil.copy(FEED, path)
            service = self.service(path)
            self.assertFalse(service.feed_state()["stale"])
            text = path.read_text(encoding="utf-8").replace("USDJPY,SPOT,150.25", "USDJPY,SPOT,151.25")
            self.assertNotEqual(text, path.read_text(encoding="utf-8"))
            path.write_text(text, encoding="utf-8")
            os.utime(path, (time.time() + 5, time.time() + 5))
            self.assertTrue(service.feed_state()["stale"])
            r = service.refresh_feed({"legs": [{"pair": "USDJPY", "expiry": "3M"}]})
            self.assertFalse(r["feed"]["stale"])
            self.assertAlmostEqual(r["legs"][0]["spot"], 151.25)

    def test_a_leg_the_feed_cannot_quote_keeps_its_place_with_the_reason(self):
        service = self.service(FEED)
        r = service.refresh_feed({"legs": [{"pair": "USDJPY", "expiry": "1M"},
                                           {"pair": "GBPNZD", "expiry": "1M"},
                                           {"pair": "USDJPY", "expiry": "not a date"}]})
        self.assertEqual(len(r["legs"]), 3)
        self.assertEqual(r["legs"][0]["error"], "")
        # Half a triangle is still a refusal: the file has GBPUSD and no
        # NZDUSD.  The reason names the pair and says both halves were tried.
        self.assertIn("GBPNZD", r["legs"][1]["error"])
        self.assertIn("legs", r["legs"][1]["error"])
        self.assertIsNone(r["legs"][1]["spot"])
        # The expiry is a separate failure and does not take the market with
        # it -- and the row still holds its place either way.
        self.assertTrue(r["legs"][2]["error"])
        self.assertEqual(r["legs"][2]["expiry"], "")
        self.assertEqual([q["index"] for q in r["legs"]], [0, 1, 2])

    def test_a_cross_is_filled_from_the_legs_the_file_does_quote(self):
        """The old bug: this route asked the feed for the pair *by name*.

        The file quotes EURUSD and USDJPY and not EURJPY, so Fill refused the
        cross while the pricing grid underneath it priced the very same leg
        off the very same file -- ``Book.market_level`` composes it.  One
        place reads a level, and this is now that place too.
        """
        service = self.service(FEED)
        r = service.refresh_feed({"legs": [{"pair": "EURJPY", "expiry": "3M"}]})
        q = r["legs"][0]
        self.assertEqual(q["error"], "")
        self.assertTrue(q["derived"])
        self.assertEqual(q["via"], "EURUSD and USDJPY")
        priced = service.price({"legs": [{"pair": "EURJPY", "expiry": "3M",
                                          "strike": "ATM", "type": "C"}]})["legs"][0]
        self.assertAlmostEqual(priced["forward"], q["forward"], places=9)

    def test_the_points_are_the_ones_the_pricer_would_use(self):
        """Filling a leg must not put a different market in front of it."""
        service = self.service(FEED)
        r = service.refresh_feed({"legs": [{"pair": "USDJPY", "expiry": "3M"}]})
        q = r["legs"][0]
        priced = service.price({"legs": [{"pair": "USDJPY", "expiry": "3M", "strike": "ATM",
                                          "type": "C"}]})
        self.assertAlmostEqual(priced["legs"][0]["spot"], q["spot"])
        self.assertAlmostEqual(priced["legs"][0]["forward"], q["forward"], places=9)

    def test_refreshing_without_a_feed_says_so(self):
        from volkit.webapp import BookService
        from volkit.feed import FeedError
        service = BookService(str(WORKBOOK), ASOF)
        with self.assertRaises(FeedError):
            service.refresh_feed({"legs": []})


class TestTheThreeMarketBoxes(unittest.TestCase):
    """The pricing screen shows one box each for spot, the forward and the expiry.

    The forward box is the **outright**, not points over a pip divisor, and
    both level boxes are filled from the feed at the leg's own expiry and are
    then editable.  ``pricing.resolve_legs`` is what fills them, and it is the
    same reading the pricer does -- one place for the calendar and one place
    for the level.
    """

    @classmethod
    def setUpClass(cls):
        from volkit.webapp import BookService
        cls.service = BookService(str(WORKBOOK), ASOF, feed_path=str(FEED))

    def rows(self, *legs):
        return self.service.legs({"legs": list(legs)})["legs"]

    def test_every_spelling_of_a_date_comes_back_as_the_one_standard_date(self):
        """Whatever is typed, the box ends up holding ``YYYY-MM-DD``.

        A desk writes a date half a dozen ways and none of them is worth
        making somebody translate by hand -- ``28May24`` least of all, since
        that is the form this package prints in a leg's own label.
        """
        for text in ("2024-05-28", "28May24", "28May2024", "28 May 24", "28 May 2024",
                     "May 28 2024", "May 28, 2024", "28-May-2024", "28-May-24",
                     "2024/05/28", "5/28/2024", "20240528", "2024.05.28"):
            with self.subTest(text):
                row = self.rows({"pair": "USDJPY", "expiry": text})[0]
                self.assertEqual(row["error"], "")
                self.assertEqual(row["expiry"], "2024-05-28")

    def test_the_box_takes_a_spelled_out_tenor_and_a_year_less_date(self):
        """'1wk' is a tenor and '28 May' is the coming twenty-eighth of May.

        Both used to be refused: the tenor because the unit had to be one
        letter, the date because the year was not optional.  The year comes
        from the book's clock (28-Feb-2024 here), never the machine's.
        """
        week = self.rows({"pair": "USDJPY", "expiry": "1wk"})[0]
        self.assertEqual(week["error"], "")
        self.assertEqual(week["expiry"],
                         self.rows({"pair": "USDJPY", "expiry": "1W"})[0]["expiry"])
        for text in ("28 May", "28May", "May 28"):
            with self.subTest(text):
                row = self.rows({"pair": "USDJPY", "expiry": text})[0]
                self.assertEqual(row["error"], "")
                self.assertEqual(row["expiry"], "2024-05-28")
        # A date already past this year is next year's.
        row = self.rows({"pair": "USDJPY", "expiry": "10 Jan"})[0]
        self.assertEqual(row["error"], "")
        self.assertEqual(row["expiry"], "2025-01-10")

    def test_a_tenor_is_resolved_once_on_the_pair_s_own_calendar(self):
        for tenor in ("1W", "8d", "3M", "2y"):
            with self.subTest(tenor):
                row = self.rows({"pair": "USDJPY", "expiry": tenor})[0]
                self.assertEqual(row["error"], "")
                self.assertTrue(self.service.book.calendars.is_business_day(
                    "USDJPY", date.fromisoformat(row["expiry"])))
        # "8d" is eight days and not the eighth of something: the tenor is
        # tried first, exactly as `resolve_expiry` has always done it.  And a
        # day tenor is eight *business* days, which is what O/N is one of --
        # so it lands eleven or twelve calendar days out, not eight.  Adding
        # calendar days to the spot date instead collapsed the short tenors
        # onto each other: dealt on a Wednesday, "1d" and "2d" both came back
        # Thursday, because the two business days taken off at the end
        # swallowed the weekend the addition had just crossed.
        row = self.rows({"pair": "USDJPY", "expiry": "8d"})[0]
        cal = self.service.book.calendars
        today = self.service.book.clock.now.date()
        self.assertEqual(row["expiry"],
                         cal.add_business_days("USDJPY", today, 8).isoformat())
        self.assertGreater(row["days"], 8)

    def test_the_short_tenors_do_not_collapse_onto_one_date(self):
        """One business day apart, each of them, however the weekend falls."""
        seen = [self.rows({"pair": "USDJPY", "expiry": t})[0]["expiry"]
                for t in ("O/N", "1d", "2d", "3d", "4d")]
        self.assertEqual(seen[0], seen[1])          # O/N is one business day
        self.assertEqual(len(set(seen)), 4)         # and the rest are distinct

    def test_a_leg_carries_its_spot_and_settlement_dates(self):
        """The settlement date is a fact the desk confirms on, not an internal.

        It is the spot lag *after* the expiry, on the pair's own calendar, and
        it is the date the forward beside it is a forward to.
        """
        cal = self.service.book.calendars
        for text in ("1M", "2024-05-28"):
            with self.subTest(text):
                row = self.rows({"pair": "USDJPY", "expiry": text})[0]
                expiry = date.fromisoformat(row["expiry"])
                self.assertEqual(row["settle"],
                                 cal.delivery_from_expiry("USDJPY", expiry).isoformat())
                self.assertEqual(row["spot_date"], cal.spot_date(
                    "USDJPY", self.service.book.clock.now.date()).isoformat())
                self.assertTrue(cal.is_settlement_day("USDJPY",
                                                      date.fromisoformat(row["settle"])))

    def test_the_level_is_the_feed_s_at_that_leg_s_own_expiry(self):
        one, three = self.rows({"pair": "USDJPY", "expiry": "1M"},
                               {"pair": "USDJPY", "expiry": "3M"})
        self.assertTrue(one["feed"] and three["feed"])
        # One spot, two forwards: the points are interpolated at each expiry,
        # which is the whole reason the box is refilled when the expiry moves.
        self.assertAlmostEqual(one["spot"], three["spot"], places=12)
        self.assertNotAlmostEqual(one["forward"], three["forward"], places=6)

    def test_a_row_that_cannot_be_read_keeps_its_place_and_its_reason(self):
        rows = self.rows({"pair": "USDJPY", "expiry": "1M"},
                         {"pair": "USDJPY", "expiry": "not a date"},
                         {"pair": "", "expiry": "1M"})
        self.assertEqual([r["index"] for r in rows], [0, 1, 2])
        self.assertEqual(rows[0]["error"], "")
        self.assertTrue(rows[1]["error"])
        self.assertEqual(rows[1]["expiry"], "")
        self.assertIn("currency pair", rows[2]["error"])

    def test_typing_does_not_re_read_the_feed_file(self):
        """The box is refilled on a keystroke; the file is read on a button.

        Going to disk every time somebody paused in the expiry box would make
        an editor's hesitation a file read, and would pick a republished feed
        up underneath a price being looked at -- which is what the auto-load
        switch exists to make a deliberate choice.
        """
        import shutil, tempfile, os, time
        from volkit.webapp import BookService
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "feed.csv"
            shutil.copy(FEED, path)
            service = BookService(str(WORKBOOK), ASOF, feed_path=str(path))
            was = service.legs({"legs": [{"pair": "USDJPY", "expiry": "3M"}]})["legs"][0]
            path.write_text(path.read_text(encoding="utf-8").replace(
                "USDJPY,SPOT,150.25", "USDJPY,SPOT,151.25"), encoding="utf-8")
            os.utime(path, (time.time() + 5, time.time() + 5))
            still = service.legs({"legs": [{"pair": "USDJPY", "expiry": "3M"}]})["legs"][0]
            self.assertAlmostEqual(still["spot"], was["spot"], places=12)
            now = service.refresh_feed(
                {"legs": [{"pair": "USDJPY", "expiry": "3M"}]})["legs"][0]
            self.assertAlmostEqual(now["spot"], 151.25, places=12)


class TestTheSettlementDateBox(unittest.TestCase):
    """The settlement date is an input, and the calendar is only its default.

    A tenor names a settlement date first and the expiry comes back from it,
    so the settlement date is the anchor of the whole construction -- but it
    is also the one date on a leg the calendar cannot always answer, because
    a broken date is a thing two counterparties agree and not a thing a
    holiday table knows.  So the box fills itself from the calendar, says so,
    and takes a date typed over it; emptying it hands it back.
    """

    @classmethod
    def setUpClass(cls):
        from volkit.webapp import BookService
        cls.service = BookService(str(WORKBOOK), ASOF, feed_path=str(FEED))
        # The service's own book, so the leg rows and the level checked
        # against them are read off one feed: `feed.load_for` places a tenor
        # pillar on its real delivery date, which is the placement the whole
        # screen is built on.
        cls.book = cls.service.book

    def rows(self, *legs):
        return self.service.legs({"legs": list(legs)})["legs"]

    def test_an_untyped_box_is_the_calendar_s_and_says_what_the_default_is(self):
        row = self.rows({"pair": "USDJPY", "expiry": "1M"})[0]
        cal = self.service.book.calendars
        expiry = date.fromisoformat(row["expiry"])
        want = cal.delivery_from_expiry("USDJPY", expiry).isoformat()
        self.assertEqual(row["settle"], want)
        # The default travels with it so the screen has somewhere to put the
        # box back to the moment it is emptied, without rebuilding a
        # construction that lives on the pair's own holidays.
        self.assertEqual(row["settle_default"], want)
        self.assertFalse(row["settle_stated"])
        self.assertEqual(row["settle_note"], "")

    def test_a_date_in_the_box_is_not_evidence_anybody_typed_it(self):
        """The screen fills this box the way it fills spot from the feed.

        It then posts what is in it, so a date arriving here means nothing on
        its own; ``settlesrc`` is the screen saying whether somebody chose
        it.  Read the other way, every leg would have claimed a broken date.
        """
        default = self.rows({"pair": "USDJPY", "expiry": "1M"})[0]["settle_default"]
        row = self.rows({"pair": "USDJPY", "expiry": "1M",
                         "settle": "2024-12-31", "settlesrc": "calc"})[0]
        self.assertEqual(row["settle"], default)
        self.assertFalse(row["settle_stated"])
        # A caller that sends a date and no flag -- the API, a script -- means
        # it, which is the reading that needs no screen to be present.
        row = self.rows({"pair": "USDJPY", "expiry": "1M", "settle": "2024-12-31"})[0]
        self.assertEqual(row["settle"], "2024-12-31")
        self.assertTrue(row["settle_stated"])

    def test_the_forward_is_read_on_the_date_in_the_box_and_the_expiry_stays(self):
        """The one thing a stated settlement date moves.

        It is the date a forward is a price for, so a forward read there is a
        different forward; the expiry is what the option is worth time on and
        it does not move, so the volatility and the time to expiry are
        untouched.
        """
        base = self.rows({"pair": "USDJPY", "expiry": "1M"})[0]
        far = self.rows({"pair": "USDJPY", "expiry": "1M", "settle": "2024-06-28"})[0]
        self.assertEqual(far["expiry"], base["expiry"])
        self.assertEqual(far["days"], base["days"])
        self.assertEqual(far["years"], base["years"])
        self.assertEqual(far["settle"], "2024-06-28")
        self.assertAlmostEqual(far["spot"], base["spot"], places=12)
        self.assertNotAlmostEqual(far["points"], base["points"], places=4)
        # And it is the *level* lookup that moved, not a second copy of it:
        # the same date asked of the one place a level is read.
        level = self.book.market_level_for("USDJPY", date.fromisoformat(base["expiry"]),
                                           date(2024, 6, 28))
        self.assertAlmostEqual(far["forward"], level["forward"], places=12)
        self.assertEqual(level["settle"], "2024-06-28")

    def test_the_priced_leg_reads_its_forward_on_the_same_date_it_shows(self):
        """One construction, so the date shown and the date read are one date."""
        from volkit.pricing import OptionLeg, price_strip
        base, moved = price_strip(self.book, [
            OptionLeg("USDJPY", "1M", "ATM"),
            OptionLeg("USDJPY", "1M", "ATM", settle="2024-06-28")])["legs"]
        self.assertEqual(moved["settle"], "2024-06-28")
        self.assertEqual(moved["expiry"], base["expiry"])
        self.assertAlmostEqual(moved["t"], base["t"], places=15)
        self.assertAlmostEqual(moved["atm_vol"], base["atm_vol"], places=12)
        self.assertNotAlmostEqual(moved["forward"], base["forward"], places=5)
        self.assertIn("as typed", moved["settle_rule"])

    def test_a_settlement_date_before_the_expiry_is_refused(self):
        """An option cannot settle before it is exercised.

        Nothing fails silently: the row keeps its place and carries the
        reason, and the message says how to get the box back.
        """
        row = self.rows({"pair": "USDJPY", "expiry": "1M", "settle": "2024-03-01"})[0]
        self.assertIn("before the expiry", row["error"])
        self.assertIn("Empty the box", row["error"])
        from volkit.pricing import OptionLeg, price_strip
        r = price_strip(self.book,
                        [OptionLeg("USDJPY", "1M", "ATM", settle="2024-03-01")])["legs"][0]
        self.assertFalse(r["ok"])
        self.assertIn("before the expiry", r["error"])

    def test_a_broken_date_is_taken_and_said_out_loud(self):
        """The case the box exists for is not the case it refuses.

        A settlement date that is not a value date for the pair is
        deliverable only by agreement -- which is exactly what somebody typing
        one is telling the screen -- so it is priced and reported, never
        rejected and never silently rolled to the next good day.
        """
        saturday = date(2024, 6, 29)
        self.assertFalse(self.service.book.calendars.is_settlement_day("USDJPY", saturday))
        row = self.rows({"pair": "USDJPY", "expiry": "1M", "settle": saturday.isoformat()})[0]
        self.assertEqual(row["error"], "")
        self.assertEqual(row["settle"], saturday.isoformat())
        self.assertIn("not a value date", row["settle_note"])
        from volkit.pricing import OptionLeg, price_strip
        r = price_strip(self.book, [OptionLeg("USDJPY", "1M", "ATM",
                                              settle=saturday.isoformat())])["legs"][0]
        self.assertTrue(r["ok"], r["error"])
        self.assertTrue(any("not a value date" in w for w in r["warnings"]))
        # A date the calendar produced says nothing: a note on every leg is a
        # note nobody reads.
        self.assertEqual(self.rows({"pair": "USDJPY", "expiry": "1M"})[0]["settle_note"], "")

    def test_the_box_takes_every_spelling_the_expiry_box_takes(self):
        """One timestamp reader, so a screen never grows a second date parser."""
        for text in ("2024-06-28", "28Jun24", "28 Jun 2024", "June 28, 2024",
                     "2024/06/28", "6/28/2024", "20240628"):
            with self.subTest(text):
                row = self.rows({"pair": "USDJPY", "expiry": "1M", "settle": text})[0]
                self.assertEqual(row["error"], "")
                self.assertEqual(row["settle"], "2024-06-28")


class TestLegMarketOverrides(unittest.TestCase):
    """Two boxes, either of which may be typed over, and the feed fills the rest.

    The screen sends both, so the ordinary case is that both are typed and
    what is priced is what is on the screen.  The interesting cases are the
    partial ones.
    """

    @classmethod
    def setUpClass(cls):
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        cls.book.feed = MarketFeed.load(FEED)

    def one(self, **kw):
        return price_strip(self.book, [OptionLeg("USDJPY", "3M", "ATM", **kw)])["legs"][0]

    def test_the_feed_fills_both_boxes_when_neither_is_typed(self):
        r = self.one()
        self.assertEqual(r["market_source"], "feed")
        self.assertTrue(r["feed_used"])
        self.assertAlmostEqual(r["spot"], 150.25, places=10)
        self.assertNotAlmostEqual(r["forward"], r["spot"], places=6)

    def test_a_typed_forward_is_the_forward_and_spot_stays_the_feed_s(self):
        r = self.one(forward=155.0)
        self.assertAlmostEqual(r["forward"], 155.0, places=12)
        self.assertAlmostEqual(r["spot"], 150.25, places=10)
        self.assertEqual(r["market_source"], "spot from the feed")

    def test_a_typed_spot_leaves_the_forward_to_the_feed(self):
        """Each box falls back on its own.

        Clearing one of two boxes is an ordinary thing to do to one leg of a
        strip, and it must not need the other cleared as well.  The old
        screen had no forward box at all -- points on top of spot -- so a
        typed spot took the whole market with it.
        """
        fed = self.one()
        r = self.one(spot=160.0)
        self.assertAlmostEqual(r["spot"], 160.0, places=12)
        self.assertAlmostEqual(r["forward"], fed["forward"], places=12)
        self.assertEqual(r["market_source"], "forward from the feed")

    def test_both_boxes_typed_are_priced_exactly_as_they_stand(self):
        r = self.one(spot=160.0, forward=159.4)
        self.assertAlmostEqual(r["spot"], 160.0, places=12)
        self.assertAlmostEqual(r["forward"], 159.4, places=12)
        self.assertEqual(r["market_source"], "typed")
        self.assertFalse(r["feed_used"])

    def test_the_points_spelling_still_says_where_the_forward_is(self):
        """``forward_points`` defaults to None and not to zero.

        Nothing else can tell "said nothing about points" from "said the
        forward is at spot" -- and the two want opposite things from the
        feed.  Defaulting to 0.0, as this did while the screen sent points,
        made every leg the second kind the moment the screen stopped sending
        them.
        """
        r = self.one(spot=160.0, forward_points=0.0, pip=100.0)
        self.assertAlmostEqual(r["forward"], 160.0, places=12)
        self.assertEqual(r["market_source"], "typed")
        r = self.one(spot=160.0, forward_points=-45.0, pip=100.0)
        self.assertAlmostEqual(r["forward"], 160.0 - 0.45, places=10)
        # And a forward given outright wins over both of them.
        r = self.one(spot=160.0, forward=159.0, forward_points=-45.0, pip=100.0)
        self.assertAlmostEqual(r["forward"], 159.0, places=12)

    def test_a_filled_box_the_screen_calls_the_feed_s_reports_as_the_feed_s(self):
        """The bug: every leg read ``typed``, on a screen nobody had typed into.

        The pricing screen fills spot and the outright from the feed and then
        posts what is in the boxes, so by the time a leg is priced a feed
        level and a hand-marked one look identical and the old inference --
        "a box with something in it was somebody's" -- called both of them
        typed.  Only the screen knows, so the screen says: ``spot_source`` /
        ``forward_source`` change no number, they name where the number came
        from.  This is what makes `Refresh spot`, which puts every box back
        on the feed, put the *Market* row back with them.
        """
        fed = self.one()                     # blank boxes: the feed fills both
        r = self.one(spot=fed["spot"], forward=fed["forward"],
                     spot_source="feed", forward_source="feed")
        self.assertEqual(r["market_source"], "feed")
        self.assertTrue(r["feed_used"])
        self.assertAlmostEqual(r["spot"], fed["spot"], places=12)
        self.assertAlmostEqual(r["forward"], fed["forward"], places=12)

    def test_a_level_the_screen_says_is_typed_is_typed_however_it_got_there(self):
        fed = self.one()
        r = self.one(spot=fed["spot"], forward=fed["forward"],
                     spot_source="typed", forward_source="typed")
        self.assertEqual(r["market_source"], "typed")
        self.assertFalse(r["feed_used"])
        # Half and half, each half named: a typed outright over the feed's
        # spot, and the other way round.
        self.assertEqual(self.one(spot=fed["spot"], forward=155.0,
                                  spot_source="feed", forward_source="typed"
                                  )["market_source"], "spot from the feed")
        # A typed spot with the feed's *swap* on top of it: the outright box
        # then holds neither party's outright, and saying "forward from the
        # feed" would claim the file published a level it did not.
        self.assertEqual(self.one(spot=160.0, forward=159.5,
                                  spot_source="typed", forward_source="feed"
                                  )["market_source"], "swap from the feed")

    def test_a_blank_box_is_the_feed_s_whatever_the_caller_says(self):
        """Neither half can be talked out of what actually happened.

        A caller that leaves the box empty and calls it typed still gets the
        feed's level, because that is where it came from; and a caller that
        says nothing at all gets the old inference, which is what the command
        line and any script hold.
        """
        self.assertEqual(self.one(spot_source="typed", forward_source="typed"
                                  )["market_source"], "feed")
        self.assertEqual(self.one()["market_source"], "feed")
        self.assertEqual(self.one(spot=160.0, forward=159.4)["market_source"], "typed")

    def test_a_forward_on_its_own_with_no_feed_is_a_whole_market(self):
        """This model carries no discount curve, so spot has nothing to add.

        The alternative was the old fallback, spot = 1.0, which would price a
        yen option 150 times away from the level in the box beside it.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        r = price_strip(book, [OptionLeg("USDJPY", "3M", "ATM", forward=155.0)])["legs"][0]
        self.assertTrue(r["ok"], r["error"])
        self.assertAlmostEqual(r["spot"], 155.0, places=12)
        self.assertAlmostEqual(r["forward"], 155.0, places=12)
        self.assertEqual(r["market_source"], "typed")


class TestWorkbooksAreNotHeldOpen(unittest.TestCase):
    """This tool reads other people's files; it must not lock them.

    ``pd.ExcelFile(path)`` keeps the file open for as long as the reader is
    alive, and openpyxl's workbook is full of parent/child cycles, so the
    handle outlived the call that made it.  On Windows that is enough to stop
    Excel saving the very sheet the tool had just read: the reported bug was
    that a loaded historical workbook could no longer be saved.
    """

    class _Tracked:
        """Patches ``pd.ExcelFile`` and records every reader that is made."""

        def __enter__(self):
            import pandas as pd
            self.pd, self.real, self.made = pd, pd.ExcelFile, []
            made = self.made

            class Tracking(self.real):
                def __init__(inner, io, *a, **kw):
                    # What was handed to the reader is kept here rather than
                    # read back off it afterwards: pandas 3.0 dropped the
                    # ``ExcelFile.io`` attribute this used to look at, and the
                    # whole suite failed on the runner with "'Tracking' object
                    # has no attribute 'io'".  The argument is the subject of
                    # the test and this is the one place that has it whatever
                    # pandas does with it next.
                    inner.volkit_io = io
                    inner.volkit_closed = False
                    made.append(inner)
                    super().__init__(io, *a, **kw)

                def close(inner):
                    inner.volkit_closed = True
                    return super().close()

            pd.ExcelFile = Tracking
            return self

        def __exit__(self, *exc):
            self.pd.ExcelFile = self.real
            return False

    def check(self, made):
        import io as _io
        self.assertTrue(made, "nothing was read, so this proves nothing")
        for reader in made:
            # Over a copy in memory, never over the path: the file itself is
            # opened, copied and closed before any parsing starts.
            self.assertIsInstance(reader.volkit_io, _io.BytesIO)
            self.assertTrue(reader.volkit_closed,
                            "a reader was left open after the file had been read")

    def test_the_marks_workbook_is_not_left_open(self):
        with self._Tracked() as t:
            ExcelSource(WORKBOOK).load()
        self.check(t.made)

    def test_the_historical_workbook_is_not_left_open(self):
        sample = Path(__file__).resolve().parents[1] / "files" / "history_sample.xlsx"
        with self._Tracked() as t:
            hist = history.load_history(sample)
        self.assertTrue(hist.pairs)
        self.check(t.made)

    def test_a_forward_sheet_is_not_left_open(self):
        sample = Path(__file__).resolve().parents[1] / "files" / "history_sample.xlsx"
        with self._Tracked() as t:
            with self.assertRaises(Exception):
                # Whatever the sheet turns out to be, the reader closes: a
                # file left open by a failed read is the same lock.
                analytics.ForwardCurve.from_excel(sample, "NOT_A_SHEET")
        self.check(t.made)

    def test_a_loaded_workbook_can_still_be_replaced(self):
        """What the desk actually does: load it here, then save it there."""
        import shutil, tempfile
        sample = Path(__file__).resolve().parents[1] / "files" / "history_sample.xlsx"
        with tempfile.TemporaryDirectory() as tmp:
            live = Path(tmp) / "vol_history.xlsx"
            shutil.copy(sample, live)
            history.load_history(live)
            spare = Path(tmp) / "next.xlsx"
            shutil.copy(sample, spare)
            spare.replace(live)              # Excel's own save is a replace
            self.assertTrue(history.load_history(live).pairs)


class TestTextFilesAreUtf8(unittest.TestCase):
    """Every text file is UTF-8, whatever the machine's locale says.

    Python decodes and encodes text with the *locale* encoding by default,
    which on the Windows desk this is built for is cp1252.  Reading
    ``volkit/web/index.html`` with it ended the Windows build at the test
    suite with ``'charmap' codec can't decode byte 0x81``, and the same
    default sits under every settings file, holiday override, band row and
    published feed the tool reads.  ``paths.read_text`` / ``paths.open_text``
    / ``paths.write_text`` are the one place that says otherwise.
    """

    ROOT = Path(__file__).resolve().parents[1]

    def sources(self):
        files = sorted((self.ROOT / "volkit").rglob("*.py"))
        files += [self.ROOT / "build_exe.py", self.ROOT / "volkit.spec",
                  Path(__file__).resolve()]
        # paths.py is where the encoding is named, so it is the one file that
        # may spell it out; launcher and the spec are build scaffolding.
        return [f for f in files if f.name != "paths.py"]

    def test_no_text_file_is_read_in_the_locale_encoding(self):
        import ast

        offenders = []
        for f in self.sources():
            tree = ast.parse(f.read_text(encoding="utf-8"), filename=str(f))
            for node in ast.walk(tree):
                if not isinstance(node, ast.Call):
                    continue
                fn = node.func
                if not isinstance(fn, ast.Attribute):
                    continue        # a bare read_text() is the helper itself
                if isinstance(fn.value, ast.Name) and fn.value.id == "paths":
                    continue        # ...and so is paths.read_text()
                if isinstance(fn.value, ast.Name) and fn.value.id == "webbrowser":
                    continue        # webbrowser.open opens a tab, not a file
                kw = {k.arg for k in node.keywords}
                bad = None
                if fn.attr in ("read_text", "write_text") and "encoding" not in kw:
                    bad = fn.attr
                elif fn.attr == "open" and "timeout" in kw:
                    bad = None      # a network opener (urllib), not a file: it
                                    # returns bytes and has no encoding to get
                                    # wrong.  A file open has no timeout.
                elif fn.attr == "open" and "encoding" not in kw and "b" not in "".join(
                        a.value for a in node.args if isinstance(a, ast.Constant)
                        and isinstance(a.value, str)):
                    bad = "open"
                elif fn.attr == "run" and "text" in kw and "encoding" not in kw:
                    bad = "subprocess.run(text=True)"
                if bad:
                    offenders.append(f"{f.relative_to(self.ROOT)}:{node.lineno}  {bad}")
        self.assertEqual(offenders, [], "these read or write text in the locale's "
                         "encoding, which is cp1252 on the desk machine:\n  "
                         + "\n  ".join(offenders))

    def test_a_settings_file_in_the_machines_own_code_page_is_read_and_said(self):
        """Notepad's default on a Chinese Windows is ANSI -- cp936, not UTF-8
        -- and a ``volkit.cfg`` saved that way used to stop the packaged exe
        at startup with a decode error instead of reading the workbook path
        in it.  It is read as the machine's own code page, which is a fact
        about the machine that saved it and not a guess, and the fallback is
        never silent."""
        import tempfile
        from volkit import config, paths

        with tempfile.TemporaryDirectory() as tmp:
            cfg = Path(tmp) / "volkit.cfg"
            cfg.write_bytes("command = check\nworkbook = \u4e0a\u6d77/\u6ce2\u52a8\u7387.xlsx\n"
                            .encode("cp936"))
            before = list(paths.ENCODING_NOTES)
            paths.ENCODING_NOTES.clear()
            self.addCleanup(lambda: (paths.ENCODING_NOTES.clear(),
                                     paths.ENCODING_NOTES.extend(before)))
            real = paths.ansi_encoding
            paths.ansi_encoding = lambda: "cp936"     # stand in for that machine
            self.addCleanup(setattr, paths, "ansi_encoding", real)

            loaded = config.load(cfg)
            self.assertEqual(loaded.argv[0], "check")
            self.assertIn("\u4e0a\u6d77/\u6ce2\u52a8\u7387.xlsx", loaded.argv)
            self.assertTrue(any("cp936" in n for n in paths.ENCODING_NOTES),
                            paths.ENCODING_NOTES)

    def test_a_file_saved_as_notepad_unicode_is_read(self):
        """Notepad's 'Unicode' is UTF-16 with a byte order mark.  The mark is
        unmistakable, so it is read; UTF-16 *without* one is not guessed at,
        because that is how an ASCII file becomes Chinese."""
        from volkit import paths
        text, note = paths.decode_text("port = 8900\n".encode("utf-16"), "volkit.cfg")
        self.assertEqual(text.strip(), "port = 8900")
        self.assertIn("UTF-16", note)

    def test_a_file_in_no_encoding_at_all_is_refused_with_the_way_out(self):
        """Read it wrong or refuse it, but say what to do either way."""
        from volkit import paths
        real = paths.ansi_encoding
        paths.ansi_encoding = lambda: "utf-8"     # a Mac: there is no second reading
        self.addCleanup(setattr, paths, "ansi_encoding", real)
        with self.assertRaises(UnicodeDecodeError) as caught:
            paths.decode_text(b"port = \xc9\xcf\n", "volkit.cfg")
        self.assertIn("save it as UTF-8", str(caught.exception))

    def test_the_launcher_speaks_utf8_before_it_prints_anything(self):
        """A settings file may name a workbook under a path written in
        Chinese.  ``cli.main`` sets the streams, but the launcher prints the
        settings three lines before it gets there, and on a cp1252 stream that
        ended the packaged exe with a traceback before it had done anything at
        all.  Pinned by reading the source: the call has to come first, and a
        test that ran the launcher would have to own the process's streams."""
        import ast
        import launcher

        src = _source("launcher.py")
        tree = ast.parse(src)
        fn = next(n for n in tree.body
                  if isinstance(n, ast.FunctionDef) and n.name == "main")
        # Everything before the first print must not itself print, and the
        # stream call must be in there.
        calls = []
        for node in ast.walk(fn):
            if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
                calls.append((node.lineno, node.func.id))
        first_print = min((ln for ln, name in calls if name == "print"), default=10**9)
        set_streams = min((ln for ln, name in calls if name == "use_utf8_streams"),
                          default=10**9)
        self.assertLess(set_streams, first_print,
                        "the launcher prints before it sets the streams")

    def test_a_file_saved_with_a_byte_order_mark_still_reads(self):
        """Notepad and Excel both write one, and it is not part of the data.

        Left in place the mark becomes part of the first key of a settings
        file and of the first pair name of a feed -- a heading nobody typed
        and nothing matches.
        """
        import codecs
        import tempfile
        from volkit import config, paths

        with tempfile.TemporaryDirectory() as tmp:
            cfg = Path(tmp) / "volkit.cfg"
            cfg.write_bytes("\ufeffcommand = serve\nport = 8900\n".encode("utf-8"))
            self.assertEqual(config.load(cfg).argv[0], "serve")

            feed = Path(tmp) / "feed.csv"
            feed.write_bytes("\ufeffUSDJPY,SPOT,150.25\n".encode("utf-8"))
            self.assertIn("USDJPY", MarketFeed.load(feed))

            # And what the tool writes carries no mark of its own.  Only the
            # first bytes are asserted: Windows translates the line ending on
            # the way out and that is its business, while a mark it did not
            # ask for would be read back as part of the first field.
            out = Path(tmp) / "written.txt"
            paths.write_text(out, "USDJPY\n")
            self.assertFalse(out.read_bytes().startswith(codecs.BOM_UTF8))
            self.assertEqual(paths.read_text(out).splitlines(), ["USDJPY"])

    def test_a_file_that_is_not_utf8_names_itself(self):
        """A decoding failure that does not say which file it was is the
        swallowed error this project exists to remove."""
        import tempfile
        from volkit import paths

        # Pinned to a machine whose own code page *is* UTF-8, so there is no
        # second reading to fall back to.  On a cp1252 box this file would be
        # read as latin-1 and noted, which is the other half of the ladder and
        # is pinned by its own test above.
        real = paths.ansi_encoding
        paths.ansi_encoding = lambda: "utf-8"
        self.addCleanup(setattr, paths, "ansi_encoding", real)
        with tempfile.TemporaryDirectory() as tmp:
            bad = Path(tmp) / "cp1252.cfg"
            bad.write_bytes(b"note = caf\xe9\n")       # latin-1, not UTF-8
            with self.assertRaises(UnicodeDecodeError) as ctx:
                paths.read_text(bad)
            self.assertIn("cp1252.cfg", str(ctx.exception))

    def test_the_page_reads_as_utf8(self):
        """The one bundled file that actually carries non-ASCII."""
        from volkit import paths

        page = paths.read_text(self.ROOT / "volkit" / "web" / "index.html")
        self.assertIn("\u2014", page)          # an em dash, which is what broke
        self.assertTrue(page.rstrip().endswith("</html>"))


class TestAutoReload(unittest.TestCase):
    """Watching the market feed, which is off unless it is asked for.

    Everything here drives ``auto_check`` directly: the watcher thread does
    nothing else, so there is no timing in the test.
    """

    def setUp(self):
        import shutil, tempfile
        from volkit.webapp import BookService
        self.tmp = tempfile.TemporaryDirectory()
        d = Path(self.tmp.name)
        self.wb, self.feed = d / "vol_marks.xlsx", d / "feed.csv"
        shutil.copy(WORKBOOK, self.wb)
        shutil.copy(Path(__file__).resolve().parents[1] / "files" / "market_feed.csv", self.feed)
        self.service = BookService(str(self.wb), ASOF, feed_path=str(self.feed),
                                   auto_reload=1.0)
        self.assertIsNone(self.service.load_error)

    def tearDown(self):
        self.service.stop_watching()
        self.tmp.cleanup()

    def touch(self, path, ahead=5.0):
        import os, time
        when = time.time() + ahead
        os.utime(path, (when, when))

    def settle(self, path, ahead=5.0):
        """Change a file and give the watcher the two passes it waits for."""
        self.touch(path, ahead)
        first = self.service.auto_check()
        self.assertEqual(first, [], "a file is read once its write time has stopped moving")
        return self.service.auto_check()

    def test_nothing_happens_while_nothing_changes(self):
        self.assertEqual(self.service.auto_check(), [])
        self.assertEqual(self.service.auto_check(), [])
        self.assertEqual(self.service.auto_state()["seq"], 0)

    def test_a_rewritten_feed_is_re_read(self):
        text = self.feed.read_text(encoding="utf-8").replace("USDJPY,SPOT,150.25", "USDJPY,SPOT,151.25")
        self.assertNotEqual(text, self.feed.read_text(encoding="utf-8"))
        self.feed.write_text(text, encoding="utf-8")
        events = self.settle(self.feed)
        self.assertEqual([e["what"] for e in events], ["feed"])
        self.assertTrue(events[0]["ok"])
        self.assertAlmostEqual(self.service.book.feed.pairs["USDJPY"].spot, 151.25)
        self.assertFalse(self.service.feed_state()["stale"])

    def test_a_file_still_being_written_is_left_for_the_next_pass(self):
        """One pass sees the write time move; the pass after reads it."""
        self.touch(self.feed, 5.0)
        self.assertEqual(self.service.auto_check(), [])
        self.touch(self.feed, 9.0)                  # still being written
        self.assertEqual(self.service.auto_check(), [])
        self.assertEqual(len(self.service.auto_check()), 1)

    def test_only_the_feed_is_watched(self):
        """The workbook and the historical sheet are deliberately not.

        Re-reading the workbook discards every mark this session has made
        (nothing writes to the workbook), and a historical sheet is a record
        of what happened rather than a market.  Both stay on their buttons;
        the feed is a publication and is the only file worth chasing.
        """
        self.assertEqual([w["what"] for w in self.service.auto_state()["watching"]], ["feed"])
        pair = self.service.book.pairs[0]
        self.service.overwrite({"pair": pair, "kind": "atm", "tenor": "1m", "value": 9.5})
        self.touch(self.wb)
        self.assertEqual(self.service.auto_check(), [])
        self.assertEqual(self.service.auto_check(), [])
        # ... and the mark this session made is exactly where it was put.
        row = next(r for r in self.service.marks({"pair": pair, "cut": "NY"})["atm"]
                   if r["tenor"] == "1m")
        self.assertAlmostEqual(row["overwrite"], 0.095, places=12)
        self.assertEqual(self.service.auto_state()["seq"], 0)

    def test_a_watcher_that_is_off_is_reported_and_does_nothing(self):
        from volkit.webapp import BookService
        quiet = BookService(str(self.wb), ASOF, feed_path=str(self.feed))
        state = quiet.auto_state()
        self.assertFalse(state["enabled"])
        self.assertFalse(quiet.start_watching())
        self.assertEqual([w["what"] for w in state["watching"]], ["feed"])
        self.touch(self.feed)
        # Off means the loop never runs; a check driven by hand still works,
        # which is what the "check the feed now" button does.
        self.assertEqual(quiet.auto_check(settle=False)[0]["what"], "feed")

    def test_the_switch_turns_the_watcher_on_and_off(self):
        """The pricing screen's checkbox, which posts to the same method."""
        from volkit.webapp import BookService
        quiet = BookService(str(self.wb), ASOF, feed_path=str(self.feed))
        self.addCleanup(quiet.stop_watching)
        self.assertFalse(quiet.auto_state()["enabled"])
        self.assertTrue(quiet.auto_state()["available"])
        state = quiet.set_auto({"enabled": True, "interval": 3})
        self.assertTrue(state["enabled"])
        self.assertEqual(state["interval"], 3)
        self.assertIsNotNone(quiet._watcher)
        # Off again, and the thread really stops rather than the flag alone.
        state = quiet.set_auto({"enabled": False})
        self.assertFalse(state["enabled"])
        self.assertIsNone(quiet._watcher)
        # The interval it was given survives being switched off and on.
        self.assertEqual(quiet.set_auto({"enabled": True})["interval"], 3)
        with self.assertRaises(ValueError):
            quiet.set_auto({"interval": 0})

    def test_no_feed_file_means_there_is_nothing_to_auto_load(self):
        """A switch that can be turned on and then does nothing is worse than
        one that says why it is greyed out."""
        from volkit.webapp import BookService
        quiet = BookService(str(self.wb), ASOF)
        self.assertFalse(quiet.auto_state()["available"])
        self.assertEqual(quiet.auto_state()["watching"], [])

    def test_the_sequence_number_moves_only_when_something_happened(self):
        before = self.service.auto_state()["seq"]
        self.assertEqual(self.service.auto_check(), [])
        self.assertEqual(self.service.auto_state()["seq"], before)
        self.settle(self.feed)
        self.assertGreater(self.service.auto_state()["seq"], before)


class TestListedPositions(unittest.TestCase):
    """Positions on the exchange-traded screen, and the risk they aggregate to.

    Everything the greeks need comes off the panel a position belongs to, so
    these build one or two panels and price against them exactly as the screen
    does -- the panels are posted whole and the server keeps none of it.
    """

    EXPIRY = "2024-06-14 19:00"
    LATER = "2024-09-13 19:00"
    CLOCK = Clock(datetime(2024, 5, 28, 12, 0, tzinfo=UTC))
    TABLE = ("1.0300\t8.90\n1.0600\t8.10\n1.0900\t7.60\n1.1200\t7.80\n1.1500\t8.60\n")

    def panel(self, code="6E", expiry=None, forward=1.09, **kw):
        return dict({"underlying": code, "expiry": expiry or self.EXPIRY,
                     "forward": forward, "text": self.TABLE}, **kw)

    def agg(self, text, panels=None, **kw):
        return listed.positions_from_request(dict(
            {"text": text, "panels": panels or [self.panel()]}, **kw)).run(clock=self.CLOCK)

    # -- the paste -------------------------------------------------------
    def test_the_layout_is_decided_once_from_the_whole_table(self):
        """Reading each row on its own width would move a quantity into a
        strike the first time somebody left a cell blank."""
        parsed = listed.parse_positions(
            "6E, 2024-06-14 19:00, 1.09, C, 25\n"
            "6E, 2024-06-14 19:00, 1.06, P, -40\n"
            "6E, 1.12, C, -15\n")
        self.assertEqual(parsed.layout,
                         ("contract", "expiry", "strike", "type", "quantity"))
        self.assertEqual(len(parsed.positions), 2)
        self.assertEqual([n for n, _, _ in parsed.skipped], [3])
        self.assertIn("4 columns", parsed.skipped[0][2])

    def test_the_short_layouts_leave_the_panel_to_be_worked_out(self):
        parsed = listed.parse_positions("1.09 C 25\n1.06 P -40\n")
        self.assertEqual(parsed.layout, ("strike", "type", "quantity"))
        self.assertEqual(parsed.positions[0].underlying, "")
        self.assertEqual(parsed.positions[0].expiry, "")
        self.assertTrue(parsed.positions[0].is_call)
        self.assertEqual(parsed.positions[1].quantity, -40.0)

    def test_a_header_row_may_name_the_columns_in_any_order(self):
        parsed = listed.parse_positions(
            "Qty\tRight\tStrike\n25\tCall\t1.09\n-40\tPut\t1.06\n")
        self.assertEqual(parsed.positions[0].strike, 1.09)
        self.assertEqual(parsed.positions[0].quantity, 25.0)
        self.assertFalse(parsed.positions[1].is_call)
        self.assertTrue(any("header row read" in n for n in parsed.notes))

    def test_atm_is_a_strike_and_a_bad_cell_keeps_its_line(self):
        parsed = listed.parse_positions("ATM C 10\n1.09 X 5\nabc P 3\n1.06 P two\n")
        self.assertEqual(len(parsed.positions), 1)
        self.assertIsNone(parsed.positions[0].strike)
        self.assertEqual([n for n, _, _ in parsed.skipped], [2, 3, 4])

    def test_a_thousands_separator_in_a_comma_paste_is_a_column(self):
        """A comma is a column boundary here as it is in a broker run, so the
        line is refused rather than read as a size of 1."""
        parsed = listed.parse_positions(
            "6E, 2024-06-14 19:00, 1.09, C, 1,000\n6E, 2024-06-14 19:00, 1.06, P, -40\n")
        self.assertEqual([n for n, _, _ in parsed.skipped], [1])
        # ... and with tabs there is no ambiguity and it is a size.
        parsed = listed.parse_positions("1.09\tC\t1,000\n")
        self.assertEqual(parsed.positions[0].quantity, 1000.0)

    # -- the greeks ------------------------------------------------------
    def test_both_columns_price_the_same_option_and_differ_only_in_the_greeks(self):
        r = self.agg("1.09 C 25\n1.06 P -40\n")
        for row in r["positions"]:
            self.assertEqual(row["error"], "")
            self.assertIsNotNone(row["premium"])
            # One volatility at one strike, so the premium cannot depend on
            # which set of sensitivities is being taken around it.
            self.assertNotEqual(row["bs"]["delta_futures"], row["smile"]["delta_futures"])
        self.assertAlmostEqual(r["totals"]["bs"]["vega"], r["totals"]["bs"]["vega"])

    def test_at_the_forward_the_smile_vega_is_the_black_scholes_vega(self):
        """The curve is lifted by a move measured at the forward, so an option
        struck there sees exactly that move and the two must agree."""
        r = self.agg("ATM C 10\n")
        row = r["positions"][0]
        self.assertAlmostEqual(row["strike"], 1.09, places=12)
        self.assertAlmostEqual(row["smile"]["vega"] / row["bs"]["vega"], 1.0, places=5)

    def test_a_call_and_a_put_at_one_strike_differ_by_exactly_one_delta(self):
        """Put/call parity, which the aggregate has to reproduce or a straddle
        is not a straddle.  Note it is *not* zero at the forward -- the
        delta-neutral strike is F exp(sigma^2 t / 2), not F -- so pinning it
        at zero would be pinning a mistake."""
        r = self.agg("ATM C 10\nATM P 10\n")
        call, put = r["positions"]
        g = r["groups"][0]
        self.assertAlmostEqual(call["bs"]["delta_futures"] / 10.0
                               - put["bs"]["delta_futures"] / 10.0, 1.0, places=12)
        self.assertAlmostEqual(g["bs"]["vega"], 2 * call["bs"]["vega"], places=8)
        # Long options decay: theta is money and is negative on both sides.
        self.assertLess(g["bs"]["theta"], 0.0)
        self.assertLess(g["smile"]["theta"], 0.0)

    def test_money_scales_with_the_contract_size_and_a_count_does_not(self):
        big = self.agg("1.09 C 25\n")["positions"][0]
        small = self.agg("1.09 C 25\n", [self.panel(contract_size=62_500)])["positions"][0]
        self.assertEqual(big["contract_size"], 125_000)
        self.assertAlmostEqual(big["vega"] if False else big["bs"]["vega"],
                               2 * small["bs"]["vega"], places=6)
        self.assertAlmostEqual(big["bs"]["delta_futures"],
                               small["bs"]["delta_futures"], places=12)

    def test_the_vol_bump_and_the_theta_window_scale_what_they_say_they_do(self):
        one = self.agg("1.09 C 25\n")["positions"][0]
        two = self.agg("1.09 C 25\n", vol_bump=2, theta_days=3)["positions"][0]
        self.assertAlmostEqual(two["bs"]["vega"], 2 * one["bs"]["vega"], places=8)
        self.assertAlmostEqual(two["bs"]["volga"], 4 * one["bs"]["volga"], places=8)
        self.assertAlmostEqual(two["bs"]["theta"], 3 * one["bs"]["theta"], places=8)

    # -- matching a position to a panel ----------------------------------
    def test_a_line_that_matches_no_panel_keeps_its_place_with_the_reason(self):
        r = self.agg("6J, 2024-06-14 19:00, 1.09, C, 25\n6E, 2024-06-14 19:00, 1.06, P, -40\n")
        self.assertEqual(len(r["positions"]), 2)
        self.assertIn("6J", r["positions"][0]["error"])
        self.assertEqual(r["positions"][1]["error"], "")
        self.assertTrue(any("could not be priced" in w for w in r["warnings"]))

    def test_a_line_that_matches_two_panels_is_refused_rather_than_guessed(self):
        """A position priced against the wrong month's curve looks perfectly
        ordinary, which is why this may never be guessed."""
        panels = [self.panel(), self.panel(expiry=self.LATER, forward=1.10)]
        r = self.agg("1.09 C 25\n", panels)
        self.assertIn("matches 2 panels", r["positions"][0]["error"])
        # Naming the expiry settles it.
        r = self.agg(f"6E, {self.LATER}, 1.09, C, 25\n", panels)
        self.assertEqual(r["positions"][0]["error"], "")
        self.assertEqual(r["positions"][0]["expiry"][:16], "2024-09-13T19:00")

    def test_a_panel_that_will_not_fit_does_not_empty_the_rest(self):
        panels = [self.panel(label="good"),
                  dict(self.panel(label="bad", expiry=self.LATER), text="nonsense")]
        r = self.agg("good, , 1.09, C, 25\nbad, , 1.09, C, 25\n", panels)
        self.assertEqual(r["positions"][0]["error"], "")
        self.assertIn("bad", r["positions"][1]["error"])
        self.assertEqual([p["ok"] for p in r["panels"]], [True, False])

    def test_money_totals_across_contracts_but_a_futures_count_does_not(self):
        """A euro future is not a yen future.  Summing the two would be a
        number with no meaning printed where a risk figure goes."""
        panels = [self.panel(),
                  self.panel(code="6J", forward=0.00645, scale=1,
                             expiry=self.EXPIRY)]
        panels[1]["text"] = ("0.00610\t9.40\n0.00628\t9.00\n0.00645\t8.70\n"
                             "0.00662\t8.85\n0.00680\t9.30\n")
        r = self.agg("6E, , 1.09, C, 25\n6J, , 0.00645, C, 10\n", panels)
        self.assertEqual([p["ok"] for p in r["panels"]], [True, True])
        self.assertEqual(len(r["groups"]), 2)
        self.assertNotIn("delta_futures", r["totals"]["bs"])
        self.assertAlmostEqual(
            r["totals"]["bs"]["vega"],
            sum(g["bs"]["vega"] for g in r["groups"]), places=6)
        self.assertTrue(any("not a future of another" in n for n in r["notes"]))
        # Both settle in US dollars, so there is one money total and it is the
        # all-in one.
        self.assertEqual([c["ccy"] for c in r["currencies"]], ["USD"])
        self.assertAlmostEqual(r["currencies"][0]["bs"]["vega"],
                               r["totals"]["bs"]["vega"], places=12)

    def test_every_column_adds_within_one_contract_across_its_expiries(self):
        """The screen aggregated per *panel* and then jumped to money-only,
        so a book of one contract over four expiries had no futures-equivalent
        delta anywhere -- the one number a desk asks for when it asks how much
        6E it is running.  §8 said "totalled per contract"; the code totalled
        per panel, and the two only coincide with one panel per contract."""
        panels = [self.panel(), self.panel(expiry=self.LATER, forward=1.10)]
        r = self.agg(f"6E, {self.EXPIRY}, 1.09, C, 25\n6E, {self.LATER}, 1.09, P, -10\n",
                     panels)
        self.assertEqual([p["error"] for p in r["positions"]], ["", ""])
        self.assertEqual(len(r["contracts"]), 1)
        con = r["contracts"][0]
        self.assertEqual((con["underlying"], con["panels"], con["n"]), ("6E", 2, 2))
        self.assertEqual(len(con["expiries"]), 2)
        for which in ("bs", "smile"):
            for key in ("delta_futures", "gamma_futures", "vega", "theta", "premium"):
                got = con[which][key] if key != "premium" else con["premium"]
                want = sum((g[which][key] if key != "premium" else g["premium"])
                           for g in r["groups"])
                self.assertAlmostEqual(got, want, places=9, msg=f"{which}.{key}")
        # And the note says what that futures total is and is not.
        self.assertTrue(any("not the same future" in n for n in r["notes"]))

    def test_money_is_not_totalled_across_two_settlement_currencies(self):
        """A sum of euros and dollars is not a number.  It was unreachable
        while the contract came off a list of CME codes, every one of which
        settles in dollars; a typed contract makes it reachable."""
        panels = [self.panel(code="XA", pair="EURUSD", contract_size=125_000),
                  self.panel(code="XB", pair="EURGBP", contract_size=125_000)]
        r = self.agg("XA, , 1.09, C, 25\nXB, , 1.09, P, -10\n", panels)
        self.assertEqual([p["error"] for p in r["positions"]], ["", ""])
        self.assertEqual([c["ccy"] for c in r["currencies"]], ["GBP", "USD"])
        self.assertIsNone(r["totals"]["premium"])
        self.assertIsNone(r["totals"]["bs"]["vega"])
        self.assertTrue(any("no all-in money total" in w for w in r["warnings"]))
        # The per-currency rows still hold every figure.
        byccy = {c["ccy"]: c for c in r["currencies"]}
        self.assertAlmostEqual(byccy["USD"]["premium"], r["groups"][0]["premium"], places=9)
        self.assertAlmostEqual(byccy["GBP"]["premium"], r["groups"][1]["premium"], places=9)

    def test_two_typed_contracts_on_one_screen_can_be_told_apart(self):
        """The old bug, and the reason the contract box is free text: with a
        dropdown, every contract missing from it was CUSTOM, both panels were
        called CUSTOM, and a position line naming one was refused as matching
        two panels -- with no field left that could settle it."""
        panels = [self.panel(code="XA", pair="EURUSD", contract_size=125_000),
                  self.panel(code="XB", pair="EURUSD", contract_size=125_000)]
        r = self.agg("XA, , 1.09, C, 25\nXB, , 1.09, P, -10\n", panels)
        self.assertEqual([p["error"] for p in r["positions"]], ["", ""])
        self.assertEqual([g["underlying"] for g in r["groups"]], ["XA", "XB"])
        self.assertEqual(sorted(c["underlying"] for c in r["contracts"]), ["XA", "XB"])
        self.assertFalse(any(c["known"] for c in r["contracts"]))
        # Two panels that really are the same thing are still refused, and the
        # refusal now names the one field that is always free to differ.
        same = self.agg("1.09 C 25\n", [self.panel(), self.panel()])
        self.assertIn("label", same["positions"][0]["error"])

    def test_a_custom_contract_with_no_size_says_so_rather_than_using_one(self):
        """The money columns are then per one unit of the base currency, which
        is a perfectly good number and a terrible one to read as dollars."""
        r = self.agg("1.09 C 25\n", [self.panel(code="CUSTOM", pair="EURUSD")])
        self.assertEqual(r["positions"][0]["contract_size"], 1.0)
        self.assertTrue(any("no contract size" in w for w in r["warnings"]))
        self.assertTrue(any("per one unit" in w for w in r["warnings"]))

    def test_the_bumps_are_refused_rather_than_silently_ignored(self):
        for bad in ({"vol_bump": -1}, {"theta_days": -3}):
            with self.assertRaises(ValueError):
                self.agg("1.09 C 25\n", **bad)

    def test_structured_rows_refuse_an_unreadable_call_put(self):
        """A short put booked as a long call is not a rounding error, so the
        already-read form does not default the side either."""
        good = listed.positions_from_request({
            "positions": [{"strike": "ATM", "type": "P", "quantity": -5}],
            "panels": [self.panel()]}).run(clock=self.CLOCK)
        self.assertFalse(good["positions"][0]["type"] == "call")
        self.assertAlmostEqual(good["positions"][0]["strike"], 1.09, places=12)
        with self.assertRaises(ValueError):
            listed.positions_from_request({"positions": [{"strike": 1.09, "quantity": 5}]})

    def test_a_theta_window_past_the_expiry_is_blank_with_the_reason(self):
        """There is no revaluation to take the decay from, and a plausible
        number in its place would be the silent zero this project removes."""
        r = self.agg("1.09 C 25\n", theta_days=90)
        row = r["positions"][0]
        self.assertIsNone(row["smile"]["theta"])
        self.assertIn("reaches past this expiry", row["error"])
        # The Black-Scholes column is closed form and is still reported.
        self.assertIsNotNone(row["bs"]["theta"])


class TestListedClock(unittest.TestCase):
    def test_a_panel_with_no_pair_takes_the_book_s_clock(self):
        """The Exchange-traded screen refused to fit a CUSTOM contract.

        The clock was looked for on the mapped surface and nowhere else, so a
        contract with no pair -- CUSTOM, or one whose pair is not in this
        workbook -- reported "a clock is required" on a screen holding a
        perfectly good one.  Only the *comparison* needs a surface.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        panel = listed.Panel(
            underlying=listed.resolve_underlying("CUSTOM"),
            expiry="2024-06-14 19:00", forward=100.0,
            quotes=tuple(listed.Quote(strike=k, vol=v) for k, v in
                         ((95.0, 0.11), (100.0, 0.10), (105.0, 0.105))),
        )
        out = panel.run(book)
        self.assertEqual(out["valuation"], ASOF.now.isoformat())
        self.assertIsNone(out["comparison"])
        self.assertGreater(out["years"], 0)


class TestCrossVegaSplit(unittest.TestCase):
    """Where a cross's at-the-money vega actually sits."""

    def rows(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["AUDJPY"])
        return book, analytics.triangle_table(book, "AUDJPY", cut="NY", with_noise=False,
                                              tenors=["3m"])

    def test_the_split_is_the_derivative_of_the_variance_triangle(self):
        book, rows = self.rows()
        r = rows[0]
        va, vb = r.leg_atm
        ca, cb = r.coefficients
        x = ca * cb * r.rho
        sigma = r.variance_triangle_atm
        self.assertAlmostEqual(r.leg_vega[0], (va + x * vb) / sigma)
        self.assertAlmostEqual(r.leg_vega[1], (vb + x * va) / sigma)
        self.assertAlmostEqual(r.rho_vega, ca * cb * va * vb / sigma)

    def test_a_bump_in_a_leg_moves_the_cross_by_the_split(self):
        """The number is a hedge ratio, so it is checked against a real bump."""
        book, rows = self.rows()
        r = rows[0]
        va, vb = r.leg_atm
        ca, cb = r.coefficients
        h = 1e-6

        def triangle(a, b):
            return math.sqrt(a * a + b * b + 2.0 * ca * cb * r.rho * a * b)

        self.assertAlmostEqual((triangle(va + h, vb) - triangle(va - h, vb)) / (2 * h),
                               r.leg_vega[0], places=6)
        self.assertAlmostEqual((triangle(va, vb + h) - triangle(va, vb - h)) / (2 * h),
                               r.leg_vega[1], places=6)

    def test_the_two_hedges_satisfy_euler_rather_than_adding_to_one(self):
        """The ratios are hedges, not shares.

        Reading them as a split of something into parts is the mistake: they
        do not add to one.  What is exact is Euler's identity -- the triangle
        is homogeneous of degree one in the two leg volatilities, so weighting
        each ratio by its own leg accounts for the whole of the cross's.
        """
        from volkit.analytics import _vega_split
        for va, vb, rho in ((0.10, 0.10, 0.30), (0.07, 0.13, -0.60), (0.09, 0.11, 0.85)):
            with self.subTest(rho=rho):
                sigma = math.sqrt(va * va + vb * vb + 2 * rho * va * vb)
                da, db, drho = _vega_split(va, vb, rho, 1, 1, sigma)
                self.assertAlmostEqual(va * da + vb * db, sigma)
                self.assertNotAlmostEqual(da + db, 1.0)
                # The correlation term is degree zero and is not in the identity.
                self.assertAlmostEqual(drho, va * vb / sigma)

    def test_the_split_matches_the_triangle_the_book_is_built_on(self):
        """Euler again, on the book's own marks rather than on made-up ones."""
        book, rows = self.rows()
        r = rows[0]
        va, vb = r.leg_atm
        self.assertAlmostEqual(va * r.leg_vega[0] + vb * r.leg_vega[1],
                               r.variance_triangle_atm)

    def test_a_zero_cross_volatility_has_no_hedge_ratio_rather_than_an_infinity(self):
        from volkit.analytics import _vega_split
        out = _vega_split(0.1, 0.1, -1.0, 1, 1, 0.0)
        self.assertTrue(all(v != v for v in out))

    def test_the_row_that_could_not_be_built_carries_no_split_either(self):
        """A failed row keeps its place; it must not carry a made-up ratio."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["AUDJPY"])
        rows = analytics.triangle_table(book, "AUDJPY", cut="NY", with_noise=False,
                                        tenors=["3m"])
        self.assertTrue(all(v == v for v in rows[0].leg_vega))


class TestDeterminism(unittest.TestCase):
    def test_same_clock_gives_identical_results(self):
        """Legacy read datetime.utcnow() inside the model on every call."""
        a = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        b = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        e = datetime(2024, 5, 28, tzinfo=UTC)
        self.assertEqual(float(a["USDJPY"].vol(1.02, e)), float(b["USDJPY"].vol(1.02, e)))

    def test_different_clocks_give_different_results(self):
        other = Clock(datetime(2024, 3, 15, 12, tzinfo=UTC))
        a = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        b = Book.from_excel(WORKBOOK, other).load_all(["USDJPY"])
        e = datetime(2024, 5, 28, tzinfo=UTC)
        self.assertNotEqual(float(a["USDJPY"].vol(1.02, e)), float(b["USDJPY"].vol(1.02, e)))


class TestScreens(unittest.TestCase):
    """Building without a screen.

    A build can be made without some of the five tabs (build_exe.py
    --exclude-tab).  What is pinned here is that an excluded screen is really
    gone and says so: the tab is not offered, its routes are refused by name,
    and its subcommands are not registered.  A screen that merely disappeared
    from the page while its routes kept answering would be the same silent
    half-measure as a swallowed error.
    """

    def setUp(self):
        from volkit import screens
        self.screens = screens
        screens.enabled.cache_clear()
        self.addCleanup(screens.enabled.cache_clear)

    def _select(self, names):
        """Run the rest of the test as a build with only *names*."""
        import os
        old = os.environ.get(self.screens.ENV_VAR)
        os.environ[self.screens.ENV_VAR] = ",".join(names)
        self.screens.enabled.cache_clear()

        def restore():
            if old is None:
                os.environ.pop(self.screens.ENV_VAR, None)
            else:
                os.environ[self.screens.ENV_VAR] = old
            self.screens.enabled.cache_clear()

        self.addCleanup(restore)

    def test_a_source_tree_has_every_screen(self):
        self.assertEqual(self.screens.enabled(), self.screens.ALL)
        self.assertEqual(self.screens.excluded(), ())
        self.assertEqual(self.screens.summary(), "")

    def test_the_manifest_in_the_bundle_beats_the_environment(self):
        """The manifest is the build's own decision.

        An environment variable that could put a screen back would make the
        exclusion a suggestion; it is meant to be a property of the build.
        """
        import os
        import tempfile
        from volkit import paths
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self.screens.write_manifest(root / "volkit" / "data", ["pricing", "mm"])
            os.environ[self.screens.ENV_VAR] = "analysis"
            self.addCleanup(os.environ.pop, self.screens.ENV_VAR, None)
            real = paths.resource_dir
            paths.resource_dir = lambda: root
            self.addCleanup(setattr, paths, "resource_dir", real)
            self.screens.enabled.cache_clear()
            self.assertEqual(self.screens.enabled(), ("pricing", "mm"))

    def test_a_misspelled_screen_is_refused_not_ignored(self):
        with self.assertRaises(self.screens.ScreenError) as ctx:
            self.screens.parse_names("pricing, markign")
        self.assertIn("markign", str(ctx.exception))
        with self.assertRaises(self.screens.ScreenError):
            self.screens.parse_names("   ")

    def test_the_manifest_round_trips_and_keeps_tab_order(self):
        text = self.screens.manifest_text(["mm", "pricing"])
        self.assertEqual(self.screens.parse_names(text, source="t"), ("pricing", "mm"))
        self.assertIn("Excluded:", text)

    def test_every_route_belongs_to_at_most_one_screen(self):
        seen = set()
        for screen in self.screens.SCREENS:
            for route in screen.routes:
                self.assertNotIn(route, seen, f"{route} is claimed twice")
                seen.add(route)

    def test_an_excluded_screen_refuses_its_routes_by_name(self):
        self._select(["pricing", "marking"])
        msg = self.screens.route_refusal("/api/mm/fit")
        self.assertIsNotNone(msg)
        self.assertIn("Market maker", msg)
        # Both agents live on this tab and leave with it: the quoting agent's
        # card and the marking agent's, routes and command alike.
        self.assertIsNotNone(self.screens.route_refusal("/api/mm/mark"))
        self.assertIsNotNone(self.screens.route_refusal("/api/mm/mark/record"))
        self.assertEqual(self.screens.command_screen("mark"), "mm")
        # The shell and the screens that stayed are untouched.
        self.assertIsNone(self.screens.route_refusal("/api/price"))
        self.assertIsNone(self.screens.route_refusal("/api/state"))
        self.assertIsNone(self.screens.route_refusal("/api/reload"))

    def test_the_vol_query_route_leaves_with_the_marking_screen(self):
        """The card sits on that tab, so `/api/vol` is that tab's route.

        It was `/api/calc` and belonged to nobody, which in a build made
        without the marking screen left the one endpoint of a card that was
        no longer there still answering.
        """
        owner = {r: sc.name for sc in self.screens.SCREENS for r in sc.routes}
        self.assertEqual(owner.get("/api/vol"), "marking")
        self._select(["pricing"])
        msg = self.screens.route_refusal("/api/vol")
        self.assertIsNotNone(msg)
        self.assertIn("Vol marking", msg)

    def test_the_server_turns_an_excluded_route_away_with_404(self):
        from volkit import webapp
        self._select(["pricing"])

        class FakeHandler(webapp.Handler):
            def __init__(self):  # no socket, no request
                self.sent = None

            def _json(self, payload, code=200):
                self.sent = (code, payload)

        h = FakeHandler()
        h.path = "/api/analysis?pair=EURJPY"
        h.do_GET()
        self.assertEqual(h.sent[0], 404)
        self.assertIn("Analysis", h.sent[1]["error"])

    def test_the_state_response_tells_the_page_which_screens_it_has(self):
        from volkit.webapp import BookService
        self._select(["pricing", "marking"])
        state = BookService(str(WORKBOOK), ASOF).state()
        self.assertEqual(state["screens"], ["pricing", "marking"])

    def test_an_excluded_screen_loses_its_subcommands(self):
        from volkit.cli import build_parser
        import argparse as _argparse
        self._select(["pricing", "marking"])
        names = set()
        for action in build_parser()._actions:
            if isinstance(action, _argparse._SubParsersAction):
                names.update(action.choices)
        self.assertNotIn("mm", names)
        self.assertNotIn("analysis", names)
        self.assertNotIn("listed", names)
        # The shell commands and the screens that stayed are untouched.
        for kept in ("check", "serve", "tenors", "vol"):
            self.assertIn(kept, names)

    def test_an_excluded_subcommand_says_why_rather_than_invalid_choice(self):
        """argparse would answer 'invalid choice', which is not what happened."""
        from volkit import cli
        self._select(["pricing"])
        self.assertEqual(cli._excluded_request(["mm", "EURUSD"]), "mm")
        # The global options may come first, and take a value.
        self.assertEqual(cli._excluded_request(["-w", "book.xlsx", "listed"]), "listed")
        self.assertEqual(cli._excluded_request(["--asof", "2024-01-01", "vol"]), None)
        # A subcommand's own arguments are never read as a subcommand: only
        # the first positional is inspected.
        self.assertEqual(cli._excluded_request(["vol", "mm", "1M", "1.0"]), None)

    def _hidden_build(self, visible, shy):
        """Run the rest of the test as a build with *shy* hidden."""
        import os
        old = os.environ.get(self.screens.ENV_VAR)
        os.environ[self.screens.ENV_VAR] = ", ".join(
            list(visible) + [f"{n} hidden" for n in shy])
        self.screens.deactivate_all()

        def restore():
            if old is None:
                os.environ.pop(self.screens.ENV_VAR, None)
            else:
                os.environ[self.screens.ENV_VAR] = old
            self.screens.deactivate_all()

        self.addCleanup(restore)

    def test_a_hidden_screen_is_off_until_it_is_asked_for(self):
        """The third state.  Off, it is turned away exactly like an excluded
        screen; the only difference is the sentence, and the sentence is the
        whole point -- one of them can be had by starting the tool
        differently."""
        self._hidden_build(["pricing", "marking"], ["analysis"])
        self.assertEqual(self.screens.built(),
                         ("pricing", "marking", "analysis"))
        self.assertEqual(self.screens.hidden(), ("analysis",))
        self.assertEqual(self.screens.enabled(), ("pricing", "marking"))
        msg = self.screens.route_refusal("/api/analysis")
        self.assertIn("--enable-tab analysis", msg)
        self.assertIn("volkit.cfg", msg)

        self.assertEqual(self.screens.activate(["analysis"]), ("analysis",))
        self.assertEqual(self.screens.enabled(), ("pricing", "marking", "analysis"))
        self.assertIsNone(self.screens.route_refusal("/api/analysis"))
        # Asking twice, or for a screen already showing, is not an error.
        self.assertEqual(self.screens.activate(["analysis", "pricing"]), ())

    def test_an_excluded_screen_cannot_be_switched_on(self):
        """Otherwise a build could be talked out of its own decision."""
        self._hidden_build(["pricing"], ["marking"])
        with self.assertRaises(self.screens.ScreenError) as ctx:
            self.screens.activate(["mm"])
        self.assertIn("excluded from this build", str(ctx.exception))
        self.assertEqual(self.screens.enabled(), ("pricing",))

    def test_a_hidden_screens_subcommand_is_registered_only_once_it_is_on(self):
        from volkit.cli import build_parser
        import argparse as _argparse

        def names():
            got = set()
            for action in build_parser()._actions:
                if isinstance(action, _argparse._SubParsersAction):
                    got.update(action.choices)
            return got

        self._hidden_build(["pricing", "marking"], ["mm"])
        self.assertNotIn("mm", names())
        self.screens.activate(["mm"])
        self.assertIn("mm", names())

    def test_the_command_line_switch_is_read_before_the_parser_is_built(self):
        """The flag has to change the parser that would otherwise reject it."""
        from volkit import cli
        self._hidden_build(["pricing", "marking"], ["analysis"])
        self.assertEqual(cli._requested_screens(["--enable-tab", "analysis", "analysis",
                                                 "EURUSD"]), ["analysis"])
        self.assertEqual(cli._requested_screens(["--enable-tab=mm"]), ["mm"])
        self.assertEqual(cli._excluded_request(["analysis", "EURUSD"]), "analysis")
        self.screens.activate(["analysis"])
        self.assertIsNone(cli._excluded_request(["analysis", "EURUSD"]))

    def test_a_selection_that_hides_everything_is_refused(self):
        """A build needs at least one tab that shows without a switch."""
        with self.assertRaises(self.screens.ScreenError):
            self.screens.parse_selection("pricing hidden, mm hidden", source="t")
        with self.assertRaises(self.screens.ScreenError):
            self.screens.parse_selection("pricing, pricing hidden", source="t")
        with self.assertRaises(self.screens.ScreenError):
            self.screens.parse_selection("pricing invisible", source="t")

    def test_the_page_hides_the_screens_it_was_not_given(self):
        """The tab, the panel and the boot work all key off the same list."""
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        self.assertIn("STATE.screens", js)
        for name in self.screens.ALL:
            self.assertIn(f"{name}:'#{self.screens.BY_NAME[name].panel}'", js)

    def test_a_pricing_leg_is_removed_from_its_column_header_only(self):
        """One delete per column, in the header; the Remove row is gone.

        The row of Remove buttons sat between the inputs and the results, so
        the click that removed a leg was one row away from the fields being
        typed into.  The header cross does the same job out of the way of the
        grid.  Pinned in both directions because the header and the rows are
        painted by the same function, and either could come back as a one-line
        change nobody would notice.
        """
        html = _source("volkit", "web", "index.html")
        js = html.split("<script>")[1].split("</script>")[0]
        grid = js.split("function renderGrid(")[1].split("\nfunction ")[0]
        head, body = grid.split("<tr class=\"sec\">", 1)
        self.assertIn("data-del", head)                  # the header cross
        self.assertNotIn("data-del", body)               # and nothing in the grid
        self.assertNotIn("rmcol", js)
        self.assertIn("class=\"x\"", head)
        self.assertIn(".colhead .x", html)               # it is still styled
        self.assertIn('id="delleg"', html)               # - Remove last
        self.assertIn('id="clearlegs"', html)            # Clear all


class TestPackageImport(unittest.TestCase):
    """The package has to be importable before its dependencies exist."""

    ROOT = Path(__file__).resolve().parents[1]

    def test_reading_the_screen_list_does_not_import_numpy(self):
        """build_exe.py asks volkit.screens what to build, and it does that
        *before* installing numpy -- it is the thing that installs numpy.

        An eager ``from .atm import AtmCurve`` in volkit/__init__.py dragged
        the whole numeric stack in behind ``from volkit import screens``, so
        the Windows build died at ``import numpy`` on atm.py before printing
        its first line. Nothing in screens, paths or config needs it.
        """
        import subprocess
        probe = ("import volkit.screens, volkit.paths, volkit.config, sys; "
                 "print(sorted(m for m in ('numpy', 'scipy', 'pandas') if m in sys.modules))")
        out = subprocess.run([sys.executable, "-c", probe], cwd=str(self.ROOT),
                             capture_output=True, text=True,
                             encoding="utf-8", errors="replace")
        self.assertEqual(out.returncode, 0, out.stderr)
        self.assertEqual(out.stdout.strip(), "[]")

    def test_every_documented_name_still_resolves(self):
        """Lazy binding must not quietly drop a name from the public API."""
        import volkit
        for name in volkit.__all__:
            with self.subTest(name):
                self.assertTrue(hasattr(volkit, name))
        from volkit import Book, Clock            # the README's own first line
        self.assertTrue(callable(Book.from_excel))
        self.assertTrue(callable(Clock.utcnow))
        self.assertIn("Book", dir(volkit))
        with self.assertRaises(AttributeError):
            getattr(volkit, "NotAThing")


class TestPackaging(unittest.TestCase):
    """The Windows build.

    None of this builds anything -- PyInstaller takes minutes and cannot
    cross-compile anyway.  What it pins is the part of the build that can go
    wrong silently in the source tree: an input the spec never picks up, or a
    launcher that stops recognising a subcommand.
    """

    ROOT = Path(__file__).resolve().parents[1]

    def test_launcher_knows_every_subcommand(self):
        """A hardcoded list here went stale when 'analysis' and 'listed' landed.

        The launcher appends 'serve' when it sees no subcommand, so an
        unrecognised one did not fail loudly -- it turned
        'volkit.exe analysis EURJPY' into 'analysis EURJPY serve' and argparse
        rejected the pair.  The names are read off the parser now; this holds
        that.
        """
        import launcher
        from volkit.cli import build_parser
        import argparse as _argparse

        expected = set()
        for action in build_parser()._actions:
            if isinstance(action, _argparse._SubParsersAction):
                expected.update(action.choices)
        self.assertTrue(expected)
        self.assertEqual(launcher._subcommands(), frozenset(expected))

    def test_launcher_defaults_to_serve_only_without_a_subcommand(self):
        import launcher
        known = launcher._subcommands()
        self.assertIn("analysis", known)
        self.assertIn("listed", known)
        self.assertNotIn("--help", known)

    def test_build_inputs_exist(self):
        """Every file the build reads, checked here rather than on a Windows box."""
        import build_exe
        for rel in build_exe.REQUIRED_SOURCES:
            with self.subTest(rel):
                self.assertTrue((build_exe.ROOT / rel).exists(), f"{rel} is missing")

    def test_the_shipped_workbooks_are_clean(self):
        """The build reads them before the suite does, and refuses on a
        problem.  This is the same reading, so a workbook edit that would stop
        a Windows build is caught here in two seconds -- rather than half an
        hour into a CI run, as "EURGBP: no smile term structure" once was.
        """
        import build_exe
        from volkit.marketdata import ExcelSource
        for rel in build_exe.CHECKED_WORKBOOKS:
            with self.subTest(rel):
                path = build_exe.ROOT / rel
                self.assertTrue(path.exists(), f"{rel} is missing")
                self.assertEqual(ExcelSource(path).load().problems, [], rel)

    def test_the_spec_bundles_the_resources_the_code_reads(self):
        """The page and the calendar travel inside the bundle; user data does not.

        paths.resource_dir() and paths.app_dir() are different places, and
        putting a file in the wrong one produces an exe that starts and then
        serves an empty page.
        """
        spec = (self.ROOT / "volkit.spec").read_text(encoding="utf-8")
        self.assertIn("volkit/web", spec)
        self.assertIn("volkit/data", spec)
        # tzdata is not optional on Windows: there is no system IANA database.
        self.assertIn("tzdata", spec)

        import build_exe
        # The user's own files must be staged beside the exe, never bundled.
        for rel in build_exe.USER_DATA:
            with self.subTest(rel):
                self.assertNotIn(rel, spec)

    def test_the_screen_selection_is_checked_before_anything_is_built(self):
        import build_exe
        from volkit import screens
        self.assertEqual(build_exe.choose_screens(None, []), (screens.ALL, ()))
        self.assertEqual(build_exe.choose_screens("pricing,marking", []),
                         (("pricing", "marking"), ()))
        # --only-tabs sets the starting set, --exclude-tab takes further ones away.
        self.assertEqual(build_exe.choose_screens("pricing,marking", ["marking"]),
                         (("pricing",), ()))
        # One flag may carry a list, and an unknown name is an error rather
        # than a screen that quietly stayed in the build.
        self.assertEqual(build_exe.choose_screens(None, ["mm,listed"]),
                         (("pricing", "marking", "monitor", "analysis"), ()))
        with self.assertRaises(screens.ScreenError):
            build_exe.choose_screens(None, ["markign"])
        with self.assertRaises(build_exe.BuildError):
            build_exe.choose_screens(None, list(screens.ALL))

    def test_a_hidden_screen_is_built_but_not_shown(self):
        """The third state: in the build, off until --enable-tab.

        Hiding a screen that was also excluded is refused: the switch could
        never work, and a build whose documented flag does nothing is the
        silent failure this project exists to remove.
        """
        import build_exe
        from volkit import screens
        shown, shy = build_exe.choose_screens(None, [], ["mm"])
        self.assertNotIn("mm", shown)
        self.assertEqual(shy, ("mm",))
        with self.assertRaises(build_exe.BuildError):
            build_exe.choose_screens(None, ["mm"], ["mm"])          # excluded and hidden
        with self.assertRaises(build_exe.BuildError):
            build_exe.choose_screens("pricing", [], ["pricing"])    # nothing left showing
        text = screens.manifest_text(shown, shy)
        self.assertEqual(screens.parse_selection(text, source="t"), (shown, shy))
        self.assertIn("--enable-tab mm", text)

    def test_the_screens_manifest_is_bundled_and_never_left_in_the_source_tree(self):
        """Written under build/, picked up by the spec, read from the bundle.

        Writing it into volkit/data would leave the source tree in a state
        where running from source silently lost a screen.
        """
        import build_exe
        from volkit import screens
        self.assertIn("build", build_exe.SCREENS_BUILD_DIR.parts)
        self.assertNotIn("data", build_exe.SCREENS_BUILD_DIR.parts)
        self.assertFalse((self.ROOT / screens.MANIFEST).exists())
        spec = (self.ROOT / "volkit.spec").read_text(encoding="utf-8")
        self.assertIn("VOLKIT_SCREENS_FILE", spec)
        self.assertIn("volkit/data", spec)
        for rel in build_exe.USER_DATA:
            self.assertNotIn("screens.txt", rel)

    def test_the_handover_zip_keeps_samples_out_of_the_exes_own_folder(self):
        """The one-file zip flattened everything to the top level.

        That put the synthetic history workbook exactly where
        ``find_data_file()`` looks, which is the failure staging it into
        samples/ exists to prevent -- the folder build got this right and the
        one-file build quietly did not.
        """
        import build_exe
        entries = {str(src): arc for src, arc in
                   build_exe.zip_entries(self.ROOT / "dist" / "volkit.exe", onefile=True)}
        self.assertTrue(entries)
        for rel in build_exe.SAMPLE_DATA:
            arc = entries.get(str(self.ROOT / rel))
            if arc is not None:
                self.assertTrue(arc.startswith("samples/"), f"{rel} lands at {arc!r}")
        for rel in build_exe.USER_DATA:
            arc = entries.get(str(self.ROOT / rel))
            if arc is not None:
                self.assertNotIn("/", arc)      # beside the exe, where app_dir() looks

    def test_samples_are_not_staged_beside_the_exe(self):
        """Synthetic data must not sit where find_data_file() would pick it up.

        Made-up numbers appearing on a screen nobody asked for is the same
        failure as a silent zero, so the sample history goes in samples/.
        """
        import build_exe
        self.assertTrue(build_exe.SAMPLE_DATA)
        for rel in build_exe.SAMPLE_DATA:
            with self.subTest(rel):
                self.assertNotIn(rel, build_exe.USER_DATA)


class TestStartupConfig(unittest.TestCase):
    """The settings file a double-clicked executable reads.

    A packaged app has no command line, so this is the only place a desk can
    fix the port, the workbook or a hidden screen without a rebuild.  What is
    pinned here is that it never applies silently and never applies twice.
    """

    def parse(self, text):
        from volkit import config
        return config.parse(text, source="test.cfg")

    def test_a_settings_file_becomes_a_command_line(self):
        cfg = self.parse("# a note\n"
                         "command = serve\n"
                         "port = 8900\n"
                         "workbook = C:\\Marks and data\\vol_marks.xlsx\n"
                         "no-browser = true\n"
                         "zip = false\n"
                         "enable-tab = analysis\n"
                         "enable-tab = mm\n")
        self.assertEqual(cfg.argv, [
            "serve", "--port", "8900",
            "--workbook", "C:\\Marks and data\\vol_marks.xlsx",
            "--no-browser", "--enable-tab", "analysis", "--enable-tab", "mm"])
        # A value is the rest of the line, so a path with spaces needs no
        # quoting; a false switch is left out entirely but still reported.
        self.assertTrue(any("--zip off" in n for n in cfg.notes))

    def test_the_command_may_carry_its_own_arguments(self):
        cfg = self.parse("command = analysis EURJPY --horizon 7\ncut = NY\n")
        self.assertEqual(cfg.argv, ["analysis", "EURJPY", "--horizon", "7", "--cut", "NY"])

    def test_a_line_that_is_not_a_setting_is_refused_by_line_number(self):
        """'port 8900' with no '=' would otherwise vanish silently."""
        from volkit import config
        with self.assertRaises(config.ConfigError) as ctx:
            self.parse("command = serve\nport 8900\n")
        self.assertIn("line 2", str(ctx.exception))
        with self.assertRaises(config.ConfigError):
            self.parse("command = serve\ncommand = check\n")

    def test_the_file_is_read_only_when_nothing_was_typed(self):
        import tempfile
        from volkit import config
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "volkit.cfg"
            path.write_text("command = serve\nport = 8900\n", encoding="utf-8")
            argv, cfg = config.startup_argv(["--config", str(path)], {"serve", "check"})
            self.assertEqual(argv, ["serve", "--port", "8900"])
            self.assertEqual(cfg.path, path)
            # Something typed, and no --config: the file stays shut.
            argv, cfg = config.startup_argv(["check"], {"serve", "check"})
            self.assertEqual(argv, ["check"])
            self.assertIsNone(cfg.path)
            # Explicitly refused.
            argv, cfg = config.startup_argv(["--no-config"], {"serve", "check"})
            self.assertEqual(argv, [])
            self.assertIsNone(cfg.path)
            with self.assertRaises(config.ConfigError):
                config.startup_argv(["--config", str(path), "--no-config"])
            with self.assertRaises(config.ConfigError):
                config.startup_argv(["--config", str(Path(tmp) / "nope.cfg")])

    def test_two_different_commands_are_refused_rather_than_resolved(self):
        import tempfile
        from volkit import config
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "volkit.cfg"
            path.write_text("command = serve\n", encoding="utf-8")
            with self.assertRaises(config.ConfigError) as ctx:
                config.startup_argv(["--config", str(path), "check"], {"serve", "check"})
            self.assertIn("serve", str(ctx.exception))
            # The same command in both places is somebody typing what the file
            # already says, and its own arguments still count.
            argv, _ = config.startup_argv(["--config", str(path), "serve", "--port", "1"],
                                          {"serve", "check"})
            self.assertEqual(argv, ["serve", "--port", "1"])

    def test_the_launcher_puts_serve_in_front_of_a_file_of_options(self):
        """Appending it left the options in front of the subcommand, where
        argparse cannot place them."""
        import tempfile
        import launcher
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "volkit.cfg"
            path.write_text("port = 8900\nno-browser = true\n", encoding="utf-8")
            argv, cfg = launcher.resolve(["--config", str(path)])
            self.assertEqual(argv[0], "serve")
            self.assertEqual(argv, ["serve", "--port", "8900", "--no-browser"])
            self.assertEqual(cfg.path, path)
        # A real command line is untouched and reads no file.
        argv, cfg = launcher.resolve(["tenors", "USDJPY"])
        self.assertEqual(argv, ["tenors", "USDJPY"])
        self.assertIsNone(cfg.path)

    def test_the_sample_settings_file_runs(self):
        """It is staged beside the exe, so a double-click reads it as it ships."""
        from volkit import config
        import build_exe
        self.assertIn("files/volkit.cfg", build_exe.USER_DATA)
        cfg = config.load(Path(__file__).resolve().parents[1] / "files" / "volkit.cfg")
        # The one live setting it ships with is where the kACE feed posts --
        # the desk's own address, confirmed 2026-09-01 -- so a double-click
        # gets the Post buttons without anybody editing the file.
        self.assertEqual(cfg.argv, ["serve", "--kace-url", "https://pfcshkwapp01:8500/pricing"])


# ===========================================================================
# market making
# ===========================================================================


class TestQuoteParsing(unittest.TestCase):
    """Reading a broker run written in English."""

    def parse(self, text, **kw):
        kw.setdefault("pair", "EURUSD")
        return quotes.parse_quotes(text, **kw)

    def test_reads_the_instruments_a_desk_actually_writes(self):
        run = self.parse(
            "1M ATM 8.20/8.60 in 100mm vega\n"
            "3M 25d RR 0.35/0.55 eur call over\n"
            "2M 25d fly 0.20/0.28\n"
            "6M 1.1000 call 7.90/8.40\n"
            "1M/3M ATM spread 0.30/0.55\n")
        self.assertEqual([q.instrument for q in run.quotes],
                         ["atm", "rr", "fly", "outright", "spread"])
        self.assertEqual(run.skipped, ())
        atm = run.quotes[0]
        self.assertAlmostEqual(atm.bid, 0.0820)
        self.assertAlmostEqual(atm.ask, 0.0860)
        self.assertEqual(atm.size, 100.0)
        self.assertEqual(atm.size_basis, "vega")
        self.assertAlmostEqual(run.quotes[1].delta, 0.25)
        self.assertAlmostEqual(run.quotes[3].strike, 1.10)
        # A volatility at an absolute strike is one number whichever side it
        # is quoted from, so the side is dropped and the two lines are one quote.
        self.assertIsNone(run.quotes[3].is_call)

    def test_a_volatility_is_read_as_the_number_it_was_written_as(self):
        """The level is not evidence of the unit (§4).

        A paste whose at-the-money is a third of a point is a managed pair, not
        a paste in decimals; sniffing the magnitude returned it as 35 points.
        A paste really in decimals is loaded by saying so.
        """
        run = self.parse("1M ATM 8.20/8.60\n3M 25d RR 0.35/0.55\n")
        self.assertEqual(run.vol_unit, "percent")
        self.assertAlmostEqual(run.quotes[1].bid, 0.0035)
        # The whole paste below 1.0: still points, and it says so.
        low = self.parse("1M ATM 0.35/0.40\n3M 25d RR 0.02/0.04\n")
        self.assertEqual(low.vol_unit, "percent")
        self.assertAlmostEqual(low.quotes[0].bid, 0.0035)
        self.assertTrue(any("as written" in n for n in low.notes))
        dec = self.parse("1M ATM 0.0820/0.0860\n3M 25d RR 0.0035/0.0055\n",
                         vol_unit="decimal")
        self.assertEqual(dec.vol_unit, "decimal")
        self.assertAlmostEqual(dec.quotes[0].ask, 0.0860)

    def test_a_paste_that_straddles_one_is_read_and_not_refused(self):
        """It used to be refused as 'percent in one place and decimal in
        another'.  There is only one reading now: both lines are points."""
        run = self.parse("1M ATM 8.20/8.60\n3M ATM 0.35/0.45\n")
        self.assertEqual(run.vol_unit, "percent")
        self.assertAlmostEqual(run.quotes[0].bid, 0.0820)
        self.assertAlmostEqual(run.quotes[1].bid, 0.0035)

    def test_a_paste_with_no_level_quote_still_reads_as_written(self):
        run = self.parse("3M 25d RR 0.35/0.55\n6M 25d fly 0.20/0.28\n")
        self.assertEqual(run.vol_unit, "percent")
        self.assertAlmostEqual(run.quotes[0].bid, 0.0035)

    def test_a_direction_word_is_resolved_against_the_pair(self):
        """JPY call over on USDJPY is a dollar put over, so it is negative.

        This is the same class of mistake as the cross-triangle sign in §5:
        the magnitude is right and the sign is not, which nothing downstream
        catches.
        """
        base = quotes.parse_quotes("1M atm 9.0/9.4\n3M 25d rr 0.40/0.60 jpy call over",
                                   pair="USDJPY")
        rr = base.quotes[1]
        self.assertLess(rr.ask, 0.0)
        self.assertAlmostEqual(rr.bid, -0.0060)
        self.assertAlmostEqual(rr.ask, -0.0040)
        other = quotes.parse_quotes("1M atm 9.0/9.4\n3M 25d rr 0.40/0.60 usd call over",
                                    pair="USDJPY")
        self.assertAlmostEqual(other.quotes[1].bid, 0.0040)

    def test_a_currency_that_is_not_a_leg_is_refused(self):
        run = self.parse("1M atm 8.2/8.6\n3M 25d rr 0.4/0.6 jpy call over")
        self.assertEqual(len(run.skipped), 1)
        self.assertIn("not a leg of EURUSD", run.skipped[0][2])

    def test_a_truncated_offer_is_refused_rather_than_repaired(self):
        """'8.2/6' means 8.20/8.60 to a human and repairing it means inventing
        the digits, so it is refused with the reason instead."""
        run = self.parse("1M atm 8.2/8.6\n2M atm 8.2/6")
        self.assertEqual(len(run.skipped), 1)
        self.assertIn("offers below its own bid", run.skipped[0][2])

    def test_calendar_and_literal_spread_orientation(self):
        """'1M/3M' is the calendar convention; '3M-1M' is read literally.

        Both end up as the same difference here, and both say which reading
        they used, because a spread quoted the other way round is a sign error
        nothing downstream would catch.
        """
        # Two different spreads, deliberately: '1M/3M' and '3M-1M' are the
        # same quote written two ways, and one now supersedes the other.
        run = self.parse("1M atm 8.2/8.6\n1M/3M atm spread 0.30/0.55\n6M-2M atm spread 0.30/0.55")
        for q, near, far in ((run.quotes[1], "1M", "3M"), (run.quotes[2], "2M", "6M")):
            self.assertEqual(q.instrument, "spread")
            self.assertEqual(str(q.expiry), near)
            self.assertEqual(str(q.expiry_far), far)
        self.assertTrue(any("calendar convention" in n for n in run.quotes[1].notes))
        self.assertTrue(any("read literally" in n for n in run.quotes[2].notes))

    def test_strangle_and_smile_fly_pin_their_own_convention(self):
        run = self.parse("1M atm 8.2/8.6\n1M 25d strangle 0.20/0.28\n"
                         "2M 25d smile fly 0.20/0.28\n3M 25d fly 0.20/0.28",
                         fly_convention="smile")
        self.assertEqual(run.quotes[1].fly_kind, "market")
        self.assertEqual(run.quotes[2].fly_kind, "smile")
        self.assertEqual(run.quotes[3].fly_kind, "smile")

    def test_nothing_unreadable_is_dropped_quietly(self):
        run = self.parse("1M atm 8.2/8.6\nrumour has it\n1M 25d rr\n")
        self.assertEqual(len(run.quotes), 1)
        self.assertEqual([n for n, _, _ in run.skipped], [2, 3])
        for _, _, why in run.skipped:
            self.assertTrue(why)

    def test_a_vega_profile_reads_tenors_and_reports_the_rest(self):
        profile, notes, skipped = quotes.parse_vega_profile("1M 250\n3M -120\nnope 4\n1M 50")
        self.assertEqual(profile, {"1M": 300.0, "3M": -120.0})
        self.assertEqual(skipped[0][0], 3)
        self.assertTrue(any("more than once" in n for n in notes))


class TestRequestParsing(unittest.TestCase):
    """Reading the other box: what is being asked for, with no price on it.

    Same grammar as a broker run with one thing taken out, read by the same
    tokeniser -- and the absence of the price is *enforced*, which is the
    whole difference between the two boxes.
    """

    def parse(self, text, **kw):
        kw.setdefault("pair", "EURUSD")
        return quotes.parse_requests(text, **kw)

    def test_reads_the_instruments_a_desk_is_asked_for(self):
        asked = self.parse(
            "1M ATM in 100mm vega\n"
            "3M 25d RR\n"
            "2M 25d fly\n"
            "6M 1.1000 call\n"
            "1M/3M ATM spread\n")
        self.assertEqual([q.instrument for q in asked.requests],
                         ["atm", "rr", "fly", "outright", "spread"])
        self.assertEqual(asked.skipped, ())
        self.assertEqual(asked.requests[0].size, 100.0)
        self.assertEqual(asked.requests[0].size_basis, "vega")
        self.assertAlmostEqual(asked.requests[1].delta, 0.25)
        self.assertAlmostEqual(asked.requests[3].strike, 1.10)
        self.assertIsNone(asked.requests[3].is_call)   # a strike quote carries no side

    def test_a_price_in_the_request_box_is_refused_not_read_as_a_strike(self):
        """A broker run pasted into the wrong box would otherwise be quoted at
        levels nobody asked about, which is the silent wrong answer this
        project exists to remove."""
        asked = self.parse("1M ATM 8.20/8.60\n3M 25d RR 0.35/0.55\n")
        self.assertEqual(asked.requests, ())
        self.assertEqual(len(asked.skipped), 2)
        for _, _, why in asked.skipped:
            self.assertIn("reads as a price", why)
            self.assertIn("market box", why)

    def test_one_number_on_a_line_that_has_not_said_what_it_is_struck_at(self):
        """'6M 1.1000' is a strike; '1M ATM 8.5' is a price on an instrument
        that already said what it is."""
        got = self.parse("6M 1.1000\n").requests[0]
        self.assertEqual(got.instrument, "outright")
        self.assertAlmostEqual(got.strike, 1.10)
        self.assertEqual(self.parse("1M ATM 8.5\n").requests, ())

    def test_a_direction_word_is_resolved_against_the_pair(self):
        """The sign lives on the request and is applied once, where the row is
        built.  §5's first entry is what a second place for a sign costs."""
        plain = self.parse("3M 25d rr\n", pair="USDJPY").requests[0]
        asked = self.parse("3M 25d rr jpy call over\n", pair="USDJPY").requests[0]
        self.assertEqual(plain.sign, 1.0)
        self.assertEqual(asked.sign, -1.0)
        self.assertEqual(asked.direction, "jpy")
        self.assertIn("JPY call over", asked.describe())
        # And a currency that is not a leg is a refusal, not a guess.
        self.assertIn("not a leg", self.parse("3M 25d rr chf call over\n",
                                              pair="USDJPY").skipped[0][2])

    def test_the_same_instrument_asked_for_twice_is_two_questions(self):
        """Unlike the market box, where a run is a conversation and a later
        quote of one thing replaces the earlier one.  Two sizes of the same
        tenor are two prices to make."""
        asked = self.parse("1M ATM in 50mm\n1M ATM in 500mm\n")
        self.assertEqual(len(asked.requests), 2)
        self.assertEqual([q.size for q in asked.requests], [50.0, 500.0])

    def test_an_instrument_that_cannot_be_read_keeps_its_reason(self):
        asked = self.parse("3M rr\n1M ATM\nnonsense\n")
        self.assertEqual(len(asked.requests), 1)
        self.assertEqual([n for n, _, _ in asked.skipped], [1, 3])
        self.assertIn("needs a delta", asked.skipped[0][2])


class TestColumnQuotes(unittest.TestCase):
    """A run written as ``expiry, strike, bid/offer`` columns.

    The same parser reads it and the broker-English form, because a run that
    mixes them -- and they do -- must not depend on which line came first.
    """

    def parse(self, text, **kw):
        kw.setdefault("pair", "EURUSD")
        return quotes.parse_quotes(text, **kw)

    def test_the_three_strike_column_spellings_all_read(self):
        run = self.parse("09:15, 1M, ATM, 8.20/8.60\n"
                         "09:15, 3M, 1.0900, 8.10/8.50\n"
                         "09:15, 2M, 25d, 8.00/8.40\n"
                         "09:15, 6M, 25dp, 7.90/8.30\n")
        self.assertEqual(len(run.quotes), 4, run.skipped)
        atm, strike, call, put = run.quotes
        self.assertEqual(atm.instrument, "atm")
        self.assertEqual((strike.instrument, strike.strike), ("outright", 1.09))
        self.assertEqual((call.instrument, call.delta, call.is_call), ("outright", 0.25, True))
        self.assertEqual((put.instrument, put.delta, put.is_call), ("outright", 0.10 * 2.5, False))
        self.assertAlmostEqual(atm.bid, 0.0820)
        self.assertAlmostEqual(strike.ask, 0.0850)

    def test_an_absolute_strike_needs_no_side_and_is_not_called_a_put(self):
        """The volatility at a strike is one number whichever side quotes it.

        ``is_call`` defaulted to None and ``describe`` read None as a put, so a
        strike-column quote came back labelled as something it was not.
        """
        q = self.parse("3M, 1.0900, 8.10/8.50").quotes[0]
        self.assertIsNone(q.is_call)
        self.assertEqual(q.describe(), "3M 1.09")
        self.assertNotIn("put", q.describe())

    def test_a_strike_with_the_side_glued_on_is_a_strike(self):
        """'7.77c' is the 7.77 call, and it beats a delta on the same line.

        It used to match nothing -- not a number, not a word -- and was
        reported as ignored, which left a line that had named its strike to be
        quoted off whatever delta was beside it, or off the at-the-money.
        """
        q = self.parse("3M, 7.77c, 8.10/8.30").quotes[0]
        self.assertEqual((q.instrument, q.strike), ("outright", 7.77))
        self.assertFalse(any("ignored" in n for n in q.notes), q.notes)

        # The strike names the option exactly and the delta only through the
        # marks, so the strike wins and the line says the delta was dropped.
        both = self.parse("1M 25d 7.77p 8.10/8.30").quotes[0]
        self.assertEqual(both.strike, 7.77)
        self.assertIsNone(both.delta)
        self.assertTrue(any("delta is dropped" in n for n in both.notes), both.notes)

        # On a premium the side is the whole difference between two prices,
        # so there it survives.
        prem = self.parse("6M 1.1000c 0.0123 prem").quotes[0]
        self.assertEqual((prem.strike, prem.is_call), (1.10, True))

        # The delta spellings above it are untouched.
        self.assertEqual(self.parse("1M 25dc 8.1/8.3").quotes[0].delta, 0.25)
        self.assertEqual(self.parse("1M 100k 8.2/8.3").quotes[0].size, 0.1)

    def test_a_bare_delta_takes_the_call_wing_and_says_so(self):
        """A delta names two strikes, one on each wing, so it has to pick.

        It picks the same one the pricing screen's strike box picks for a bare
        '25d', and reports it rather than letting the choice be invisible.
        """
        q = self.parse("2M, 25d, 8.00/8.40").quotes[0]
        self.assertTrue(q.is_call)
        self.assertTrue(any("bare delta" in n for n in q.notes), q.notes)

    def test_a_comma_is_a_column_boundary_and_a_price_never_straddles_one(self):
        """This is the whole difference between two readings of three numbers.

        ``3M, 7.75, 8.30`` is a choice at the 7.75 strike; ``3M 7.75 8.30``,
        with no columns, is the two-way at-the-money it has always been. With
        the commas thrown away, as they used to be, the two are the same line.
        """
        columned = self.parse("1M atm 8.2/8.6\n3M, 7.75, 8.30").quotes[1]
        self.assertEqual(columned.instrument, "outright")
        self.assertEqual(columned.strike, 7.75)
        self.assertAlmostEqual(columned.bid, 0.0830)
        self.assertAlmostEqual(columned.ask, 0.0830)

        plain = self.parse("1M atm 8.2/8.6\n3M 7.75 8.30").quotes[1]
        self.assertEqual(plain.instrument, "atm")
        self.assertAlmostEqual(plain.bid, 0.0775)
        self.assertAlmostEqual(plain.ask, 0.0830)

    def test_a_thousands_separator_is_not_a_column(self):
        q = self.parse("1M ATM 8.20/8.60 in 1,000mm vega").quotes[0]
        self.assertEqual(q.instrument, "atm")
        self.assertEqual(q.size, 1000.0)
        self.assertAlmostEqual(q.bid, 0.0820)

    def test_two_numbers_before_the_price_column_are_refused(self):
        run = self.parse("1M atm 8.2/8.6\n3M, 7.75, 7.80, 8.10/8.50")
        self.assertEqual(len(run.quotes), 1)
        self.assertIn("strike column holds one strike", run.skipped[0][2])

    def test_a_column_header_is_recognised_rather_than_reported_as_a_bad_line(self):
        """A run pasted out of a spreadsheet brings its header with it.

        Listing it as a line that could not be read is noise on top of a paste
        that worked; it is passed over and said so instead. Two header words at
        least, because one stray word is more likely a quote that failed.
        """
        run = self.parse("time, expiry, strike, bid/offer\n09:15, 1M, ATM, 8.20/8.60\n")
        self.assertEqual(len(run.quotes), 1)
        self.assertEqual(run.skipped, ())
        self.assertTrue(any("column header" in n for n in run.notes))

        # One word is not a header, and a line with numbers in it never is.
        broken = self.parse("1M atm 8.2/8.6\nstrike\n")
        self.assertEqual(len(broken.skipped), 1)

    def test_broker_english_still_reads_the_way_it_did(self):
        """The columnar reading must not have moved the old one."""
        run = self.parse("1M ATM 8.20/8.60 in 100mm vega\n"
                         "3M 25d RR 0.35/0.55 eur call over\n"
                         "2M 25d fly 0.20/0.28\n"
                         "6M 1.1000 call 7.90/8.40\n"
                         "1M/3M ATM spread 0.30/0.55\n")
        self.assertEqual([q.instrument for q in run.quotes],
                         ["atm", "rr", "fly", "outright", "spread"])
        self.assertAlmostEqual(run.quotes[3].strike, 1.1000)
        # A volatility at an absolute strike is one number whichever side it
        # is quoted from, so the side is dropped and the two lines are one quote.
        self.assertIsNone(run.quotes[3].is_call)


class TestQuoteTimestamps(unittest.TestCase):
    """A run is a conversation: the same thing is quoted again as it moves."""

    def parse(self, text, **kw):
        kw.setdefault("pair", "EURUSD")
        return quotes.parse_quotes(text, **kw)

    def test_a_later_timestamp_wins_whatever_order_it_was_pasted_in(self):
        """The point of reading the timestamp at all.

        Line 3 is the newest quote and line 5 is an older one pasted after it;
        without timestamps the last line would win and the screen would show a
        stale market as the live one.
        """
        run = self.parse("09:15, 1M, ATM, 8.20/8.60\n"
                         "09:41, 1M, ATM, 8.25/8.65\n"
                         "09:05, 1M, ATM, 8.10/8.50\n")
        self.assertEqual(len(run.quotes), 1)
        self.assertAlmostEqual(run.quotes[0].bid, 0.0825)
        self.assertEqual(run.quotes[0].line, 2)
        self.assertEqual({q.line for q in run.superseded}, {1, 3})
        self.assertTrue(all(q.replaced_by == 2 for q in run.superseded))

    def test_without_timestamps_the_later_line_wins(self):
        """The only ordering an untimed line carries is where it was written."""
        run = self.parse("1M ATM 8.20/8.60\n1M ATM 8.25/8.65\n")
        self.assertEqual(len(run.quotes), 1)
        self.assertAlmostEqual(run.quotes[0].bid, 0.0825)
        self.assertEqual(run.superseded[0].line, 1)

    def test_only_the_same_thing_is_superseded(self):
        """An update replaces its own quote and nothing else."""
        run = self.parse("09:15, 1M, ATM, 8.20/8.60\n"
                         "09:41, 1M, ATM, 8.25/8.65\n"
                         "09:41, 3M, ATM, 8.40/8.80\n"
                         "09:41, 1M, 25d, 8.30/8.70\n")
        self.assertEqual(len(run.quotes), 3)
        self.assertEqual(len(run.superseded), 1)
        self.assertEqual({q.describe() for q in run.quotes},
                         {"1M ATM", "3M ATM", "1M 25d call"})

    def test_a_market_strangle_and_a_smile_fly_are_not_the_same_quote(self):
        run = self.parse("1M atm 8.2/8.6\n1M 25d strangle 0.20/0.28\n"
                         "1M 25d smile fly 0.20/0.28\n")
        self.assertEqual(len(run.quotes), 3)
        self.assertEqual(run.superseded, ())

    def test_the_survivor_keeps_the_first_position(self):
        """An updated run reads in the order it was written."""
        run = self.parse("09:15, 1M, ATM, 8.20/8.60\n"
                         "09:15, 3M, ATM, 8.40/8.80\n"
                         "09:41, 1M, ATM, 8.25/8.65\n")
        self.assertEqual([q.describe() for q in run.quotes], ["1M ATM", "3M ATM"])

    def test_a_time_only_line_takes_the_last_date_above_it(self):
        run = self.parse("2024-02-28 09:15, 1M, ATM, 8.20/8.60\n"
                         "09:41, 1M, ATM, 8.25/8.65\n")
        self.assertEqual(len(run.quotes), 1)
        self.assertAlmostEqual(run.quotes[0].bid, 0.0825)
        self.assertEqual(run.quotes[0].timestamp.strftime("%Y-%m-%d %H:%M"), "2024-02-28 09:41")
        self.assertTrue(any("took the last date above them" in n for n in run.notes))

    def test_an_undated_run_says_it_is_ordered_as_one_day(self):
        """That ordering is wrong across midnight, so it is stated."""
        run = self.parse("09:15, 1M, ATM, 8.20/8.60\n23:50, 3M, ATM, 8.40/8.80\n")
        self.assertTrue(any("one day" in n and "midnight" in n for n in run.notes))
        # And the nominal day is never shown: the text is what was written.
        self.assertEqual([q.timestamp_text for q in run.quotes], ["09:15", "23:50"])

    def test_a_date_alone_is_an_expiry_and_a_date_with_a_time_is_a_stamp(self):
        """Reading one as the other moves a quote to a tenor nobody asked for."""
        expiry = self.parse("2024-05-28 ATM 8.15/8.55").quotes[0]
        self.assertEqual(expiry.timestamp_text, "")
        self.assertEqual(str(expiry.expiry)[:10], "2024-05-28")

        stamped = self.parse("2024-02-28T10:05Z, 1M, ATM, 8.30/8.70").quotes[0]
        self.assertEqual(str(stamped.expiry), "1M")
        self.assertEqual(stamped.timestamp.strftime("%Y-%m-%d %H:%M"), "2024-02-28 10:05")

    def test_a_bracketed_time_is_a_time_and_not_a_label(self):
        run = self.parse("[08:00] 3M ATM 8.00/8.40\n[broker A] 1M ATM 8.20/8.60")
        self.assertEqual(run.quotes[0].timestamp_text, "08:00")
        self.assertEqual(run.quotes[0].label, "")
        self.assertEqual(run.quotes[1].label, "broker a")
        self.assertEqual(run.quotes[1].timestamp_text, "")

    def test_a_superseded_quote_is_kept_rather_than_dropped(self):
        """A line read, understood and then silently discarded is the failure
        this module exists to remove."""
        run = self.parse("09:15, 1M, ATM, 8.20/8.60\n09:41, 1M, ATM, 8.25/8.65\n")
        self.assertEqual(len(run.all_quotes), 2)
        self.assertEqual([q.line for q in run.all_quotes], [1, 2])
        self.assertTrue(any("replaced by a later quote" in n for n in run.notes))

    def test_the_panel_reports_the_time_and_what_it_replaced(self):
        from volkit import marketmaker as mm
        panel = mm.panel_from_request({
            "pair": "EURUSD", "cut": "NY", "method": "SVI",
            "text": ("09:15, 1M, ATM, 8.20/8.60\n09:41, 1M, ATM, 8.25/8.65\n"
                     "09:20, 2M, 25d, 8.00/8.40\n"),
            "fit_curve": False, "tune_wings": False,
        })
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        sheet = panel.run(book)["market"]
        self.assertEqual([r["timestamp"] for r in sheet["rows"]], ["09:41", "09:20"])
        self.assertEqual(len(sheet["superseded"]), 1)
        self.assertEqual(sheet["superseded"][0]["replaced_by"], 2)
        self.assertAlmostEqual(sheet["superseded"][0]["bid"], 8.20)


class TestQuoteGrammar(unittest.TestCase):
    """The looser reading of a run: what wins when a line says two things,
    which lines are somebody else's, and structures of more than two legs.

    Every rule here is a precedence stated once: a strike over a delta, a
    date over a tenor, the side only where the side changes the number.
    """

    def parse(self, text, **kw):
        kw.setdefault("pair", "EURUSD")
        return quotes.parse_quotes(text, **kw)

    def test_a_strike_beats_a_delta_and_says_so(self):
        q = self.parse("6M 25d 1.1200 call 7.8/8.3\n").quotes[0]
        self.assertEqual(q.instrument, "outright")
        self.assertAlmostEqual(q.strike, 1.12)
        self.assertIsNone(q.delta)
        self.assertTrue(any("the strike is used" in n for n in q.notes))
        # And in columns, where this used to be a refusal.
        col = self.parse("3M, 25d 1.0900, 8.10/8.50\n")
        self.assertEqual(col.skipped, ())
        self.assertAlmostEqual(col.quotes[0].strike, 1.09)
        self.assertIsNone(col.quotes[0].delta)

    def test_a_date_beats_a_tenor(self):
        q = self.parse("1M 2026-09-30 atm 8.1/8.5\n").quotes[0]
        self.assertEqual(str(q.expiry)[:10], "2026-09-30")
        self.assertTrue(any("the date is used" in n for n in q.notes))
        # Dates in the spellings a desk types, not only ISO.
        for spelt in ("30sep26", "30-Sep-2026", "2026/09/30", "09/30/2026"):
            got = self.parse(f"{spelt} atm 8.1/8.5\n")
            self.assertEqual(got.skipped, (), spelt)
            self.assertEqual(str(got.quotes[0].expiry)[:10], "2026-09-30", spelt)
        # Two tenors on a line that is not a spread stay a refusal.
        self.assertIn("2 tenors", self.parse("1M 3M atm 8.1/8.5\n").skipped[0][2])

    def test_the_side_matters_only_with_a_delta_or_on_a_premium(self):
        """'6M 1.10 call' and '6M 1.10 put' are one volatility, so they are one
        quote and the later one supersedes the earlier.  '25d call' and
        '25d put' are two strikes.  A premium at a strike needs the side,
        because a call and a put there are two different prices."""
        run = self.parse("6M 1.10 call 7.9/8.4\n6M 1.10 put 7.95/8.45\n")
        self.assertEqual(len(run.quotes), 1)
        self.assertIsNone(run.quotes[0].is_call)
        self.assertEqual(len(run.superseded), 1)
        wings = self.parse("1M 25d call 8.9/9.3\n1M -25d 8.9/9.3\n1M 25dp 8.9/9.3\n")
        self.assertEqual([q.is_call for q in wings.quotes], [True, False])
        self.assertEqual(len(wings.superseded), 1)
        live = self.parse("6M 1.10 live 0.0125/0.0135\n")
        self.assertEqual(live.quotes, ())
        self.assertIn("needs the side", live.skipped[0][2])

    def test_a_premium_is_read_in_its_unit_and_never_scaled(self):
        run = self.parse("1M atm 8.2/8.6\n"
                         "6M 1.10 call 125/135 pips\n"
                         "6M 1.10 put 1.25%/1.35% prem\n"
                         "3M 1.10 call 0.0125/0.0135 usd\n"
                         "3M 1.10 put 0.0125/0.0135 eur\n")
        self.assertEqual(run.skipped, ())
        self.assertEqual(run.vol_unit, "percent")
        kinds = [(q.quote_kind, q.premium_unit) for q in run.quotes]
        self.assertEqual(kinds, [("vol", None), ("premium", "pips"), ("premium", "pct"),
                                 ("premium", "price"), ("premium", "pct")])
        self.assertAlmostEqual(run.quotes[1].bid, 125.0)       # not 1.25
        self.assertAlmostEqual(run.quotes[2].ask, 1.35)
        self.assertAlmostEqual(run.quotes[3].bid, 0.0125)
        self.assertAlmostEqual(run.quotes[4].bid, 1.25)        # a fraction of base, as a per cent
        self.assertTrue(all(q.is_call is not None for q in run.quotes[1:]))
        self.assertIn("premium", run.quotes[1].describe())
        # A premium on something that is not an option is refused.
        self.assertIn("premium", self.parse("1M atm 0.5/0.6 prem\n").skipped[0][2])
        # A currency that is neither leg is refused rather than guessed.
        self.assertIn("not a leg", self.parse("3M 1.10 call 0.01/0.02 chf\n").skipped[0][2])

    def test_lines_for_another_pair_are_passed_over_not_refused(self):
        run = self.parse("1M atm 8.2/8.6\n"
                         "USDJPY 1M atm 9/9.4\n"
                         "GBPUSD\n"
                         "1M atm 7/7.4\n"
                         "eur/usd:\n"
                         "3M atm 8.1/8.5\n")
        self.assertEqual([q.describe() for q in run.quotes], ["1M ATM", "3M ATM"])
        self.assertEqual(run.skipped, ())
        self.assertEqual([(n, why) for n, _, why in run.ignored],
                         [(2, "quotes USDJPY, not EURUSD"), (4, "quotes GBPUSD, not EURUSD")])
        self.assertTrue(any("passed over" in n for n in run.notes))
        self.assertEqual(run.quotes[1].pair, "EURUSD")
        # With no pair given nothing is filtered and every line carries its pair.
        every = quotes.parse_quotes("USDJPY 1M atm 9/9.4\nGBPUSD\n1M atm 7/7.4\n")
        self.assertEqual([q.pair for q in every.quotes], ["USDJPY", "GBPUSD"])
        # The same on the request box.
        asked = quotes.parse_requests("1M atm\nUSDJPY 1M atm\n", pair="EURUSD")
        self.assertEqual(len(asked.requests), 1)
        self.assertEqual(asked.ignored[0][0], 2)

    def test_two_legs_of_one_instrument_are_the_calendar_spread_as_before(self):
        run = self.parse("1M/3M atm spread 0.30/0.55\n"
                         "1M vs 3M 25d rr 0.10/0.20\n"
                         "1M atm vs 3M atm 0.30/0.55\n"
                         "buy 1M atm sell 3M atm 0.30/0.55\n")
        self.assertEqual(run.skipped, ())
        # Line 3 is the same quote as line 1 and supersedes it; line 4 is a
        # structure because it carries signs.
        self.assertEqual([q.instrument for q in run.quotes], ["spread", "spread", "structure"])
        rr = run.quotes[1]
        self.assertEqual((str(rr.expiry), str(rr.expiry_far), rr.leg, rr.delta),
                         ("1M", "3M", "rr", 0.25))
        self.assertTrue(any("took" in n or "same instrument" in n for n in rr.notes))
        self.assertEqual(quotes.instrument_key(run.superseded[0]),
                         quotes.instrument_key(run.quotes[0]))
        self.assertEqual([leg.weight for leg in run.quotes[2].legs], [1.0, -1.0])

    def test_a_structure_is_the_signed_sum_of_its_legs(self):
        run = self.parse("6M 1.10 call vs 1.15 call 0.35/0.55\n"
                         "+1M atm vs -2x 3M atm vs +6M atm 0.05/0.15\n"
                         "1M vs 3M vs 6M atm 0.1/0.2\n"
                         "sell 3M 25d rr jpy call over buy 6M 25d rr 0.1/0.2\n", pair="USDJPY")
        self.assertEqual([n for n, _, _ in run.skipped], [3])
        self.assertIn("needs a sign on each", run.skipped[0][2])
        cs, fly, rr = run.quotes
        self.assertEqual(cs.instrument, "structure")
        self.assertEqual([(str(l.expiry), l.strike, l.weight) for l in cs.legs],
                         [("6M", 1.10, -1.0), ("6M", 1.15, 1.0)])
        self.assertIsNone(cs.expiry_far)                  # one tenor, so one expiry
        self.assertEqual([str(x) for x in cs.expiries()], ["6M"])
        self.assertEqual([l.weight for l in fly.legs], [1.0, -2.0, 1.0])
        self.assertEqual((str(fly.expiry), str(fly.expiry_far)), ("1M", "6M"))
        self.assertEqual([str(x) for x in fly.expiries()], ["1M", "3M", "6M"])
        # A direction word on a leg is folded into that leg's weight.
        self.assertEqual([l.weight for l in rr.legs], [1.0, 1.0])
        self.assertEqual(rr.describe(), "-3M 25d RR (JPY call over) +6M 25d RR")
        self.assertAlmostEqual(cs.bid, 0.0035)
        # The request box reads the same structures with no price.
        asked = quotes.parse_requests("6M 1.10 call vs 1.15 call\n", pair="EURUSD")
        self.assertEqual(asked.requests[0].instrument, "structure")
        self.assertEqual(len(asked.requests[0].legs), 2)


class TestQuoteGrammarOnTheBook(unittest.TestCase):
    """The new grammar through the fit and the quote."""

    @classmethod
    def setUpClass(cls):
        from volkit.book import Book
        from volkit.feed import MarketFeed
        cls.book = Book.from_excel(WORKBOOK, ASOF).build(["EURUSD"])
        cls.book.feed = MarketFeed.load(FEED)

    def test_a_premium_becomes_the_volatility_that_reprices_it(self):
        from volkit import marketmaker as mm
        run = quotes.parse_quotes("3M 1.0900 call 125/135 pips\n3M 1.0900 put 0.9%/1.0% prem\n",
                                  pair="EURUSD")
        expiries = mm.resolve_expiries(self.book.clock, run.quotes)
        levels = mm._levels_for(self.book, "EURUSD", expiries)
        out, errors = mm.premiums_as_vols(list(run.quotes), expiries, levels, "EURUSD")
        self.assertEqual(errors, ["", ""])
        F, spot, pip = (levels["3M"][k] for k in ("forward", "spot", "pip"))
        _, t = expiries["3M"]
        for q, px, factor in ((out[0], (125.0, 135.0), 1.0 / pip),
                              (out[1], (0.9, 1.0), spot / 100.0)):
            self.assertEqual(q.quote_kind, "vol")
            for v, p in zip((q.bid, q.ask), px):
                self.assertAlmostEqual(float(black.price(F, 1.09, v, t, bool(q.is_call))),
                                       p * factor, places=12)
        # No feed, no conversion -- and a reason rather than a forward of 1.
        bare, why = mm.premiums_as_vols(list(run.quotes), expiries, {}, "EURUSD")
        self.assertEqual(bare[0].quote_kind, "premium")
        self.assertIn("no forward feed", why[0])

    def test_a_structure_values_as_the_signed_sum_of_its_legs(self):
        from volkit import marketmaker as mm
        run = quotes.parse_quotes("+1M atm vs -2x 3M atm vs +6M atm 0.05/0.15\n"
                                  "1M atm 8/8.4\n3M atm 8/8.4\n6M atm 8/8.4\n", pair="EURUSD")
        surface = self.book["EURUSD"]
        ev = mm.Evaluator(surface, "SVI", "NY")
        expiries = mm.resolve_expiries(self.book.clock, run.quotes)
        got = ev.value(run.quotes[0], expiries, {})
        legs = [ev.value(q, expiries, {}) for q in run.quotes[1:]]
        self.assertAlmostEqual(got, legs[0] - 2 * legs[1] + legs[2], places=14)
        # Every expiry a structure names is resolved, not only its two ends.
        self.assertEqual(sorted(expiries), ["1M", "3M", "6M"])


class TestKnowledgeBank(unittest.TestCase):
    """Desk knowledge as an overlay, and the rules for resolving it."""

    def bank(self):
        pk = PairKnowledge(rules=[
            Rule("spread", 0.30, "any"),
            Rule("spread", 0.25, "atm", max_days=31),
            Rule("spread", 0.45, "atm", max_days=31, max_size=200),
            Rule("floor", 0.15),
            Rule("floor", 0.20, "fly"),
            Rule("shift", 0.05, "atm", tenor="1W"),
            Rule("note", text="wings always wider into an ECB week"),
        ])
        return pk

    def test_the_narrowest_matching_rule_wins_and_is_named(self):
        pk = self.bank()
        wide = pk.overlay(instrument="atm", days=20, tenor="1M", size=150, size_basis="vega")
        self.assertEqual(wide.spread, 0.45)
        self.assertIn("200mm", wide.spread_rule)
        # Over the size band, the size-conditioned rule no longer matches.
        small = pk.overlay(instrument="atm", days=20, tenor="1M", size=500, size_basis="vega")
        self.assertEqual(small.spread, 0.25)
        far = pk.overlay(instrument="atm", days=200, tenor="6M")
        self.assertEqual(far.spread, 0.30)

    def test_a_floor_is_the_widest_matching_one_not_the_narrowest(self):
        """Every floor applies and the widest wins, because that is what a
        floor means; widths take the narrowest rule instead."""
        pk = PairKnowledge(rules=[Rule("spread", 0.05, "fly"), Rule("floor", 0.15),
                                  Rule("floor", 0.22, "fly")])
        got = pk.overlay(instrument="fly", days=30, tenor="1M", delta=0.25)
        self.assertEqual(got.floor, 0.22)
        self.assertEqual(got.spread, 0.22)

    def test_a_note_is_shown_and_never_applied(self):
        got = self.bank().overlay(instrument="rr", days=30, tenor="1M", delta=0.25)
        self.assertEqual(got.notes, ("wings always wider into an ECB week",))
        self.assertEqual(got.shift, 0.0)

    def test_no_matching_rule_means_no_width_not_an_invented_one(self):
        """There is no built-in default anywhere.  A number on a screen with no
        source is the same failure as a silent zero."""
        pk = PairKnowledge(rules=[Rule("spread", 0.25, "atm")])
        got = pk.overlay(instrument="rr", days=30, tenor="1M", delta=0.25)
        self.assertIsNone(got.spread)
        self.assertIn("no width rule", got.reason)
        fell_back = pk.overlay(instrument="rr", days=30, tenor="1M", delta=0.25, fallback=0.4)
        self.assertEqual(fell_back.spread, 0.4)
        self.assertIsNone(fell_back.spread_rule)
        self.assertIn("panel fallback", fell_back.reason)

    def test_a_bad_rule_set_is_rejected_whole(self):
        bank = KnowledgeBank()
        problems = bank.set_pair("EURUSD", [Rule("spread", 0.25, "atm"), Rule("spread", -1.0)],
                                 ASOF.now)
        self.assertTrue(problems)
        self.assertEqual(bank.pairs, {})

    def test_the_bank_round_trips_through_its_file(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "mm_knowledge.json"
            bank = KnowledgeBank()
            self.assertEqual(bank.set_pair("USDJPY", self.bank().rules, ASOF.now, "test"), [])
            bank.save(path)
            back = KnowledgeBank.load(path)
            self.assertEqual([r.describe() for r in back.for_pair("usdjpy").rules],
                             [r.describe() for r in self.bank().rules])
            self.assertEqual(back.for_pair("USDJPY").source_note, "test")

    def test_a_missing_bank_is_empty_and_says_so_rather_than_failing(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            bank = KnowledgeBank.load(Path(tmp) / "nothing.json")
            self.assertEqual(bank.pairs, {})
            self.assertTrue(any("will be created" in p for p in bank.problems))

    def test_learning_measures_the_paste_and_ignores_choice_prices(self):
        """A quote written as a single mid has no width; averaging its zero in
        would quietly tighten the whole ladder.

        The two 1M lines are the same quote twice, so only the later one is
        live -- but both are evidence of how wide this market is shown, which
        is why the bank reads ``all_quotes`` and the fit reads ``quotes``.
        """
        run = quotes.parse_quotes(
            "1M atm 8.20/8.60\n1M atm 8.30/8.70\n2M atm 9.00\n", pair="EURUSD")
        self.assertEqual(len(run.quotes), 2)
        self.assertEqual(len(run.superseded), 1)
        rules, notes = suggest_rules(run.all_quotes, days_of=lambda q: 30.0)
        self.assertEqual(len(rules), 1)
        self.assertAlmostEqual(rules[0].value, 0.0040)
        self.assertIn("median of 2", rules[0].text)


class TestMarketMakerModel(unittest.TestCase):
    """The three stages: the curve, the wings, and the quote."""

    @classmethod
    def setUpClass(cls):
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD", "USDJPY"])

    # -- the fast path is measured, never assumed --------------------------
    def test_the_wing_reproduces_the_interpolated_smile_where_the_check_says_it_does(self):
        """The whole fine tune rests on this.

        The interpolators are fitted *through* five anchor points taken off the
        two SABR wings, so at those deltas the wing and the interpolation are
        normally the same number -- which is what lets the fit skip a 19ms SVI
        solve per expiry per evaluation.  Where ``anchor_gap`` says they agree,
        this holds them to it.
        """
        for pair in ("EURUSD", "USDJPY"):
            surface = self.book[pair]
            for tenor in ("1w", "1m", "3m", "1y"):
                t = tenor_to_years(tenor)
                dt = self.book.clock.datetime_from_years(t)
                gap = marketmaker.anchor_gap(surface, dt, t, "SVI", "NY")
                with self.subTest(pair=pair, tenor=tenor):
                    self.assertLessEqual(gap, marketmaker.ANCHOR_TOLERANCE)
                sl = surface.slice_at(dt, "SVI", "NY")
                ev = marketmaker.Evaluator(surface, "SVI", "NY",
                                           fast_at=frozenset({round(t, 10)}))
                for delta, is_call in ((0.25, True), (0.25, False),
                                       (0.10, True), (0.10, False)):
                    with self.subTest(pair=pair, tenor=tenor, delta=delta, call=is_call):
                        fast = ev.delta_vol(dt, t, delta, is_call)
                        slow = sl.strike_from_delta(delta if is_call else -delta, is_call)[1]
                        self.assertAlmostEqual(fast, slow, delta=marketmaker.ANCHOR_TOLERANCE)
                self.assertEqual(ev.slices_built, 0, "the fast path should build no slice")

    def test_a_slice_that_cannot_pass_through_its_anchors_is_refused_the_shortcut(self):
        """SVI here is arbitrage constrained, so five parameters through five
        points is not a free interpolation.  On USDCNY -- a managed pair whose
        marked wings are the least well behaved in the book -- the constraint
        binds and the fitted smile lands more than a tenth of a volatility
        point off its own anchors.  Assuming the shortcut there would have the
        fit marking to a smile the rest of the tool does not price on.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDCNY"])
        surface = book["USDCNY"]
        t = tenor_to_years("1m")
        dt = book.clock.datetime_from_years(t)
        self.assertGreater(marketmaker.anchor_gap(surface, dt, t, "SVI", "NY"), 1e-4)
        q = quotes.MarketQuote(instrument="rr", expiry="1M", bid=0.001, ask=0.002, delta=0.25)
        expiries = marketmaker.resolve_expiries(book.clock, [q])
        fast_at, notes = marketmaker.verified_fast_expiries(
            surface, [q], expiries, "SVI", "NY")
        self.assertEqual(fast_at, frozenset())
        self.assertTrue(any("does not pass through its own anchor points" in n for n in notes))
        # And with no shortcut, the evaluator agrees with the surface itself.
        ev = marketmaker.Evaluator(surface, "SVI", "NY", fast_at=fast_at)
        self.assertAlmostEqual(ev.value(q, expiries, {}),
                               surface.risk_reversal(dt, 0.25, "SVI", "NY"), places=12)

    def test_the_fast_path_is_declined_where_it_would_not_be_exact(self):
        """Vanna-volga at 25 delta reproduces its own three anchors and says
        nothing about the 10 delta ones, so the shortcut is refused there."""
        self.assertEqual(marketmaker.anchor_wing("SVI", 0.25), 25)
        self.assertEqual(marketmaker.anchor_wing("SVI", 0.10), 10)
        self.assertEqual(marketmaker.anchor_wing("VV25", 0.25), 25)
        self.assertIsNone(marketmaker.anchor_wing("VV25", 0.10))
        self.assertIsNone(marketmaker.anchor_wing("VV10", 0.25))
        self.assertIsNone(marketmaker.anchor_wing("SVI", 0.15))
        self.assertEqual(marketmaker.anchor_wing("SABR25", 0.15), 25)

    def test_the_evaluator_is_exact_unless_it_is_told_otherwise(self):
        """An unverified fast path is how a fit ends up marking to a smile the
        rest of the tool does not use, so it is off by default."""
        surface = self.book["EURUSD"]
        t = tenor_to_years("3m")
        dt = self.book.clock.datetime_from_years(t)
        ev = marketmaker.Evaluator(surface, "SVI", "NY")
        q = quotes.MarketQuote(instrument="rr", expiry="3M", bid=0.0, ask=0.001, delta=0.25)
        got = ev.value(q, {"3M": (dt, t)}, {})
        self.assertGreater(ev.slices_built, 0)
        self.assertAlmostEqual(got, surface.risk_reversal(dt, 0.25, "SVI", "NY"), places=12)

    # -- the curve ---------------------------------------------------------
    def test_segmented_accumulation_matches_integrating_from_zero(self):
        """The fit accumulates the term structure segment by segment rather
        than integrating from zero once per tenor.  Daily variances summing to
        the term variance is what makes that identical, so it is pinned."""
        atm = self.book["EURUSD"].atm
        ts = [tenor_to_years(x) for x in ("1w", "1m", "3m", "6m", "1y")]
        fast = marketmaker._curve_vols(atm, ts)
        slow = [atm.curve_vol(t) for t in ts]
        for a, b in zip(fast, slow):
            self.assertAlmostEqual(a, b, places=14)

    def test_the_curve_fit_recovers_a_curve_it_was_moved_away_from(self):
        import copy
        atm = copy.deepcopy(self.book["EURUSD"].atm)
        tenors = ("1w", "2w", "1m", "2m", "3m", "6m", "9m", "1y")
        want = [marketmaker.CurveTarget(t.upper(), tenor_to_years(t),
                                        atm.curve_vol(tenor_to_years(t)))
                for t in tenors]
        original = dict(vars(atm.params))
        atm.set_params(initial_vol=0.03, long_term_vol=0.15, mean_reversion=40.0)
        fit = marketmaker.fit_atm_curve(atm, want)
        for target, got in zip(want, fit.achieved_after):
            self.assertAlmostEqual(got, target.vol, places=6)
        self.assertAlmostEqual(fit.after["long_term_vol"], original["long_term_vol"], places=4)

    def test_the_curve_fit_leaves_the_curve_it_was_given_alone(self):
        """It runs on a copy, so a failed or unwanted fit cannot leave a
        half-marked curve behind."""
        atm = self.book["EURUSD"].atm
        before = dict(vars(atm.params))
        targets = [marketmaker.CurveTarget(t.upper(), tenor_to_years(t), 0.09)
                   for t in ("1w", "1m", "3m", "6m", "1y")]
        marketmaker.fit_atm_curve(atm, targets)
        self.assertEqual(dict(vars(atm.params)), before)

    def test_a_pinned_parameter_is_not_moved_by_the_sweep(self):
        """The starting-point sweep may only vary free parameters.  Sweeping a
        pinned one and keeping whatever the best node held would change a
        parameter the caller deliberately froze."""
        import copy
        atm = copy.deepcopy(self.book["EURUSD"].atm)
        targets = [marketmaker.CurveTarget(t.upper(), tenor_to_years(t), v)
                   for t, v in (("1m", 0.062), ("3m", 0.064), ("6m", 0.0675), ("1y", 0.0715))]
        fit = marketmaker.fit_atm_curve(atm, targets, free=("initial_vol", "long_term_vol"))
        self.assertEqual(fit.after["short_decay"], atm.params.short_decay)
        self.assertEqual(fit.after["mean_reversion"], atm.params.mean_reversion)

    def test_more_free_parameters_than_targets_is_refused(self):
        import copy
        atm = copy.deepcopy(self.book["EURUSD"].atm)
        targets = [marketmaker.CurveTarget("1M", tenor_to_years("1m"), 0.08)]
        with self.assertRaises(ValueError) as ctx:
            marketmaker.fit_atm_curve(atm, targets)
        self.assertIn("cannot determine", str(ctx.exception))

    def test_the_mean_reversion_is_fitted_inside_the_marked_range(self):
        """The range is a marking judgement, so the fit stays in it and the
        sweep nodes are taken from it.  A node the polish may not reach can
        still win the sweep on cost and is then clipped into the bound, which
        is a different curve from the one that was measured."""
        import copy
        lo, hi = marketmaker.MEAN_REVERSION_RANGE
        self.assertEqual(marketmaker._BOUNDS["mean_reversion"], (lo, hi))
        for node in marketmaker.reversion_nodes():
            self.assertGreaterEqual(node, lo)
            self.assertLessEqual(node, hi)
        atm = copy.deepcopy(self.book["EURUSD"].atm)
        # A target curve that flattens far faster than the range allows: the
        # unconstrained fit would run the reversion well past the top.
        want = [marketmaker.CurveTarget(t.upper(), tenor_to_years(t), v)
                for t, v in (("1w", 0.055), ("2w", 0.062), ("1m", 0.068),
                             ("3m", 0.0715), ("6m", 0.072), ("1y", 0.0721))]
        fit = marketmaker.fit_atm_curve(atm, want)
        self.assertLessEqual(fit.after["mean_reversion"], hi)
        self.assertGreaterEqual(fit.after["mean_reversion"], lo)
        # Held back, it says so, and names the constant that would let it go.
        rest = [w for w in fit.warnings if "mean_reversion came to rest" in w]
        self.assertTrue(rest)
        self.assertIn("marking judgement", rest[0])
        self.assertIn("Widen the range on the fit panel", rest[0])

    def test_a_parameter_on_its_bound_that_met_its_targets_is_not_warned_about(self):
        """AUDUSD is marked at exactly the top of the range, so an ungated
        check warned on every refit of the curve the desk already had.  A
        parameter resting on a bound limits nothing while the targets are
        still met, and a warning that fires when nothing is wrong is one
        nobody reads."""
        import copy
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["AUDUSD"])
        atm = copy.deepcopy(book["AUDUSD"].atm)
        self.assertEqual(atm.params.mean_reversion, marketmaker.MEAN_REVERSION_RANGE[1])
        want = [marketmaker.CurveTarget(t.upper(), tenor_to_years(t),
                                        atm.curve_vol(tenor_to_years(t)))
                for t in ("1m", "2m", "3m", "6m", "1y")]
        fit = marketmaker.fit_atm_curve(atm, want)
        self.assertLessEqual(fit.rmse, marketmaker._BOUND_BINDING_RMSE)
        self.assertAlmostEqual(fit.after["mean_reversion"],
                               marketmaker.MEAN_REVERSION_RANGE[1], places=6)
        self.assertEqual([w for w in fit.warnings if "came to rest" in w], [])

    def test_the_range_is_a_marking_judgement_a_panel_may_override(self):
        """It is the one bound a caller may move, because it is a judgement
        about what a desk marks rather than a property of the model -- and the
        sweep nodes move with it, so the nodes and the polish can never be
        taken from two different ranges."""
        import copy
        atm = copy.deepcopy(self.book["EURUSD"].atm)
        want = [marketmaker.CurveTarget(t.upper(), tenor_to_years(t), v)
                for t, v in (("1w", 0.055), ("2w", 0.062), ("1m", 0.068),
                             ("3m", 0.0715), ("6m", 0.072), ("1y", 0.0721))]
        house = marketmaker.fit_atm_curve(atm, want)
        wide = marketmaker.fit_atm_curve(atm, want, reversion_range=(1.0, 40.0))
        self.assertAlmostEqual(house.after["mean_reversion"],
                               marketmaker.MEAN_REVERSION_RANGE[1], places=6)
        self.assertGreater(wide.after["mean_reversion"],
                           marketmaker.MEAN_REVERSION_RANGE[1])
        # Given room, the same targets are reached better -- which is the whole
        # reason the range is a judgement and not a fact.
        self.assertLess(wide.rmse, house.rmse)
        self.assertEqual(marketmaker.reversion_nodes((1.0, 5.0))[0], 1.0)
        self.assertEqual(marketmaker.reversion_nodes((1.0, 5.0))[-1], 5.0)

    def test_a_half_typed_mean_reversion_range_is_refused(self):
        """Two blanks are the house range -- the same reading as an empty
        market box handing the field back to the feed.  One blank is a range
        somebody meant to type and did not finish, and reading it half way
        would fit in a range nobody chose."""
        base = {"pair": "EURUSD"}
        self.assertIsNone(marketmaker.panel_from_request(base).reversion_range)
        self.assertIsNone(marketmaker.panel_from_request(
            {**base, "reversion_lo": "", "reversion_hi": ""}).reversion_range)
        self.assertEqual(marketmaker.panel_from_request(
            {**base, "reversion_lo": "2", "reversion_hi": "9"}).reversion_range, (2.0, 9.0))
        for bad, why in ((("2", ""), "both"), (("", "9"), "both"),
                         (("0", "9"), "above zero"), (("9", "2"), "above its floor")):
            with self.assertRaises(ValueError) as ctx:
                marketmaker.panel_from_request(
                    {**base, "reversion_lo": bad[0], "reversion_hi": bad[1]})
            self.assertIn(why, str(ctx.exception))

    def test_a_cross_fits_its_correlation_not_a_level_it_does_not_own(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["AUDUSD", "USDJPY", "AUDJPY"])
        knobs = marketmaker._Knobs(book["AUDJPY"].atm)
        self.assertTrue(knobs.is_cross)
        self.assertIn("corr_initial", knobs.available)
        self.assertNotIn("initial_vol", knobs.available)
        with self.assertRaises(ValueError) as ctx:
            marketmaker.fit_atm_curve(
                book["AUDJPY"].atm,
                [marketmaker.CurveTarget("1M", tenor_to_years("1m"), 0.09)],
                free=("initial_vol",))
        self.assertIn("cross", str(ctx.exception))

    # -- the hinge ---------------------------------------------------------
    def test_the_hinge_is_flat_inside_the_market_and_signed_outside_it(self):
        self.assertEqual(marketmaker._hinge(0.082, 0.080, 0.086), 0.0)
        self.assertAlmostEqual(marketmaker._hinge(0.078, 0.080, 0.086), -0.002)
        self.assertAlmostEqual(marketmaker._hinge(0.090, 0.080, 0.086), 0.004)

    def test_the_wing_tune_pulls_a_quote_it_can_reach_inside_the_market(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        surface = book["EURUSD"]
        t = tenor_to_years("3m")
        dt = book.clock.datetime_from_years(t)
        model = surface.risk_reversal(dt, 0.25, "SVI", "NY")
        target = model - 0.004                      # 0.4 vol points away
        q = quotes.MarketQuote(instrument="rr", expiry="3M", bid=target - 0.0005,
                               ask=target + 0.0005, delta=0.25)
        expiries = marketmaker.resolve_expiries(book.clock, [q])
        # One quote determines one parameter.  Freeing the pair as well would be
        # refused, which is a separate test.
        res = marketmaker.tune_smile_shifts(surface, [q], expiries, {}, method="SVI", cut="NY",
                                            free=("rho25",))
        self.assertEqual(res.inside_before, 0)
        self.assertEqual(res.inside_after, 1)
        self.assertLess(res.worst_after, res.worst_before)

    def test_more_free_parameters_than_wing_quotes_is_refused(self):
        """The same rule the curve fit applies to its targets.  An
        under-determined hinge has a flat manifold of answers and the optimiser
        wanders along it burning its whole budget on the tie-breakers."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        q = quotes.MarketQuote(instrument="rr", expiry="3M", bid=-0.002, ask=-0.001, delta=0.25)
        expiries = marketmaker.resolve_expiries(book.clock, [q])
        with self.assertRaises(ValueError) as ctx:
            marketmaker.tune_smile_shifts(book["EURUSD"], [q], expiries, {},
                                          method="SVI", cut="NY")
        self.assertIn("cannot determine", str(ctx.exception))

    def test_a_parameter_no_quote_reads_off_is_left_where_it_is(self):
        """A 25-delta quote reads the 25-delta anchor, which is built from
        rho25 and slog25 alone.  Freeing the ten-delta pair would not inform
        them; it would only make the objective flat in two more directions."""
        informed, _ = marketmaker.informative_params(
            [quotes.MarketQuote(instrument="rr", expiry="3M", bid=-0.002, ask=-0.001,
                                delta=0.25)], "SVI")
        self.assertEqual(informed, {"rho25", "slog25"})
        both, _ = marketmaker.informative_params(
            [quotes.MarketQuote(instrument="outright", expiry="3M", bid=0.06, ask=0.065,
                                strike=1.1, is_call=True)], "SVI")
        self.assertEqual(both, set(PARAM_NAMES))

    def test_the_wing_tune_reports_what_a_curve_wide_shift_cannot_reach(self):
        """Two tenors asking for opposite skews cannot both be met by one
        shift.  Saying so beats bending the surface to whichever quote the
        optimiser happened to weight most."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        surface = book["EURUSD"]
        made = []
        for tenor, offset in (("1M", -0.006), ("1Y", +0.006)):
            t = tenor_to_years(tenor)
            dt = book.clock.datetime_from_years(t)
            base = surface.risk_reversal(dt, 0.25, "SVI", "NY") + offset
            made.append(quotes.MarketQuote(instrument="rr", expiry=tenor, bid=base - 0.0002,
                                           ask=base + 0.0002, delta=0.25))
        expiries = marketmaker.resolve_expiries(book.clock, made)
        res = marketmaker.tune_smile_shifts(surface, made, expiries, {}, method="SVI", cut="NY")
        self.assertLess(res.inside_after, 2)
        self.assertTrue(any("still outside their market" in w for w in res.warnings))

    # -- shading the mid ----------------------------------------------------
    def test_richness_and_a_long_position_both_shade_the_mid_down(self):
        """Both are reasons to want to sell, and you attract a seller's trade
        by shading the price down.  A sign error here quotes the wrong way
        round on every row."""
        q = quotes.MarketQuote(instrument="atm", expiry="1M", bid=0.080, ask=0.086)
        rich = marketmaker.skew_for(q, 0.08, half_width=0.003, richness=0.004, axe=None,
                                    fair_weight=0.25, axe_weight=0.5, cap_ratio=1.0,
                                    bank_shift=0.0)
        self.assertLess(rich.fair, 0.0)
        long_vega = marketmaker.skew_for(q, 0.08, half_width=0.003, richness=None, axe=1.0,
                                         fair_weight=0.25, axe_weight=0.5, cap_ratio=1.0,
                                         bank_shift=0.0)
        self.assertAlmostEqual(long_vega.axe, -0.5 * 0.003)
        short_vega = marketmaker.skew_for(q, 0.08, half_width=0.003, richness=None, axe=-1.0,
                                          fair_weight=0.25, axe_weight=0.5, cap_ratio=1.0,
                                          bank_shift=0.0)
        self.assertGreater(short_vega.axe, 0.0)

    def test_the_shading_is_capped_at_a_fraction_of_the_width(self):
        """An axe may lean the price inside the market; it may not walk it out
        of the market on its own, which stops being a quote."""
        q = quotes.MarketQuote(instrument="atm", expiry="1M", bid=0.080, ask=0.086)
        got = marketmaker.skew_for(q, 0.08, half_width=0.003, richness=0.20, axe=1.0,
                                   fair_weight=0.25, axe_weight=0.5, cap_ratio=1.0,
                                   bank_shift=0.0)
        self.assertTrue(got.capped)
        self.assertAlmostEqual(abs(got.total), 0.003)

    def test_neither_lean_is_applied_to_a_risk_reversal(self):
        """A break-even against realized volatility and a vega position are
        both statements about the level.  Neither says where the skew belongs,
        so the row says so instead of inventing a lean."""
        q = quotes.MarketQuote(instrument="rr", expiry="1M", bid=-0.003, ask=-0.001, delta=0.25)
        got = marketmaker.skew_for(q, 0.08, half_width=0.001, richness=0.004, axe=1.0,
                                   fair_weight=0.25, axe_weight=0.5, cap_ratio=1.0,
                                   bank_shift=0.0005)
        self.assertEqual(got.fair, 0.0)
        self.assertEqual(got.axe, 0.0)
        self.assertAlmostEqual(got.total, 0.0005)
        self.assertIn("not a level", got.reason)


class TestMarketMakerPanel(unittest.TestCase):
    """The screen as a whole: what each of its two stages reports, and what
    they leave behind.

    Fitting and quoting are two panels and two routes, so they are two sets of
    tests here.  Most of these switch the wing fine tune off; it is exercised
    properly in ``TestMarketMakerModel`` and in the tests here that need it,
    and running a full one in every case would spend a minute of the suite
    re-proving it.
    """

    TEXT = ("1M ATM 6.05/6.35 in 100mm vega\n"
            "3M ATM 6.25/6.55\n"
            "6M ATM 6.60/6.90\n"
            "1Y atm 7.00/7.30\n"
            "1M 25d rr -0.30/-0.10 eur call over\n"
            "3M 25d fly 0.12/0.20\n")
    ASKED = ("1M ATM in 100mm\n"
             "1M 25d rr\n"
             "3M 25d fly\n")

    @classmethod
    def setUpClass(cls):
        # Shared by everything that only reports: both panels put the book
        # back, which is itself two of the tests below.
        cls.book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])

    def panel(self, **kw):
        payload = {"pair": "EURUSD", "text": self.TEXT, "target_source": "quotes",
                   "tune_wings": False}
        payload.update(kw)
        return marketmaker.panel_from_request(payload)

    def quote(self, **kw):
        payload = {"pair": "EURUSD", "request_text": self.ASKED, "fallback_spread": "0.30"}
        payload.update(kw)
        return marketmaker.quote_panel_from_request(payload)

    def bank(self):
        bank = KnowledgeBank()
        bank.set_pair("EURUSD", [Rule("spread", 0.28, "atm"), Rule("spread", 0.20, "rr"),
                                 Rule("spread", 0.12, "fly")], ASOF.now)
        return bank

    # -- the fit ----------------------------------------------------------
    def test_reporting_puts_the_book_back_exactly(self):
        """The default is to report.  A screen that quietly re-marked the book
        every time somebody typed in it would be unusable.  This runs the wing
        tune as well, because the shifts have to come back too."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        surface = book["EURUSD"]
        before = list(surface.atm.tenor_table())
        shifts = dict(surface.param_shifts)
        out = self.panel(tune_wings=True).run(book)
        self.assertFalse(out["applied"])
        self.assertIsNotNone(out["wings"])
        self.assertNotEqual(out["wings"]["after"], out["wings"]["before"])
        self.assertEqual(list(surface.atm.tenor_table()), before)
        self.assertEqual(dict(surface.param_shifts), shifts)

    def test_keeping_the_marks_writes_them_and_says_it_did(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        surface = book["EURUSD"]
        before = surface.atm.term_vol(tenor_to_years("1m"))
        out = self.panel(apply=True).run(book)
        self.assertTrue(out["applied"])
        self.assertAlmostEqual(surface.atm.term_vol(surface.tenor_years("1m")), 0.062, places=5)
        self.assertNotAlmostEqual(surface.atm.term_vol(tenor_to_years("1m")), before)
        self.assertTrue(any("in memory only" in w for w in out["warnings"]))

    def test_the_fit_puts_a_price_on_nothing(self):
        """The whole point of the split.  A fit that also quoted the run it was
        fitted to made a price in every instrument a broker happened to show,
        which is not what anybody asked for -- and it meant a request could
        only be priced by re-running a fit against a market that had nothing
        to do with it."""
        out = self.panel().run(self.book)
        self.assertNotIn("sheet", out)
        for row in out["market"]["rows"]:
            self.assertNotIn("our_bid", row)
            self.assertNotIn("width", row)

    def test_a_section_that_cannot_run_empties_only_itself(self):
        """No pinned tenor means no target curve.  The market table is still
        built, the same way the analysis screen keeps its sections apart."""
        out = self.panel(target_source="overwrites").run(self.book)
        self.assertIsNone(out["curve"])
        self.assertIn("no tenor is pinned", out["unavailable"]["curve"])
        self.assertIsNotNone(out["market"])
        self.assertEqual(out["market"]["n_quotes"], 6)

    def test_the_market_table_reports_where_the_model_sits_against_theirs(self):
        out = self.panel().run(self.book)
        for row in out["market"]["rows"]:
            self.assertIn(row["position"], ("inside", "below", "above"))
            if row["position"] == "inside":
                self.assertEqual(row["edge"], 0.0)
            else:
                self.assertNotEqual(row["edge"], 0.0)
        self.assertEqual(out["market"]["n_quotes"], 6)

    def test_a_paste_the_reader_cannot_use_is_listed_not_silently_shortened(self):
        out = self.panel(text=self.TEXT + "3M 25d rr 0.4/0.6 jpy call over\n").run(self.book)
        self.assertEqual(out["market"]["n_quotes"], 6)
        self.assertEqual(len(out["market"]["skipped"]), 1)
        self.assertIn("not a leg of EURUSD", out["market"]["skipped"][0]["why"])

    # -- the hand-off -----------------------------------------------------
    def test_the_handoff_reproduces_the_fit_exactly(self):
        """The two halves of the split have to meet at the same numbers.

        Quoting off the marks a fit handed back must give what quoting off a
        book the same fit was *applied* to gives.  Anything less and the
        screen's price would depend on whether somebody ticked "keep the
        marks", which is a decision about the book and not about the price.
        """
        applied_book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        fit = self.panel(apply=True, tune_wings=True).run(applied_book)
        on_book = self.quote().run(applied_book, bank=self.bank())

        reported_book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        reported = self.panel(tune_wings=True).run(reported_book)
        handed = self.quote(marks=reported["marks"]).run(reported_book, bank=self.bank())

        self.assertEqual([r["our_bid"] for r in handed["sheet"]["rows"]],
                         [r["our_bid"] for r in on_book["sheet"]["rows"]])
        self.assertEqual([r["our_ask"] for r in handed["sheet"]["rows"]],
                         [r["our_ask"] for r in on_book["sheet"]["rows"]])
        self.assertEqual(fit["marks"]["knobs"], reported["marks"]["knobs"])

    def test_keeping_a_fit_keeps_the_marks_it_handed_back(self):
        """The book and the panel must hold one number, not two spellings of it.

        A knob leaves the fit in volatility points and comes back divided by a
        hundred, and ``x * 100 / 100`` differs from ``x`` in the last place for
        about an eighth of all values.  Keeping the raw fitted numbers on the
        surface therefore left the book a bit away from the marks the quote
        panel was posting, and the price depended on whether "keep the marks"
        had been ticked -- a nanovol apart, which is nothing to a market and
        everything to a screen that has to reproduce itself.
        """
        applied = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        fit = self.panel(apply=True, tune_wings=True).run(applied)
        self.assertEqual(marketmaker.capture_marks(applied["EURUSD"]),
                         {"knobs": fit["marks"]["knobs"], "shifts": fit["marks"]["shifts"]})

    def test_a_quote_standing_on_a_fit_puts_the_book_back(self):
        """The marks go on for the length of one call and come off again.  A
        surface left half-marked by a price nobody kept, priced off all
        morning, is the worst outcome available to this tool."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        before = marketmaker.capture_marks(book["EURUSD"])
        fit = self.panel(tune_wings=True).run(book)
        self.assertNotEqual(fit["marks"]["knobs"], before["knobs"])
        out = self.quote(marks=fit["marks"]).run(book, bank=self.bank())
        self.assertEqual(marketmaker.capture_marks(book["EURUSD"]), before)
        self.assertEqual([w for w in out["warnings"] if "put back" in w], [])

    def test_a_quote_says_which_marks_it_stood_on(self):
        """A price made on this morning's fit and one made on last night's
        marks must never read the same."""
        fit = self.panel().run(self.book)
        handed = self.quote(marks=fit["marks"]).run(self.book, bank=self.bank())
        self.assertTrue(handed["marks"]["on_the_fit"])
        self.assertIn("handed", handed["marks"]["note"])
        plain = self.quote().run(self.book, bank=self.bank())
        self.assertFalse(plain["marks"]["on_the_fit"])
        self.assertIn("as they stand", plain["marks"]["note"])
        # And they are different prices, which is the point of saying which.
        self.assertNotEqual([r["our_bid"] for r in handed["sheet"]["rows"]],
                            [r["our_bid"] for r in plain["sheet"]["rows"]])

    def test_marks_fitted_on_another_pair_are_refused(self):
        """The browser holds the fit and the pair selector apart, and the two
        can be moved apart.  Quoting EURUSD off a USDJPY fit is a wrong answer
        that reads perfectly well."""
        fit = self.panel().run(self.book)
        marks = dict(fit["marks"], pair="USDJPY")
        with self.assertRaises(ValueError) as got:
            marketmaker.quote_panel_from_request(
                {"pair": "EURUSD", "request_text": self.ASKED, "marks": marks})
        self.assertIn("fitted on USDJPY", str(got.exception))

    def test_a_knob_the_curve_does_not_have_is_refused_rather_than_skipped(self):
        marks = marketmaker.capture_marks(self.book["EURUSD"])
        marks["knobs"]["corr_initial"] = 0.5
        with self.assertRaises(ValueError) as got:
            marketmaker.apply_marks(self.book["EURUSD"], marks)
        self.assertIn("corr_initial", str(got.exception))

    # -- the quote --------------------------------------------------------
    def test_the_width_a_rule_states_is_the_width_that_comes_out(self):
        """The bank is written in volatility points and the model in decimals.
        Reading a 0.28 rule as a decimal produced a 28 vol point market."""
        out = self.quote().run(self.book, bank=self.bank())
        atm = next(r for r in out["sheet"]["rows"] if r["instrument"] == "atm")
        self.assertAlmostEqual(atm["width"], 0.28)
        self.assertAlmostEqual(atm["our_ask"] - atm["our_bid"], 0.28, places=9)
        self.assertIn("ATM", atm["width_source"])

    def test_a_request_with_no_rule_and_no_fallback_gets_no_price(self):
        out = self.quote(fallback_spread="").run(self.book, bank=KnowledgeBank())
        for row in out["sheet"]["rows"]:
            self.assertIsNone(row["our_bid"])
            self.assertEqual(row["verdict"], "no width")
            self.assertTrue(any("no width rule" in w for w in row["warnings"]))

    def test_an_absolute_strike_without_a_feed_is_reported_not_priced_at_one(self):
        """Without a forward there is no moneyness, and pricing it at a forward
        of 1 would be a silent, badly wrong answer."""
        out = self.quote(request_text="6M 1.1000 call\n").run(self.book, bank=self.bank())
        row = out["sheet"]["rows"][0]
        self.assertEqual(row["verdict"], "not priced")
        self.assertIsNone(row["model"])
        self.assertTrue(any("forward feed" in w for w in row["warnings"]))

    def test_a_quote_needs_no_market_at_all(self):
        """The reason the request box exists.  A request does not arrive with a
        broker run attached to it, and the price is a property of the marks and
        the bank rather than of what somebody happened to show."""
        out = self.quote(text="").run(self.book, bank=self.bank())
        self.assertEqual(out["sheet"]["n_quotes"], 3)
        self.assertEqual(out["sheet"]["priced"], 3)
        self.assertEqual(out["sheet"]["matched"], 0)
        for row in out["sheet"]["rows"]:
            self.assertIsNone(row["market_mid"])
            self.assertEqual(row["verdict"], "quoted")

    def test_a_request_the_market_also_quoted_carries_their_market(self):
        """So "inside their market" survives the split.  The match is on the
        instrument, which is what makes two lines the same quote -- not on the
        text, which is written differently in the two boxes."""
        out = self.quote(text=self.TEXT).run(self.book, bank=self.bank())
        rows = {r["describe"]: r for r in out["sheet"]["rows"]}
        atm = rows["1M ATM"]
        self.assertAlmostEqual(atm["market_bid"], 6.05)
        self.assertAlmostEqual(atm["market_ask"], 6.35)
        self.assertIn(atm["position"], ("inside", "below", "above"))
        self.assertNotEqual(atm["verdict"], "quoted")
        self.assertEqual(out["sheet"]["matched"], 3)

    def test_a_risk_reversal_is_quoted_in_the_convention_it_was_asked_in(self):
        """Asked as 'JPY call over', answered as 'JPY call over'.  Quoting a
        skew back in the opposite sign is the §5 class of error, so the flip is
        applied once, at the row, and every number on the row turns with it."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["USDJPY"])
        bank = KnowledgeBank()
        bank.set_pair("USDJPY", [Rule("spread", 0.20, "rr")], ASOF.now)
        ours = marketmaker.quote_panel_from_request(
            {"pair": "USDJPY", "request_text": "3M 25d rr\n"}).run(book, bank=bank)
        theirs = marketmaker.quote_panel_from_request(
            {"pair": "USDJPY", "request_text": "3M 25d rr jpy call over\n"}).run(book, bank=bank)
        a, b = ours["sheet"]["rows"][0], theirs["sheet"]["rows"][0]
        self.assertEqual(a["sign"], 1.0)
        self.assertEqual(b["sign"], -1.0)
        self.assertAlmostEqual(b["model"], -a["model"], places=12)
        # A bid is still the low side of what we show, in whichever convention.
        self.assertLess(b["our_bid"], b["our_ask"])
        self.assertAlmostEqual(b["our_bid"], -a["our_ask"], places=12)
        self.assertIn("JPY call over", b["describe"])

    def test_the_panel_and_the_command_line_share_one_entry_point(self):
        """A panel set up in the browser and the same panel run from a shell
        must produce the same numbers, which is only guaranteed if there is one
        function -- and now two stages, so two of them."""
        import inspect
        from volkit import cli
        source = inspect.getsource(cli.cmd_mm)
        self.assertIn("panel_from_request", source)
        self.assertIn("quote_panel_from_request", source)


class TestCurveInvalidation(unittest.TestCase):
    def test_a_parameter_change_keeps_the_weight_profile_it_cannot_have_changed(self):
        """The intraday weight profile is a pure function of the pair, the
        clock and the horizon.  Dropping it on every backbone change cost about
        20ms against 2ms of integrals that genuinely had to go -- a 17x tax on
        every re-mark, which the market-maker fit pays thousands of times.
        """
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        atm = book["EURUSD"].atm
        atm.term_vol(1.0)
        cached = len(atm.weighting._cache)
        self.assertGreater(cached, 0)
        before = atm.term_vol(0.25)
        atm.set_params(initial_vol=atm.params.initial_vol * 1.10)
        self.assertEqual(len(atm.weighting._cache), cached)
        self.assertNotAlmostEqual(atm.term_vol(0.25), before)


class TestSmileParameterShifts(unittest.TestCase):
    def test_a_shift_moves_the_level_and_keeps_the_term_structure(self):
        """An overwrite replaces a parameter and flattens its term structure;
        re-marking a wing against a broker run should move its level and leave
        the shape alone."""
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        surface = book["EURUSD"]
        before = [surface.params_at(t)["rho25"] for t in (0.05, 0.25, 1.0)]
        self.assertEqual(surface.set_param_shifts({"rho25": 0.03}), [])
        after = [surface.params_at(t)["rho25"] for t in (0.05, 0.25, 1.0)]
        for a, b in zip(after, before):
            self.assertAlmostEqual(a - b, 0.03, places=12)
        surface.clear_param_shifts()
        self.assertEqual([surface.params_at(t)["rho25"] for t in (0.05, 0.25, 1.0)], before)

    def test_an_unknown_parameter_is_refused(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        problems = book["EURUSD"].set_param_shifts({"vega": 0.1})
        self.assertTrue(problems)
        self.assertIn("unknown smile parameter", problems[0])

    def test_a_clamped_shift_is_reported_rather_than_absorbed(self):
        book = Book.from_excel(WORKBOOK, ASOF).load_all(["EURUSD"])
        surface = book["EURUSD"]
        surface.set_param_shifts({"rho25": 1.4})
        self.assertLessEqual(abs(surface.params_at(0.25)["rho25"]), 0.999)
        self.assertTrue(any("clamped" in w for w in surface.shift_warnings()))


class TestKaceFeed(unittest.TestCase):
    """The kACE feed (`kace.py`): the XML_poster workbook, done from the book.

    The sheet is the specification, so the first test pins strings the sheet
    itself produced -- `XML_poster_DailyVol_v3.1_USDCNH_JL.xlsx`, recalculated
    and read back -- against a Feed built from the same inputs.  The rest pin
    the three things done differently on purpose (the horizon, `horDate`,
    the spread table naming the pillars) and the desk's conventions.
    """

    # The pillars and widths are a tab of the workbook now, not a file beside it.
    SPREADS = Path(__file__).resolve().parents[1] / "files" / "vol_marks.xlsx"

    # The sheet's nine pillars (USDCNHData!K3:S11), as of a 2026-01-22 valuation.
    SHEET_PILLARS = [
        ("O/N", date(2026, 1, 23), 1.0, -0.25, -0.45, 0.125, 0.39375),
        ("1W", date(2026, 1, 29), 0.8, -0.225, -0.405, 0.125, 0.39375),
        ("2W", date(2026, 2, 5), 0.6, -0.2125, -0.3825, 0.125, 0.39375),
        ("1M", date(2026, 2, 24), 0.4, -0.2, -0.36, 0.125, 0.39375),
        ("2M", date(2026, 3, 24), 0.3, -0.175, -0.315, 0.1275, 0.401625),
        ("3M", date(2026, 4, 23), 0.3, -0.175, -0.315, 0.15, 0.4725),
        ("6M", date(2026, 7, 23), 0.2, -0.175, -0.35, 0.1875, 0.590625),
        ("9M", date(2026, 10, 22), 0.2, -0.175, -0.3675, 0.2125, 0.669375),
        ("1Y", date(2027, 1, 22), 0.2, -0.156, -0.37, 0.25, 0.7875),
    ]
    # Rows of the sheet's daily series (USDCNHData!A:B), and what its formulas
    # wrote for them.  Row 1 is before every pillar (the ISERROR fallback to
    # the O/N spread); 28 Jan is still on O/N's 1.0; 29 Jan is the 1W expiry
    # and takes 0.8; 5 Feb is the 2W expiry; 30 Jan 2027 is the last row.
    SHEET_DAYS = [
        (date(2026, 1, 23), 2.25820980020178, "0.0175820980020178/0.0275820980020178"),
        (date(2026, 1, 24), 2.38916548538905, "0.0188916548538905/0.0288916548538905"),
        (date(2026, 1, 28), 1.80166787406075, "0.0130166787406075/0.0230166787406075"),
        (date(2026, 1, 29), 2.29288067023934, "0.0189288067023934/0.0269288067023934"),
        (date(2026, 1, 30), 2.25351386681335, "0.0185351386681335/0.0265351386681335"),
        (date(2026, 2, 5), 2.18646548219792, "0.0188646548219792/0.0248646548219792"),
        (date(2026, 2, 6), 2.19092359623445, "0.0189092359623445/0.0249092359623445"),
        (date(2027, 1, 30), 3.77727687403344, "0.0367727687403344/0.0387727687403344"),
    ]

    def _sheet_feed(self):
        from volkit import kace
        daily = {d: v for d, v, _ in self.SHEET_DAYS}
        # Every pillar has to be a row of the series; the sheet's were.
        for tenor, expiry, spread, rr25, rr10, fly25, fly10 in self.SHEET_PILLARS:
            daily.setdefault(expiry, 3.7)
        daily[date(2027, 1, 22)] = 3.76476480984279      # S41 in the sheet
        pillars = [kace.Pillar(tenor=t, expiry=e, spread=sp, atm=daily[e], rr25=a, rr10=b,
                               fly25=c, fly10=d, wings="marks")
                   for t, e, sp, a, b, c, d in self.SHEET_PILLARS]
        return kace.Feed(pair="USDCNH", hor_date=date(2026, 1, 22), cut="NY",
                         source="marks", daily=daily, pillars=pillars)

    @staticmethod
    def _nodes(text):
        import xml.etree.ElementTree as ET
        root = ET.fromstring(text.encode("utf-8"))
        return root, {n.get("name"): {f.get("name"): f.get("value") for f in n.findall("field")}
                      for n in root.iter("node")}

    def test_the_message_is_what_the_sheet_wrote(self):
        """Strings the workbook's own formulas produced, reproduced exactly."""
        feed = self._sheet_feed()
        text = feed.xml("feeuser", "password1", timestamp=datetime(2026, 1, 22, 9, tzinfo=UTC))
        root, nodes = self._nodes(text)
        by_day = {n["Maturity"]: n for k, n in nodes.items() if not k.startswith("S")}
        for day, _, want in self.SHEET_DAYS:
            label = f"{day.day:02d} {day:%b} {day.year}"
            self.assertEqual(by_day[label]["Volity"], want, label)
            self.assertEqual(by_day[label]["VolType"], "ATM")
        # The pillar block: five nodes each, in the sheet's order and naming.
        self.assertEqual(nodes["S1"], {"RateType": "Volatility", "Currency": "USD",
                                       "CtrCcy": "CNH", "Maturity": "23 Jan 2026",
                                       "VolType": "ATM",
                                       "Volity": "0.0175820980020178/0.0275820980020178"})
        self.assertEqual(nodes["S2"]["Volity"], "-0.0025")
        self.assertEqual((nodes["S2"]["PctDelta"], nodes["S2"]["VolType"]), ("0.25", "RR"))
        self.assertEqual(nodes["S3"]["Volity"], "-0.0045")
        self.assertEqual((nodes["S3"]["PctDelta"], nodes["S3"]["VolType"]), ("0.10", "RR"))
        self.assertEqual(nodes["S4"]["Volity"], "0.00125")
        self.assertEqual((nodes["S4"]["PctDelta"], nodes["S4"]["VolType"]), ("0.25", "S"))
        self.assertEqual(nodes["S5"]["Volity"], "0.0039375")
        self.assertEqual(nodes["S41"]["Volity"], "0.0366476480984279/0.0386476480984279")
        self.assertEqual(nodes["S42"]["Volity"], "-0.00156")
        self.assertEqual(nodes["S45"]["Volity"], "0.007875")
        self.assertEqual(nodes["S45"]["Maturity"], "22 Jan 2027")
        self.assertEqual(len(nodes), len(feed.daily) + 45)
        # The envelope the poster page takes.
        header = {e.tag: (e.text or "").strip() for e in root.find("header")}
        self.assertEqual(header["username"], "feeuser")
        self.assertEqual(header["password"], "password1")
        self.assertEqual(header["timestamp"], "2026-01-22T09:00:00+00:00")
        action = root.find("body").find("action")
        self.assertEqual(action.get("function"), "RATE_FEED")
        opts = {o.get("name"): o.get("value") for o in action}
        self.assertEqual(opts["scenario"], "Xyz")
        self.assertEqual(opts["horDate"], "22 Jan 2026")
        self.assertNotIn("clearRate", opts)

    def test_the_day_takes_the_spread_of_the_last_pillar_on_or_before_it(self):
        """The sheet's approximate VLOOKUP, and its ISERROR fallback, as a rule."""
        from volkit import kace
        pillars = self._sheet_feed().pillars
        self.assertEqual(kace.spread_for(date(2026, 1, 1), pillars), 1.0)    # before O/N
        self.assertEqual(kace.spread_for(date(2026, 1, 23), pillars), 1.0)   # the O/N expiry
        self.assertEqual(kace.spread_for(date(2026, 1, 28), pillars), 1.0)   # still O/N's
        self.assertEqual(kace.spread_for(date(2026, 1, 29), pillars), 0.8)   # the 1W expiry
        self.assertEqual(kace.spread_for(date(2026, 2, 4), pillars), 0.8)    # 1W's until 2W
        self.assertEqual(kace.spread_for(date(2026, 2, 5), pillars), 0.6)
        self.assertEqual(kace.spread_for(date(2027, 6, 1), pillars), 0.2)    # past 1Y: 1Y's

    def test_decimals_are_plain_and_dates_are_english(self):
        from volkit import kace
        self.assertEqual(kace._decimal(0.0175820980020178), "0.0175820980020178")
        self.assertEqual(kace._decimal(5e-05), "0.00005")
        self.assertEqual(kace._decimal(-0.0025), "-0.0025")
        self.assertEqual(kace._date(date(2026, 2, 5)), "05 Feb 2026")

    def test_a_non_positive_bid_is_refused(self):
        from volkit import kace
        feed = self._sheet_feed()
        feed.daily[date(2026, 1, 24)] = 0.4          # spread 1.0 straddles zero
        with self.assertRaises(kace.KaceError) as ctx:
            feed.xml("u", "p")
        self.assertIn("24 Jan 2026", str(ctx.exception))

    def test_no_username_no_message(self):
        from volkit import kace
        with self.assertRaises(kace.KaceError) as ctx:
            self._sheet_feed().xml("", "")
        self.assertIn(kace.ENV_USER, str(ctx.exception))
        with self.assertRaises(kace.KaceError):
            kace.clear_message("USDCNH", date(2026, 1, 22), "", "")

    def test_the_clear_message(self):
        from volkit import kace
        text = kace.clear_message("USDCNH", date(2026, 1, 22), "u", "p",
                                  timestamp=datetime(2026, 1, 22, tzinfo=UTC))
        root, nodes = self._nodes(text)
        opts = {o.get("name"): o.get("value") for o in root.find("body").find("action")}
        self.assertEqual(opts["clearRate"], "true")
        self.assertEqual(opts["horDate"], "22 Jan 2026")
        self.assertEqual(list(nodes), ["USDCNH"])
        self.assertEqual(nodes["USDCNH"], {"RateType": "Volatility", "Currency": "USD",
                                           "CtrCcy": "CNH"})

    # -- the spread table -------------------------------------------------
    def test_the_shipped_table_is_the_sheets_column_l(self):
        from volkit import kace
        table = kace.SpreadTable.load(self.SPREADS)
        self.assertEqual(table.for_pair("usdcnh"),
                         {"O/N": 1.0, "1W": 0.8, "2W": 0.6, "1M": 0.4, "2M": 0.3, "3M": 0.3,
                          "6M": 0.2, "9M": 0.2, "1Y": 0.2})
        with self.assertRaises(kace.KaceError) as ctx:
            table.for_pair("EURUSD")
        self.assertIn("EURUSD", str(ctx.exception))

    def test_a_bad_table_is_refused_by_row(self):
        """The table is a tab now, so a bad cell is reported by the row Excel
        shows -- which is the number somebody goes and looks at."""
        import tempfile
        import openpyxl
        from volkit import kace

        def workbook(path, rows):
            wb = openpyxl.Workbook()
            wb.active.title = kace.SPREADS_SHEET
            for row in rows:
                wb.active.append(row)
            wb.save(path)
            return path

        with tempfile.TemporaryDirectory() as tmp:
            p = workbook(Path(tmp) / "marks.xlsx", [
                ["pair", "tenor", "spread"],
                ["USDCNH", "on", 1],
                ["USDCNH", "1W", "wide"],
                ["USDCNH", "1W", 0.8],
                ["USDCNH", "1W", None],
                ["USDCNH", "7Q", 0.1],
            ])
            with self.assertRaises(kace.KaceError) as ctx:
                kace.SpreadTable.load(p)
            msg = str(ctx.exception)
            self.assertIn(kace.SPREADS_SHEET, msg)
            self.assertIn("row 3", msg)           # not a number
            self.assertIn("row 5", msg)           # no spread
            self.assertIn("row 6", msg)           # not a tenor

            # Notes above the header, a '#' row anywhere, and a heading or a
            # cell in whatever case somebody typed it.
            good = workbook(Path(tmp) / "good.xlsx", [
                ["# the desk's pillars"],
                ["Pair", "Tenor", "Spread"],
                ["USDCNH", " on ", 1],
                ["# and the rest"],
                ["usdcnh", "1w", 0.8],
            ])
            self.assertEqual(kace.SpreadTable.load(good).for_pair("USDCNH"),
                             {"O/N": 1.0, "1W": 0.8})

            # A workbook without the tab, and no workbook at all, are both
            # said by name rather than answered with an empty table.
            bare = Path(tmp) / "bare.xlsx"
            openpyxl.Workbook().save(bare)
            with self.assertRaises(kace.KaceError) as ctx:
                kace.SpreadTable.load(bare)
            self.assertIn(kace.SPREADS_SHEET, str(ctx.exception))
            with self.assertRaises(kace.KaceError) as ctx:
                kace.SpreadTable.load(Path(tmp) / "missing.xlsx")
            self.assertIn("missing.xlsx", str(ctx.exception))

    # -- off the book ---------------------------------------------------------
    @classmethod
    def _book(cls):
        if not hasattr(cls, "_cached_book"):
            cls._cached_book = Book.from_excel(str(WORKBOOK), ASOF).load_all(["USDCNH"])
        return cls._cached_book

    def _table(self, rows):
        from volkit import kace
        t = kace.SpreadTable(path="test")
        t.rows = {"USDCNH": dict(rows)}
        return t

    def test_the_series_reaches_the_last_pillar_and_hordate_is_the_books(self):
        """The sheet's two quiet failures: #N/A at 1Y, and TODAY() in horDate."""
        from volkit import kace
        feed = kace.build(self._book(), "USDCNH", kace.SpreadTable.load(self.SPREADS))
        self.assertEqual(feed.hor_date, date(2024, 2, 28))       # ASOF, not date.today()
        self.assertNotEqual(feed.hor_date, date.today())
        for p in feed.pillars:
            self.assertIn(p.expiry, feed.daily, p.tenor)
            self.assertAlmostEqual(p.atm, feed.daily[p.expiry])
        last = max(p.expiry for p in feed.pillars)
        self.assertEqual(last, date(2025, 2, 27))                # the 1Y expiry
        self.assertGreaterEqual(max(feed.daily), last)
        self.assertLessEqual((max(feed.daily) - last).days, kace.MARGIN_DAYS + 1)
        # Valued at 12:00 UTC, before the 14:00 NY cut: today's own bucket
        # is not a cumulative vol, and is left out with a note.
        self.assertNotIn(date(2024, 2, 28), feed.daily)
        self.assertEqual(min(feed.daily), date(2024, 2, 29))
        self.assertTrue(any("28 Feb 2024" in n for n in feed.notes))
        self.assertEqual(feed.node_count(), len(feed.daily) + 45)

    def test_overnight_is_the_next_business_day_and_borrows_the_shortest_wings(self):
        from volkit import kace
        feed = kace.build(self._book(), "USDCNH", kace.SpreadTable.load(self.SPREADS))
        on, w1 = feed.pillars[0], feed.pillars[1]
        self.assertEqual(on.tenor, "O/N")
        self.assertEqual(on.expiry, date(2024, 2, 29))
        self.assertEqual(w1.tenor, "1W")
        self.assertEqual(w1.expiry, date(2024, 3, 6))
        self.assertEqual((on.rr25, on.rr10, on.fly25, on.fly10),
                         (w1.rr25, w1.rr10, w1.fly25, w1.fly10))
        self.assertEqual(on.wings, "marks at 1W")
        self.assertEqual(w1.wings, "marks")
        self.assertTrue(any("O/N" in n and "1W" in n for n in feed.notes))
        # The desk's convention: RR is the USD call over the put, in vol
        # points, straight off the marks sheet; S is the strangle mark.
        self.assertAlmostEqual(w1.rr25, 0.385)
        self.assertAlmostEqual(w1.fly25, 0.1825)
        self.assertAlmostEqual(feed.pillars[-1].rr10, 2.645)

    def test_fitted_wings_come_off_the_surface_near_the_marks(self):
        from volkit import kace
        feed = kace.build(self._book(), "USDCNH", kace.SpreadTable.load(self.SPREADS),
                          source="fitted")
        marks = kace.build(self._book(), "USDCNH", kace.SpreadTable.load(self.SPREADS))
        for f, m in zip(feed.pillars, marks.pillars):
            self.assertEqual(f.wings, "fitted")
            for v in (f.rr25, f.rr10, f.fly25, f.fly10):
                self.assertTrue(math.isfinite(v))
            if f.tenor != "O/N":
                self.assertLess(abs(f.rr25 - m.rr25), 0.15, f.tenor)
                self.assertLess(abs(f.fly25 - m.fly25), 0.1, f.tenor)
        self.assertEqual(feed.notes, [n for n in feed.notes if "O/N" not in n])

    def test_a_pillar_with_no_mark_is_refused_by_name(self):
        from volkit import kace
        with self.assertRaises(kace.KaceError) as ctx:
            kace.build(self._book(), "USDCNH", self._table([("1W", 0.8), ("3W", 0.5)]))
        self.assertIn("3W", str(ctx.exception))
        with self.assertRaises(kace.KaceError):
            kace.build(self._book(), "USDCNH", self._table([("O/N", 1.0)]))
        with self.assertRaises(kace.KaceError):
            kace.build(self._book(), "USDCNH", self._table([("1W", 0.8)]), source="murex")

    def test_the_web_service_and_the_download(self):
        from volkit import kace
        from volkit.webapp import BookService
        service = BookService(str(WORKBOOK), ASOF, kace_spreads_path=str(self.SPREADS),
                              kace_user="feeuser", kace_password="pw")
        state = service.state()["kace"]
        self.assertEqual(state["pairs"], ["USDCNH"])
        self.assertTrue(state["credentials"])
        self.assertIsNone(state["error"])
        out = service.kace({"pair": "USDCNH"})
        self.assertEqual(out["nodes"], out["xml"].count("<node "))
        self.assertEqual(out["hor_date"], "2024-02-28")
        self.assertEqual([p["tenor"] for p in out["pillars"]][:2], ["O/N", "1W"])
        self.assertEqual(out["scenario"], "Xyz")                 # the start-up default
        self.assertIn('<option name="scenario" value="Xyz"/>', out["xml"])
        name, text = service.export_kace({"pair": "USDCNH"})
        self.assertEqual(name, "USDCNH_kace_vols_Xyz_2024-02-28.xml")
        self.assertEqual(text, out["xml"])
        name, text = service.export_kace({"pair": "USDCNH", "clear": "1"})
        self.assertEqual(name, "USDCNH_kace_clear_Xyz_2024-02-28.xml")
        self.assertIn('name="clearRate"', text)
        # The scenario is the page's box: it goes into the message, the
        # file name, and the clear message alike; blank is refused, and a
        # character XML cannot carry in an attribute is escaped.
        out = service.kace({"pair": "USDCNH", "scenario": " UAT-2 "})
        self.assertEqual(out["scenario"], "UAT-2")
        self.assertIn('<option name="scenario" value="UAT-2"/>', out["xml"])
        name, _ = service.export_kace({"pair": "USDCNH", "scenario": "UAT-2"})
        self.assertEqual(name, "USDCNH_kace_vols_UAT_2_2024-02-28.xml")
        _, text = service.export_kace({"pair": "USDCNH", "clear": "1", "scenario": "UAT-2"})
        self.assertIn('<option name="scenario" value="UAT-2"/>', text)
        self.assertIn('<option name="clearRate" value="true"/>', text)
        with self.assertRaises(kace.KaceError) as ctx:
            service.kace({"pair": "USDCNH", "scenario": "  "})
        self.assertIn("blank", str(ctx.exception))
        out = service.kace({"pair": "USDCNH", "scenario": 'A&B"c'})
        self.assertIn('value="A&amp;B&quot;c"', out["xml"])
        self._nodes(out["xml"])                                   # still parses
        # Without a username the table is still shown; the message is not.
        import os
        had = os.environ.pop(kace.ENV_USER, None)
        if had is not None:
            self.addCleanup(os.environ.__setitem__, kace.ENV_USER, had)
        bare = BookService(str(WORKBOOK), ASOF, kace_spreads_path=str(self.SPREADS))
        self.assertFalse(bare.state()["kace"]["credentials"])
        out = bare.kace({"pair": "USDCNH"})
        self.assertIsNone(out["xml"])
        self.assertIn(kace.ENV_USER, out["xml_error"])
        self.assertEqual(len(out["pillars"]), 9)
        with self.assertRaises(kace.KaceError):
            bare.export_kace({"pair": "USDCNH"})
        # A table that is wrong is the card's error, not a crash at startup.
        broken = BookService(str(WORKBOOK), ASOF, kace_spreads_path="/nowhere/kace.csv")
        self.assertIn("kace.csv", broken.state()["kace"]["error"])
        with self.assertRaises(kace.KaceError):
            broken.kace({"pair": "USDCNH"})

    def test_the_routes_and_the_command_belong_to_the_marking_screen(self):
        from volkit import screens
        from volkit.cli import build_parser
        owner = {r: sc.name for sc in screens.SCREENS for r in sc.routes}
        self.assertEqual(owner["/api/kace"], "marking")
        self.assertEqual(owner["/api/export/kace"], "marking")
        self.assertEqual(screens.command_screen("kace"), "marking")
        args = build_parser().parse_args(["kace", "USDCNH", "--clear", "--kace-user", "u"])
        self.assertTrue(args.clear)
        self.assertEqual(args.kace_user, "u")
        self.assertEqual(args.source, "marks")
        args = build_parser().parse_args(["serve", "--kace-spreads", "s.csv",
                                          "--kace-scenario", "Prod"])
        self.assertEqual((args.kace_spreads, args.kace_scenario), ("s.csv", "Prod"))
        self.assertIn("kace", (Path(__file__).resolve().parents[1] / "volkit" / "web"
                               / "index.html").read_text(encoding="utf-8"))

    # -- posting ----------------------------------------------------------
    REPLY_OK = ("<?xml version='1.0' encoding='UTF-8'?>\n"
                '<gfi_message version="2.0">\n  <header>\n    <transactionId>1234567890</transactionId>\n'
                "    <timestamp>2023-06-12T16:15:26+08:00</timestamp>\n"
                "    <processingTime>0.298</processingTime>\n  </header>\n  <body>\n"
                '    <response name="action1" function="RATE_FEED" version="1.0" />\n'
                "  </body>\n</gfi_message>\n")

    def test_the_reply_is_read_as_the_poster_page_shows_it(self):
        """The one shape known to mean success, and everything else refused."""
        from volkit import kace
        ok, took, msg = kace.read_reply(self.REPLY_OK)
        self.assertTrue(ok)
        self.assertAlmostEqual(took, 0.298)
        self.assertIn("0.298", msg)
        ok, _, msg = kace.read_reply("<html><body>Please log in</body></html>")
        self.assertFalse(ok)
        self.assertIn("not a gfi_message", msg)
        ok, _, msg = kace.read_reply("502 Bad Gateway\nnginx")
        self.assertFalse(ok)
        self.assertIn("not XML", msg)
        self.assertIn("502 Bad Gateway", msg)
        ok, _, msg = kace.read_reply('<gfi_message><header/><body><error>unknown scenario '
                                     'Prod</error></body></gfi_message>')
        self.assertFalse(ok)
        self.assertIn("unknown scenario Prod", msg)
        ok, _, msg = kace.read_reply('<gfi_message><header/><body><response status="ERROR: bad"/>'
                                     '</body></gfi_message>')
        self.assertFalse(ok)
        ok, _, msg = kace.read_reply('<gfi_message><header><processingTime>0.1</processingTime>'
                                     '</header><body/></gfi_message>')
        self.assertFalse(ok)
        self.assertIn("no <response>", msg)

    def test_the_body_is_the_forms_the_vba_sent(self):
        from volkit import kace
        body = kace.form_body('<a b="1"> x&y</a>')
        self.assertTrue(body.startswith(b"xml="))
        import urllib.parse
        self.assertEqual(urllib.parse.parse_qs(body.decode())["xml"], ['<a b="1"> x&y</a>'])
        self.assertEqual(len(kace.message_hash("m")), 16)
        self.assertNotEqual(kace.message_hash("m"), kace.message_hash("n"))

    def test_post_message_through_an_injected_network(self):
        from volkit import kace
        seen = {}

        def fake(url, body, headers, *, timeout, ca, insecure):
            seen.update(url=url, body=body, headers=headers, timeout=timeout, ca=ca,
                        insecure=insecure)
            return 200, self.REPLY_OK.encode()

        r = kace.post_message("<gfi_message/>", "https://kace:8500/pricing", opener=fake,
                              ca="/desk/ca.pem")
        self.assertTrue(r.ok)
        self.assertEqual(r.status, 200)
        self.assertAlmostEqual(r.processing_time, 0.298)
        self.assertEqual(seen["headers"]["Content-Type"], kace.FORM_CONTENT_TYPE)
        self.assertEqual(seen["body"], b"xml=%3Cgfi_message%2F%3E")
        self.assertEqual(seen["headers"]["Content-Length"], str(len(seen["body"])))
        self.assertEqual((seen["ca"], seen["insecure"], seen["timeout"]),
                         ("/desk/ca.pem", False, kace.POST_TIMEOUT))
        self.assertEqual(r.bytes_sent, len(seen["body"]))
        # A status that is not 200 is the failure, with the first line of the body.
        r = kace.post_message("<m/>", "http://kace:8500/x",
                              opener=lambda *a, **k: (500, b"<html>Internal Server Error</html>"))
        self.assertFalse(r.ok)
        self.assertEqual(r.status, 500)
        self.assertIn("HTTP 500", r.message)
        self.assertIn("Internal Server Error", r.message)
        # No address, or not an http one, is refused before anything is sent.
        with self.assertRaises(kace.KacePostError) as ctx:
            kace.post_message("<m/>", "", opener=fake)
        self.assertIn(kace.ENV_URL, str(ctx.exception))
        with self.assertRaises(kace.KacePostError):
            kace.post_message("<m/>", "ftp://kace/x", opener=fake)

    def test_the_post_log_and_a_refused_post(self):
        import tempfile
        from volkit import kace
        with tempfile.TemporaryDirectory() as tmp:
            log = kace.PostLog.at(Path(tmp) / "posts.jsonl")
            when = datetime(2024, 2, 28, 9, 12, tzinfo=UTC)
            # A dry run says what would go and writes nothing.
            e = kace.post_feed("<m/>", pair="USDCNH", scenario="Xyz", clear=False,
                               hor_date=date(2024, 2, 28), nodes=3, url="http://kace/x",
                               log=log, when=when, dry_run=True)
            self.assertIsNone(e["ok"])
            self.assertIn("dry run", e["message"])
            self.assertIn("http://kace/x", e["message"])
            self.assertEqual(log.entries(), [])
            # A post that cannot reach the server is recorded, as a failure.
            def down(*a, **k):
                raise kace.KacePostError("could not reach http://kace/x: refused")
            e = kace.post_feed("<m/>", pair="USDCNH", scenario="Xyz", clear=True,
                               hor_date=date(2024, 2, 28), nodes=1, url="http://kace/x",
                               log=log, when=when, opener=down)
            self.assertFalse(e["ok"])
            self.assertIn("refused", e["message"])
            self.assertEqual(e["logged"], log.path)
            # And one that lands.
            e = kace.post_feed("<m/>", pair="usdcnh", scenario="Live", clear=False,
                               hor_date=date(2024, 2, 28), nodes=413, url="http://kace/x",
                               log=log, when=when,
                               opener=lambda *a, **k: (200, self.REPLY_OK.encode()))
            self.assertTrue(e["ok"])
            self.assertEqual(e["pair"], "USDCNH")
            rows = log.entries()
            self.assertEqual([r["ok"] for r in rows], [False, True])
            self.assertEqual(rows[1]["scenario"], "Live")
            self.assertEqual(rows[1]["hash"], kace.message_hash("<m/>"))
            self.assertEqual(rows[1]["at"], "2024-02-28T09:12:00+00:00")
            # A line that will not parse is skipped, not fatal; the pair filter works.
            with Path(log.path).open("a", encoding="utf-8") as fh:
                fh.write("{not json\n")
            self.assertEqual(len(log.entries()), 2)
            self.assertEqual(log.entries(pair="EURUSD"), [])
            self.assertEqual(len(log.entries(pair="usdcnh", limit=1)), 1)

    def test_the_post_button_sends_what_the_table_shows_and_nothing_else(self):
        import tempfile
        from volkit import kace
        from volkit.webapp import BookService
        with tempfile.TemporaryDirectory() as tmp:
            log_path = Path(tmp) / "kace_posts.jsonl"
            service = BookService(str(WORKBOOK), ASOF, kace_spreads_path=str(self.SPREADS),
                                  kace_user="feeuser", kace_password="pw",
                                  kace_url="https://kace:8500/pricing", kace_insecure=True,
                                  kace_log_path=str(log_path))
            sent = []

            def fake(url, body, headers, *, timeout, ca, insecure):
                sent.append((url, body, headers, insecure))
                return 200, self.REPLY_OK.encode()

            service.kace_opener = fake
            state = service.state()["kace"]
            self.assertEqual(state["url"], "https://kace:8500/pricing")
            self.assertTrue(state["insecure"])
            self.assertEqual(state["posts"], [])
            shown = service.kace({"pair": "USDCNH", "scenario": "UAT"})
            r = service.kace_post({"pair": "USDCNH", "scenario": "UAT"})
            self.assertTrue(r["ok"])
            self.assertAlmostEqual(r["processing_time"], 0.298)
            self.assertEqual(r["nodes"], 413)
            self.assertEqual(r["scenario"], "UAT")
            self.assertEqual(r["hor_date"], "2024-02-28")
            self.assertEqual(r["at"], "2024-02-28T12:00:00+00:00")   # the book's clock
            import urllib.parse
            url, body, headers, insecure = sent[0]
            self.assertEqual(url, "https://kace:8500/pricing")
            self.assertTrue(insecure)
            self.assertEqual(headers["Content-Type"], kace.FORM_CONTENT_TYPE)
            posted = urllib.parse.parse_qs(body.decode())["xml"][0]
            self.assertEqual(posted, shown["xml"])               # exactly the table's message
            self.assertEqual(r["hash"], kace.message_hash(shown["xml"]))
            self.assertEqual(len(r["posts"]), 1)
            self.assertEqual(log_path.read_text(encoding="utf-8").count("\n"), 1)
            # The clear goes the same way, and says so in the record.
            r = service.kace_post({"pair": "USDCNH", "scenario": "UAT", "clear": "1"})
            self.assertTrue(r["clear"])
            self.assertIn('name="clearRate"', urllib.parse.parse_qs(sent[1][1].decode())["xml"][0])
            # A dry run reaches no network and writes no line.
            r = service.kace_post({"pair": "USDCNH", "dry_run": True})
            self.assertTrue(r["dry_run"])
            self.assertEqual(len(sent), 2)
            self.assertEqual(len(service.kace_log.entries()), 2)
            self.assertEqual(service.state()["kace"]["posts"][-1]["clear"], True)
            # A server that answers with something else is a failure the page shows.
            service.kace_opener = lambda *a, **k: (200, b"<html>Session expired</html>")
            r = service.kace_post({"pair": "USDCNH"})
            self.assertFalse(r["ok"])
            self.assertIn("not a gfi_message", r["message"])
            self.assertIn("Session expired", r["reply"])
            # No address: refused by name, and recorded as refused.
            bare = BookService(str(WORKBOOK), ASOF, kace_spreads_path=str(self.SPREADS),
                               kace_user="u", kace_log_path=str(Path(tmp) / "bare.jsonl"))
            import os
            os.environ.pop(kace.ENV_URL, None)
            bare.kace_url = ""
            self.assertFalse(bare.state()["kace"]["url"])
            r = bare.kace_post({"pair": "USDCNH"})
            self.assertFalse(r["ok"])
            self.assertIn(kace.ENV_URL, r["message"])
            # Without a username there is no message to post at all.
            nouser = BookService(str(WORKBOOK), ASOF, kace_spreads_path=str(self.SPREADS),
                                 kace_url="http://kace/x")
            nouser.kace_user = ""
            with self.assertRaises(kace.KaceError):
                nouser.kace_post({"pair": "USDCNH"})

    def test_the_post_route_and_options(self):
        from volkit import screens
        from volkit.cli import build_parser
        owner = {r: sc.name for sc in screens.SCREENS for r in sc.routes}
        self.assertEqual(owner["/api/kace/post"], "marking")
        args = build_parser().parse_args(["kace", "USDCNH", "--post", "--dry-run",
                                          "--kace-url", "https://k:8500/p", "--kace-insecure"])
        self.assertTrue(args.post and args.dry_run and args.kace_insecure)
        self.assertEqual(args.kace_url, "https://k:8500/p")
        args = build_parser().parse_args(["serve", "--kace-url", "http://k/x", "--kace-ca",
                                          "ca.pem", "--kace-log", "p.jsonl"])
        self.assertEqual((args.kace_url, args.kace_ca, args.kace_log),
                         ("http://k/x", "ca.pem", "p.jsonl"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
